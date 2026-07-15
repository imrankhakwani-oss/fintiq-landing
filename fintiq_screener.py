"""
FINTIQ — Intelligent Trading Screener  v3.0
Strategies 1 (Quality Value) + 2 (Catalyst Alerts) + 3 (Pairs Trading)
+ Intrinsic Value Calculator + Trading Journal + Professional UI
Global Markets | Custom Pairs | Powered by yfinance + FMP
Author: Built for Imran Khakwani | July 2026
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import yfinance as yf
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
import sqlite3, os, math, io, json
import warnings
warnings.filterwarnings("ignore")

# ── Supabase auth ──────────────────────────────────────────────
try:
    from supabase import create_client, Client as SupabaseClient
    _SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
    _SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
    _SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
    _sb: SupabaseClient = create_client(_SUPABASE_URL, _SUPABASE_KEY) if _SUPABASE_URL else None
    # Admin client uses service role key — bypasses RLS for profile reads/writes
    _sb_admin: SupabaseClient = create_client(_SUPABASE_URL, _SUPABASE_SERVICE_KEY) if (_SUPABASE_URL and _SUPABASE_SERVICE_KEY) else _sb
except Exception:
    _sb = None
    _sb_admin = None

# ── Stripe ─────────────────────────────────────────────────────
try:
    import stripe as _stripe
    _STRIPE_SECRET  = os.environ.get("STRIPE_SECRET_KEY", "")
    _STRIPE_PUB     = os.environ.get("STRIPE_PUBLISHABLE_KEY", "")
    _PRICE_MONTHLY  = "price_1Tt8fXFSQ5tKKNExQf7T2t92"
    _PRICE_ANNUAL   = "price_1Tt8fXFSQ5tKKNExXuntkPsI"
    _APP_URL        = "https://fintiq.uk"
    if _STRIPE_SECRET:
        _stripe.api_key = _STRIPE_SECRET
except Exception:
    _stripe = None  # type: ignore

# ─────────────────────────────────────────────────────────────
# WATCHLIST — JSON persistence helpers
# ─────────────────────────────────────────────────────────────
_WL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fintiq_watchlist.json")

def _wl_load() -> dict:
    """Load watchlist from JSON file. Returns dict keyed by ticker."""
    if os.path.exists(_WL_FILE):
        try:
            with open(_WL_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _wl_save(wl: dict):
    """Persist watchlist dict to JSON file."""
    try:
        with open(_WL_FILE, "w") as f:
            json.dump(wl, f, indent=2)
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────
# PAIRS WATCHLIST — JSON persistence helpers
# ─────────────────────────────────────────────────────────────
_PWL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fintiq_pairs_watchlist.json")

def _pwl_load() -> list:
    """Load pairs watchlist from JSON. Returns list of dicts."""
    if os.path.exists(_PWL_FILE):
        try:
            with open(_PWL_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def _pwl_save(pairs: list):
    """Persist pairs watchlist list to JSON file."""
    try:
        with open(_PWL_FILE, "w") as f:
            json.dump(pairs, f, indent=2)
    except Exception:
        pass

# (watchlist session state seeded after st.set_page_config below)
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────

FMP_KEY  = "c3gRy6dPp8uETaNIYoFJj83J7hm998bB"
FMP_BASE = "https://financialmodelingprep.com/api"

# ── Currency by exchange ──────────────────────────────────────
EXCHANGE_CURRENCY = {
    "LSE": "£", "AIM": "£",
    "NYSE": "$", "NASDAQ": "$",
    "XETRA": "€", "EURONEXT": "€",
    "TSX": "C$", "ASX": "A$",
    "NSE": "₹", "HKEX": "HK$",
}

# ── Stock Universe ────────────────────────────────────────────
STOCK_UNIVERSE = {
    "LSE": [
        # FTSE 100 — Large Cap
        "AZN.L","SHEL.L","HSBA.L","ULVR.L","BP.L","RIO.L","GSK.L","DGE.L",
        "LSEG.L","BATS.L","NG.L","REL.L","PRU.L","LLOY.L","BARC.L","NWG.L",
        "ABF.L","IMB.L","BHP.L","ANTO.L","STAN.L","WPP.L","JD.L","SMT.L",
        "CNA.L","HLMA.L","BA.L","IAG.L","TSCO.L","SBRY.L","VOD.L","BT-A.L",
        "TW.L","PSN.L","RKT.L","EXPN.L","CRH.L","AAL.L","AUTO.L","FRES.L",
        "ITV.L","SGE.L","WEIR.L","SSE.L","MKS.L","OCDO.L","LAND.L","AV.L",
        "MNG.L","ADM.L","PHNX.L","III.L","DCC.L","CPI.L","ENT.L","SPX.L",
        "MNDI.L","SKG.L","EZJ.L","DPLM.L","RTO.L","SMIN.L","INF.L",
        "RMV.L","RCP.L","GRG.L","HWDN.L","KGF.L","BME.L","MRO.L","SDR.L",
        "BNZL.L","FUTR.L","DLN.L","TRIG.L","UU.L","PNN.L","SVT.L",
        "CCH.L","CRDA.L","GLEN.L","HIK.L","ITRK.L","JMAT.L","LGEN.L",
        "NXT.L","PSON.L","RS1.L","SGRO.L","SN.L","STJ.L",
        "ATG.L","BBY.L","BKG.L","BRBY.L","BVIC.L","DNLM.L","DOM.L",
        "EMG.L","FOUR.L","GRND.L","HILS.L","HLN.L","HMSO.L","HOME.L",
        "ICG.L","IMI.L","IWG.L","LRE.L","MONY.L","MTRO.L",
        "PAGE.L","PETS.L","PHP.L","RBN.L","RDW.L","SMDS.L","TRN.L",
        "TUI.L","VLX.L","VTY.L","WMH.L","ZIP.L",
        # FTSE 250 — Mid Cap additions
        "ABDN.L","AGR.L","AIR.L","ALW.L","APF.L","APAX.L","APS.L",
        "AWE.L","BBOX.L","BCPT.L","BEMO.L","BGEO.L","BHMG.L","BIG.L",
        "BKW.L","BLND.L","BOY.L","BYIT.L","CAL.L","CBG.L","CLDN.L",
        "CLLN.L","CNN.L","COB.L","CPG.L","CWK.L","DARK.L","DCO.L",
        "DEST.L","DFS.L","DGOC.L","DIG.L","DISCV.L","DLG.L","DMGT.L",
        "DPLM.L","DSM.L","DTY.L","ELCO.L","ETL.L","EVET.L","EWI.L",
        "FCIT.L","FDM.L","FGP.L","FLTV.L","FNX.L","FORT.L","GAH.L",
        "GCP.L","GENL.L","GHE.L","GPE.L","GRI.L","HARK.L","HBOS.L",
        "HFG.L","HIML.L","HMSO.L","HNE.L","HSL.L","HTG.L","HUW.L",
        "HYVE.L","IGG.L","IHG.L","INPP.L","INVP.L","IP.L","ITH.L",
        "JFJ.L","JMAT.L","JMG.L","JUP.L","KIE.L","LBOW.L","LIO.L",
        "LLPD.L","LON.L","LOOK.L","LPA.L","MAB.L","MACF.L","MCB.L",
        "MCLS.L","MCS.L","MGGT.L","MGP.L","MHLD.L","MIL.L","MNGS.L",
        "MON.L","MPAC.L","MTC.L","MTVW.L","MWB.L","NAH.L","NBH.L",
        "NCT.L","NEX.L","NEWA.L","NFG.L","NHD.L","NMC.L","OAK.L",
        "OCA.L","OCN.L","OML.L","ONT.L","OSB.L","OTB.L","OXIG.L",
        "PAG.L","PAHB.L","PCA.L","PDG.L","PEG.L","PGH.L","PIC.L",
        "PLUS.L","PMO.L","POW.L","PRS.L","PSH.L","PTEC.L","PUR.L",
        "QQ.L","RBD.L","RBGP.L","RCDO.L","RDL.L","REDD.L","RELX.L",
        "RGD.L","RGS.L","RHIM.L","RHM.L","RICA.L","RIDE.L","RIO.L",
        "RLT.L","ROO.L","RPC.L","RPT.L","RRR.L","RNWH.L","SBT.L",
        "SCF.L","SCST.L","SEM.L","SHED.L","SHFD.L","SHI.L","SIA.L",
        "SIG.L","SLND.L","SMWH.L","SNX.L","SOHO.L","SOPH.L","SPDI.L",
        "SREI.L","SRP.L","SRS.L","STB.L","STO.L","SVS.L","SXS.L",
        "TAN.L","TCG.L","TIFS.L","TMG.L","TNI.L","TPKD.L","TRI.L",
        "TRST.L","TSG.L","TUPU.L","TWI.L","UANC.L","UBM.L","UCI.L",
        "UIL.L","ULVR.L","UTL.L","VANL.L","VCP.L","VEIL.L","VIP.L",
        "VLE.L","VOF.L","VRS.L","VSVS.L","WAG.L","WAND.L","WEYS.L",
        "WIN.L","WJG.L","WKC.L","WTAN.L","WWH.L","YCA.L","YMK.L",
    ],
    "AIM": [
        "BOO.L","ASOS.L","CEY.L","CHG.L","CNE.L","THG.L","GAW.L","SDI.L",
        "STEM.L","YCA.L","ZOO.L","BVXP.L","DOTD.L","FLO.L","GHH.L","HFD.L",
        "MXCT.L","NXR.L","PFC.L","SHOE.L","TUNE.L","AVON.L","CBOX.L",
        "BVS.L","CARD.L","CFX.L","CLG.L","COA.L","COG.L","IGP.L",
        "RBG.L","SCT.L","SLP.L","TRX.L","AGK.L","AJB.L","APH.L","MBH.L",
        "ALFA.L","ANP.L","ATM.L","BAB.L","BLV.L","BOOM.L","CAD.L",
        "CBG.L","CHL.L","CMH.L","CRB.L","CSG.L","CVS.L","DSG.L",
        "DTG.L","EKF.L","EVO.L","FDEV.L","FEVR.L","FIF.L","FRP.L",
        "GBG.L","GKP.L","GOG.L","HAT.L","HGT.L","IDOX.L","IQE.L",
        "JET2.L","JTC.L","KWS.L","LTG.L","MAB.L","MGNS.L","MIN.L",
        "NAHL.L","OCA.L","PAY.L","PDL.L","PEN.L","PHE.L","RWS.L",
        "SEPL.L","SHI.L","SIM.L","SKP.L","SNR.L","STB.L","SYS1.L",
        "TGR.L","TPFG.L","TXP.L","ULS.L","VCP.L","WGB.L","WIL.L",
        "WSG.L","XPD.L","ZINC.L",
        # Additional AIM stocks
        "ACSO.L","AFC.L","AFM.L","AGA.L","AGH.L","AHT.L","ALBA.L",
        "ALPH.L","ALT.L","AMAP.L","AMC.L","AMFW.L","AMGO.L","AMP.L",
        "AMRK.L","ANCR.L","AND.L","ANPX.L","AOF.L","APC.L","APGT.L",
        "APMO.L","APTD.L","AQX.L","ARB.L","ARCH.L","ARG.L","ARW.L",
        "ASA.L","ASC.L","ASCL.L","ASPL.L","AST.L","ASTR.L","ATK.L",
        "ATIS.L","ATL.L","ATMT.L","ATO.L","ATQT.L","ATST.L","ATT.L",
        "AUE.L","AUTG.L","AVN.L","AVO.L","AWE.L","AWIN.L","AXON.L",
        "AYM.L","AZL.L","BADM.L","BAGL.L","BAIR.L","BAND.L","BAPE.L",
        "BARK.L","BASG.L","BBH.L","BBRG.L","BBX.L","BCN.L","BCPT.L",
        "BCSV.L","BEG.L","BELL.L","BELR.L","BEN.L","BFIG.L","BGO.L",
        "BGUK.L","BHR.L","BHSL.L","BID.L","BIDS.L","BIG.L","BIO.L",
        "BIOM.L","BIOG.L","BION.L","BISH.L","BKG.L","BKMA.L","BKS.L",
        "BLC.L","BLCR.L","BLND.L","BLV.L","BMK.L","BMS.L","BMTS.L",
        "BNC.L","BNK.L","BNR.L","BOK.L","BOLT.L","BONA.L","BPC.L",
        "BPM.L","BPR.L","BPTY.L","BRG.L","BRIO.L","BRK.L","BRL.L",
    ],
    "NYSE": [
        # Financials
        "JPM","BAC","WFC","GS","MS","C","BRK-B","V","MA","AXP","BLK","SCHW",
        "CB","MET","AFL","PGR","TRV","HIG","ALL","AIG","PRU","LNC","UNM",
        "COF","DFS","SYF","ALLY","RF","KEY","CFG","FITB","HBAN","ZION","MTB",
        "USB","PNC","TFC","SIVB","NTRS","STT","BK","ICE","CME","CBOE",
        # Healthcare
        "JNJ","PFE","MRK","ABT","BMY","UNH","LLY","TMO","DHR","MDT",
        "BSX","SYK","ZBH","EW","BDX","BAX","CAH","MCK","ABC","CNC",
        "HUM","CVS","MOH","ANTM","ELV","CI","HCA","THC","UHS","CRL",
        "IQV","IQVIA","PPD","MEDP","NEOG","XRAY","HSIC","PKI","A",
        # Energy
        "XOM","CVX","COP","SLB","EOG","PSX","VLO","MPC","OXY","HES",
        "DVN","FANG","PXD","APA","MRO","HAL","BKR","NOV","FTI","TDW",
        "KMI","WMB","OKE","ET","EPD","MMP","MPLX","PAA","LNG","RRC",
        # Industrials
        "BA","GE","MMM","HON","CAT","DE","LMT","RTX","NOC","GD","L",
        "UPS","FDX","CSX","NSC","UNP","CP","CNI","XPO","JBHT","SAIA",
        "PH","ITW","EMR","ROK","AME","GWW","MSC","FAST","WSO","HUBB",
        "FTV","ROP","IEX","CFX","FLS","GNRC","IR","OTIS","CARR","JCI",
        # Consumer
        "WMT","TGT","HD","LOW","COST","DG","CVS","MCD","SBUX","NKE",
        "KO","PEP","PG","CL","GIS","MO","PM","DIS","CMCSA","T","VZ",
        "YUM","QSR","WING","CMG","DPZ","SHAK","CAKE","DRI","MKC","K",
        "CPB","HRL","SJM","CAG","KHC","TSN","MOS","ADM","BG","INGR",
        "RL","PVH","HBI","VFC","TJX","ROST","M","KSS","JWN","AN","KMX",
        # Utilities & REITs
        "NEE","DUK","SO","D","AEP","EXC","AMT","PLD","SPG","O","WELL",
        "ES","ETR","FE","PPL","CMS","NI","ATO","OGE","PNM","CLECO",
        "PSA","EQR","AVB","MAA","UDR","CPT","NNN","STORE","VICI","MGM",
        "ARE","BXP","HIW","KIM","REG","FRT","WRI","SKT","CBL","SRG",
        # Materials
        "NEM","FCX","NUE","STLD","RS","CMC","ATI","CLF","AA","X","MT",
        "PPG","SHW","RPM","ECL","DOW","LYB","HUN","CC","WLK","OLN",
        "APD","LIN","PX","ALB","SQM","MP","FSLR","ENPH","RUN","NOVA",
    ],
    "NASDAQ": [
        # Mega-cap tech
        "AAPL","MSFT","GOOGL","AMZN","META","TSLA","NVDA","AMD","INTC","QCOM",
        # Semiconductors
        "TXN","MU","AVGO","AMAT","LRCX","KLAC","MRVL","ON","NXPI","SWKS",
        "WOLF","MPWR","SLAB","CREE","DIOD","IXYS","RMBS","POWI","SITM","AEIS",
        "ACLS","AXTI","BRKS","CCMP","COHU","FORM","ICHR","IPGP","MKSI","UCTT",
        # Software — Enterprise
        "NFLX","ADBE","CRM","ORCL","INTU","NOW","SNOW","WDAY","TEAM","HUBS",
        "ZM","DOCU","SPLK","VEEV","PAYC","PCTY","APPF","JAMF","WEX","COUP",
        "BILL","SMAR","BOX","ESTC","FROG","GTLB","MNDY","MTTR","NCNO","PEGA",
        "PTC","QLIK","RNG","SCWX","TOST","TTD","TYL","VNET","WIX","XPEL",
        # Software — Consumer / Saas
        "LULU","ABNB","BKNG","EBAY","PYPL","SQ","COIN","EXPE","LYFT","UBER",
        "DASH","RBLX","SNAP","PINS","SPOT","MTCH","IAC","ZG","OPEN","ANGI",
        "CARS","CARG","DKNG","GENI","PENN","SKLZ","SPCE","ACMR","AFRM","UPST",
        # Biotech / Pharma
        "GILD","AMGN","BIIB","REGN","VRTX","MRNA","DXCM","IDXX","ISRG","ALGN",
        "BMRN","EXEL","HALO","INCY","IONS","JAZZ","LGND","MDGL","NBIX","NKTR",
        "PCVX","RARE","RCKT","RETA","RVMD","SAGE","SMMT","SRPT","TECH","UTHR",
        "ACAD","ACHC","ADPT","ADUS","ADVM","AFMD","AGIO","AGTC","AGEN","AGLE",
        # Cybersecurity / Cloud
        "CSCO","ANET","FTNT","PANW","CRWD","ZS","OKTA","DDOG","NET","MDB",
        "CYBR","QLYS","VRNS","TENB","RDWR","SAIL","RGEN","SCWX","SAIC","SCCF",
        "JNPR","NTAP","NTNX","PSTG","SMCI","STX","WDC","XLNX","KEYS","LITE",
        # Fintech / Payments
        "FISV","FIS","GPN","TRMK","PCVT","NDAQ","SPGI","MCO","MSCI","VRSK",
        "SEIC","LPLA","SSNC","PFG","VOYA","NAVI","SLM","CACC","DT","WEX",
        # E-commerce / Consumer
        "ETSY","CHWY","W","OSTK","PDD","JD","BABA","BILI","IQ","MELI",
        "SE","GRAB","GLOB","STNE","PAGS","VTEX","VTOL","XMTR","YEXT","ZETA",
        # Healthcare Tech
        "ILMN","HOLX","MASI","MMSI","NVCR","NVST","ONEM","OSUR","PCRX","PDCO",
        "PRAX","PRGO","PRVA","PSMT","QDEL","QGEN","RARE","RCKT","RDUS","RCUS",
    ],
    "XETRA": [
        # DAX 40 — Full list
        "SAP.DE","SIE.DE","ALV.DE","MUV2.DE","BMW.DE","MBG.DE","BAYN.DE",
        "BAS.DE","VOW3.DE","ADS.DE","HEN3.DE","LIN.DE","DTE.DE","RWE.DE",
        "EOAN.DE","MRK.DE","DB1.DE","DBK.DE","HEI.DE","FRE.DE","DHL.DE",
        "AIR.DE","ZAL.DE","PUM.DE","CON.DE","BOSS.DE","IFX.DE","MTX.DE",
        "LEG.DE","VNA.DE","SHL.DE","ENR.DE","DHER.DE","SY1.DE","BEI.DE",
        "SMHN.DE","1COV.DE","QIA.DE","P911.DE","MBB.DE",
        # MDAX additions
        "AFX.DE","ARL.DE","ARND.DE","BC8.DE","BFSA.DE","BING.DE","BOIF.DE",
        "BR3.DE","CEV.DE","CLIQ.DE","COP.DE","CWC.DE","DBAN.DE","DE.DE",
        "DEQ.DE","DIC.DE","DMG.DE","DMGK.DE","ECK.DE","EVD.DE","EVK.DE",
        "EVO.DE","EWG.DE","FPE3.DE","FRA.DE","GBF.DE","GFK.DE","GLJ.DE",
        "HAW.DE","HBH.DE","HDD.DE","HDI.DE","HLAG.DE","HOT.DE","HYQ.DE",
        "IFX.DE","ILM1.DE","IVU.DE","JEN.DE","JKHY.DE","JUN3.DE","K+S.DE",
        "KGX.DE","KLR.DE","KNEBV.DE","KSB.DE","KU2.DE","KWHP.DE",
        "LNSX.DE","LPKF.DE","LXS.DE","MAN.DE","MBB.DE","MLP.DE","MOR.DE",
        "MSF.DE","MTE.DE","MVMB.DE","NET1.DE","NXU.DE","O2D.DE","OHB.DE",
        "OLINK.DE","OMV.DE","OPTI.DE","PAH3.DE","PBB.DE","PC8.DE","PFV.DE",
        "PMOX.DE","POC.DE","PSM.DE","PUT.DE","RAA.DE","REP.DE","RHM.DE",
        "RKET.DE","RNL.DE","RSL2.DE","RTL.DE","RWE.DE","S92.DE","SDAX.DE",
        "SFQ.DE","SGL.DE","SIE.DE","SOBA.DE","SOW.DE","SPI.DE","SRTX.DE",
        "SSL.DE","STO3.DE","STR.DE","STRN.DE","SY1.DE","SZG.DE","TE.DE",
        "TLX.DE","TMV.DE","TOM.DE","TPE.DE","TRR.DE","TUI1.DE","UNI3.DE",
        "VBK.DE","VIB3.DE","VNA.DE","VOE.DE","WAF.DE","WCH.DE","WDI.DE",
        "WIN.DE","WKL.DE","WMT.DE","XONA.DE","ZIL2.DE",
    ],
    "EURONEXT": [
        # CAC 40 — Full list
        "AI.PA","OR.PA","MC.PA","BNP.PA","SAN.PA","TTE.PA","ENGI.PA",
        "DG.PA","VIE.PA","ORA.PA","SGO.PA","RI.PA","CAP.PA","HO.PA",
        "SU.PA","BN.PA","KER.PA","ATO.PA","STMPA.PA","CS.PA","DSY.PA",
        "EL.PA","GLE.PA","ML.PA","PUB.PA","RMS.PA","SAF.PA","SW.PA",
        "URW.PA","VIE.PA","WLN.PA","DSFIR.PA","TEP.PA","TFI.PA",
        "SEB.PA","SGEF.PA","SOLB.PA","SPG.PA","STM.PA","TCL.PA",
        # Amsterdam (AEX) — Full
        "ASML.AS","HEIA.AS","NN.AS","RAND.AS","PHG.AS","WKL.AS",
        "ABN.AS","AD.AS","ADYEN.AS","AGN.AS","AKZA.AS","BESI.AS",
        "DSMF.AS","EXOR.AS","GLPG.AS","IMCD.AS","INGA.AS","MT.AS",
        "NSG.AS","PHIA.AS","PROSUS.AS","REN.AS","TKWY.AS","UNA.AS",
        "VPK.AS","WEB.AS","WHA.AS","AALB.AS","ABN.AS","ACOMO.AS",
        # Brussels (BEL20)
        "ABI.BR","AGS.BR","ANG.BR","ARGX.BR","BPOST.BR","COLR.BR",
        "CFE.BR","COIL.BR","ECONB.BR","EVS.BR","GBL.BR","GBLB.BR",
        "KBC.BR","LOTB.BR","MELX.BR","ONTEX.BR","PROX.BR","SOF.BR",
        "TNET.BR","UCB.BR","UMI.BR","WDP.BR","XIOR.BR",
        # Madrid (IBEX 35)
        "ACS.MC","AENA.MC","AMS.MC","ANA.MC","BBVA.MC","BKT.MC",
        "CABK.MC","CIE.MC","COL.MC","ELE.MC","ENG.MC","FER.MC",
        "GRF.MC","IAG.MC","IBE.MC","IDR.MC","ITX.MC","LOG.MC",
        "MAP.MC","MEL.MC","MRL.MC","MTS.MC","NTGY.MC","PHM.MC",
        "RED.MC","REE.MC","REP.MC","ROVI.MC","SAB.MC","SAN.MC",
        "SGRE.MC","SLR.MC","SOL.MC","TEF.MC","VIS.MC",
    ],
    "TSX": [
        "RY.TO","TD.TO","BNS.TO","BMO.TO","CM.TO","MFC.TO","SLF.TO",
        "TRI.TO","CNR.TO","CP.TO","ENB.TO","TRP.TO","SU.TO","CNQ.TO",
        "ABX.TO","WPM.TO","T.TO","BCE.TO","SHOP.TO","CSU.TO",
        "ATD.TO","L.TO","WN.TO","MRU.TO","EMA.TO","FTS.TO",
    ],
    "ASX": [
        "CBA.AX","BHP.AX","CSL.AX","NAB.AX","WBC.AX","ANZ.AX","MQG.AX",
        "WES.AX","WOW.AX","RIO.AX","FMG.AX","TLS.AX","WDS.AX",
        "STO.AX","QAN.AX","COL.AX","ALL.AX","REA.AX","SEK.AX","CPU.AX",
        "WTC.AX","XRO.AX","PMV.AX","JHX.AX","LLC.AX","DXS.AX","GPT.AX",
    ],
    "NSE": [
        "RELIANCE.NS","TCS.NS","HDFCBANK.NS","INFY.NS","ICICIBANK.NS",
        "HINDUNILVR.NS","ITC.NS","SBIN.NS","BAJFINANCE.NS","BHARTIARTL.NS",
        "KOTAKBANK.NS","LT.NS","ASIANPAINT.NS","AXISBANK.NS","MARUTI.NS",
        "SUNPHARMA.NS","TITAN.NS","ULTRACEMCO.NS","NESTLEIND.NS","WIPRO.NS",
    ],
    "HKEX": [
        "0700.HK","0005.HK","0941.HK","1299.HK","2318.HK","0388.HK",
        "1398.HK","3988.HK","0939.HK","2628.HK","0883.HK","0001.HK",
        "0016.HK","0003.HK","0002.HK","0011.HK","0012.HK","0019.HK",
    ],
}

ALL_EXCHANGES = {
    "🏛 London (LSE)":          "LSE",
    "📊 London AIM":            "AIM",
    "🗽 New York (NYSE)":       "NYSE",
    "📈 NASDAQ":                "NASDAQ",
    "🏦 Germany (XETRA)":       "XETRA",
    "🌐 Euronext (Paris/AMS)":  "EURONEXT",
    "🍁 Toronto (TSX)":         "TSX",
    "🌏 Australia (ASX)":       "ASX",
    "💹 India (NSE)":           "NSE",
    "🏮 Hong Kong (HKEX)":      "HKEX",
}

SECTORS = [
    "All Sectors","Technology","Healthcare","Financials",
    "Consumer Discretionary","Consumer Staples","Energy","Materials",
    "Industrials","Utilities","Real Estate","Communication Services",
]

PRESET_PAIRS = {
    # ── UK (LSE) ──────────────────────────────────────────────────────────────
    "🇬🇧 Lloyds / Barclays (UK Banks)":                  ("LLOY.L",  "BARC.L"),
    "🇬🇧 Lloyds / NatWest (UK Banks)":                   ("LLOY.L",  "NWG.L"),
    "🇬🇧 Barclays / NatWest (UK Banks)":                 ("BARC.L",  "NWG.L"),
    "🇬🇧 HSBC / Standard Chartered (Global Banks)":      ("HSBA.L",  "STAN.L"),
    "🇬🇧 Taylor Wimpey / Persimmon (Housebuilders)":     ("TW.L",    "PSN.L"),
    "🇬🇧 Taylor Wimpey / Barratt (Housebuilders)":       ("TW.L",    "BA.L"),
    "🇬🇧 BT Group / Vodafone (Telecoms)":                ("BT-A.L",  "VOD.L"),
    "🇬🇧 Rio Tinto / Anglo American (Mining)":           ("RIO.L",   "AAL.L"),
    "🇬🇧 Rio Tinto / BHP (Diversified Mining)":          ("RIO.L",   "BHP.L"),
    "🇬🇧 Anglo American / Antofagasta (Copper)":         ("AAL.L",   "ANTO.L"),
    "🇬🇧 Tesco / Sainsbury's (Grocery)":                 ("TSCO.L",  "SBRY.L"),
    "🇬🇧 Shell / BP (UK Energy Majors)":                 ("SHEL.L",  "BP.L"),
    "🇬🇧 AstraZeneca / GSK (UK Pharma)":                 ("AZN.L",   "GSK.L"),
    "🇬🇧 Diageo / Haleon (UK Consumer)":                 ("DGE.L",   "HLN.L"),
    "🇬🇧 Prudential / Legal & General (UK Insurance)":   ("PRU.L",   "LGEN.L"),
    "🇬🇧 Rentokil / Bunzl (UK Support Svcs)":            ("RTO.L",   "BNZL.L"),
    "🇬🇧 Experian / RELX (UK Data/Analytics)":           ("EXPN.L",  "REL.L"),
    "🇬🇧 IAG / easyJet (UK Aviation)":                   ("IAG.L",   "EZJ.L"),
    "🇬🇧 WPP / Publicis — (UK vs 🇫🇷 FR Advertising)":  ("WPP.L",   "PUB.PA"),
    "🇬🇧 Greggs / JD Sports (UK Consumer Growth)":       ("GRG.L",   "JD.L"),
    # ── US ────────────────────────────────────────────────────────────────────
    "🇺🇸 JPMorgan / Bank of America (US Banks)":         ("JPM",     "BAC"),
    "🇺🇸 Goldman Sachs / Morgan Stanley (Investment Bk)":("GS",      "MS"),
    "🇺🇸 Wells Fargo / Citigroup (US Retail Banks)":     ("WFC",     "C"),
    "🇺🇸 Visa / Mastercard (Payments)":                  ("V",       "MA"),
    "🇺🇸 PayPal / Block (Fintech)":                      ("PYPL",    "SQ"),
    "🇺🇸 Coca-Cola / PepsiCo (Beverages)":               ("KO",      "PEP"),
    "🇺🇸 ExxonMobil / Chevron (US Energy)":              ("XOM",     "CVX"),
    "🇺🇸 Pfizer / Johnson & Johnson (Pharma)":           ("PFE",     "JNJ"),
    "🇺🇸 Merck / AbbVie (US Biopharma)":                 ("MRK",     "ABBV"),
    "🇺🇸 Amazon / Walmart (Retail)":                     ("AMZN",    "WMT"),
    "🇺🇸 Target / Costco (US Retail)":                   ("TGT",     "COST"),
    "🇺🇸 Apple / Microsoft (US Mega Cap Tech)":          ("AAPL",    "MSFT"),
    "🇺🇸 Meta / Alphabet (Social/Search)":               ("META",    "GOOGL"),
    "🇺🇸 Netflix / Disney (Streaming)":                  ("NFLX",    "DIS"),
    "🇺🇸 Nvidia / AMD (Semiconductors)":                 ("NVDA",    "AMD"),
    "🇺🇸 Intel / Qualcomm (Chips)":                      ("INTC",    "QCOM"),
    "🇺🇸 Boeing / Lockheed Martin (Aerospace/Defense)":  ("BA",      "LMT"),
    "🇺🇸 Caterpillar / Deere (Industrials)":             ("CAT",     "DE"),
    "🇺🇸 Ford / General Motors (US Autos)":              ("F",       "GM"),
    "🇺🇸 Delta / United Airlines (US Aviation)":         ("DAL",     "UAL"),
    "🇺🇸 Duke Energy / NextEra (US Utilities)":          ("DUK",     "NEE"),
    # ── Europe ────────────────────────────────────────────────────────────────
    "🇩🇪 SAP / Siemens (German Tech/Industrial)":        ("SAP.DE",  "SIE.DE"),
    "🇩🇪 Volkswagen / BMW (German Autos)":               ("VOW3.DE", "BMW.DE"),
    "🇩🇪 Deutsche Bank / Commerzbank (German Banks)":    ("DBK.DE",  "CBK.DE"),
    "🇫🇷 LVMH / Hermès (French Luxury)":                 ("MC.PA",   "RMS.PA"),
    "🇫🇷 TotalEnergies / Equinor (European Energy)":     ("TTE.PA",  "EQNR"),
    "🇨🇭 Nestlé / Unilever (Consumer Staples)":          ("NESN.SW", "ULVR.L"),
    "🇨🇭 Novartis / Roche (Swiss Pharma)":               ("NOVN.SW", "ROG.SW"),
    "🇳🇱 ASML / Infineon (European Semis)":              ("ASML",    "IFX.DE"),
    # ── Asia / Emerging ───────────────────────────────────────────────────────
    "🇯🇵 Toyota / Honda (Japanese Autos)":               ("7203.T",  "7267.T"),
    "🇯🇵 Sony / Panasonic (Japanese Electronics)":       ("6758.T",  "6752.T"),
    "🇦🇺 BHP / Rio Tinto (Australian Mining)":           ("BHP.AX",  "RIO.AX"),
}

# ─────────────────────────────────────────────────────────────
# TRADING JOURNAL DATABASE
# ─────────────────────────────────────────────────────────────

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fintiq_journal.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS trades (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT,
            ticker      TEXT,
            company     TEXT,
            direction   TEXT,
            strategy    TEXT,
            entry_price REAL,
            exit_price  REAL,
            shares      REAL,
            currency    TEXT,
            status      TEXT,
            notes       TEXT,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def db_add_trade(date, ticker, company, direction, strategy,
                 entry_price, exit_price, shares, currency, status, notes):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        INSERT INTO trades (date,ticker,company,direction,strategy,
                            entry_price,exit_price,shares,currency,status,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (date, ticker, company, direction, strategy,
          entry_price, exit_price, shares, currency, status, notes))
    conn.commit()
    conn.close()

def db_get_trades() -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    df   = pd.read_sql("SELECT * FROM trades ORDER BY date DESC", conn)
    conn.close()
    return df

def db_delete_trade(trade_id: int):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM trades WHERE id=?", (trade_id,))
    conn.commit()
    conn.close()

init_db()

# ─────────────────────────────────────────────────────────────
# HELPERS — FORMATTING
# ─────────────────────────────────────────────────────────────

def fmt_currency(value, symbol="$", decimals=2):
    """Format a number as currency with commas and symbol."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "–"
    try:
        v = float(value)
        if abs(v) >= 1_000_000_000:
            return f"{symbol}{v/1_000_000_000:,.1f}B"
        if abs(v) >= 1_000_000:
            return f"{symbol}{v/1_000_000:,.1f}M"
        if abs(v) >= 1_000:
            return f"{symbol}{v:,.{decimals}f}"
        return f"{symbol}{v:,.{decimals}f}"
    except Exception:
        return "–"

def fmt_number(value, decimals=2):
    """Format plain number with commas."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "–"
    try:
        return f"{float(value):,.{decimals}f}"
    except Exception:
        return "–"

def fmt_pct(value, decimals=1):
    if value is None:
        return "–"
    try:
        return f"{float(value)*100:.{decimals}f}%"
    except Exception:
        return "–"

def get_currency_symbol(ticker: str) -> str:
    for exch, syms in STOCK_UNIVERSE.items():
        if ticker in syms:
            return EXCHANGE_CURRENCY.get(exch, "$")
    return "$"

def get_price_display(value, ticker: str, info: dict = None) -> str:
    """
    LSE/AIM stocks: yfinance returns prices in PENCE (GBp).
    Always show dual format: £X.XX (Xp) for any value where abs >= 100p.
    Negative DCF values shown as –£X.XX (–Xp) so format is consistent.
    All other markets: use normal fmt_currency.
    """
    if value is None:
        return "–"
    raw_currency = (info or {}).get("currency", "") if info else ""
    sym = get_currency_symbol(ticker)
    if raw_currency == "GBp":
        p = _f(value)
        if p is None:
            return "–"
        # Always show dual £/p format for any meaningful pence value
        if abs(p) >= 100:
            return f"£{p/100:,.2f}  ({p:,.0f}p)"
        return f"{p:,.2f}p"
    return fmt_currency(value, sym)

def _f(v):
    """Module-level type-safe float converter."""
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Fintiq | Alpha Securities Intelligence",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# (Guest ID tracking removed — login required before any action)

# ─────────────────────────────────────────────────────────────
# AUTH GATE — Login / Sign-up wall
# ─────────────────────────────────────────────────────────────

def _auth_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap');
    html, body, .stApp { background: #0F1923 !important; font-family: 'Inter', sans-serif !important; }
    .auth-wrap {
        max-width: 420px; margin: 60px auto 0 auto;
        background: linear-gradient(135deg, #0D2137 0%, #0A1628 100%);
        border: 1px solid rgba(245,158,11,0.25); border-radius: 16px;
        padding: 40px 36px; box-shadow: 0 8px 40px rgba(0,0,0,0.6);
    }
    .auth-logo {
        font-size: 2.6rem; font-weight: 900; color: #F59E0B;
        letter-spacing: -2px; text-align: center; margin-bottom: 4px;
        text-shadow: 0 0 20px rgba(245,158,11,0.5);
    }
    .auth-sub { text-align: center; color: #64748B; font-size: 0.82rem;
                margin-bottom: 28px; font-style: italic; }
    .auth-err { background: rgba(239,68,68,0.12); border: 1px solid rgba(239,68,68,0.4);
                color: #F87171; border-radius: 8px; padding: 10px 14px;
                font-size: 0.85rem; margin-bottom: 12px; }
    .auth-ok  { background: rgba(34,197,94,0.12); border: 1px solid rgba(34,197,94,0.4);
                color: #4ADE80; border-radius: 8px; padding: 10px 14px;
                font-size: 0.85rem; margin-bottom: 12px; }
    </style>
    """, unsafe_allow_html=True)

def _show_auth():
    """Render login / signup page. Returns True when user is authenticated."""
    _auth_css()

    st.markdown("""
    <div class="auth-wrap">
      <div class="auth-logo">📊 Fintiq</div>
      <div class="auth-sub">Intelligent Trading Screener · From speculation to strategy</div>
    </div>
    """, unsafe_allow_html=True)

    # Centre the form
    _, col, _ = st.columns([1, 2, 1])
    with col:
        mode = st.radio("", ["Login", "Sign up"], horizontal=True,
                        label_visibility="collapsed", key="auth_mode")
        email = st.text_input("Email", placeholder="you@example.com", key="auth_email")
        password = st.text_input("Password", type="password",
                                 placeholder="Min 6 characters", key="auth_pw")

        if mode == "Sign up":
            confirm = st.text_input("Confirm password", type="password",
                                    placeholder="Repeat password", key="auth_pw2")

        btn_label = "Create account" if mode == "Sign up" else "Log in"
        if st.button(btn_label, use_container_width=True, type="primary"):
            if not email or not password:
                st.markdown('<div class="auth-err">Please enter your email and password.</div>',
                            unsafe_allow_html=True)
                return False

            if _sb is None:
                st.markdown('<div class="auth-err">Auth service unavailable — check configuration.</div>',
                            unsafe_allow_html=True)
                return False

            try:
                if mode == "Sign up":
                    if password != st.session_state.get("auth_pw2", ""):
                        st.markdown('<div class="auth-err">Passwords do not match.</div>',
                                    unsafe_allow_html=True)
                        return False
                    if len(password) < 6:
                        st.markdown('<div class="auth-err">Password must be at least 6 characters.</div>',
                                    unsafe_allow_html=True)
                        return False
                    res = _sb.auth.sign_up({"email": email, "password": password})
                    if res.user:
                        # Auto-login immediately (email confirmation disabled in Supabase)
                        try:
                            lr = _sb.auth.sign_in_with_password({"email": email, "password": password})
                            if lr.user:
                                _tok = lr.session.access_token if lr.session else None
                                st.session_state["fintiq_user"] = {"email": lr.user.email, "id": lr.user.id, "session": _tok}
                                if _tok:
                                    st.query_params["_t"] = _tok
                                st.rerun()
                                return True
                        except Exception:
                            pass
                        st.markdown('<div class="auth-ok">✅ Account created! Please log in.</div>',
                                    unsafe_allow_html=True)
                        return False
                    else:
                        st.markdown('<div class="auth-err">Sign-up failed. Please try again.</div>',
                                    unsafe_allow_html=True)
                        return False
                else:
                    res = _sb.auth.sign_in_with_password({"email": email, "password": password})
                    if res.user:
                        _tok = res.session.access_token if res.session else None
                        st.session_state["fintiq_user"] = {
                            "email": res.user.email,
                            "id": res.user.id,
                            "session": _tok,
                        }
                        # Save token in URL so browser refresh restores session
                        if _tok:
                            st.query_params["_t"] = _tok
                        # Process any pending Stripe payment
                        _pss = st.session_state.pop("_pending_stripe_session", "")
                        if _pss:
                            _verify_stripe_session(_pss, res.user.id)
                        st.rerun()
                    else:
                        st.markdown('<div class="auth-err">Invalid email or password.</div>',
                                    unsafe_allow_html=True)
                        return False
            except Exception as e:
                err_msg = str(e)
                if "Invalid login" in err_msg or "invalid_grant" in err_msg or "Email not confirmed" in err_msg:
                    st.markdown('<div class="auth-err">Invalid email or password.</div>',
                                unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="auth-err">Error: {err_msg}</div>', unsafe_allow_html=True)
                return False
    return False

# ── Usage limits ─────────────────────────────────────────────
if "free_searches" not in st.session_state:
    st.session_state["free_searches"] = 0
_MONTHLY_LIMIT = 5   # free-account searches per calendar month

# ── Supabase profile helpers ──────────────────────────────────
def _get_profile(user_id: str) -> dict:
    if not _sb_admin or not user_id:
        return {}
    try:
        r = _sb_admin.table("profiles").select("*").eq("id", user_id).maybe_single().execute()
        return r.data or {}
    except Exception:
        return {}

def _upsert_profile(user_id: str, data: dict):
    if not _sb_admin or not user_id:
        return
    try:
        _sb_admin.table("profiles").upsert({"id": user_id, **data}).execute()
    except Exception:
        pass

def _increment_search(user_id: str, profile: dict) -> dict:
    now_month = datetime.now().strftime("%Y-%m")
    # Always read CURRENT count from Supabase to avoid stale increments
    try:
        if _sb:
            r = _sb.table("profiles").select("monthly_searches,search_month").eq("id", user_id).execute()
            if r.data:
                db_row = r.data[0]
                searches = db_row.get("monthly_searches", 0) if db_row.get("search_month") == now_month else 0
            else:
                searches = 0
        else:
            searches = profile.get("monthly_searches", 0) if profile.get("search_month") == now_month else 0
    except Exception:
        searches = profile.get("monthly_searches", 0) if profile.get("search_month") == now_month else 0
    searches += 1
    updated = {**profile, "monthly_searches": searches, "search_month": now_month}
    _upsert_profile(user_id, {"monthly_searches": searches, "search_month": now_month})
    return updated

# ── Stripe helpers ────────────────────────────────────────────
_stripe_last_error: str = ""   # populated on failure for diagnostics

def _create_checkout(plan: str, user_email: str, user_id: str) -> str | None:
    global _stripe_last_error
    _stripe_last_error = ""
    if not _stripe:
        _stripe_last_error = "stripe library not imported"
        return None
    if not _STRIPE_SECRET:
        _stripe_last_error = "STRIPE_SECRET_KEY env var is empty"
        return None
    price_id = _PRICE_ANNUAL if plan == "annual" else _PRICE_MONTHLY
    try:
        session = _stripe.checkout.Session.create(
            mode="subscription",
            customer_email=user_email,
            line_items=[{"price": price_id, "quantity": 1}],
            success_url=f"{_APP_URL}?stripe_session={{CHECKOUT_SESSION_ID}}",
            cancel_url=f"{_APP_URL}",
            metadata={"user_id": user_id},
        )
        return session.url
    except Exception as _e:
        _stripe_last_error = str(_e)
        return None

def _verify_stripe_session(session_id: str, user_id: str) -> bool:
    if not _stripe or not _STRIPE_SECRET:
        return False
    try:
        session = _stripe.checkout.Session.retrieve(session_id)
        if session.payment_status == "paid":
            _upsert_profile(user_id, {
                "is_pro": True,
                "stripe_customer_id": session.customer,
                "stripe_subscription_id": session.subscription,
            })
            st.session_state["fintiq_user"]["is_pro"] = True
            if "fintiq_profile" in st.session_state:
                st.session_state["fintiq_profile"]["is_pro"] = True
            return True
    except Exception:
        pass
    return False

# ── Guest search tracking (persistent via Supabase) ──────────
def _get_guest_count(guest_id: str) -> int:
    """Return how many searches this guest has done (from Supabase)."""
    if not _sb or not guest_id:
        return 0
    try:
        r = _sb.table("guest_searches").select("count").eq("id", guest_id).execute()
        if r.data:
            return r.data[0].get("count", 0)
        return 0
    except Exception:
        return 0

def _increment_guest(guest_id: str) -> int:
    """Increment guest search count and return new total."""
    if not _sb or not guest_id:
        return 1
    try:
        r = _sb.table("guest_searches").select("count").eq("id", guest_id).execute()
        current = r.data[0].get("count", 0) if r.data else 0
        new_count = current + 1
        _sb.table("guest_searches").upsert({"id": guest_id, "count": new_count}).execute()
        return new_count
    except Exception:
        return 1

# ── Auth / upgrade gate ───────────────────────────────────────
def _check_auth_gate() -> bool:
    """Returns True if allowed to run a search."""
    user = st.session_state.get("fintiq_user", {})

    # Pro — unlimited
    if user.get("is_pro"):
        return True

    # Guest — must sign up before any action
    if not user:
        for _k in ["screened_df", "screened_symbols"]:
            if _k in st.session_state: del st.session_state[_k]
        st.session_state["_show_auth_wall"] = True
        st.rerun()

    # Free registered user — always read count from Supabase (source of truth)
    user_id = user.get("id", "")
    now_month = datetime.now().strftime("%Y-%m")

    # Read current count via admin client (bypasses RLS)
    current_searches = 0
    _dbg_read_err = None
    _dbg_write_err = None
    _dbg_row = None
    try:
        if _sb_admin:
            r = _sb_admin.table("profiles").select("monthly_searches,search_month,is_pro").eq("id", user_id).execute()
            _dbg_row = r.data
            if r.data:
                row = r.data[0]
                if row.get("is_pro"):
                    st.session_state["fintiq_user"]["is_pro"] = True
                    return True
                current_searches = row.get("monthly_searches", 0) if row.get("search_month") == now_month else 0
    except Exception as e:
        _dbg_read_err = str(e)

    st.info(f"DEBUG: admin_client={'YES' if _sb_admin and _sb_admin is not _sb else 'FALLBACK'} | row={_dbg_row} | count={current_searches} | limit={_MONTHLY_LIMIT} | read_err={_dbg_read_err}")

    if current_searches < _MONTHLY_LIMIT:
        try:
            _sb_admin.table("profiles").upsert({"id": user_id, "monthly_searches": current_searches + 1, "search_month": now_month}).execute()
        except Exception as e:
            _dbg_write_err = str(e)
            st.warning(f"DEBUG write err: {_dbg_write_err}")
        return True

    # Limit reached — show upgrade wall
    for _k in ["screened_df", "screened_symbols"]:
        if _k in st.session_state: del st.session_state[_k]
    st.session_state["_show_upgrade_wall"] = (user.get("email", ""), user_id)
    st.rerun()

def _show_auth_wall():
    """Signup wall shown to guests when they try to use the app."""
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0D2137,#0A1628);
        border:1px solid rgba(245,158,11,0.4);border-radius:16px;
        padding:36px;text-align:center;max-width:500px;margin:20px auto;
        box-shadow:0 8px 40px rgba(0,0,0,0.6);">
      <div style="font-size:2rem;font-weight:900;color:#F59E0B;letter-spacing:-1px;margin-bottom:8px">
        📊 Fintiq</div>
      <div style="color:#F1F5F9;font-size:1.1rem;font-weight:700;margin-bottom:8px">
        Sign up to start using Fintiq</div>
      <div style="color:#94A3B8;font-size:0.88rem;margin-bottom:8px">
        Create a free account to get <b style="color:#F59E0B">5 searches/month</b> — no card required.<br>
        Upgrade to Pro for unlimited access.
      </div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1,2,1])
    with col:
        mode = st.radio("", ["Sign up free", "I have an account"], horizontal=True,
                        label_visibility="collapsed", key="wall_mode")
        email    = st.text_input("Email", placeholder="you@example.com", key="wall_email")
        password = st.text_input("Password", type="password",
                                 placeholder="Min 6 characters", key="wall_pw")
        if mode == "Sign up free":
            st.text_input("Confirm password", type="password",
                          placeholder="Repeat password", key="wall_pw2")
        if st.button("Continue →", use_container_width=True, type="primary", key="wall_btn"):
            if not email or not password:
                st.error("Please enter your email and password."); return
            if _sb is None:
                st.error("Auth service unavailable."); return
            try:
                if mode == "Sign up free":
                    if password != st.session_state.get("wall_pw2",""):
                        st.error("Passwords do not match."); return
                    if len(password) < 6:
                        st.error("Password must be at least 6 characters."); return
                    res = _sb.auth.sign_up({"email": email, "password": password})
                    if res.user:
                        try:
                            lr = _sb.auth.sign_in_with_password({"email": email, "password": password})
                            if lr.user:
                                _tok = lr.session.access_token if lr.session else None
                                st.session_state["fintiq_user"] = {"email": lr.user.email, "id": lr.user.id, "session": _tok}
                                st.session_state.pop("_show_auth_wall", None)
                                st.session_state.pop("_show_upgrade_wall", None)
                                if _tok:
                                    st.query_params["_t"] = _tok
                                st.rerun()
                        except Exception:
                            st.success("✅ Account created! Please log in.")
                    else:
                        st.error("Sign-up failed. Please try again.")
                else:
                    res = _sb.auth.sign_in_with_password({"email": email, "password": password})
                    if res.user:
                        _tok2 = res.session.access_token if res.session else None
                        st.session_state["fintiq_user"] = {"email": res.user.email, "id": res.user.id}
                        st.session_state["free_searches"] = 0
                        st.session_state.pop("_show_auth_wall", None)
                        st.session_state.pop("_show_upgrade_wall", None)
                        if _tok2:
                            st.query_params["_t"] = _tok2
                        # Process any pending Stripe payment
                        _pss = st.session_state.pop("_pending_stripe_session", "")
                        if _pss:
                            _verify_stripe_session(_pss, res.user.id)
                        st.rerun()
                    else:
                        st.error("Invalid email or password.")
            except Exception as e:
                err = str(e)
                if "Email not confirmed" in err:
                    st.warning("Please confirm your email — check your inbox.")
                elif "Invalid login" in err or "invalid_grant" in err:
                    st.error("Invalid email or password.")
                else:
                    st.error(f"Error: {err}")

def _show_upgrade_wall(user_email: str, user_id: str):
    """Upgrade wall shown to free users after 5 searches/month.
    Stripe URLs pre-generated on first render and cached — no button-click API calls."""

    # Only cache successful URLs — retry every render if empty (avoids stale "" from earlier failure)
    if not st.session_state.get("_wall_monthly_url"):
        _mu = _create_checkout("monthly", user_email, user_id)
        if _mu:
            st.session_state["_wall_monthly_url"] = _mu
    if not st.session_state.get("_wall_annual_url"):
        _au = _create_checkout("annual", user_email, user_id)
        if _au:
            st.session_state["_wall_annual_url"] = _au

    _m_url = st.session_state.get("_wall_monthly_url", "")
    _a_url = st.session_state.get("_wall_annual_url", "")
    _last_err = _stripe_last_error  # capture after create_checkout calls above

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0D2137,#0A1628);
        border:1px solid rgba(245,158,11,0.4);border-radius:16px;
        padding:36px;text-align:center;max-width:560px;margin:20px auto;
        box-shadow:0 8px 40px rgba(0,0,0,0.6);">
      <div style="font-size:2rem;font-weight:900;color:#F59E0B;letter-spacing:-1px;margin-bottom:8px">
        📊 Fintiq Pro</div>
      <div style="color:#F1F5F9;font-size:1.1rem;font-weight:700;margin-bottom:8px">
        You've used all 5 free searches this month</div>
      <div style="color:#94A3B8;font-size:0.88rem;margin-bottom:20px">
        Upgrade to Pro for <b style="color:#F1F5F9">unlimited searches</b>, all global markets,
        and priority data.
      </div>
      <div style="display:flex;gap:16px;justify-content:center;flex-wrap:wrap;margin-bottom:8px">
        <div style="background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.4);
            border-radius:12px;padding:18px 28px;min-width:160px">
          <div style="color:#F59E0B;font-weight:700;font-size:1.4rem">£10</div>
          <div style="color:#CBD5E1;font-size:0.85rem">per month</div>
        </div>
        <div style="background:rgba(34,197,94,0.1);border:1px solid rgba(34,197,94,0.4);
            border-radius:12px;padding:18px 28px;min-width:160px;position:relative">
          <div style="position:absolute;top:-16px;left:50%;transform:translateX(-50%);
              background:#22C55E;color:#fff;font-size:0.7rem;font-weight:700;
              padding:3px 12px;border-radius:10px;white-space:nowrap">SAVE 2 MONTHS</div>
          <div style="color:#4ADE80;font-weight:700;font-size:1.4rem;margin-top:8px">£100</div>
          <div style="color:#CBD5E1;font-size:0.85rem">per year</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    _, col, _ = st.columns([1, 2, 1])
    with col:
        if _m_url and _a_url:
            # st.link_button = plain <a href> — no Streamlit rerun, no logout
            c1, c2 = st.columns(2)
            with c1:
                st.link_button("▶ Monthly — £10/mo", _m_url,
                               use_container_width=True, type="primary")
            with c2:
                st.link_button("⭐ Annual — £100/yr", _a_url,
                               use_container_width=True)
            st.caption("🔒 Secure payment via Stripe · Cancel anytime · Card never stored by Fintiq")
        else:
            st.error(f"Payment system error: {_last_err or _stripe_last_error or 'unknown — check Railway logs'}")

# ── Restore session from URL token on browser refresh ────────
# After login we store the Supabase access_token in ?_t= URL param.
# On page refresh Streamlit loses session_state but the URL param survives.
# We validate it with Supabase and restore the session silently.
_qp_token = st.query_params.get("_t", "")
if _qp_token and "fintiq_user" not in st.session_state and _sb:
    try:
        _tok_resp = _sb.auth.get_user(_qp_token)
        if _tok_resp and _tok_resp.user:
            st.session_state["fintiq_user"] = {
                "email": _tok_resp.user.email,
                "id": _tok_resp.user.id,
                "session": _qp_token,
            }
            # Also load pro status immediately
            _tok_prof = _get_profile(_tok_resp.user.id)
            if _tok_prof.get("is_pro"):
                st.session_state["fintiq_user"]["is_pro"] = True
            st.session_state["fintiq_profile"] = _tok_prof
    except Exception:
        # Token expired — clear it silently
        try: st.query_params.pop("_t")
        except Exception: pass

# ── Capture stripe_session BEFORE any login redirect clears it ──
_early_stripe = st.query_params.get("stripe_session", "")
if _early_stripe:
    st.session_state["_pending_stripe_session"] = _early_stripe

# ── Logged-in user email (empty string if guest) ─────────────
_user_email = st.session_state.get("fintiq_user", {}).get("email", "")

# ─────────────────────────────────────────────────────────────
# GLOBAL CSS — Professional Navy/Gold Theme
# ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
  /* ── Touch / mobile base ── */
  * { -webkit-tap-highlight-color: rgba(245,158,11,0.15); box-sizing: border-box; }
  img, svg, iframe { max-width: 100% !important; }
  /* Prevent horizontal overflow on mobile */
  body, .stApp, .main { overflow-x: hidden !important; }
  /* Smoother scrolling */
  html { scroll-behavior: smooth; -webkit-overflow-scrolling: touch; }

  /* ── HIDE STREAMLIT DEFAULT HEADER / TOOLBAR ── */
  header[data-testid="stHeader"] { display: none !important; height: 0 !important; min-height: 0 !important; }
  #MainMenu { visibility: hidden !important; display: none !important; }
  .stDeployButton { display: none !important; }
  footer { display: none !important; }

  /* ── FINTIQ GLOBAL FOOTER ── */
  .fintiq-global-footer {
    position: fixed; bottom: 0; left: 0; right: 0; z-index: 9999;
    background: rgba(8,18,32,0.85);
    border-top: 1px solid rgba(245,158,11,0.15);
    padding: 3px 24px;
    display: flex; align-items: center; justify-content: space-between;
    font-size: 0.65rem; color: #334155;
    backdrop-filter: blur(4px);
    opacity: 0.6;
    transition: opacity 0.2s ease;
  }
  .fintiq-global-footer:hover { opacity: 1; }
  .fintiq-global-footer a { color: #475569; text-decoration: none; }
  .fintiq-global-footer a:hover { color: #F59E0B; }
  .fintiq-global-footer .fg-brand { color: #64748B; font-weight: 700; }
  section[data-testid="stAppViewContainer"] { padding-bottom: 28px !important; }
  /* Remove ALL top padding from every Streamlit wrapper — covers v1.30+ and v1.40+ element names */
  section[data-testid="stAppViewContainer"] > div:first-child { padding-top: 0 !important; }
  div[data-testid="stAppViewBlockContainer"] { padding-top: 0 !important; }
  div[data-testid="stMainBlockContainer"] { padding-top: 0 !important; }
  div[data-testid="stMain"] { padding-top: 0 !important; }
  div[data-testid="stVerticalBlock"] > div:first-child { padding-top: 0 !important; }
  .appview-container .main .block-container { padding-top: 0 !important; margin-top: 0 !important; }
  .main .block-container { padding-top: 0 !important; margin-top: 0 !important; }
  [data-testid="stAppViewContainer"] { padding-top: 0 !important; }
  .stApp > header { display: none !important; }
  .stApp { padding-top: 0 !important; margin-top: 0 !important; }

  /* ── DARK PROFESSIONAL BASE ── */
  html, body, .stApp {
    background-color: #0F1923 !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 16px !important;
    color: #CBD5E1 !important;
  }
  .main .block-container {
    background-color: #0F1923 !important;
    padding-top: 0 !important;
    padding-bottom: 1rem !important;
    max-width: 1400px !important;
  }
  p, span, label, .stMarkdown { color: #CBD5E1 !important; font-size: 15px !important; }
  div { color: #CBD5E1 !important; }
  h1, h2, h3 { color: #F8FAFC !important; font-size: 1.4rem !important; }

  /* Subtle market chart background pattern */
  .stApp::before {
    content: '';
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background:
      radial-gradient(ellipse at 20% 50%, rgba(245,158,11,0.04) 0%, transparent 50%),
      radial-gradient(ellipse at 80% 20%, rgba(30,58,95,0.15) 0%, transparent 50%),
      linear-gradient(rgba(245,158,11,0.025) 1px, transparent 1px),
      linear-gradient(90deg, rgba(245,158,11,0.025) 1px, transparent 1px);
    background-size: 100% 100%, 100% 100%, 50px 50px, 50px 50px;
    pointer-events: none;
    z-index: 0;
  }

  /* ── TOP NAV BAR ── */
  .fintiq-nav {
    background: linear-gradient(135deg, #0A1628 0%, #0D2137 50%, #0A1628 100%);
    padding: 5px 22px;
    border-radius: 0 0 8px 8px;
    margin-bottom: 2px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 2px solid #F59E0B;
    box-shadow: 0 4px 30px rgba(245,158,11,0.15), 0 2px 60px rgba(0,0,0,0.8);
  }
  .fintiq-logo {
    font-size: 2.1rem !important;
    font-weight: 900 !important;
    color: #F59E0B !important;
    letter-spacing: -2px !important;
    font-family: 'Inter', 'Arial Black', sans-serif !important;
    text-shadow: 0 0 20px rgba(245,158,11,0.7), 0 2px 4px rgba(0,0,0,0.6) !important;
    line-height: 1 !important;
    display: inline !important;
  }
  .fintiq-tagline {
    font-size: 0.7rem;
    color: #64748B;
    margin-top: 1px;
    font-style: italic;
    letter-spacing: 0.3px;
  }
  .nav-badge {
    background: rgba(245,158,11,0.12);
    border: 1px solid rgba(245,158,11,0.4);
    color: #F59E0B;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.72rem;
    font-weight: 600;
  }

  /* ── TICKER TAPE ── */
  .ticker-tape {
    background: #0A1520;
    border-top: 1px solid rgba(245,158,11,0.2);
    border-bottom: 1px solid rgba(245,158,11,0.2);
    padding: 1px 0;
    overflow: hidden;
    white-space: nowrap;
    margin-bottom: 1px;
    font-size: 0.7rem;
    font-family: 'Inter', monospace;
  }
  .ticker-tape-inner {
    display: inline-block;
    animation: scroll-left 40s linear infinite;
  }
  .ticker-item {
    display: inline-block;
    margin: 0 28px;
    color: #94A3B8;
  }
  .ticker-item .t-sym { color: #F59E0B; font-weight: 700; margin-right: 5px; }
  .ticker-item .t-up  { color: #4ADE80; font-weight: 600; }
  .ticker-item .t-dn  { color: #F87171; font-weight: 600; }
  @keyframes scroll-left {
    0%   { transform: translateX(0); }
    100% { transform: translateX(-50%); }
  }

  /* ── METRIC CARDS ── */
  .metric-card {
    background: linear-gradient(135deg, #162032 0%, #1A2840 100%);
    border-radius: 10px;
    padding: 16px 18px;
    border: 1px solid rgba(255,255,255,0.07);
    border-left: 3px solid #2A4A6B;
    margin-bottom: 10px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.3);
    height: 110px;
    min-height: 110px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    overflow: hidden;
    transition: height 0.25s ease, box-shadow 0.25s ease, border-color 0.2s ease;
    cursor: default;
  }
  .metric-card:hover {
    height: auto !important;
    min-height: 110px;
    overflow: visible;
    box-shadow: 0 6px 28px rgba(245,158,11,0.2), 0 2px 12px rgba(0,0,0,0.5);
    border-color: rgba(245,158,11,0.4);
    z-index: 10;
    position: relative;
  }
  .metric-card-gold  { border-left-color: #F59E0B !important; }
  .metric-card-green { border-left-color: #22C55E !important; }
  .metric-card-red   { border-left-color: #EF4444 !important; }
  .metric-label {
    font-size: 0.68rem !important; color: #64748B !important; font-weight: 700 !important;
    text-transform: uppercase; letter-spacing: 0.8px;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }
  .metric-card:hover .metric-label {
    white-space: normal; overflow: visible;
  }
  .metric-value {
    font-size: 1.45rem !important; font-weight: 800 !important; color: #F1F5F9 !important;
    margin-top: 4px; font-family: 'Inter', monospace;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }
  .metric-card:hover .metric-value {
    white-space: normal; overflow: visible;
  }
  .metric-sub {
    font-size: 0.75rem !important; color: #94A3B8 !important; margin-top: 3px;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }
  .metric-card:hover .metric-sub {
    white-space: normal; overflow: visible;
  }

  /* ── DOWNLOAD BUTTON ── */
  .stDownloadButton > button {
    background: linear-gradient(135deg, #D97706 0%, #F59E0B 100%) !important;
    color: #0F1923 !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 10px 22px !important;
    box-shadow: 0 3px 12px rgba(245,158,11,0.35) !important;
    transition: all 0.2s ease !important;
  }
  .stDownloadButton > button:hover {
    background: linear-gradient(135deg, #F59E0B 0%, #FBBF24 100%) !important;
    box-shadow: 0 5px 20px rgba(245,158,11,0.5) !important;
    transform: translateY(-1px) !important;
  }

  /* ── SECTION HEADERS ── */
  .section-header {
    background: linear-gradient(135deg, #0D2137 0%, #1A3355 100%);
    color: #F59E0B;
    padding: 10px 18px;
    border-radius: 8px;
    font-size: 1rem;
    font-weight: 700;
    margin: 18px 0 12px 0;
    border-left: 4px solid #F59E0B;
    letter-spacing: 0.3px;
  }

  /* ── SIGNAL BADGES ── */
  .signal-long  { background:rgba(34,197,94,0.15); color:#4ADE80; padding:6px 18px;
                  border-radius:20px; font-weight:700; font-size:1rem;
                  border: 1px solid #22C55E; }
  .signal-short { background:rgba(239,68,68,0.15); color:#F87171; padding:6px 18px;
                  border-radius:20px; font-weight:700; font-size:1rem;
                  border: 1px solid #EF4444; }
  .signal-none  { background:rgba(100,116,139,0.2); color:#94A3B8; padding:6px 18px;
                  border-radius:20px; font-weight:700; font-size:1rem;
                  border: 1px solid #475569; }

  /* ── TIER BOXES (dark) ── */
  .tier-3 { background:rgba(34,197,94,0.08); border-left:4px solid #22C55E;
            padding:12px 18px; border-radius:8px; margin:6px 0;
            border: 1px solid rgba(34,197,94,0.2); }
  .tier-2 { background:rgba(245,158,11,0.08); border-left:4px solid #F59E0B;
            padding:12px 18px; border-radius:8px; margin:6px 0;
            border: 1px solid rgba(245,158,11,0.2); }
  .tier-1 { background:rgba(59,130,246,0.08); border-left:4px solid #3B82F6;
            padding:12px 18px; border-radius:8px; margin:6px 0;
            border: 1px solid rgba(59,130,246,0.2); }

  /* ── VALUATION CARDS (dark) ── */
  .val-card {
    background: linear-gradient(135deg, #0D1F35 0%, #0A1828 100%);
    border-radius: 12px;
    padding: 22px;
    border: 1px solid rgba(245,158,11,0.2);
    margin: 8px 0;
    text-align: center;
    box-shadow: 0 4px 20px rgba(0,0,0,0.4);
    /* Equal height + clip overflow in collapsed state */
    height: 220px;
    overflow: hidden;
    transition: height 0.3s ease, box-shadow 0.3s ease, border-color 0.3s ease;
    cursor: default;
  }
  .val-card:hover {
    height: auto !important;
    overflow: visible;
    box-shadow: 0 8px 40px rgba(245,158,11,0.25);
    border-color: rgba(245,158,11,0.5);
    z-index: 10;
    position: relative;
  }
  .val-method { font-size: 0.78rem; color: #64748B; font-weight: 700;
                text-transform: uppercase; letter-spacing: 1px; }
  .val-price  { font-size: 1.9rem; font-weight: 900; color: #F8FAFC; margin: 8px 0; }
  .val-upside-pos { color: #4ADE80; font-weight: 700; font-size: 1.1rem; }
  .val-upside-neg { color: #F87171; font-weight: 700; font-size: 1.1rem; }

  /* ── HERO TITLE on Home page ── */
  .fintiq-hero-title {
    font-size: 2.8rem !important;
    font-weight: 900 !important;
    color: #F59E0B !important;
    letter-spacing: 3px !important;
    text-align: center !important;
    margin: 0 0 12px 0 !important;
    padding: 0 !important;
    line-height: 1.1 !important;
    text-shadow: 0 0 30px rgba(245,158,11,0.3) !important;
  }

  /* ── TABS (dark) ── */
  .stTabs [data-baseweb="tab-list"] {
    background: #0D1F35;
    border-radius: 10px;
    padding: 4px;
    border: 1px solid rgba(245,158,11,0.15);
    gap: 4px;
  }
  .stTabs [data-baseweb="tab"] {
    border-radius: 8px !important;
    font-weight: 600 !important;
    color: #64748B !important;
    font-size: 0.95rem !important;
  }
  .stTabs [aria-selected="true"] {
    background: linear-gradient(135deg,#1E3A5F,#0D2137) !important;
    color: #F59E0B !important;
    border: 1px solid rgba(245,158,11,0.3) !important;
  }

  /* ── INPUT WIDGETS ── */
  .stTextInput input, .stNumberInput input, .stTextArea textarea {
    background: #162032 !important;
    color: #E2E8F0 !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 8px !important;
    font-size: 1rem !important;
  }

  /* ── SELECTBOX & MULTISELECT — dark bg, white text ── */
  [data-baseweb="select"] > div {
    background-color: #162032 !important;
    border-color: rgba(255,255,255,0.15) !important;
  }
  [data-baseweb="select"] > div > div,
  [data-baseweb="select"] > div > div > div {
    background-color: #162032 !important;
    color: #E2E8F0 !important;
  }
  /* Tag value container — ensure left padding so first tag isn't flush with border */
  [data-baseweb="select"] > div > div:first-child {
    padding-left: 6px !important;
    overflow: visible !important;
  }
  /* Placeholder text */
  [data-baseweb="select"] input { color: #E2E8F0 !important; }

  /* ── DROPDOWN POPUP LIST ── */
  [data-baseweb="menu"],
  [data-baseweb="popover"] [data-baseweb="menu"] {
    background: #162032 !important;
    border: 1px solid rgba(245,158,11,0.2) !important;
    border-radius: 8px !important;
  }
  [data-baseweb="option"] {
    background: #162032 !important;
    color: #E2E8F0 !important;
  }
  [data-baseweb="option"]:hover,
  [data-baseweb="option"][aria-selected="true"] {
    background: #1E3A5F !important;
    color: #F59E0B !important;
  }
  ul[role="listbox"] li, div[role="option"] {
    background: #162032 !important;
    color: #E2E8F0 !important;
  }

  /* ── EXPANDER — all Streamlit versions ── */
  .streamlit-expanderHeader,
  div[data-testid="stExpander"] > details > summary,
  div[data-testid="stExpanderToggleIcon"],
  details > summary {
    background: #162032 !important;
    color: #CBD5E1 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    border: 1px solid rgba(245,158,11,0.2) !important;
  }
  .streamlit-expanderContent,
  div[data-testid="stExpander"] > details > div[data-testid="stExpanderDetails"],
  div[data-testid="stExpander"] details[open] > div {
    background: #111E2E !important;
    border: 1px solid rgba(245,158,11,0.1) !important;
    border-top: none !important;
    border-radius: 0 0 8px 8px !important;
    color: #CBD5E1 !important;
  }
  /* Make sure expander text is visible */
  div[data-testid="stExpander"] summary p,
  div[data-testid="stExpander"] summary span,
  div[data-testid="stExpander"] summary {
    color: #CBD5E1 !important;
    background: #162032 !important;
  }

  /* ── COMPACT VERTICAL GAPS (Fundamental Screen above-fold) ── */
  div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"],
  div[data-testid="stVerticalBlock"] > div {
    gap: 0.25rem !important;
  }
  div[data-testid="stCaptionContainer"] {
    margin-top: 0 !important; margin-bottom: 0 !important;
    padding-top: 0 !important; padding-bottom: 0 !important;
    line-height: 1.2 !important;
  }
  div[data-testid="stHorizontalBlock"] {
    gap: 0.5rem !important;
  }
  /* Compact expander header */
  div[data-testid="stExpander"] > details > summary {
    padding: 3px 12px !important;
    min-height: 26px !important;
    font-size: 0.82rem !important;
    line-height: 1.2 !important;
  }
  /* Tighten multiselect + selectbox labels */
  div[data-testid="stMultiSelect"] > label,
  div[data-testid="stSelectbox"] > label {
    margin-bottom: 1px !important;
    font-size: 0.82rem !important;
    line-height: 1.2 !important;
  }
  /* Multiselect selected tags — target via role="button" since data-baseweb="tag"
     is not present in newer Streamlit/BaseWeb versions */
  div[data-testid="stMultiSelect"] span[role="button"] {
    background-color: rgba(245,158,11,0.25) !important;
    border: 1px solid rgba(245,158,11,0.55) !important;
    border-radius: 6px !important;
    max-width: none !important;
    overflow: visible !important;
    margin: 2px 4px 2px 6px !important;
    flex-shrink: 0 !important;
  }
  /* Inner text label — remove the 128px max-width Streamlit applies */
  div[data-testid="stMultiSelect"] span[role="button"] > span:first-child {
    max-width: none !important;
    overflow: visible !important;
    white-space: nowrap !important;
    color: #FDE68A !important;
  }
  /* × close icon */
  div[data-testid="stMultiSelect"] span[role="button"] span[aria-hidden="true"],
  div[data-testid="stMultiSelect"] span[role="button"] svg {
    color: #FDE68A !important;
    fill: #FDE68A !important;
  }
  /* Also keep original selector as fallback */
  span[data-baseweb="tag"] {
    background-color: rgba(245,158,11,0.25) !important;
    border: 1px solid rgba(245,158,11,0.55) !important;
    border-radius: 6px !important;
    max-width: none !important;
    overflow: visible !important;
    margin: 2px 4px 2px 6px !important;
  }
  span[data-baseweb="tag"] span { color: #FDE68A !important; }
  span[data-baseweb="tag"] svg  { fill: #FDE68A !important; }
  /* Reduce tab bar top margin */
  div[data-testid="stTabs"] {
    margin-top: 4px !important;
  }
  /* Tighten tab content area top padding */
  div[data-testid="stTabsContent"] {
    padding-top: 8px !important;
  }

  /* ── TOAST NOTIFICATIONS ── */
  div[data-testid="stToast"],
  div[data-testid="stToast"] p,
  div[data-testid="stToast"] span {
    background: #1E3A52 !important;
    color: #F1F5F9 !important;
    border: 1px solid rgba(245,158,11,0.4) !important;
    border-radius: 8px !important;
  }

  /* ── DATAFRAME (dark) ── */
  .stDataFrame { border-radius: 10px !important; overflow: hidden !important; }

  /* ── ALERTS (dark) ── */
  .stAlert {
    border-radius: 10px !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
  }

  /* ── BUTTONS — target all Streamlit button variants ── */
  .stButton > button,
  div[data-testid="stButton"] > button,
  button[data-testid="baseButton-secondary"],
  button[data-testid="baseButton-primary"],
  button[kind="secondary"],
  button[kind="primary"] {
    background: #0D1F35 !important;
    color: #F59E0B !important;
    border: 1.5px solid #F59E0B !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    font-size: 0.9rem !important;
    padding: 10px 22px !important;
    letter-spacing: 0.4px !important;
    transition: all 0.2s !important;
  }
  .stButton > button:hover,
  div[data-testid="stButton"] > button:hover,
  button[data-testid="baseButton-secondary"]:hover,
  button[data-testid="baseButton-primary"]:hover,
  button[kind="secondary"]:hover,
  button[kind="primary"]:hover {
    background: rgba(245,158,11,0.15) !important;
    border-color: #FBBF24 !important;
    color: #FBBF24 !important;
    transform: translateY(-1px) !important;
  }

  /* ── DIVIDER ── */
  hr { border-color: rgba(245,158,11,0.15) !important; }

  /* ── CAPTION text ── */
  .stCaption { color: #94A3B8 !important; font-size: 0.88rem !important; }

  /* dropdown styles consolidated above in INPUT WIDGETS section */

  /* ── SVG BACKGROUND — subtle candlestick pattern ── */
  .fintiq-bg {
    position: fixed;
    top: 0; left: 0; width: 100%; height: 100%;
    pointer-events: none;
    z-index: -1;
    opacity: 0.10;
  }

  /* ── DISCLAIMER FOOTER ── */
  .disclaimer-footer {
    background: rgba(13,33,55,0.6);
    border: 1px solid rgba(245,158,11,0.15);
    border-radius: 8px;
    padding: 14px 20px;
    margin-top: 40px;
    font-size: 0.78rem;
    color: #64748B;
    line-height: 1.6;
  }
  .disclaimer-footer strong { color: #94A3B8; }

  /* Hide sidebar toggle */
  section[data-testid="stSidebar"] { display:none; }

  /* ══════════════════════════════════════════════
     MOBILE RESPONSIVE — full overhaul ≤768px
     ══════════════════════════════════════════════ */
  @media (max-width: 768px) {

    /* ── Base layout ── */
    html, body, .stApp { font-size: 14px !important; }
    .main .block-container {
      padding-left: 10px !important;
      padding-right: 10px !important;
      padding-bottom: 60px !important;
      max-width: 100vw !important;
    }
    /* Streamlit columns: stack to full width on mobile */
    div[data-testid="stHorizontalBlock"] {
      flex-direction: column !important;
      gap: 8px !important;
    }
    div[data-testid="stHorizontalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"],
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
      width: 100% !important;
      min-width: 100% !important;
      flex: 1 1 100% !important;
    }

    /* ── Nav bar ── */
    .fintiq-nav {
      flex-direction: column !important;
      align-items: flex-start !important;
      padding: 8px 12px !important;
      gap: 4px !important;
    }
    .fintiq-logo { font-size: 1.6rem !important; }
    .fintiq-tagline { font-size: 0.65rem !important; display: none !important; }
    .nav-badge { display: none !important; }

    /* ── Ticker tape ── */
    .ticker-tape { font-size: 0.62rem !important; padding: 4px 0 !important; }
    .ticker-item { margin: 0 12px !important; }

    /* ── Tabs — horizontally scrollable ── */
    .stTabs [data-baseweb="tab-list"] {
      overflow-x: auto !important;
      flex-wrap: nowrap !important;
      -webkit-overflow-scrolling: touch !important;
      scrollbar-width: none !important;
      padding-bottom: 2px !important;
    }
    .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar { display: none !important; }
    .stTabs [data-baseweb="tab"] {
      font-size: 0.68rem !important;
      padding: 5px 7px !important;
      white-space: nowrap !important;
    }

    /* ── Hero ── */
    .fintiq-hero-title {
      font-size: 1.5rem !important;
      letter-spacing: 0.5px !important;
    }

    /* ── Inline HTML grids — collapse to 1 or 2 cols ── */
    /* 4-col stat grid → 2×2 */
    div[style*="grid-template-columns:repeat(4"] {
      grid-template-columns: 1fr 1fr !important;
    }
    /* 6-col KPI row → 2×3 */
    div[style*="grid-template-columns:repeat(6"] {
      grid-template-columns: 1fr 1fr !important;
    }
    /* 3-col grids → 1 col */
    div[style*="grid-template-columns:1fr 1fr 1fr"],
    div[style*="grid-template-columns: 1fr 1fr 1fr"] {
      grid-template-columns: 1fr !important;
    }
    /* 2-col grids with min-width children → stack */
    div[style*="grid-template-columns:repeat(2"] {
      grid-template-columns: 1fr !important;
    }
    /* Decision framework 5-col → 1-col */
    div[style*="grid-template-columns:repeat(5"] {
      grid-template-columns: 1fr !important;
    }
    /* Flex rows with gap — wrap */
    div[style*="display:flex"][style*="gap"],
    div[style*="display: flex"][style*="gap"] {
      flex-wrap: wrap !important;
    }

    /* ── Metric / KPI cards ── */
    .metric-card {
      height: auto !important;
      min-height: 70px !important;
      padding: 10px !important;
    }
    .metric-value { font-size: 1rem !important; }
    .metric-label { font-size: 0.6rem !important; }
    .metric-sub   { font-size: 0.65rem !important; }

    /* ── Valuation cards ── */
    .val-card {
      height: auto !important;
      padding: 12px !important;
      margin: 4px 0 !important;
      font-size: 0.8rem !important;
    }
    .val-price { font-size: 1.2rem !important; }
    .val-method { font-size: 0.7rem !important; }

    /* ── Section headers ── */
    .section-header {
      font-size: 0.8rem !important;
      padding: 7px 10px !important;
    }

    /* ── Buttons — full width, big tap targets ── */
    .stButton > button,
    div[data-testid="stButton"] > button,
    button[data-testid="baseButton-primary"],
    button[data-testid="baseButton-secondary"] {
      width: 100% !important;
      padding: 12px 16px !important;
      font-size: 0.83rem !important;
      min-height: 44px !important;
    }
    /* Link buttons */
    a[data-testid="stLinkButton"],
    div[data-testid="stLinkButton"] a {
      display: block !important;
      width: 100% !important;
      text-align: center !important;
      min-height: 44px !important;
      line-height: 44px !important;
    }

    /* ── Inputs — full width ── */
    div[data-testid="stTextInput"],
    div[data-testid="stNumberInput"],
    div[data-testid="stSelectbox"],
    div[data-testid="stMultiSelect"],
    div[data-testid="stSlider"] {
      width: 100% !important;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input {
      font-size: 16px !important; /* prevents iOS zoom on focus */
    }

    /* ── Charts — responsive, no overflow ── */
    .js-plotly-plot, .plotly, .plot-container {
      max-width: 100% !important;
      overflow-x: hidden !important;
    }
    .js-plotly-plot .plotly .main-svg {
      max-width: 100% !important;
    }

    /* ── Dataframes — horizontal scroll ── */
    .stDataFrame, div[data-testid="stDataFrame"] {
      overflow-x: auto !important;
      font-size: 0.72rem !important;
      -webkit-overflow-scrolling: touch !important;
    }

    /* ── Expanders ── */
    div[data-testid="stExpander"] summary {
      font-size: 0.82rem !important;
      padding: 8px 10px !important;
    }

    /* ── Home page inline HTML ── */
    /* Hero description */
    div[style*="max-width:920px"] { padding: 0 4px !important; }
    /* Portfolio optimizer card: flex → column */
    div[style*="display:flex"][style*="flex-wrap:wrap"] > div {
      flex: 1 1 100% !important;
      min-width: 100% !important;
    }

    /* ── MC KPI rows — force 2 cols ── */
    /* The MC results grid is inline style — override via child count heuristic */
    div[style*="grid-template-columns:repeat(6,1fr)"] {
      grid-template-columns: repeat(2,1fr) !important;
      gap: 6px !important;
    }
    div[style*="grid-template-columns:repeat(4,1fr)"] {
      grid-template-columns: repeat(2,1fr) !important;
      gap: 6px !important;
    }

    /* ── Decision dashboard cards ── */
    div[style*="grid-template-columns:repeat(4,1fr)"] {
      grid-template-columns: 1fr 1fr !important;
      gap: 8px !important;
    }

    /* ── Section padding ── */
    div[style*="padding:28px 32px"] { padding: 14px 12px !important; }
    div[style*="padding:24px 28px"] { padding: 12px 10px !important; }
    div[style*="padding:22px"]      { padding: 12px !important; }
    div[style*="padding:18px 14px"] { padding: 10px 8px !important; }

    /* ── Typography ── */
    .section-header,
    div[style*="font-size:1.1rem"],
    div[style*="font-size:1.25rem"] { font-size: 0.9rem !important; }
    div[style*="font-size:1.8rem"],
    div[style*="font-size:2rem"]    { font-size: 1.4rem !important; }
    div[style*="font-size:0.95rem"] { font-size: 0.82rem !important; }
    div[style*="font-size:0.9rem"]  { font-size: 0.8rem !important; }
    div[style*="font-size:0.88rem"] { font-size: 0.78rem !important; }
    div[style*="font-size:0.85rem"] { font-size: 0.76rem !important; }
    div[style*="font-size:0.82rem"] { font-size: 0.74rem !important; }

    /* ── Disclaimer ── */
    .disclaimer-footer {
      font-size: 0.68rem !important;
      padding: 10px 12px !important;
    }

    /* ── Global footer — stack on mobile ── */
    .fintiq-global-footer {
      flex-direction: column !important;
      gap: 4px !important;
      padding: 6px 12px !important;
      font-size: 0.65rem !important;
      text-align: center !important;
    }
    .fintiq-global-footer > div:last-child {
      gap: 8px !important;
      flex-wrap: wrap !important;
      justify-content: center !important;
    }

    /* ── Streamlit radio buttons / checkboxes — bigger tap targets ── */
    div[data-testid="stRadio"] label,
    div[data-testid="stCheckbox"] label {
      min-height: 36px !important;
      display: flex !important;
      align-items: center !important;
    }

    /* ── Remove excessive margins on mobile ── */
    div[style*="margin-bottom:28px"] { margin-bottom: 16px !important; }
    div[style*="margin-bottom:20px"] { margin-bottom: 12px !important; }
    div[style*="margin-top:12px"]    { margin-top: 8px !important; }

    /* ── Streamlit column gaps ── */
    div[data-testid="stVerticalBlock"] {
      gap: 0.5rem !important;
    }
  }

  /* ── Named grid classes ── */
  @media (max-width: 768px) {
    .fiq-stat-grid  { grid-template-columns: 1fr 1fr !important; }
    .fiq-3col-grid  { grid-template-columns: 1fr !important; }
  }

  /* ── Small phones (≤400px) ── */
  @media (max-width: 400px) {
    .fintiq-logo { font-size: 1.3rem !important; }
    .stTabs [data-baseweb="tab"] {
      font-size: 0.6rem !important;
      padding: 4px 5px !important;
    }
    div[style*="grid-template-columns:repeat(2"] {
      grid-template-columns: 1fr !important;
    }
    div[style*="grid-template-columns:repeat(6,1fr)"],
    div[style*="grid-template-columns:repeat(4,1fr)"] {
      grid-template-columns: 1fr 1fr !important;
    }
    .main .block-container {
      padding-left: 6px !important;
      padding-right: 6px !important;
    }
  }
</style>
""", unsafe_allow_html=True)

# ── Seed watchlist from file ──
if "fintiq_watchlist" not in st.session_state:
    st.session_state["fintiq_watchlist"] = _wl_load()

# ── Seed pairs watchlist from file ──
if "fintiq_pairs_watchlist" not in st.session_state:
    st.session_state["fintiq_pairs_watchlist"] = _pwl_load()

# ── Google Analytics GA4 ──
import streamlit.components.v1 as _stc_ga
_stc_ga.html("""
<script async src="https://www.googletagmanager.com/gtag/js?id=G-ZPP2L744HP"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){dataLayer.push(arguments);}
  gtag('js', new Date());
  gtag('config', 'G-ZPP2L744HP');
</script>
""", height=0)

# ── Inject dropdown styles into PARENT document (escapes iframe) ──
import streamlit.components.v1 as _stc
_stc.html("""
<script>
(function() {
  var css = `
    [data-baseweb="popover"],
    [data-baseweb="menu"],
    [data-baseweb="select"] [role="listbox"],
    [data-baseweb="select"] ul {
      background: #1A2840 !important;
      border: 1px solid rgba(245,158,11,0.3) !important;
    }
    [data-baseweb="option"],
    [data-baseweb="select"] li,
    li[role="option"],
    div[role="option"] {
      background: #1A2840 !important;
      color: #E8EDF4 !important;
      font-size: 0.97rem !important;
    }
    [data-baseweb="option"]:hover,
    li[role="option"]:hover,
    div[role="option"]:hover,
    [data-baseweb="option"][aria-selected="true"] {
      background: #2A4060 !important;
      color: #F59E0B !important;
    }
  `;
  function inject(doc) {
    if (!doc) return;
    var s = doc.createElement('style');
    s.id = 'fintiq-dropdown-fix';
    s.textContent = css;
    if (!doc.getElementById('fintiq-dropdown-fix')) doc.head.appendChild(s);
  }
  inject(window.parent.document);
  inject(document);
  // Re-inject after Streamlit rerenders
  var obs = new MutationObserver(function() { inject(window.parent.document); inject(document); });
  obs.observe(window.parent.document.body, {childList: true, subtree: true});
})();
</script>
""", height=0)

# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT HELPER
# ─────────────────────────────────────────────────────────────

def build_fintiq_excel(df: "pd.DataFrame", sheet_name: str = "Fintiq Screen") -> bytes:
    """
    Returns bytes of a styled .xlsx file.
    Design: dark green background (#1A5C2A), black gridlines, white text.
    Quality Score ≥ 80 → yellow highlight; negative/red metrics → red text.
    """
    if not _OPENPYXL:
        return df.to_csv(index=False).encode()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Colours ──────────────────────────────────────────────
    BG_GREEN   = PatternFill("solid", fgColor="1A5C2A")   # dark forest green
    BG_HEADER  = PatternFill("solid", fgColor="0D3B18")   # deeper green for header
    BG_YELLOW  = PatternFill("solid", fgColor="F59E0B")   # gold highlight
    BG_RED     = PatternFill("solid", fgColor="7F1D1D")   # dark red for negatives

    FT_WHITE   = Font(name="Calibri", bold=False, color="FFFFFF", size=11)
    FT_HEADER  = Font(name="Calibri", bold=True,  color="F59E0B", size=11)
    FT_YELLOW  = Font(name="Calibri", bold=True,  color="0D3B18", size=11)  # dark on gold
    FT_RED     = Font(name="Calibri", bold=True,  color="FFAAAA", size=11)  # light red text

    THIN_BLACK = Side(border_style="thin", color="000000")
    BORDER     = Border(left=THIN_BLACK, right=THIN_BLACK,
                        top=THIN_BLACK,  bottom=THIN_BLACK)

    ALIGN_C    = Alignment(horizontal="center", vertical="center")
    ALIGN_L    = Alignment(horizontal="left",   vertical="center")

    cols = list(df.columns)

    # ── Header row ───────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill   = BG_HEADER
        cell.font   = FT_HEADER
        cell.border = BORDER
        cell.alignment = ALIGN_C

    # ── Data rows ────────────────────────────────────────────
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, (col, val) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = BORDER
            cell.alignment = ALIGN_L

            # Determine highlight
            is_qs    = col == "Quality Score"
            try:
                num = float(val)
            except (TypeError, ValueError):
                num = None

            if is_qs and num is not None and num >= 80:
                cell.fill = BG_YELLOW
                cell.font = FT_YELLOW
            elif num is not None and num < 0:
                cell.fill = BG_RED
                cell.font = FT_RED
            else:
                cell.fill = BG_GREEN
                cell.font = FT_WHITE

    # ── Column widths ────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        max_len = max(len(str(col)), *(len(str(r)) for r in df[col].astype(str)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 30)

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_deepdive_excel(ticker: str, company: str, curr_px,
                          dcf_val, graham_val, pe_val, avg_iv,
                          multiples: dict, sensitivity: dict,
                          assumptions: dict, hist_rows: list,
                          commentary: str = "") -> bytes:
    """
    Multi-sheet Excel export for a single stock deep-dive.
    Sheet 1: Summary  |  Sheet 2: Sensitivity Matrix  |
    Sheet 3: Multiples  |  Sheet 4: Historical Financials
    """
    if not _OPENPYXL:
        return b""

    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    wb = Workbook()

    # Colour palette
    BG_DARK   = PatternFill("solid", fgColor="0D1F35")
    BG_HEAD   = PatternFill("solid", fgColor="0A1628")
    BG_GREEN3 = PatternFill("solid", fgColor="14532D")
    BG_GREEN2 = PatternFill("solid", fgColor="166534")
    BG_GREEN1 = PatternFill("solid", fgColor="1A5C2A")
    BG_AMBER  = PatternFill("solid", fgColor="78350F")
    BG_RED1   = PatternFill("solid", fgColor="7F1D1D")
    BG_RED2   = PatternFill("solid", fgColor="6B1414")
    BG_GOLD   = PatternFill("solid", fgColor="F59E0B")
    BG_NEUTRAL= PatternFill("solid", fgColor="1E293B")

    FT_WHITE  = Font(name="Calibri", color="F1F5F9", size=11)
    FT_GOLD   = Font(name="Calibri", bold=True, color="F59E0B", size=11)
    FT_GREEN  = Font(name="Calibri", bold=True, color="4ADE80", size=11)
    FT_RED    = Font(name="Calibri", bold=True, color="F87171", size=11)
    FT_DARK   = Font(name="Calibri", bold=True, color="0D1F35", size=11)
    FT_TITLE  = Font(name="Calibri", bold=True, color="F59E0B", size=14)
    FT_HDR    = Font(name="Calibri", bold=True, color="F59E0B", size=11)

    THIN  = Side(border_style="thin", color="1E293B")
    BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    AC    = Alignment(horizontal="center", vertical="center")
    AL    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = BG_HEAD; c.font = FT_HDR; c.border = BDR; c.alignment = AC
        if width:
            ws.column_dimensions[c.column_letter].width = width
        return c

    def _cell(ws, row, col, val, bg=None, ft=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = bg or BG_DARK; c.font = ft or FT_WHITE
        c.border = BDR; c.alignment = align or AL
        return c

    # ── SHEET 1: SUMMARY ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    title_c = ws1.cell(row=1, column=1, value=f"Fintiq Deep-Dive: {company} ({ticker})")
    title_c.font = FT_TITLE; title_c.fill = BG_HEAD; title_c.alignment = AL
    ws1.merge_cells("A1:B1"); ws1.row_dimensions[1].height = 28

    rows = [
        ("Current Price",               curr_px),
        ("DCF Intrinsic Value",          dcf_val),
        ("Graham Number",               graham_val),
        ("Industry P/E Fair Value",      pe_val),
        ("Average Intrinsic Value",      avg_iv),
        ("DCF Upside / Downside",
         f"{((avg_iv-curr_px)/curr_px*100):.1f}%" if avg_iv and curr_px else "—"),
        ("", ""),
        ("— DCF ASSUMPTIONS —",          ""),
        ("WACC / Discount Rate",         f"{assumptions.get('discount_r','—')}%"),
        ("Terminal Growth Rate",         f"{assumptions.get('terminal_growth','—')}%"),
        ("ST Revenue Growth (Yrs 1-3)",  f"{assumptions.get('rg_short','—')}%"),
        ("ST Operating Margin",          f"{assumptions.get('om_short','—')}%"),
        ("MT Revenue Growth (Yrs 4-7)",  f"{assumptions.get('rg_med','—')}%"),
        ("MT Operating Margin",          f"{assumptions.get('om_med','—')}%"),
        ("LT Revenue Growth (Yrs 8-10)", f"{assumptions.get('rg_long','—')}%"),
        ("LT Operating Margin",          f"{assumptions.get('om_long','—')}%"),
        ("Effective Tax Rate",           f"{assumptions.get('tax_rate','—')}%"),
        ("Reinvestment Rate — Yrs 1–3",  f"{assumptions.get('inv_short','—')}%"),
        ("Reinvestment Rate — Yrs 4–7",  f"{assumptions.get('inv_med','—')}%"),
        ("Reinvestment Rate — Yrs 8–10", f"{assumptions.get('inv_long','—')}%"),
        ("RONIC — Return on New Invested Capital", f"{assumptions.get('ronic','—')}%"),
        ("Continuing Value Formula",     "NOPAT(t+1) × (1 − g/RONIC) / (WACC − g)"),
    ]
    for i, (lbl, val) in enumerate(rows, start=2):
        lbl_c = ws1.cell(row=i, column=1, value=lbl)
        lbl_c.fill = BG_NEUTRAL; lbl_c.font = FT_WHITE; lbl_c.border = BDR; lbl_c.alignment = AL
        val_c = ws1.cell(row=i, column=2, value=val)
        val_c.border = BDR; val_c.alignment = AC
        if "Upside" in lbl and val != "—":
            try:
                pct = float(str(val).replace("%",""))
                val_c.fill = BG_GREEN1 if pct >= 0 else BG_RED1
                val_c.font = FT_GREEN if pct >= 0 else FT_RED
            except Exception:
                val_c.fill = BG_DARK; val_c.font = FT_WHITE
        elif lbl.startswith("—"):
            lbl_c.fill = BG_HEAD; lbl_c.font = FT_GOLD
            val_c.fill = BG_HEAD
        else:
            val_c.fill = BG_DARK; val_c.font = FT_WHITE

    if commentary:
        ws1.row_dimensions[len(rows)+3].height = 14
        com_lbl = ws1.cell(row=len(rows)+4, column=1, value="AI Commentary")
        com_lbl.fill = BG_HEAD; com_lbl.font = FT_GOLD; com_lbl.border = BDR; com_lbl.alignment = AL
        com_cell = ws1.cell(row=len(rows)+4, column=2, value=commentary[:2000])
        com_cell.fill = BG_DARK; com_cell.font = FT_WHITE; com_cell.border = BDR
        com_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[len(rows)+4].height = 120

    # ── SHEET 2: SENSITIVITY MATRIX ──────────────────────────────
    ws2 = wb.create_sheet("DCF Sensitivity")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 16

    ws2.cell(row=1, column=1,
             value=f"DCF Sensitivity: {ticker} — Value per share at WACC × Terminal Growth"
             ).font = FT_TITLE
    ws2.cell(row=1, column=1).fill = BG_HEAD
    ws2.merge_cells(f"A1:{chr(65+len(sensitivity.get('tg_range',[0]*6)))}1")

    wacc_range = sensitivity.get("wacc_range", [])
    tg_range   = sensitivity.get("tg_range",   [])
    matrix     = sensitivity.get("matrix",      {})
    price      = sensitivity.get("price",       0)
    unit       = sensitivity.get("unit",        "")

    _hdr(ws2, 2, 1, "WACC \\ Term.g")
    for j, tg in enumerate(tg_range, start=2):
        c = _hdr(ws2, 2, j, f"{tg}%", width=13)
        ws2.column_dimensions[c.column_letter].width = 13

    for i, w in enumerate(wacc_range, start=3):
        wc = ws2.cell(row=i, column=1, value=f"{w}%")
        wc.fill = BG_HEAD; wc.font = FT_HDR; wc.border = BDR; wc.alignment = AC
        for j, tg in enumerate(tg_range, start=2):
            iv = matrix.get((w, tg))
            c  = ws2.cell(row=i, column=j)
            c.border = BDR; c.alignment = AC
            if iv is None:
                c.value = "—"; c.fill = BG_NEUTRAL; c.font = FT_WHITE
            else:
                c.value = f"{iv:.0f}{unit}" if unit == "p" else f"{unit}{iv:.2f}"
                ratio = (iv - price) / price if price else 0
                if   ratio >  0.30: c.fill = BG_GREEN3; c.font = FT_GREEN
                elif ratio >  0.10: c.fill = BG_GREEN2; c.font = FT_WHITE
                elif ratio >  0.00: c.fill = BG_GREEN1; c.font = FT_WHITE
                elif ratio > -0.10: c.fill = BG_AMBER;  c.font = FT_WHITE
                elif ratio > -0.30: c.fill = BG_RED1;   c.font = FT_WHITE
                else:               c.fill = BG_RED2;   c.font = FT_RED
                # Highlight current WACC + tg
                if w == assumptions.get("discount_r") and tg == assumptions.get("terminal_growth"):
                    c.fill = BG_GOLD; c.font = FT_DARK

    ws2.freeze_panes = "B3"

    # ── SHEET 3: VALUATION MULTIPLES ─────────────────────────────
    ws3 = wb.create_sheet("Valuation Multiples")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 26

    _hdr(ws3, 1, 1, "Multiple")
    _hdr(ws3, 1, 2, f"{ticker} Value")
    _hdr(ws3, 1, 3, "Sector Benchmark")
    _hdr(ws3, 1, 4, "vs Sector")
    _hdr(ws3, 1, 5, "Interpretation")

    mult_items = multiples.get("items", [])
    for i, m in enumerate(mult_items, start=2):
        _cell(ws3, i, 1, m.get("label",""), BG_NEUTRAL)
        val = m.get("value"); bm = m.get("benchmark")
        val_str = f"{val:.1f}{m.get('fmt','x')}" if val is not None else "—"
        _cell(ws3, i, 2, val_str)
        _cell(ws3, i, 3, f"{bm}{m.get('fmt','x')}" if bm else "—")
        if val is not None and bm is not None:
            pct = (val - bm) / bm * 100
            cheap = val < bm if m.get("lower_is_cheaper", True) else val > bm
            vc = ws3.cell(row=i, column=4, value=f"{'▼' if val<bm else '▲'}{abs(pct):.0f}%")
            vc.fill = BG_GREEN1 if cheap else BG_RED1
            vc.font = FT_GREEN if cheap else FT_RED
            vc.border = BDR; vc.alignment = AC
            _cell(ws3, i, 5, "Cheaper than peers" if cheap else "Expensive vs peers")
        else:
            _cell(ws3, i, 4, "—")
            _cell(ws3, i, 5, "No benchmark")

    # ── SHEET 4: HISTORICAL FINANCIALS ───────────────────────────
    if hist_rows:
        ws4 = wb.create_sheet("Historical Financials")
        ws4.sheet_view.showGridLines = False
        if hist_rows:
            h_cols = [c for c in hist_rows[0].keys() if c != "_row_type"]
            for j, col in enumerate(h_cols, start=1):
                _hdr(ws4, 1, j, col, width=max(len(col)+2, 14))
            for i, row in enumerate(hist_rows, start=2):
                is_sect = row.get("_row_type") == "sector"
                for j, col in enumerate(h_cols, start=1):
                    val = row.get(col, "")
                    c = ws4.cell(row=i, column=j, value=val)
                    c.border = BDR; c.alignment = AL
                    if is_sect:
                        c.fill = BG_HEAD; c.font = FT_GOLD
                    else:
                        try:
                            num = float(str(val).replace("%","").replace("£","").replace("p","").replace(",",""))
                            c.fill = BG_RED1 if num < 0 else BG_DARK
                            c.font = FT_RED  if num < 0 else FT_WHITE
                        except Exception:
                            c.fill = BG_DARK; c.font = FT_WHITE

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def fmp_get(endpoint: str, params: dict = None):
    p = params or {}
    p["apikey"] = FMP_KEY
    try:
        r = requests.get(f"{FMP_BASE}{endpoint}", params=p, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception:
        return {}


@st.cache_data(ttl=3600, show_spinner=False)
def get_yf_info(ticker: str) -> dict:
    try:
        info = yf.Ticker(ticker).info
        return info if info else {}
    except Exception:
        return {}


@st.cache_data(ttl=1800, show_spinner=False)
def get_price_history(ticker: str, period: str = "1y") -> pd.DataFrame:
    try:
        df = yf.download(ticker, period=period, auto_adjust=True,
                         progress=False, multi_level_index=False)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        for col in ["Open","High","Low","Close","Volume"]:
            if col in df.columns:
                s = df[col]
                if isinstance(s, pd.DataFrame):
                    df[col] = s.iloc[:,0]
        df = df.dropna(subset=["Close"])
        return df
    except Exception:
        return pd.DataFrame()

# ─────────────────────────────────────────────────────────────
# FUNDAMENTAL QUALITY SCORE
# ─────────────────────────────────────────────────────────────

def quality_score(info: dict) -> dict:
    roe          = _f(info.get("returnOnEquity"))
    pe           = _f(info.get("trailingPE")) or _f(info.get("forwardPE"))
    gross_margin = _f(info.get("grossMargins"))
    net_margin   = _f(info.get("profitMargins"))
    de_raw       = _f(info.get("debtToEquity"))
    debt_equity  = de_raw / 100 if de_raw is not None else None
    market_cap   = _f(info.get("marketCap")) or 0
    ocf          = _f(info.get("operatingCashflow")) or 0
    ni           = _f(info.get("netIncomeToCommon")) or 1
    cash_conv    = round(ocf / ni, 2) if ni != 0 else None

    res = {
        "roe": roe, "pe": pe, "gross_margin": gross_margin,
        "net_margin": net_margin, "debt_equity": debt_equity,
        "cash_conv": cash_conv, "market_cap": market_cap,
        "sector":   info.get("sector", "") or "",
        "industry": info.get("industry", "") or "",
        "currency": info.get("currency", "") or "",
        "score": 0,
    }
    score = 0
    if roe          and roe          > 0.15: score += 20
    if gross_margin and gross_margin > 0.20: score += 20
    if debt_equity  is not None and 0 <= debt_equity < 2.0: score += 20
    if cash_conv    and cash_conv    > 0.80: score += 20
    if pe           and 0 < pe       < 25:   score += 20
    res["score"] = score
    return res

# ─────────────────────────────────────────────────────────────
# INTRINSIC VALUE CALCULATOR
# ─────────────────────────────────────────────────────────────

def calc_dcf(eps, growth_rate, discount_rate, terminal_pe, years=10):
    """Simple DCF based on EPS growth."""
    if not eps or eps <= 0:
        return None
    try:
        future_eps = eps * ((1 + growth_rate) ** years)
        terminal_value = future_eps * terminal_pe
        # Discount back to present
        pv = terminal_value / ((1 + discount_rate) ** years)
        # Add PV of dividends/earnings stream
        pv_earnings = sum(
            eps * ((1 + growth_rate) ** t) / ((1 + discount_rate) ** t)
            for t in range(1, years + 1)
        )
        return round(pv + pv_earnings, 2)
    except Exception:
        return None

def calc_graham_number(eps, bvps):
    """Graham Number = sqrt(22.5 × EPS × Book Value per Share)"""
    try:
        if eps and bvps and eps > 0 and bvps > 0:
            return round(math.sqrt(22.5 * eps * bvps), 2)
        return None
    except Exception:
        return None

def calc_pe_intrinsic(eps, sector_pe):
    """Industry average P/E × EPS"""
    try:
        if eps and sector_pe and eps > 0:
            return round(eps * sector_pe, 2)
        return None
    except Exception:
        return None

SECTOR_PE_AVERAGES = {
    "Technology": 28, "Healthcare": 22, "Financials": 13,
    "Consumer Discretionary": 20, "Consumer Staples": 18,
    "Energy": 12, "Materials": 15, "Industrials": 18,
    "Utilities": 16, "Real Estate": 22,
    "Communication Services": 20, "Other": 17,
}

# ─────────────────────────────────────────────────────────────
# TECHNICAL INDICATORS
# ─────────────────────────────────────────────────────────────

def _to_series(x) -> pd.Series:
    if isinstance(x, pd.DataFrame):
        x = x.iloc[:, 0]
    return pd.to_numeric(x, errors="coerce")

def calc_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    delta = series.diff()
    gain  = delta.clip(lower=0).rolling(period).mean()
    loss  = (-delta.clip(upper=0)).rolling(period).mean()
    rs    = gain / loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

def calc_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in list(df.columns):
        s = df[col]
        if isinstance(s, pd.DataFrame):
            df[col] = s.iloc[:, 0]
        df[col] = pd.to_numeric(df[col], errors="coerce")
    close  = _to_series(df["Close"])
    volume = _to_series(df["Volume"])
    df["MA50"]     = close.rolling(50).mean()
    df["MA200"]    = close.rolling(200).mean()
    df["RSI"]      = calc_rsi(close)
    df["Vol20"]    = volume.rolling(20).mean()
    ma20           = close.rolling(20).mean()
    std20          = close.rolling(20).std()
    df["BB_upper"] = ma20 + 2 * std20
    df["BB_lower"] = ma20 - 2 * std20
    return df

def detect_signals(df: pd.DataFrame) -> dict:
    if df.empty or len(df) < 50:
        return {}
    latest = df.iloc[-1]
    prev   = df.iloc[-2]
    rsi    = latest.get("RSI", np.nan)
    sigs   = {
        "golden_cross":   bool(latest.get("MA50",0) > latest.get("MA200",0) and
                               prev.get("MA50",0)   <= prev.get("MA200",0)),
        "ma50_above_200": bool(latest.get("MA50",0) > latest.get("MA200",0)),
        "rsi":            round(float(rsi),1) if not np.isnan(rsi) else None,
        "rsi_ok":         bool(40 <= rsi <= 65) if not np.isnan(rsi) else False,
        "volume_spike":   bool(latest.get("Volume",0) > 1.5 * latest.get("Vol20",0) and
                               latest.get("Close",0)  > prev.get("Close",0)),
    }
    if len(df) >= 252:
        high_52 = df["High"].rolling(252).max().iloc[-1]
        pct     = (high_52 - latest["Close"]) / high_52 * 100
        sigs["pct_from_52w_high"] = round(float(pct), 1)
        sigs["near_breakout"]     = bool(pct < 10)
    else:
        sigs["pct_from_52w_high"] = None
        sigs["near_breakout"]     = False
    sigs["setup_score"] = sum([
        sigs["ma50_above_200"], sigs["rsi_ok"],
        sigs["volume_spike"],   sigs["near_breakout"],
        sigs["golden_cross"],
    ])
    return sigs

# ─────────────────────────────────────────────────────────────
# PAIRS CALCULATIONS
# ─────────────────────────────────────────────────────────────

def calc_spread(price_a, price_b, lookback=60):
    df = pd.DataFrame({"A": price_a, "B": price_b}).dropna()
    df["spread"]      = np.log(df["A"]) - np.log(df["B"])
    df["spread_mean"] = df["spread"].rolling(lookback).mean()
    df["spread_std"]  = df["spread"].rolling(lookback).std()
    df["zscore"]      = (df["spread"] - df["spread_mean"]) / df["spread_std"]
    return df

def pair_signal(z):
    if z < -2.0:      return "LONG A / SHORT B",  "signal-long"
    if z >  2.0:      return "SHORT A / LONG B",  "signal-short"
    if abs(z) < 0.5:  return "CLOSE / NEUTRAL",   "signal-none"
    return "HOLD / MONITOR", "signal-none"

def backtest_pair(df, entry_threshold=2.0):
    res = {"trades":0,"wins":0,"losses":0,"total_return":0.0}
    in_trade, entry_z, entry_spread = False, 0.0, 0.0
    for i in range(len(df)):
        z = df["zscore"].iloc[i]
        s = df["spread"].iloc[i]
        if np.isnan(z): continue
        if not in_trade:
            if abs(z) >= entry_threshold:
                in_trade, entry_z, entry_spread = True, z, s
        else:
            if (entry_z > 0 and z <= 0) or (entry_z < 0 and z >= 0):
                res["trades"]       += 1
                res["wins"]         += 1
                res["total_return"] += abs(entry_spread - s)
                in_trade = False
            elif abs(z) >= 3.5:
                res["trades"] += 1
                res["losses"] += 1
                in_trade = False
    res["win_rate"] = res["wins"] / res["trades"] * 100 if res["trades"] else 0
    return res

# ─────────────────────────────────────────────────────────────
# NAVIGATION HEADER (replaces sidebar)
# ─────────────────────────────────────────────────────────────

# ── Handle Stripe success redirect ───────────────────────────
_qp_stripe = st.query_params.get("stripe_session", "")
if _qp_stripe and "fintiq_user" in st.session_state:
    _uid = st.session_state["fintiq_user"].get("id", "")
    if _verify_stripe_session(_qp_stripe, _uid):
        st.query_params.clear()
        st.success("🎉 Welcome to Fintiq Pro! Unlimited searches unlocked.")
        st.rerun()
    else:
        st.query_params.clear()

# ── Nav bar — Login button lives inside the HTML ─────────────
_qp_action = st.query_params.get("action", "")

_qp_t_current = st.query_params.get("_t", "")
_pricing_href = f"?page=pricing&_t={_qp_t_current}" if _qp_t_current else "?page=pricing"
_pricing_link = (
    f'<a href="{_pricing_href}" style="background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.4);'
    'color:#F59E0B;padding:5px 18px;border-radius:20px;font-size:0.8rem;font-weight:600;'
    'text-decoration:none;letter-spacing:0.3px">Pricing</a>'
)

if _user_email:
    _is_pro = st.session_state.get("fintiq_user", {}).get("is_pro", False)
    _pro_badge = (' <span style="background:#F59E0B;color:#0F1923;font-size:0.65rem;font-weight:800;'
                  'padding:1px 7px;border-radius:8px;vertical-align:middle">PRO</span>'
                  if _is_pro else "")
    _nav_right_html = (
        _pricing_link +
        f'<span style="color:#94A3B8;font-size:0.8rem;margin-right:8px">👤 {_user_email}{_pro_badge}</span>'
        '<a href="?_logout=1" target="_self" style="background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.4);'
        'color:#F59E0B;padding:5px 16px;border-radius:20px;font-size:0.78rem;font-weight:600;'
        'text-decoration:none;white-space:nowrap">Logout</a>'
    )
else:
    # If arriving from Stripe redirect, carry the stripe_session through the login URL
    # so it's still available as a query param after the user logs in
    _login_qs = st.query_params.get("stripe_session", "")
    _login_href = f"?action=login&stripe_session={_login_qs}" if _login_qs else "?action=login"
    _nav_right_html = (
        _pricing_link +
        f'<a href="{_login_href}" style="background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.4);'
        'color:#F59E0B;padding:5px 18px;border-radius:20px;font-size:0.8rem;font-weight:600;'
        'text-decoration:none;letter-spacing:0.3px">Login</a>'
    )

_candle_svg = (
    '<svg width="42" height="36" viewBox="0 0 42 36" xmlns="http://www.w3.org/2000/svg" '
    'style="display:inline-block;vertical-align:middle;margin-right:8px">'
    # Candle 1 — left, lowest
    '<line x1="6" y1="27" x2="6" y2="24" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    '<rect x="2.5" y="24" width="7" height="7" fill="#22C55E" rx="1"/>'
    '<line x1="6" y1="31" x2="6" y2="33" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    # Candle 2 — middle
    '<line x1="21" y1="15" x2="21" y2="12" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    '<rect x="17.5" y="12" width="7" height="10" fill="#22C55E" rx="1"/>'
    '<line x1="21" y1="22" x2="21" y2="25" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    # Candle 3 — right, highest
    '<line x1="36" y1="3" x2="36" y2="1" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    '<rect x="32.5" y="3" width="7" height="14" fill="#22C55E" rx="1"/>'
    '<line x1="36" y1="17" x2="36" y2="20" stroke="#22C55E" stroke-width="2" stroke-linecap="round"/>'
    '</svg>'
)

_nav_html = (
    '<div class="fintiq-nav">'
    '<div style="display:flex;align-items:center;gap:14px">'
    '<div class="fintiq-logo">' + _candle_svg + 'Fintiq</div>'
    '<div class="fintiq-tagline">From speculation to strategy · Alpha Securities Intelligence</div>'
    '</div>'
    '<div style="display:flex;align-items:center;gap:10px">'
    + _nav_right_html +
    '</div></div>'
)
st.markdown(_nav_html, unsafe_allow_html=True)

# ── Pricing page (?page=pricing) ─────────────────────────────
_qp_page = st.query_params.get("page", "")
if _qp_page == "pricing":
    _pu = st.session_state.get("fintiq_user", {})
    _pu_email = _pu.get("email", "")
    _pu_id    = _pu.get("id", "")
    _pu_pro   = _pu.get("is_pro", False)

    st.markdown("""
    <div style="max-width:760px;margin:32px auto 0 auto;text-align:center">
      <div style="font-size:2.2rem;font-weight:900;color:#F59E0B;margin-bottom:6px">
        Simple, transparent pricing</div>
      <div style="color:#64748B;font-size:0.95rem;margin-bottom:36px">
        Start free · No card required · Cancel anytime</div>
    </div>
    """, unsafe_allow_html=True)

    _, _pc, _ = st.columns([1, 6, 1])
    with _pc:
        _col_free, _col_pro = st.columns(2, gap="large")

        with _col_free:
            st.markdown("""
            <div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.3);
                border-radius:16px;padding:28px;height:100%">
              <div style="color:#94A3B8;font-size:0.8rem;font-weight:700;
                  letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Free</div>
              <div style="font-size:2.4rem;font-weight:900;color:#F1F5F9;margin-bottom:4px">
                £0</div>
              <div style="color:#64748B;font-size:0.85rem;margin-bottom:24px">forever</div>
              <hr style="border-color:rgba(100,116,139,0.2);margin-bottom:20px">
              <div style="color:#CBD5E1;font-size:0.88rem;line-height:2">
                ✓ &nbsp;2 searches as guest<br>
                ✓ &nbsp;5 searches / month after sign-up<br>
                ✓ &nbsp;All global markets<br>
                ✓ &nbsp;Quality Value screening<br>
                ✓ &nbsp;Catalyst alerts<br>
                ✗ &nbsp;<span style="color:#475569">Unlimited searches</span><br>
                ✗ &nbsp;<span style="color:#475569">Pairs trading</span><br>
                ✗ &nbsp;<span style="color:#475569">Trading journal</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

        with _col_pro:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#1a2d1a,#0D1F33);
                border:1.5px solid rgba(245,158,11,0.6);
                border-radius:16px;padding:28px;position:relative;height:100%">
              <div style="position:absolute;top:-12px;left:50%;transform:translateX(-50%);
                  background:#F59E0B;color:#0F1923;font-size:0.72rem;font-weight:800;
                  padding:3px 16px;border-radius:12px;white-space:nowrap">MOST POPULAR</div>
              <div style="color:#F59E0B;font-size:0.8rem;font-weight:700;
                  letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Pro</div>
              <div style="display:flex;align-items:baseline;gap:8px;margin-bottom:4px">
                <span style="font-size:2.4rem;font-weight:900;color:#F1F5F9">£10</span>
                <span style="color:#64748B;font-size:0.85rem">/month</span>
              </div>
              <div style="color:#4ADE80;font-size:0.82rem;margin-bottom:24px">
                or £100/year — save 2 months free</div>
              <hr style="border-color:rgba(245,158,11,0.2);margin-bottom:20px">
              <div style="color:#CBD5E1;font-size:0.88rem;line-height:2">
                ✓ &nbsp;<b style="color:#F59E0B">Unlimited</b> searches<br>
                ✓ &nbsp;All global markets<br>
                ✓ &nbsp;Quality Value screening<br>
                ✓ &nbsp;Catalyst alerts<br>
                ✓ &nbsp;<b style="color:#F59E0B">Monte Carlo simulation</b> (10,000 paths)<br>
                ✓ &nbsp;<b style="color:#F59E0B">Portfolio Optimiser</b> (Sharpe, MPT)<br>
                ✓ &nbsp;<b style="color:#F59E0B">Pairs trading</b><br>
                ✓ &nbsp;<b style="color:#F59E0B">Trading journal &amp; P&amp;L</b><br>
                ✓ &nbsp;Priority data refresh<br>
                ✓ &nbsp;Cancel anytime
              </div>
            </div>
            """, unsafe_allow_html=True)

        # ── All 3 buttons in one row ──────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        if _pu_pro:
            st.button("You're on Pro ⭐", use_container_width=True,
                      disabled=True, key="price_pro_cur")
        else:
            _btn_free, _btn_mo, _btn_yr = st.columns(3, gap="medium")
            with _btn_free:
                if not _pu_email:
                    if st.button("Sign up free →", use_container_width=True, key="price_signup"):
                        st.query_params.clear()
                        st.query_params["action"] = "login"
                        st.rerun()
                else:
                    st.button("Free plan ✓", use_container_width=True,
                              disabled=True, key="price_free_cur")
            with _btn_mo:
                if st.button("Monthly — £10/mo", use_container_width=True,
                             type="primary", key="price_monthly"):
                    if not _pu_email:
                        st.query_params.clear()
                        st.query_params["action"] = "login"
                        st.rerun()
                    else:
                        url = _create_checkout("monthly", _pu_email, _pu_id)
                        if url:
                            st.session_state["_checkout_ready"] = url
                        else:
                            st.error("Could not start checkout — check Stripe keys.")
            with _btn_yr:
                if st.button("Annual — £100/yr ⭐", use_container_width=True,
                             key="price_annual"):
                    if not _pu_email:
                        st.query_params.clear()
                        st.query_params["action"] = "login"
                        st.rerun()
                    else:
                        url = _create_checkout("annual", _pu_email, _pu_id)
                        if url:
                            st.session_state["_checkout_ready"] = url
                        else:
                            st.error("Could not start checkout — check Stripe keys.")
            if "_checkout_ready" in st.session_state:
                st.success("✅ Checkout ready!")
                st.link_button("🔒 Proceed to Secure Stripe Payment →",
                               st.session_state["_checkout_ready"],
                               use_container_width=True, type="primary")

    st.markdown("""
    <div style="text-align:center;color:#334155;font-size:0.8rem;margin-top:32px">
      🔒 Payments secured by Stripe · Cancel anytime from your account ·
      <a href="/" style="color:#3B82F6;text-decoration:none">← Back to screener</a>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Post-Stripe banner (stripe_session in URL but user not logged in) ──
_banner_stripe_session = st.query_params.get("stripe_session", "")
if _banner_stripe_session and not _user_email:
    st.success("✅ Payment received! Log in below to activate your Fintiq Pro account.")

# ── Login / Sign-up form (shown when ?action=login OR stripe_session present) ──
if (_qp_action == "login" or _banner_stripe_session) and not _user_email:
    st.markdown("<hr style='border-color:rgba(245,158,11,0.15);margin:6px 0 16px 0'>",
                unsafe_allow_html=True)
    _, _lc, _ = st.columns([1, 2, 1])
    with _lc:
        _lf_mode = st.radio("", ["Login", "Sign up"], horizontal=True,
                            label_visibility="collapsed", key="lf_mode")
        st.markdown(f"#### {'Welcome back' if _lf_mode == 'Login' else 'Create your free account'}")
        _lf_email = st.text_input("Email", placeholder="you@example.com", key="lf_email")
        _lf_pw    = st.text_input("Password", type="password",
                                  placeholder="Min 6 characters", key="lf_pw")
        if _lf_mode == "Sign up":
            _lf_pw2 = st.text_input("Confirm password", type="password",
                                    placeholder="Repeat password", key="lf_pw2")
        _lfa, _lfb = st.columns(2)
        with _lfa:
            if st.button("Login" if _lf_mode == "Login" else "Create account",
                         use_container_width=True, type="primary", key="lf_submit"):
                if not _lf_email or not _lf_pw:
                    st.error("Please fill in all fields.")
                elif _sb is None:
                    st.error("Auth service unavailable.")
                else:
                    try:
                        if _lf_mode == "Sign up":
                            if _lf_pw != st.session_state.get("lf_pw2", ""):
                                st.error("Passwords do not match.")
                            elif len(_lf_pw) < 6:
                                st.error("Password must be at least 6 characters.")
                            else:
                                res = _sb.auth.sign_up({"email": _lf_email, "password": _lf_pw})
                                if res.user:
                                    try:
                                        lr = _sb.auth.sign_in_with_password({"email": _lf_email, "password": _lf_pw})
                                        if lr.user:
                                            _tok = lr.session.access_token if lr.session else None
                                            st.session_state["fintiq_user"] = {"email": lr.user.email, "id": lr.user.id, "session": _tok}
                                            if _tok:
                                                st.query_params["_t"] = _tok
                                            st.rerun()
                                    except Exception:
                                        st.success("✅ Account created! Please log in.")
                                else:
                                    st.error("Sign-up failed. Please try again.")
                        else:
                            res = _sb.auth.sign_in_with_password({"email": _lf_email, "password": _lf_pw})
                            if res.user:
                                _tok3 = res.session.access_token if res.session else None
                                st.session_state["fintiq_user"] = {"email": res.user.email, "id": res.user.id}
                                st.session_state["free_searches"] = 0
                                # Process Stripe payment carried through login URL
                                _qp_ss = st.query_params.get("stripe_session", "") or st.session_state.pop("_pending_stripe_session", "")
                                if _qp_ss:
                                    _verify_stripe_session(_qp_ss, res.user.id)
                                _next_page = st.query_params.get("next", "")
                                st.query_params.clear()
                                if _tok3:
                                    st.query_params["_t"] = _tok3
                                if _next_page:
                                    st.query_params["page"] = _next_page
                                st.rerun()
                            else:
                                st.error("Invalid email or password.")
                    except Exception as _e:
                        _em = str(_e)
                        if "Email not confirmed" in _em:
                            st.warning("Please confirm your email — check your inbox.")
                        elif "Invalid login" in _em or "invalid_grant" in _em:
                            st.error("Invalid email or password.")
                        else:
                            st.error(f"Error: {_em}")
        with _lfb:
            if st.button("Cancel", use_container_width=True, key="lf_cancel"):
                st.query_params.clear()
                st.rerun()
    st.stop()  # don't render the rest of the app behind the login form

# ── Background: CSS pseudo-element on stApp — most reliable Streamlit approach ──
st.markdown("""
<style>
[data-testid="stAppViewContainer"]::before {
  content: "";
  position: fixed;
  top: 0; left: 0; width: 100%; height: 100%;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 1400 900' preserveAspectRatio='xMidYMid slice'%3E%3Cdefs%3E%3CradialGradient id='g1' cx='70%25' cy='30%25' r='50%25'%3E%3Cstop offset='0%25' stop-color='%2300D4FF' stop-opacity='0.18'/%3E%3Cstop offset='100%25' stop-color='%23001830' stop-opacity='0'/%3E%3C/radialGradient%3E%3C/defs%3E%3Crect width='1400' height='900' fill='url(%23g1)'/%3E%3Ccircle cx='120' cy='680' r='5' fill='%2300D4FF'/%3E%3Ccircle cx='280' cy='580' r='4' fill='%2300D4FF'/%3E%3Ccircle cx='450' cy='490' r='6' fill='%2300D4FF'/%3E%3Ccircle cx='600' cy='540' r='4' fill='%2300D4FF'/%3E%3Ccircle cx='750' cy='420' r='5' fill='%2300D4FF'/%3E%3Ccircle cx='900' cy='360' r='7' fill='%2300D4FF'/%3E%3Ccircle cx='1050' cy='280' r='4' fill='%2300D4FF'/%3E%3Ccircle cx='1180' cy='200' r='6' fill='%2300D4FF'/%3E%3Ccircle cx='340' cy='720' r='3' fill='%2300D4FF'/%3E%3Ccircle cx='680' cy='650' r='5' fill='%2300D4FF'/%3E%3Ccircle cx='980' cy='500' r='4' fill='%2300D4FF'/%3E%3Ccircle cx='1240' cy='380' r='5' fill='%2300D4FF'/%3E%3Ccircle cx='200' cy='400' r='4' fill='%2300D4FF'/%3E%3Ccircle cx='560' cy='300' r='6' fill='%2300D4FF'/%3E%3Cline x1='120' y1='680' x2='280' y2='580' stroke='%2300D4FF' stroke-width='0.8' opacity='0.6'/%3E%3Cline x1='280' y1='580' x2='450' y2='490' stroke='%2300D4FF' stroke-width='0.8' opacity='0.6'/%3E%3Cline x1='450' y1='490' x2='750' y2='420' stroke='%2300D4FF' stroke-width='0.8' opacity='0.5'/%3E%3Cline x1='750' y1='420' x2='900' y2='360' stroke='%2300D4FF' stroke-width='1' opacity='0.6'/%3E%3Cline x1='900' y1='360' x2='1050' y2='280' stroke='%2300D4FF' stroke-width='0.8' opacity='0.6'/%3E%3Cline x1='1050' y1='280' x2='1240' y2='380' stroke='%2300D4FF' stroke-width='0.5' opacity='0.3'/%3E%3Cg stroke='%2300D4FF' stroke-width='1.5' opacity='0.8'%3E%3Cline x1='80' y1='730' x2='80' y2='620'/%3E%3Crect x='72' y='660' width='16' height='40' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='180' y1='690' x2='180' y2='570'/%3E%3Crect x='172' y='615' width='16' height='45' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='280' y1='635' x2='280' y2='510'/%3E%3Crect x='272' y='558' width='16' height='48' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='380' y1='575' x2='380' y2='455'/%3E%3Crect x='372' y='502' width='16' height='44' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='480' y1='520' x2='480' y2='390'/%3E%3Crect x='472' y='445' width='16' height='45' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='580' y1='465' x2='580' y2='340'/%3E%3Crect x='572' y='390' width='16' height='46' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='680' y1='480' x2='680' y2='340'/%3E%3Crect x='672' y='400' width='16' height='52' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='780' y1='420' x2='780' y2='285'/%3E%3Crect x='772' y='345' width='16' height='45' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='880' y1='375' x2='880' y2='245'/%3E%3Crect x='872' y='298' width='16' height='46' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='980' y1='355' x2='980' y2='225'/%3E%3Crect x='972' y='280' width='16' height='46' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='1080' y1='310' x2='1080' y2='190'/%3E%3Crect x='1072' y='240' width='16' height='44' fill='%2300D4FF' fill-opacity='0.25'/%3E%3Cline x1='1180' y1='290' x2='1180' y2='170'/%3E%3Crect x='1172' y='220' width='16' height='42' fill='%23E0F8FF' fill-opacity='0.15'/%3E%3Cline x1='1280' y1='255' x2='1280' y2='140'/%3E%3Crect x='1272' y='188' width='16' height='42' fill='%2300D4FF' fill-opacity='0.25'/%3E%3C/g%3E%3Cpath d='M80%2C680 C200%2C620 350%2C550 480%2C490 S700%2C400 900%2C340 S1150%2C240 1380%2C140' fill='none' stroke='%2300D4FF' stroke-width='2' opacity='0.5'/%3E%3C/svg%3E");
  background-size: cover;
  background-position: center;
  opacity: 0.12;
  pointer-events: none;
  z-index: 0;
}
</style>
""", unsafe_allow_html=True)

# ── Live market ticker — indices, commodities, crypto, FX ─────
_MARKET_ITEMS = [
    ("^FTSE",   "FTSE 100",  "pt",    1),     # points
    ("^GSPC",   "S&P 500",   "pt",    1),
    ("^DJI",    "Dow Jones", "pt",    1),
    ("^IXIC",   "Nasdaq",    "pt",    1),
    ("^FTAS",   "FTSE All",  "pt",    1),
    ("GC=F",    "Gold",      "$/oz",  1),
    ("CL=F",    "Oil (WTI)", "$/bbl", 1),
    ("BTC-USD", "Bitcoin",   "$",     1),
    ("ETH-USD", "Ethereum",  "$",     1),
    ("GBPUSD=X","GBP/USD",   "",      1),
    ("EURUSD=X","EUR/USD",   "",      1),
]
_ticker_html_items = []
for _sym, _label, _unit, _ in _MARKET_ITEMS:
    try:
        _ti  = yf.Ticker(_sym).fast_info
        _px  = _f(getattr(_ti, "last_price", None))
        _pcp = _f(getattr(_ti, "three_month_return", None))   # fallback
        # prefer day change
        try:
            _prev = _f(getattr(_ti, "previous_close", None))
            if _px and _prev and _prev != 0:
                _pcp = (_px / _prev - 1) * 100
        except Exception:
            pass
        if _px:
            # format price
            if _sym in ("GBPUSD=X","EURUSD=X"):
                _px_str = f"{_px:.4f}"
            elif _px >= 1000:
                _px_str = f"{_px:,.0f}"
            elif _px >= 10:
                _px_str = f"{_px:,.2f}"
            else:
                _px_str = f"{_px:.4f}"
            if _unit:
                _px_str = f"{_px_str} {_unit}"
            _pcp = _pcp or 0
            _cls = "t-up" if _pcp >= 0 else "t-dn"
            _arr = "▲" if _pcp >= 0 else "▼"
            _ticker_html_items.append(
                f'<span class="ticker-item">'
                f'<span class="t-sym">{_label}</span> '
                f'{_px_str} <span class="{_cls}">{_arr}{abs(_pcp):.2f}%</span>'
                f'</span>'
            )
    except Exception:
        pass

if _ticker_html_items:
    _tape = "".join(_ticker_html_items * 3)   # triple for seamless loop
    st.markdown(f'<div class="ticker-tape"><span class="ticker-tape-inner">{_tape}</span></div>',
                unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────

tab0, tab_brief, tab1, tab2, tab3, tab_mc, tab5, tab_opt, tab4 = st.tabs([
    "🏠  Home",
    "🌍  Morning Brief",
    "🔍  Fundamental",
    "⚡  Catalyst",
    "📈  Technical",
    "🎲  Monte Carlo",
    "📒  Journal",
    "📐  Optimiser",
    "⚖️  Pairs Trading",
])

# ═══════════════════════════════════════════════════════════════
# TAB 0 — HOME / LANDING PAGE
# ═══════════════════════════════════════════════════════════════

with tab0:
    st.markdown("""
    <div style="max-width:920px;margin:0 auto;padding:10px 0">

      <!-- Hero — title + tagline only, clean and fully visible above fold -->
      <div style="padding:32px 40px 36px;
                  background:linear-gradient(135deg,rgba(10,22,40,0.95) 0%,rgba(8,18,32,0.99) 100%);
                  border-radius:20px;border:1px solid rgba(245,158,11,0.25);
                  box-shadow:0 8px 50px rgba(0,0,0,0.7);margin-bottom:28px;text-align:center">
        <div class="fintiq-hero-title">Invest like an institution</div>
        <div style="font-size:1rem;color:#94A3B8;max-width:680px;margin:12px auto 0;line-height:1.7">
          Fintiq doesn't just screen stocks — it guides your entire investment decision.
          Find quality businesses · Value them with <strong style="color:#F1F5F9">DCF, Graham &amp; Industry models</strong> ·
          Map statistical price scenarios with <strong style="color:#F1F5F9">Monte Carlo simulation</strong> ·
          Confirm timing with Technical analysis · Optimise allocation with
          <strong style="color:#F1F5F9">Nobel Prize-winning MPT</strong> —
          <strong style="color:#F59E0B">every signal in one place, one decision at a time.</strong>
        </div>
      </div>

      <!-- Feature stat grid — below the fold, scroll to discover -->
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:28px" class="fiq-stat-grid">
        <div style="background:rgba(245,158,11,0.07);border:1px solid rgba(245,158,11,0.25);
                    border-radius:12px;padding:18px;text-align:center">
          <div style="font-size:2rem;font-weight:900;color:#F59E0B">10+</div>
          <div style="font-size:0.8rem;color:#94A3B8;margin-top:4px">Global Markets<br>incl. LSE, NYSE, NASDAQ</div>
        </div>
        <div style="background:rgba(74,222,128,0.07);border:1px solid rgba(74,222,128,0.25);
                    border-radius:12px;padding:18px;text-align:center">
          <div style="font-size:2rem;font-weight:900;color:#4ADE80">3</div>
          <div style="font-size:0.8rem;color:#94A3B8;margin-top:4px">Valuation Methods<br>DCF · Graham · Industry P/E</div>
        </div>
        <div style="background:rgba(96,165,250,0.07);border:1px solid rgba(96,165,250,0.25);
                    border-radius:12px;padding:18px;text-align:center">
          <div style="font-size:2rem;font-weight:900;color:#60A5FA">Live</div>
          <div style="font-size:0.8rem;color:#94A3B8;margin-top:4px">Market Data<br>Powered by Yahoo Finance</div>
        </div>
        <div style="background:rgba(167,139,250,0.07);border:1px solid rgba(167,139,250,0.25);
                    border-radius:12px;padding:18px;text-align:center">
          <div style="font-size:2rem;font-weight:900;color:#A78BFA">🎲</div>
          <div style="font-size:0.8rem;color:#94A3B8;margin-top:4px">Monte Carlo<br>10,000 price simulations</div>
        </div>
      </div>

      <!-- Philosophy -->
      <div style="background:rgba(13,31,53,0.7);border-radius:14px;padding:28px 32px;
                  border-left:4px solid #F59E0B;margin-bottom:28px">
        <div style="color:#F59E0B;font-weight:700;font-size:1.1rem;margin-bottom:12px">
          💡 Our Philosophy — More Than a Stock Screener
        </div>
        <div style="color:#CBD5E1;line-height:1.85;font-size:0.95rem">
          Most retail investors lose money not because markets are unfair, but because they act on
          <em>one signal in isolation</em> — a hot tip, a moving average, a gut feeling.
          Fintiq was built on one conviction:
          <strong style="color:#F1F5F9">process beats impulse, every time.</strong><br><br>
          We believe in owning high-quality businesses at reasonable prices, backed by multiple
          independent lines of evidence before a single pound or dollar is committed.
          Fintiq is designed to smooth the decision — taking you from
          <em style="color:#F59E0B">"is this stock interesting?"</em> all the way to
          <em style="color:#F59E0B">"here is the probability-weighted case for entering, and here is how much to size"</em>
          in one seamless workflow. No spreadsheets, no fragmented tools, no guesswork.
        </div>
      </div>

    </div>
    """, unsafe_allow_html=True)

    # ── Decision Framework — separate st.markdown to avoid Markdown code-block heuristic ──
    st.markdown("""<div style="max-width:920px;margin:0 auto">
<div style="color:#F59E0B;font-weight:700;font-size:1.1rem;margin-bottom:16px">🧭 The Fintiq Investment Decision Framework</div>
<div style="background:rgba(13,31,53,0.6);border-radius:14px;padding:24px 28px;margin-bottom:28px;border:1px solid rgba(245,158,11,0.15)">
<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:0;align-items:stretch">
<div style="background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.3);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">🔍</div><div style="color:#F59E0B;font-weight:700;font-size:0.85rem;margin-bottom:6px">Step 1<br>Screen</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">Filter 10+ global markets for quality: ROE, margins, debt, cash flow</div><div style="font-size:0.65rem;color:#475569;margin-top:8px">Fundamental tab</div></div>
<div style="display:flex;align-items:center;justify-content:center;font-size:1.4rem;color:#475569;padding:0 4px">→</div>
<div style="background:rgba(96,165,250,0.1);border:1px solid rgba(96,165,250,0.3);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">💎</div><div style="color:#60A5FA;font-weight:700;font-size:0.85rem;margin-bottom:6px">Step 2<br>Value</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">Is the business cheap? DCF, Graham Number &amp; Industry P/E give intrinsic value</div><div style="font-size:0.65rem;color:#475569;margin-top:8px">Fundamental → Deep Dive</div></div>
<div style="display:flex;align-items:center;justify-content:center;font-size:1.4rem;color:#475569;padding:0 4px">→</div>
<div style="background:rgba(167,139,250,0.1);border:1px solid rgba(167,139,250,0.3);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">🎲</div><div style="color:#A78BFA;font-weight:700;font-size:0.85rem;margin-bottom:6px">Step 3<br>Simulate</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">Run 10,000 Monte Carlo paths. What prices are statistically possible?</div><div style="font-size:0.65rem;color:#475569;margin-top:8px">Monte Carlo tab</div></div>
</div>
<div style="display:flex;justify-content:center;margin:8px 0;font-size:1.4rem;color:#475569">↓</div>
<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:0;align-items:stretch">
<div style="background:rgba(74,222,128,0.1);border:1px solid rgba(74,222,128,0.3);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">📈</div><div style="color:#4ADE80;font-weight:700;font-size:0.85rem;margin-bottom:6px">Step 4<br>Time</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">Technical chart + catalyst alerts — when is the right moment to enter?</div><div style="font-size:0.65rem;color:#475569;margin-top:8px">Technical + Catalyst tabs</div></div>
<div style="display:flex;align-items:center;justify-content:center;font-size:1.4rem;color:#475569;padding:0 4px">→</div>
<div style="background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.3);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">📐</div><div style="color:#F59E0B;font-weight:700;font-size:0.85rem;margin-bottom:6px">Step 5<br>Size</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">MPT Portfolio Optimiser — how much to allocate to each position?</div><div style="font-size:0.65rem;color:#475569;margin-top:8px">Optimiser tab ⭐ Pro</div></div>
<div style="display:flex;align-items:center;justify-content:center;font-size:1.4rem;color:#475569;padding:0 4px">→</div>
<div style="background:rgba(34,197,94,0.15);border:2px solid rgba(34,197,94,0.4);border-radius:12px;padding:18px 14px;text-align:center"><div style="font-size:1.6rem;margin-bottom:6px">🎯</div><div style="color:#22C55E;font-weight:700;font-size:0.85rem;margin-bottom:6px">Decision<br>Dashboard</div><div style="color:#94A3B8;font-size:0.75rem;line-height:1.5">Fundamental + Monte Carlo + Technical signals combined → one recommendation</div><div style="font-size:0.65rem;color:#22C55E;margin-top:8px">Monte Carlo → Decision Dashboard</div></div>
</div>
<div style="text-align:center;margin-top:16px;font-size:0.8rem;color:#64748B">Fintiq is not just a screener — it guides you from discovery all the way to a conviction decision.</div>
</div>
</div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style="max-width:920px;margin:0 auto;padding:0">

      <!-- Three Strategies -->
      <div style="color:#F59E0B;font-weight:700;font-size:1.1rem;margin-bottom:16px">
        📋 Three Strategies, One Platform
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:28px" class="fiq-3col-grid">
        <div style="background:rgba(13,31,53,0.7);border-radius:12px;padding:22px;
                    border:1px solid rgba(245,158,11,0.15);border-top:3px solid #F59E0B">
          <div style="font-size:1.5rem;margin-bottom:8px">🔍</div>
          <div style="color:#F1F5F9;font-weight:700;font-size:1rem;margin-bottom:8px">
            Strategy 1 — Quality Value Screen
          </div>
          <div style="color:#94A3B8;font-size:0.88rem;line-height:1.65">
            Screen for businesses with strong ROE, healthy margins, manageable debt and
            consistent cash conversion. Run a 3-method intrinsic value analysis (DCF,
            Graham Number, Industry P/E) to find when quality is priced attractively.
          </div>
        </div>
        <div style="background:rgba(13,31,53,0.7);border-radius:12px;padding:22px;
                    border:1px solid rgba(245,158,11,0.15);border-top:3px solid #60A5FA">
          <div style="font-size:1.5rem;margin-bottom:8px">⚡</div>
          <div style="color:#F1F5F9;font-weight:700;font-size:1rem;margin-bottom:8px">
            Strategy 2 — Catalyst Alerts
          </div>
          <div style="color:#94A3B8;font-size:0.88rem;line-height:1.65">
            Track upcoming earnings, analyst upgrades, insider buying and technical breakouts
            for your screened universe. A good business at a fair price with a near-term
            catalyst is the highest-conviction setup.
          </div>
        </div>
        <div style="background:rgba(13,31,53,0.7);border-radius:12px;padding:22px;
                    border:1px solid rgba(245,158,11,0.15);border-top:3px solid #4ADE80">
          <div style="font-size:1.5rem;margin-bottom:8px">⚖️</div>
          <div style="color:#F1F5F9;font-weight:700;font-size:1rem;margin-bottom:8px">
            Strategy 3 — Statistical Pairs
          </div>
          <div style="color:#94A3B8;font-size:0.88rem;line-height:1.65">
            Identify cointegrated stock pairs whose price ratio has diverged beyond
            statistical norms. Go long the underperformer, short the outperformer, and
            earn from mean-reversion — market-neutral, lower-risk alpha.
          </div>
        </div>
      </div>

      <!-- Portfolio Optimizer Feature Highlight -->
      <div style="background:linear-gradient(135deg,rgba(13,31,53,0.95),rgba(8,18,32,0.99));
                  border-radius:16px;padding:28px 32px;margin-bottom:28px;
                  border:1px solid rgba(245,158,11,0.3)">
        <div style="display:flex;align-items:flex-start;gap:20px;flex-wrap:wrap">
          <div style="flex:1;min-width:260px">
            <div style="font-size:0.7rem;font-weight:700;color:#F59E0B;text-transform:uppercase;
                        letter-spacing:1.5px;margin-bottom:8px">⭐ Pro Feature</div>
            <div style="font-size:1.25rem;font-weight:900;color:#F1F5F9;margin-bottom:10px">
              📐 Portfolio Optimiser
            </div>
            <div style="color:#94A3B8;font-size:0.9rem;line-height:1.75;margin-bottom:16px">
              Once you've found great stocks, the hardest question is: <em style="color:#F1F5F9">how much to put in each?</em>
              Equal-weighting leaves return on the table. Guessing takes on hidden risk.
              The Fintiq Portfolio Optimiser uses <strong style="color:#F1F5F9">Nobel Prize-winning
              Modern Portfolio Theory</strong> to calculate the mathematically optimal allocation —
              maximising return per unit of risk for your exact set of stocks.
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:0.82rem;color:#94A3B8">
              <div>✅ 6 optimisation objectives</div>
              <div>✅ Efficient Frontier chart</div>
              <div>✅ Sharpe · Sortino · VaR</div>
              <div>✅ Max Drawdown analysis</div>
              <div>✅ Correlation matrix</div>
              <div>✅ Live weight drift monitor</div>
              <div>✅ Risk Parity (All Weather)</div>
              <div>✅ AI portfolio commentary</div>
            </div>
          </div>
          <div style="flex:0 0 220px;background:rgba(245,158,11,0.06);border:1px solid rgba(245,158,11,0.2);
                      border-radius:12px;padding:16px">
            <div style="font-size:0.7rem;color:#64748B;text-transform:uppercase;margin-bottom:10px">
              Example — Equal vs Optimised</div>
            <div style="margin-bottom:10px">
              <div style="font-size:0.75rem;color:#94A3B8;margin-bottom:3px">Equal Weight (5 stocks)</div>
              <div style="display:flex;align-items:center;gap:6px">
                <div style="flex:1;background:#1E3A5F;border-radius:3px;height:6px">
                  <div style="background:#475569;width:60%;height:6px;border-radius:3px"></div></div>
                <span style="font-size:0.8rem;color:#64748B;font-weight:700">0.6 Sharpe</span>
              </div>
            </div>
            <div style="margin-bottom:14px">
              <div style="font-size:0.75rem;color:#F59E0B;margin-bottom:3px">Optimised Weights ⭐</div>
              <div style="display:flex;align-items:center;gap:6px">
                <div style="flex:1;background:#1E3A5F;border-radius:3px;height:6px">
                  <div style="background:#F59E0B;width:92%;height:6px;border-radius:3px"></div></div>
                <span style="font-size:0.8rem;color:#F59E0B;font-weight:700">1.1 Sharpe</span>
              </div>
            </div>
            <div style="font-size:0.72rem;color:#64748B;text-align:center;border-top:1px solid rgba(100,116,139,0.2);padding-top:10px">
              Same 5 stocks · Lower volatility<br>+83% better risk-adjusted return
            </div>
          </div>
        </div>
      </div>

      <!-- GBM Monte Carlo Feature Highlight -->
      <div style="background:linear-gradient(135deg,rgba(13,31,53,0.95),rgba(8,18,32,0.99));
                  border-radius:16px;padding:28px 32px;margin-bottom:28px;
                  border:1px solid rgba(167,139,250,0.3)">
        <div style="display:flex;align-items:flex-start;gap:20px;flex-wrap:wrap">
          <div style="flex:1;min-width:260px">
            <div style="font-size:0.7rem;font-weight:700;color:#A78BFA;text-transform:uppercase;
                        letter-spacing:1.5px;margin-bottom:8px">🆕 New Feature</div>
            <div style="font-size:1.25rem;font-weight:900;color:#F1F5F9;margin-bottom:10px">
              🎲 Monte Carlo Risk Simulator
            </div>
            <div style="color:#94A3B8;font-size:0.9rem;line-height:1.75;margin-bottom:16px">
              Professional investment banks don't rely on a single price target —
              they model the <em style="color:#F1F5F9">entire distribution of possible futures</em>.
              The Fintiq Monte Carlo Simulator brings this institutional technique to retail investors.
              Run <strong style="color:#F1F5F9">10,000 simulated price paths</strong> in seconds,
              see your probability of hitting a target, and understand your real downside before you commit.
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:0.82rem;color:#94A3B8">
              <div>✅ 3 simulation models</div>
              <div>✅ Fan chart with confidence bands</div>
              <div>✅ Probability of hitting target</div>
              <div>✅ VaR &amp; Expected Shortfall</div>
              <div>✅ Jump Diffusion (crash modelling)</div>
              <div>✅ EWMA volatility clustering</div>
              <div>✅ Integrates with your DCF value</div>
              <div>✅ Decision Dashboard (3 signals)</div>
            </div>
          </div>
          <div style="flex:0 0 210px;background:rgba(167,139,250,0.06);border:1px solid rgba(167,139,250,0.2);
                      border-radius:12px;padding:16px;font-size:0.78rem">
            <div style="color:#64748B;text-transform:uppercase;font-size:0.65rem;margin-bottom:10px">
              Example output — AAPL · 1 year · 10,000 sims</div>
            <div style="margin-bottom:8px">
              <div style="color:#94A3B8;margin-bottom:2px">P(profit)</div>
              <div style="background:#1E3A5F;border-radius:3px;height:7px">
                <div style="background:#22C55E;width:63%;height:7px;border-radius:3px"></div></div>
              <div style="color:#22C55E;font-weight:700;font-size:0.82rem">63%</div>
            </div>
            <div style="margin-bottom:8px">
              <div style="color:#94A3B8;margin-bottom:2px">P(≥ target price)</div>
              <div style="background:#1E3A5F;border-radius:3px;height:7px">
                <div style="background:#A78BFA;width:38%;height:7px;border-radius:3px"></div></div>
              <div style="color:#A78BFA;font-weight:700;font-size:0.82rem">38%</div>
            </div>
            <div>
              <div style="color:#94A3B8;margin-bottom:2px">VaR 95% (max loss)</div>
              <div style="background:#1E3A5F;border-radius:3px;height:7px">
                <div style="background:#EF4444;width:22%;height:7px;border-radius:3px"></div></div>
              <div style="color:#EF4444;font-weight:700;font-size:0.82rem">-22% in worst 5%</div>
            </div>
            <div style="font-size:0.65rem;color:#475569;text-align:center;margin-top:12px;
                        border-top:1px solid rgba(100,116,139,0.2);padding-top:8px">
              Illustrative only — based on historical data
            </div>
          </div>
        </div>
      </div>

      <!-- How to use -->
      <div style="background:rgba(13,31,53,0.7);border-radius:14px;padding:28px 32px;
                  border-left:4px solid #4ADE80;margin-bottom:28px">
        <div style="color:#4ADE80;font-weight:700;font-size:1.1rem;margin-bottom:16px">
          🚀 How to Use Fintiq — Step by Step
        </div>
        <div style="color:#CBD5E1;font-size:0.9rem;line-height:1.85">
          <strong style="color:#F1F5F9">Step 1 — Run the Fundamental Screen</strong> (Fundamental tab)<br>
          Select your market (e.g. GB London LSE), set quality thresholds, and click Run Screen.
          Fintiq scores every stock and surfaces the highest-quality names.<br><br>
          <strong style="color:#F1F5F9">Step 2 — Deep-Dive &amp; Value the business</strong><br>
          Click any row in the results table to open the Stock Deep-Dive. Adjust your DCF assumptions
          and read the 3-method intrinsic value estimate. Note your target price.<br><br>
          <strong style="color:#F1F5F9">Step 3 — Run Monte Carlo Simulation</strong> (Monte Carlo tab) 🆕<br>
          Enter the ticker and paste your DCF target price as the "Target". The simulator runs 10,000
          price paths and tells you the probability of reaching your target, your VaR, and the
          Decision Dashboard combines all signals into a buy/hold/avoid verdict.<br><br>
          <strong style="color:#F1F5F9">Step 4 — Check Catalysts &amp; Technicals</strong><br>
          Your screened stocks auto-populate into Catalyst and Technical tabs. Look for
          earnings dates, analyst revisions, RSI levels and volume signals to time your entry.<br><br>
          <strong style="color:#F1F5F9">Step 5 — Optimise your allocation</strong> ⭐ Pro<br>
          Enter your chosen tickers into the Portfolio Optimiser. Select an objective
          (Maximise Sharpe, Risk Parity, Target Return and more) and get the mathematically
          optimal weight for each position.<br><br>
          <strong style="color:#F1F5F9">Step 6 — Log your trades</strong> (Journal tab)<br>
          Record entries, exits and notes. Track your equity curve and win rate over time.
        </div>
      </div>

      <!-- Disclaimer -->
      <div style="background:rgba(127,29,29,0.15);border:1px solid rgba(239,68,68,0.2);
                  border-radius:10px;padding:16px 22px;text-align:center;font-size:0.82rem;
                  color:#94A3B8">
        ⚠️ <strong style="color:#F87171">For educational and informational purposes only.</strong>
        Fintiq is not a regulated investment adviser. Nothing on this platform constitutes
        financial advice. Past performance does not guarantee future results. Always conduct
        your own due diligence before making any investment decision.
        <div style="margin-top:10px;padding-top:10px;border-top:1px solid rgba(239,68,68,0.15);
                    font-size:0.75rem;color:#475569;display:flex;justify-content:center;gap:20px;flex-wrap:wrap">
          <span><strong style="color:#F59E0B">Fintiq</strong> &nbsp;·&nbsp; © 2025 Fintiq Ltd &nbsp;·&nbsp; Registered in England &amp; Wales</span>
          <span>✉ <a href="mailto:contactfintiq@gmail.com" style="color:#64748B;text-decoration:none">contactfintiq@gmail.com</a></span>
          <span>💡 <a href="mailto:contactfintiq@gmail.com?subject=Feature%20Suggestion" style="color:#64748B;text-decoration:none">Suggest a feature</a></span>
          <a href="https://fintiq.uk" style="color:#64748B;text-decoration:none">fintiq.uk</a>
        </div>
      </div>

    </div>
    """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# TAB — MORNING BRIEF
# ═══════════════════════════════════════════════════════════════

# ── Index universe ────────────────────────────────────────────
_BRIEF_INDICES = [
    # US
    ("^GSPC",    "S&P 500",        "🇺🇸", "US"),
    ("^DJI",     "Dow Jones",      "🇺🇸", "US"),
    ("^IXIC",    "Nasdaq",         "🇺🇸", "US"),
    ("^RUT",     "Russell 2000",   "🇺🇸", "US"),
    # US Futures
    ("ES=F",     "S&P Futures",    "🇺🇸", "Futures"),
    ("NQ=F",     "Nasdaq Futures", "🇺🇸", "Futures"),
    ("YM=F",     "Dow Futures",    "🇺🇸", "Futures"),
    # UK
    ("^FTSE",    "FTSE 100",       "🇬🇧", "UK"),
    ("^FTMC",    "FTSE 250",       "🇬🇧", "UK"),
    # Europe
    ("^GDAXI",   "DAX",            "🇩🇪", "Europe"),
    ("^FCHI",    "CAC 40",         "🇫🇷", "Europe"),
    ("^STOXX50E","Euro Stoxx 50",  "🇪🇺", "Europe"),
    # Asia
    ("^N225",    "Nikkei 225",     "🇯🇵", "Asia"),
    ("000001.SS","Shanghai",       "🇨🇳", "Asia"),
    ("^HSI",     "Hang Seng",      "🇭🇰", "Asia"),
    ("^AXJO",    "ASX 200",        "🇦🇺", "Asia"),
    ("^BSESN",   "Sensex",         "🇮🇳", "Asia"),
]

_BRIEF_INSTRUMENTS = [
    ("^VIX",     "VIX",            "Fear Index"),
    ("GC=F",     "Gold",           "$/oz"),
    ("BZ=F",     "Brent Oil",      "$/bbl"),
    ("CL=F",     "WTI Oil",        "$/bbl"),
    ("DX-Y.NYB", "Dollar Index",   "DXY"),
    ("GBPUSD=X", "GBP/USD",        "FX"),
    ("EURUSD=X", "EUR/USD",        "FX"),
    ("USDJPY=X", "USD/JPY",        "FX"),
    ("^TNX",     "10Y Treasury",   "Yield %"),
]

@st.cache_data(ttl=300)
def _fetch_brief_data(tickers: list[str]) -> dict:
    """Fetch latest price + % change for a list of tickers."""
    out = {}
    for sym in tickers:
        try:
            ti = yf.Ticker(sym).fast_info
            price = getattr(ti, "last_price", None)
            prev  = getattr(ti, "previous_close", None)
            if price and prev and prev != 0:
                chg    = price - prev
                chg_pct = chg / prev * 100
            else:
                chg = chg_pct = None
            out[sym] = {"price": price, "chg": chg, "chg_pct": chg_pct}
        except Exception:
            out[sym] = {"price": None, "chg": None, "chg_pct": None}
    return out

@st.cache_data(ttl=3600)
def _fetch_econ_calendar() -> list:
    """Fetch economic calendar from FMP for next 5 days."""
    today = datetime.now().strftime("%Y-%m-%d")
    end   = (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%d")
    try:
        url = f"{FMP_BASE}/v3/economic_calendar?from={today}&to={end}&apikey={FMP_KEY}"
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return [e for e in data if e.get("impact") in ("High", "Medium")] if data else []
    except Exception:
        pass
    return []

@st.cache_data(ttl=600)
def _fetch_market_news() -> list:
    """Fetch market news from FMP general news endpoint."""
    try:
        url = f"{FMP_BASE}/v4/general_news?page=0&apikey={FMP_KEY}"
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and data:
                return data[:8]
    except Exception:
        pass
    # Fallback: yfinance news (handle both old and new schema)
    try:
        raw = yf.Ticker("^GSPC").news or []
        out = []
        for n in raw[:8]:
            # new yfinance schema nests under 'content'
            c = n.get("content", n)
            title = c.get("title", n.get("title", ""))
            url_  = (c.get("canonicalUrl", {}).get("url")
                     or c.get("clickThroughUrl", {}).get("url")
                     or n.get("link", "#"))
            pub   = (c.get("provider", {}).get("displayName")
                     or n.get("publisher", ""))
            dt    = c.get("pubDate", "") or ""
            if title:
                out.append({"title": title, "url": url_, "site": pub,
                            "publishedDate": dt})
        return out
    except Exception:
        return []

@st.cache_data(ttl=3600)
def _fetch_vix_chart():
    """Fetch 30-day VIX history for mini chart."""
    try:
        df = yf.download("^VIX", period="30d", interval="1d", progress=False, auto_adjust=True)
        if not df.empty:
            close = df["Close"].dropna()
            return close.reset_index()
    except Exception:
        pass
    return None

def _generate_risk_explanation(sentiment: str, vix: float | None,
                                spx_pct: float | None, gold_pct: float | None,
                                dxy_pct: float | None) -> str:
    """Generate Goldman-style morning intelligence brief for retail investors."""
    import datetime as _dt
    vix_str  = f"{vix:.1f}" if vix is not None else "N/A"
    spx_dir  = "up" if (spx_pct or 0) >= 0 else "down"
    spx_str  = f"{spx_dir} {abs(spx_pct or 0):.2f}%" if spx_pct is not None else "flat"
    gold_dir = "rising" if (gold_pct or 0) >= 0 else "falling"
    gold_str = f"{gold_dir} {abs(gold_pct or 0):.2f}%" if gold_pct is not None else "flat"
    dxy_dir  = "strengthening" if (dxy_pct or 0) >= 0 else "weakening"
    dxy_str  = f"{dxy_dir} {abs(dxy_pct or 0):.2f}%" if dxy_pct is not None else "flat"

    # VIX regime classification
    if vix is not None:
        if vix < 12:
            vix_regime = "extreme complacency — the market is pricing in near-zero risk"
        elif vix < 16:
            vix_regime = "low volatility — institutions are comfortable holding risk"
        elif vix < 20:
            vix_regime = "below average — calm but not euphoric"
        elif vix < 25:
            vix_regime = "above average — some hedging activity, investors cautious"
        elif vix < 30:
            vix_regime = "elevated stress — significant institutional hedging underway"
        elif vix < 40:
            vix_regime = "fear — markets pricing in a meaningful negative event"
        else:
            vix_regime = "panic — historically rare; major dislocation underway"
    else:
        vix_regime = "unavailable"

    if "Risk-On" in sentiment:
        mood = (
            "### 📊 Market Regime: Risk-On\n\n"
            "**Macro read:** The tape is constructive. Equities are bid, volatility is "
            f"subdued (VIX {vix_str} — {vix_regime}), and capital is rotating into growth "
            f"and cyclical names. S&P 500 is **{spx_str}**. Gold is **{gold_str}** — "
            f"safe-haven demand is limited. The US Dollar is **{dxy_str}**, consistent "
            "with capital flowing toward risk assets globally.\n\n"
            "**What's driving this:** In risk-on environments, investors are willing to "
            "accept uncertainty in exchange for return. Corporate earnings expectations "
            "are likely intact, credit spreads are tight, and there is no systemic shock "
            "being priced in. Liquidity is the friend of equities.\n\n"
            "---\n\n"
            "**🎯 Tactical Playbook (Risk-On)**\n\n"
            "**Where to look:**\n"
            "- **High-beta cyclicals** — Industrials, Consumer Discretionary, Materials "
            "outperform when risk appetite is strong\n"
            "- **Momentum names** — stocks breaking to 52-week highs with volume "
            "confirmation tend to continue; institutions are chasing performance\n"
            "- **Small and mid-caps** — outperform large caps in genuine risk-on; "
            "reflect domestic confidence\n"
            "- **Emerging market equities** — benefit from USD weakness and global growth optimism\n\n"
            "**What to avoid:**\n"
            "- Holding excessive cash — inflation erodes it, and you miss the move\n"
            "- Over-weighting defensives (Utilities, Staples) — they lag in rallies\n"
            "- Chasing low-quality speculative names without fundamental backing\n\n"
            "**Position sizing:** Risk-on is not an invitation to abandon discipline. "
            "Keep individual positions ≤5% of portfolio. Use the momentum to trim "
            "underperformers, not to double down on losers.\n\n"
            "---\n\n"
            "**⚠️ Key risks to monitor:**\n"
            "- A VIX spike above 20 intraday signals the mood is shifting — reassess\n"
            "- Watch the 10-year US Treasury yield: a rapid rise (>10bps in a session) "
            "can derail equity multiples fast\n"
            "- Geopolitical events or surprise macro data can flip risk-on to risk-off "
            "within hours — always know where your stops are\n\n"
            "**Goldman rule of thumb:** *In risk-on, the job is not to predict the top — "
            "it's to ride the trend with discipline and exit when the data changes, not "
            "when your emotions do.*"
        )
    elif "Risk-Off" in sentiment:
        mood = (
            "### 📊 Market Regime: Risk-Off\n\n"
            "**Macro read:** Capital is in defensive mode. Equity markets are under "
            f"pressure (S&P 500 **{spx_str}**), and the fear gauge tells the story: "
            f"VIX is at **{vix_str}** — {vix_regime}. "
            f"Gold is **{gold_str}** — the classic safe-haven bid is active. "
            f"The US Dollar is **{dxy_str}** — consistent with a flight-to-quality "
            "into USD-denominated assets.\n\n"
            "**What this means structurally:** Institutions are reducing gross exposure. "
            "Hedge funds are buying put options (insurance) or shorting futures. "
            "The cost of hedging is rising. In severe risk-off, correlations collapse "
            "to 1 — meaning nearly everything falls together before differentiation returns.\n\n"
            "---\n\n"
            "**🎯 Tactical Playbook (Risk-Off)**\n\n"
            "**Defensive positioning:**\n"
            "- **Cash is a position** — having dry powder in a drawdown is not failure, "
            "it's preparation. The best trades come after the fear peaks.\n"
            "- **Defensive sectors** — Healthcare, Consumer Staples, Utilities hold value "
            "better in downturns; revenues are non-cyclical\n"
            "- **Gold (XAU)** — historically retains or gains value in genuine risk-off; "
            "not a trade, a hedge\n"
            "- **Short-duration government bonds** — price appreciates when equities fall "
            "hard (flight-to-quality)\n\n"
            "**What professionals do right now:**\n"
            "- Tighten stop losses on existing positions — protect capital first\n"
            "- Build a watchlist of quality names you want to own cheaper\n"
            "- Do NOT average down into falling positions without a clear thesis change\n"
            "- Look for divergences: stocks that are holding up in a down tape are "
            "showing relative strength — they lead on the recovery\n\n"
            "---\n\n"
            "**⚠️ VIX as a contrarian signal:**\n"
            f"VIX at **{vix_str}**. History shows:\n"
            "- VIX 30-35: Fear is elevated — start building watchlists\n"
            "- VIX 40+: Peak panic territory — historically a buying opportunity within 3-6 months\n"
            "- VIX 80+ (Covid 2020, GFC 2008): Generational entry points for patient capital\n\n"
            "**Goldman rule of thumb:** *In risk-off, institutions don't panic — they "
            "prepare. The amateur sells into fear. The professional builds a shopping list "
            "and waits for the VIX to roll over as the signal to re-engage.*"
        )
    else:
        mood = (
            "### 📊 Market Regime: Neutral / Mixed Signals\n\n"
            "**Macro read:** The market is not giving a clean directional signal today. "
            f"VIX is at **{vix_str}** — {vix_regime}. "
            f"Equities are **{spx_str}**, gold is **{gold_str}**, "
            f"and the dollar is **{dxy_str}**. These inputs are not aligned — "
            "suggesting the market is digesting competing forces.\n\n"
            "**What creates mixed regimes:** Conflicting macro data (e.g. strong jobs "
            "but weak manufacturing), central bank uncertainty, sector rotation without "
            "a clear theme, or markets waiting on a known upcoming catalyst (Fed meeting, "
            "earnings season, geopolitical development).\n\n"
            "---\n\n"
            "**🎯 Tactical Playbook (Neutral)**\n\n"
            "**The professional approach in mixed markets:**\n"
            "- **Raise the bar for new entries** — only the highest-conviction, "
            "fundamentally-backed ideas qualify. If you're not sure, you're out.\n"
            "- **Reduce position sizes** — sizing down in uncertainty is not timidity, "
            "it's risk management. You can always add when clarity returns.\n"
            "- **Focus on stock-specific catalysts** — when the macro is unclear, "
            "alpha comes from individual company events: earnings, management changes, "
            "new contracts, regulatory approvals\n"
            "- **Review and prune** — use the pause to exit positions that aren't "
            "working and free up capital for better opportunities\n\n"
            "**Sectors that perform in neutral regimes:**\n"
            "- Quality compounders with strong free cash flow — they don't need macro tailwinds\n"
            "- Dividend payers with growing payouts — income while you wait\n"
            "- Sectors with idiosyncratic drivers (biotech catalyst, energy supply story)\n\n"
            "---\n\n"
            "**⚠️ What to watch for the regime to clarify:**\n"
            "- VIX breaking below 15 (risk-on confirmed) or above 25 (risk-off confirmed)\n"
            "- S&P 500 closing decisively above/below its 50-day moving average\n"
            "- Fed language, inflation prints, or payrolls data shifting expectations\n\n"
            "**Goldman rule of thumb:** *Mixed markets punish the impatient and reward "
            "the disciplined. When the market doesn't have a view, neither should you. "
            "Wait for the data. Capital preservation in uncertainty is the precondition "
            "for outperformance when clarity returns.*"
        )
    return mood

def _brief_card(sym, label, flag, price, chg_pct):
    """Render one index card."""
    if price is None:
        color, arrow, pct_str, price_str = "#64748B", "", "—", "—"
    else:
        up = chg_pct >= 0 if chg_pct is not None else True
        color   = "#22C55E" if up else "#EF4444"
        arrow   = "▲" if up else "▼"
        pct_str = f"{arrow} {abs(chg_pct):.2f}%" if chg_pct is not None else "—"
        price_str = f"{price:,.2f}" if price > 100 else f"{price:.4f}"
    return (
        f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.25);'
        f'border-radius:10px;padding:12px 14px;min-width:130px">'
        f'<div style="font-size:0.72rem;color:#64748B;margin-bottom:2px">{flag} {label}</div>'
        f'<div style="font-size:1rem;font-weight:700;color:#F1F5F9">{price_str}</div>'
        f'<div style="font-size:0.82rem;font-weight:600;color:{color}">{pct_str}</div>'
        f'</div>'
    )

def _risk_sentiment(vix, gold_pct, dxy_pct):
    """Derive simple risk-on / risk-off signal."""
    score = 0
    if vix is not None:
        if vix < 15:   score += 2
        elif vix < 20: score += 1
        elif vix > 25: score -= 1
        elif vix > 30: score -= 2
    if gold_pct is not None:
        if gold_pct > 0.5:  score -= 1   # gold up = risk-off
        elif gold_pct < -0.5: score += 1
    if dxy_pct is not None:
        if dxy_pct > 0.3:  score -= 1   # dollar up = risk-off
        elif dxy_pct < -0.3: score += 1
    if score >= 2:
        return "🟢 Risk-On", "#22C55E", "Markets in risk-on mode — appetite for equities is strong."
    elif score >= 0:
        return "🟡 Neutral", "#F59E0B", "Mixed signals — proceed with selective conviction."
    else:
        return "🔴 Risk-Off", "#EF4444", "Risk-off environment — caution warranted, check your stops."

with tab_brief:
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;padding:2px 0 6px 0">'
        '<span style="font-size:0.95rem;font-weight:700;color:#F1F5F9">🌍 Morning Market Intelligence Briefing</span>'
        '</div>', unsafe_allow_html=True)

    # ── Fetch all data ────────────────────────────────────────
    _all_syms  = [s for s,_,_,_ in _BRIEF_INDICES] + [s for s,_,_ in _BRIEF_INSTRUMENTS]
    _brief_col, _econ_col = st.columns([3, 1])

    with _brief_col:
        with st.spinner("Loading global market data…"):
            _bd = _fetch_brief_data(_all_syms)

        # ── Sentiment banner ─────────────────────────────────
        _vix_d  = _bd.get("^VIX",     {})
        _gold_d = _bd.get("GC=F",     {})
        _dxy_d  = _bd.get("DX-Y.NYB", {})
        _sentiment, _sent_color, _sent_msg = _risk_sentiment(
            _vix_d.get("price"), _gold_d.get("chg_pct"), _dxy_d.get("chg_pct"))

        _vix_val = _vix_d.get("price")
        _vix_color = "#EF4444" if (_vix_val or 0) > 25 else "#F59E0B" if (_vix_val or 0) > 18 else "#22C55E"
        _sent_parts = _sentiment.split(" ", 1)

        # ── Sentiment banner (full width in left col) ────────
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,rgba(13,31,51,0.95),rgba(8,18,32,0.99));
            border:1.5px solid {_sent_color}55;border-radius:14px;
            padding:10px 16px;margin-bottom:6px;
            display:flex;align-items:center;gap:12px">
          <div style="font-size:1.8rem;line-height:1">{_sent_parts[0]}</div>
          <div>
            <div style="font-size:1rem;font-weight:800;color:{_sent_color}">
              {_sent_parts[1] if len(_sent_parts)>1 else ''}</div>
            <div style="font-size:0.78rem;color:#94A3B8;margin-top:2px">{_sent_msg}</div>
          </div>
          <div style="margin-left:auto;text-align:right;min-width:72px">
            <div style="font-size:0.65rem;color:#475569;text-transform:uppercase;letter-spacing:0.5px">VIX Fear Index</div>
            <div style="font-size:1.6rem;font-weight:900;color:{_vix_color};line-height:1.2">
              {f"{_vix_val:.1f}" if _vix_val else "—"}</div>
            <div style="font-size:0.65rem;color:#475569">
              {"🟢 Low fear" if (_vix_val or 0)<18 else "🟡 Elevated" if (_vix_val or 0)<25 else "🔴 High fear"}</div>
          </div>
        </div>""", unsafe_allow_html=True)

        # ── VIX chart — full width, directly under banner ────
        _vix_hist = _fetch_vix_chart()
        if _vix_hist is not None and not _vix_hist.empty:
            _vfig = go.Figure()
            # Zone background bands
            _vix_zones = [
                (0,  12,  "rgba(34,197,94,0.07)",  "Calm"),
                (12, 20,  "rgba(251,191,36,0.07)", "Mean"),
                (20, 30,  "rgba(249,115,22,0.09)", "Anxiety"),
                (30, 40,  "rgba(239,68,68,0.10)",  "Fear"),
                (40, 80,  "rgba(139,0,0,0.10)",    "Panic"),
            ]
            for _z0, _z1, _zcol, _zlbl in _vix_zones:
                _vfig.add_hrect(y0=_z0, y1=_z1, fillcolor=_zcol, layer="below",
                                line_width=0,
                                annotation_text=_zlbl,
                                annotation_position="right",
                                annotation_font_size=9,
                                annotation_font_color="#94A3B8")
            # Zone boundary lines — no annotation text (avoids overlap with y-axis ticks)
            for _yval, _col in [(20, "#F59E0B"), (30, "#EF4444"), (40, "#991B1B")]:
                _vfig.add_hline(y=_yval, line_dash="dot", line_color=_col, line_width=1)
            # VIX line
            _fill_rgba = f"rgba({int(_vix_color[1:3],16)},{int(_vix_color[3:5],16)},{int(_vix_color[5:7],16)},0.15)"
            _vfig.add_trace(go.Scatter(
                x=_vix_hist.iloc[:,0], y=_vix_hist.iloc[:,1],
                fill="tozeroy", line=dict(color=_vix_color, width=2.5),
                fillcolor=_fill_rgba, name="VIX",
                hovertemplate="<b>%{x|%d %b}</b><br>VIX: %{y:.1f}<extra></extra>"
            ))
            _dates = _vix_hist.iloc[:,0].tolist()
            _tick_dates = [_dates[0], _dates[len(_dates)//4], _dates[len(_dates)//2],
                           _dates[3*len(_dates)//4], _dates[-1]]
            _vfig.update_layout(
                height=160,
                margin=dict(l=10, r=70, t=22, b=32),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                showlegend=False,
                title=dict(text="VIX Fear Index — 30 Days", font=dict(size=11, color="#94A3B8"), x=0),
                xaxis=dict(
                    tickvals=_tick_dates,
                    tickformat="%d %b",
                    tickfont=dict(size=9, color="#64748B"),
                    showgrid=False,
                    title=dict(text="Date", font=dict(size=9, color="#64748B")),
                ),
                yaxis=dict(
                    range=[0, max(45, float(_vix_hist.iloc[:,1].max()) + 5)],
                    showgrid=True,
                    gridcolor="rgba(100,116,139,0.15)",
                    tickfont=dict(size=9, color="#64748B"),
                    title=dict(text="VIX Level", font=dict(size=9, color="#64748B")),
                ),
            )
            st.plotly_chart(_vfig, use_container_width=True, config={"displayModeBar": False})

        # ── AI Explanation expander (always visible, just below chart) ──
        _spx_pct = _bd.get("^GSPC", {}).get("chg_pct")
        _explanation = _generate_risk_explanation(
            _sentiment, _vix_val, _spx_pct,
            _gold_d.get("chg_pct"), _dxy_d.get("chg_pct"))
        with st.expander("🤖 What does this mean for me? (AI Market Explanation)"):
            st.markdown(_explanation)

        # ── Indices by region ─────────────────────────────────
        for _region in ["US", "Futures", "UK", "Europe", "Asia"]:
            _region_items = [(s,l,f) for s,l,f,r in _BRIEF_INDICES if r == _region]
            _label_map = {"US":"🇺🇸 United States", "Futures":"📊 Index Futures",
                          "UK":"🇬🇧 United Kingdom", "Europe":"🌍 Europe", "Asia":"🌏 Asia Pacific"}
            st.markdown(f'<div style="font-size:0.78rem;font-weight:700;color:#F59E0B;'
                        f'letter-spacing:1px;text-transform:uppercase;margin:14px 0 8px 0">'
                        f'{_label_map[_region]}</div>', unsafe_allow_html=True)
            _cards_html = '<div style="display:flex;flex-wrap:wrap;gap:10px">'
            for _sym, _lbl, _flg in _region_items:
                _d = _bd.get(_sym, {})
                _cards_html += _brief_card(_sym, _lbl, _flg, _d.get("price"), _d.get("chg_pct"))
            _cards_html += '</div>'
            st.markdown(_cards_html, unsafe_allow_html=True)

        # ── Key instruments ───────────────────────────────────
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#F59E0B;'
                    'letter-spacing:1px;text-transform:uppercase;margin:18px 0 8px 0">'
                    '⚙️ Key Instruments</div>', unsafe_allow_html=True)
        _inst_html = '<div style="display:flex;flex-wrap:wrap;gap:10px">'
        for _sym, _lbl, _unit in _BRIEF_INSTRUMENTS:
            if _sym == "^VIX": continue  # already shown in banner
            _d = _bd.get(_sym, {})
            _inst_html += _brief_card(_sym, f"{_lbl} ({_unit})", "", _d.get("price"), _d.get("chg_pct"))
        _inst_html += '</div>'
        st.markdown(_inst_html, unsafe_allow_html=True)

        # ── Market News ───────────────────────────────────────
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#F59E0B;'
                    'letter-spacing:1px;text-transform:uppercase;margin:22px 0 10px 0">'
                    '📰 Latest Market News</div>', unsafe_allow_html=True)
        with st.spinner("Loading news…"):
            _news = _fetch_market_news()
        if _news:
            for _n in _news:
                # Handle both FMP and yfinance schemas
                _title     = _n.get("title", "")
                _link      = _n.get("url") or _n.get("link", "#")
                _publisher = _n.get("site") or _n.get("publisher", "")
                _raw_dt    = _n.get("publishedDate") or _n.get("date", "")
                _dt        = _raw_dt[:16].replace("T", " ") if _raw_dt else ""
                if not _title:
                    continue
                st.markdown(
                    f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.2);'
                    f'border-radius:8px;padding:10px 14px;margin-bottom:8px">'
                    f'<a href="{_link}" target="_blank" style="color:#E2E8F0;font-size:0.88rem;'
                    f'font-weight:600;text-decoration:none">{_title}</a>'
                    f'<div style="color:#475569;font-size:0.72rem;margin-top:4px">'
                    f'{_publisher} · {_dt}</div></div>',
                    unsafe_allow_html=True)
        else:
            st.info("News temporarily unavailable.")

    # ── Economic Calendar (right column) ─────────────────────
    with _econ_col:
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#F59E0B;'
                    'letter-spacing:1px;text-transform:uppercase;margin-bottom:10px">'
                    '📅 Economic Calendar</div>', unsafe_allow_html=True)
        with st.spinner("Loading calendar…"):
            _econ = _fetch_econ_calendar()
        if _econ:
            _prev_date = None
            for _ev in _econ[:20]:
                _ev_date  = _ev.get("date","")[:10]
                _ev_time  = _ev.get("date","")[11:16]
                _ev_name  = _ev.get("event","")
                _ev_ctry  = _ev.get("country","")
                _ev_imp   = _ev.get("impact","")
                _ev_act   = _ev.get("actual","")
                _ev_est   = _ev.get("estimate","")
                _ev_prev  = _ev.get("previous","")
                _imp_col  = "#EF4444" if _ev_imp=="High" else "#F59E0B"
                _imp_dot  = f'<span style="color:{_imp_col};font-size:0.8rem">●</span>'
                if _ev_date != _prev_date:
                    try:
                        _dobj = datetime.strptime(_ev_date, "%Y-%m-%d")
                        _dlbl = "Today" if _dobj.date()==datetime.now().date() else \
                                "Tomorrow" if (_dobj.date()-datetime.now().date()).days==1 else \
                                _dobj.strftime("%a %d %b")
                    except Exception:
                        _dlbl = _ev_date
                    st.markdown(f'<div style="font-size:0.75rem;font-weight:700;color:#64748B;'
                                f'margin:12px 0 6px 0;text-transform:uppercase">{_dlbl}</div>',
                                unsafe_allow_html=True)
                    _prev_date = _ev_date
                st.markdown(
                    f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.2);'
                    f'border-radius:8px;padding:8px 10px;margin-bottom:6px">'
                    f'<div style="display:flex;justify-content:space-between;align-items:center">'
                    f'{_imp_dot} <span style="font-size:0.7rem;color:#64748B">{_ev_time} · {_ev_ctry}</span></div>'
                    f'<div style="font-size:0.78rem;font-weight:600;color:#E2E8F0;margin:3px 0">{_ev_name}</div>'
                    f'<div style="font-size:0.7rem;color:#475569">'
                    f'Act: <span style="color:#F1F5F9">{_ev_act or "—"}</span> &nbsp;'
                    f'Est: {_ev_est or "—"} &nbsp;Prev: {_ev_prev or "—"}</div>'
                    f'</div>',
                    unsafe_allow_html=True)
        else:
            st.markdown('<div style="color:#64748B;font-size:0.85rem">'
                        'No high/medium impact events in the next 5 days.</div>',
                        unsafe_allow_html=True)

        # ── Useful links ──────────────────────────────────────
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#F59E0B;'
                    'letter-spacing:1px;text-transform:uppercase;margin:20px 0 10px 0">'
                    '🔗 Professional Resources</div>', unsafe_allow_html=True)
        _links = [
            ("Financial Juice", "https://www.financialjuice.com", "Real-time news squawk"),
            ("Investing.com", "https://www.investing.com/economic-calendar/", "Economic calendar"),
            ("TradingView", "https://www.tradingview.com", "Charts & technicals"),
            ("Reuters", "https://www.reuters.com/markets/", "Market news"),
            ("Bloomberg", "https://www.bloomberg.com/markets", "Global markets"),
        ]
        for _name, _url, _desc in _links:
            st.markdown(
                f'<a href="{_url}" target="_blank" style="display:block;'
                f'background:#0D1F33;border:1px solid rgba(100,116,139,0.2);'
                f'border-radius:8px;padding:8px 12px;margin-bottom:6px;text-decoration:none">'
                f'<div style="font-size:0.82rem;font-weight:600;color:#3B82F6">{_name}</div>'
                f'<div style="font-size:0.7rem;color:#475569">{_desc}</div></a>',
                unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# TAB 1 — FUNDAMENTAL SCREEN
# ═══════════════════════════════════════════════════════════════

with tab1:
    # ── Persistent walls — survive reruns caused by browser events ──
    _tab1_user = st.session_state.get("fintiq_user", {})
    # Guest wall: show if session state flag set (triggered by _check_auth_gate)
    if not _tab1_user and st.session_state.get("_show_auth_wall"):
        _show_auth_wall()
        st.stop()
    # Upgrade wall
    if st.session_state.get("_show_upgrade_wall") and not _tab1_user.get("is_pro"):
        _uw = st.session_state["_show_upgrade_wall"]
        _show_upgrade_wall(_uw[0], _uw[1])
        st.stop()

    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;padding:2px 0 2px 0;margin-bottom:2px">'
        '<span style="font-size:0.95rem;font-weight:700;color:#F1F5F9">🔍 Fundamental Quality Screen</span>'
        '<span style="color:#64748B;font-size:0.74rem">Score 80–100 = strong conviction · 60–79 = good · &lt;60 = marginal</span>'
        '</div>',
        unsafe_allow_html=True)

    # ── WATCHLIST PANEL ────────────────────────────────────────────
    _wl = st.session_state.get("fintiq_watchlist", {})
    if _wl:
        with st.expander(f"⭐ My Watchlist ({len(_wl)})", expanded=False):
            _wl_cols = st.columns([2,2,2,2,2,1])
            _wl_cols[0].markdown("**Ticker**")
            _wl_cols[1].markdown("**Company**")
            _wl_cols[2].markdown("**Price**")
            _wl_cols[3].markdown("**DCF Upside**")
            _wl_cols[4].markdown("**Quality Score**")
            _wl_cols[5].markdown("**Remove**")
            _tickers_to_remove = []
            for _wtick, _wdata in list(_wl.items()):
                _wc0, _wc1, _wc2, _wc3, _wc4, _wc5 = st.columns([2,2,2,2,2,1])
                _wc0.write(f"**{_wtick}**")
                _wc1.write(_wdata.get("name", "—")[:22])
                _wpx   = _wdata.get("price")
                _wiv   = _wdata.get("dcf_iv")
                _wscore= _wdata.get("quality_score")
                _wcur = _wdata.get("currency", "")
                # Backwards-compat: old entries may not have currency stored
                if not _wcur and _wtick.endswith(".L"):
                    _wcur = "GBp"
                if _wpx:
                    if _wcur == "GBp":
                        _wpx_str = f"£{_wpx/100:,.2f}"   # convert pence → £
                    elif _wcur in ("GBP",):
                        _wpx_str = f"£{_wpx:,.2f}"
                    elif _wcur == "USD":
                        _wpx_str = f"${_wpx:,.2f}"
                    else:
                        _wpx_str = f"{_wcur} {_wpx:,.2f}" if _wcur else f"{_wpx:,.2f}"
                else:
                    _wpx_str = "—"
                _wc2.write(_wpx_str)
                if _wpx and _wiv:
                    _wup = (_wiv - _wpx) / _wpx * 100
                    _col = "🟢" if _wup > 20 else ("🔴" if _wup < -10 else "🟡")
                    _wc3.write(f"{_col} {_wup:+.1f}%")
                else:
                    _wc3.write("—")
                if _wscore:
                    _qc = "🥇" if _wscore >= 80 else ("🥈" if _wscore >= 60 else "🥉")
                    _wc4.write(f"{_qc} {_wscore}/100")
                else:
                    _wc4.write("—")
                if _wc5.button("✕", key=f"wl_rm_{_wtick}"):
                    _tickers_to_remove.append(_wtick)
            if _tickers_to_remove:
                for _tr in _tickers_to_remove:
                    del st.session_state["fintiq_watchlist"][_tr]
                _wl_save(st.session_state["fintiq_watchlist"])
                st.rerun()
            # Quick-jump
            _wl_jump_col, _wl_clear_col = st.columns([3, 1])
            with _wl_jump_col:
                _wl_pick = st.selectbox("Jump to stock →", [""] + list(_wl.keys()), key="wl_jump_select")
                if _wl_pick:
                    st.session_state["deepdive_pick"] = _wl_pick   # sync Screen 1 deep-dive + Screen 3
                    st.rerun()
            with _wl_clear_col:
                if st.button("🗑 Clear All", key="wl_clear_all"):
                    st.session_state["fintiq_watchlist"] = {}
                    _wl_save({})
                    st.rerun()

    # ── Exchange selector + Sector filter ───────────────────────
    top1, top2 = st.columns([3, 2])
    with top1:
        selected_labels = st.multiselect(
            "🌍 Select Exchanges:",
            options=list(ALL_EXCHANGES.keys()),
            default=["🏛 London (LSE)"],
        )
        selected_exchanges = [ALL_EXCHANGES[l] for l in selected_labels]
        if selected_exchanges:
            total_tickers = sum(len(STOCK_UNIVERSE.get(e,[])) for e in selected_exchanges)
            st.caption(f"📊 **{total_tickers:,} stocks** · {len(selected_exchanges)} exchange(s)")
        else:
            st.error("Select at least one exchange.")
            total_tickers = 0
    with top2:
        selected_sector = st.selectbox("Sector Filter:", SECTORS)

    # ── Advanced filters (collapsed by default) ─────────────────
    # Market cap buckets: label → (min_M, max_M) in local currency millions
    _CAP_BUCKETS = {
        "Any":             (0,        99_999_999),
        "Micro  (< 300M)": (0,        300),
        "Small  (300M – 2B)": (300,   2_000),
        "Mid    (2B – 10B)":  (2_000, 10_000),
        "Large  (10B – 200B)":(10_000,200_000),
        "Mega   (> 200B)": (200_000,  99_999_999),
    }

    # ── Preset screeners ─────────────────────────────────────────
    _PRESETS = {
        "— Custom (set filters manually) —": {},
        "🚀 1. Momentum Breakout Leaders": {
            "af_cap": "Large  (10B – 200B)", "af_pe": (5, 60), "af_pb": 20.0,
            "af_roe": 15, "af_gm": 20, "af_nm": 5, "af_de": 1.5,
            "af_cc": 0.5, "af_div": False, "af_vol": 1.0,
        },
        "🔄 2. Early Trend Reversal": {
            "af_cap": "Any", "af_pe": (0, 50), "af_pb": 15.0,
            "af_roe": 8, "af_gm": 15, "af_nm": 0, "af_de": 2.5,
            "af_cc": 0.4, "af_div": False, "af_vol": 0.5,
        },
        "📈 3. High Growth Companies": {
            "af_cap": "Mid    (2B – 10B)", "af_pe": (0, 80), "af_pb": 20.0,
            "af_roe": 20, "af_gm": 40, "af_nm": 10, "af_de": 0.5,
            "af_cc": 0.8, "af_div": False, "af_vol": 0.5,
        },
        "💎 4. Undervalued Quality (Buffett-Style)": {
            "af_cap": "Mid    (2B – 10B)", "af_pe": (0, 20), "af_pb": 5.0,
            "af_roe": 18, "af_gm": 30, "af_nm": 8, "af_de": 0.4,
            "af_cc": 0.8, "af_div": True, "af_vol": 0.2,
        },
        "💥 5. Short Squeeze Candidates": {
            "af_cap": "Small  (300M – 2B)", "af_pe": (0, 100), "af_pb": 20.0,
            "af_roe": 0, "af_gm": 0, "af_nm": 0, "af_de": 3.0,
            "af_cc": 0.0, "af_div": False, "af_vol": 0.5,
        },
        "📊 6. Swing Trading Candidates (5–15 Days)": {
            "af_cap": "Any", "af_pe": (0, 50), "af_pb": 15.0,
            "af_roe": 10, "af_gm": 20, "af_nm": 3, "af_de": 2.0,
            "af_cc": 0.4, "af_div": False, "af_vol": 1.0,
        },
        "⚡ 7. Earnings Momentum Plays": {
            "af_cap": "Mid    (2B – 10B)", "af_pe": (0, 60), "af_pb": 20.0,
            "af_roe": 15, "af_gm": 25, "af_nm": 8, "af_de": 1.5,
            "af_cc": 0.6, "af_div": False, "af_vol": 1.0,
        },
        "🏛 8. Institutional Accumulation": {
            "af_cap": "Large  (10B – 200B)", "af_pe": (0, 40), "af_pb": 10.0,
            "af_roe": 15, "af_gm": 25, "af_nm": 5, "af_de": 1.0,
            "af_cc": 0.7, "af_div": False, "af_vol": 1.0,
        },
        "🏆 9. High Relative Strength Sector Leaders": {
            "af_cap": "Large  (10B – 200B)", "af_pe": (5, 50), "af_pb": 20.0,
            "af_roe": 20, "af_gm": 35, "af_nm": 10, "af_de": 1.0,
            "af_cc": 0.7, "af_div": False, "af_vol": 1.0,
        },
        "🎯 10. Multi-Factor Hedge Fund Screen": {
            "af_cap": "Mega   (> 200B)", "af_pe": (0, 45), "af_pb": 15.0,
            "af_roe": 18, "af_gm": 30, "af_nm": 10, "af_de": 0.5,
            "af_cc": 0.8, "af_div": False, "af_vol": 2.0,
        },
        "🦁 Buffett + Graham + Munger Master": {
            "af_cap": "Mid    (2B – 10B)", "af_pe": (0, 20), "af_pb": 3.0,
            "af_roe": 18, "af_gm": 35, "af_nm": 10, "af_de": 0.5,
            "af_cc": 0.85, "af_div": True, "af_vol": 0.2,
        },
    }
    _PRESET_NOTES = {
        "🚀 1. Momentum Breakout Leaders":
            "Institutional money flowing into strong large-caps. Filters: Large cap, ROE>15%, GM>20%, NM>5%, D/E<1.5, Vol>1M. "
            "Pair with Technical tab to confirm price above 20/50/200-day MAs.",
        "🔄 2. Early Trend Reversal":
            "Catch trends before the crowd. Moderate quality thresholds with rising volume. "
            "Pair with Technical tab to confirm 50-day MA crossover and bullish MACD.",
        "📈 3. High Growth Companies":
            "Future market leaders — high ROE, thick margins, low debt, strong FCF. "
            "Technical confirmation: check price trend and volume expansion.",
        "💎 4. Undervalued Quality (Buffett-Style)":
            "Classic value — low PE, low P/B, high ROE, low debt, dividend payers. "
            "Check 5-year earnings consistency and insider activity manually.",
        "💥 5. Short Squeeze Candidates":
            "Explosive potential — small float, high short interest. Looser fundamentals by design. "
            "Check short interest ratio and float in the Technical tab manually.",
        "📊 6. Swing Trading Candidates (5–15 Days)":
            "High-probability swing setups. Look for stocks above 50-day MA with pullback to 20 EMA. "
            "Confirm ATR>3% and bullish reversal candles in Technical tab.",
        "⚡ 7. Earnings Momentum Plays":
            "Post-earnings continuation. Filters for quality companies with strong margins. "
            "Confirm beat vs estimates and guidance raise via news/earnings tab.",
        "🏛 8. Institutional Accumulation":
            "Follow smart money — large-caps with strong quality metrics. "
            "Confirm rising OBV and accumulation/distribution in Technical tab.",
        "🏆 9. High Relative Strength Sector Leaders":
            "Best stocks in the strongest sectors — high ROE, thick margins, high volume. "
            "Cross-check sector performance to confirm sector leadership.",
        "🎯 10. Multi-Factor Hedge Fund Screen":
            "Professional-grade composite — Mega-cap quality with strict fundamentals and liquidity. "
            "Highest-conviction names combining growth, quality, and momentum.",
        "🦁 Buffett + Graham + Munger Master":
            "Wide moat compounders at fair price. Ultra-strict: PE<20, ROE>18%, D/E<0.5, "
            "high cash conversion, dividend payers. Long-term hold, not a trade.",
    }

    with st.expander("⚙️  Advanced Filters", expanded=False):
            # ── Preset selector ──────────────────────────────────────
            _pr_col, _ap_col = st.columns([3, 1])
            with _pr_col:
                _preset_sel = st.selectbox(
                    "📋 Load Preset Screener",
                    list(_PRESETS.keys()), key="af_preset_sel",
                    help="Select a preset to auto-fill all filters instantly.")
            with _ap_col:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                _is_custom = (_preset_sel == "— Custom (set filters manually) —")
                if st.button("Apply →", key="af_apply_preset", use_container_width=True,
                             disabled=_is_custom, type="primary"):
                    for _pk, _pv in _PRESETS[_preset_sel].items():
                        st.session_state[_pk] = _pv
                    st.rerun()
            if not _is_custom and _preset_sel in _PRESET_NOTES:
                st.info(_PRESET_NOTES[_preset_sel])
            st.markdown('<p style="color:#64748B;font-size:0.8rem;margin:8px 0 12px 0">'
                'Adjust sliders below, then click <b>▶ Run Screen</b>.</p>',
                unsafe_allow_html=True)

            _af1, _af2, _af3, _af4, _af5 = st.columns(5)

            with _af1:
                st.markdown("**📏 Size & Liquidity**")
                cap_bucket = st.selectbox(
                    "Market Cap Size", list(_CAP_BUCKETS.keys()), index=0,
                    key="af_cap",
                    help="Filters by market capitalisation in the stock's local currency.")
                min_cap, max_cap = _CAP_BUCKETS[cap_bucket]
                min_vol_m = st.slider("Min Avg Daily Vol (M)", 0.0, 5.0, 0.0, 0.25,
                    key="af_vol",
                    help="0 = no minimum. 1.0 = at least 1M shares/day. Higher = more liquid.")

            with _af2:
                st.markdown("**💹 Valuation**")
                pe_range = st.slider("P/E Ratio (range)", 0, 100, (0, 35),
                    key="af_pe",
                    help="0 = no lower limit. Stocks with negative P/E are excluded if you set min > 0.")
                min_pe, max_pe = pe_range
                pb_max = st.slider("Max Price/Book", 0.0, 20.0, 10.0, 0.5,
                    key="af_pb",
                    help="Set to 20 to disable. Low P/B can indicate undervaluation.")

            with _af3:
                st.markdown("**📈 Profitability**")
                min_roe = st.slider("Min ROE (%)", 0, 50, 10,
                    key="af_roe",
                    help="Return on Equity. 0 = no minimum. >15% is good quality.") / 100
                min_gm  = st.slider("Min Gross Margin (%)", 0, 70, 15,
                    key="af_gm",
                    help="0 = no minimum. Higher margins = pricing power.") / 100
                min_nm  = st.slider("Min Net Margin (%)", 0, 40, 0,
                    key="af_nm",
                    help="0 = no minimum. Filters loss-making companies when >0.") / 100

            with _af4:
                st.markdown("**🏦 Financial Health**")
                max_de = st.slider("Max Debt/Equity", 0.0, 10.0, 3.0, 0.25,
                    key="af_de",
                    help="Set to 10 to disable. Lower = less financial risk.")
                min_cc = st.slider("Min Cash Conversion", 0.0, 1.5, 0.5, 0.05,
                    key="af_cc",
                    help="Operating cash flow / net income. >0.8 = high quality earnings. 0 = no minimum.")
                div_filter = st.checkbox("Dividend payers only", key="af_div",
                    help="Only show stocks that paid a dividend in the last 12 months.")

            with _af5:
                st.markdown("**📐 Technical Signal**")
                st.markdown(
                    '<p style="color:#64748B;font-size:0.78rem;line-height:1.4">'
                    'Technicals are shown as <b>indicators</b> in results — not hard filters.<br>'
                    '✅ Strong &nbsp;⚠️ Neutral &nbsp;🔴 Weak<br>'
                    'Computed from: MA position, 52-wk proximity, volume spike.</p>',
                    unsafe_allow_html=True)

            _rst1, _rst2 = st.columns([1, 5])
            with _rst1:
                if st.button("↺ Reset to Defaults", key="af_reset"):
                    for _k in ["af_cap","af_pe","af_pb","af_roe","af_gm","af_nm",
                                "af_de","af_cc","af_div","af_vol","af_preset_sel"]:
                        if _k in st.session_state: del st.session_state[_k]
                    st.rerun()

    # Read filter values from session state (set by sliders above but not causing reruns
    # because we read them only when Run Screen is pressed)
    cap_bucket  = st.session_state.get("af_cap", "Any")
    min_cap, max_cap = _CAP_BUCKETS[cap_bucket]
    _pe_r   = st.session_state.get("af_pe",  (0, 35))
    min_pe, max_pe = _pe_r
    pb_max  = st.session_state.get("af_pb",  10.0)
    min_roe = st.session_state.get("af_roe", 10) / 100
    min_gm  = st.session_state.get("af_gm",  15) / 100
    min_nm  = st.session_state.get("af_nm",  0)  / 100
    max_de  = st.session_state.get("af_de",  3.0)
    min_cc  = st.session_state.get("af_cc",  0.5)
    div_filter    = st.session_state.get("af_div",   False)
    min_vol_m     = st.session_state.get("af_vol",   0.0)    # millions of shares

    # ── Run Screen button ────────────────────────────────────────
    _rb1, _rb2, _rb3 = st.columns([2, 1, 2])
    with _rb2:
        run_screen = st.button("▶  Run Screen", type="primary",
                               use_container_width=True, key="run_screen_bot")

    if run_screen and selected_exchanges and _check_auth_gate():
        tickers = []
        for exch in selected_exchanges:
            tickers.extend(STOCK_UNIVERSE.get(exch,[]))
        tickers = list(dict.fromkeys(tickers))

        st.info(f"Fetching fundamentals for {len(tickers):,} stocks… "
                f"First run takes 2–5 min (cached 1hr after).")
        progress = st.progress(0)
        status   = st.empty()
        results  = []

        for i, sym in enumerate(tickers):
            progress.progress((i+1)/len(tickers))
            status.caption(f"Checking {sym}… ({i+1}/{len(tickers)})")
            info = get_yf_info(sym)
            if not info or ("symbol" not in info and "longName" not in info):
                continue
            if selected_sector != "All Sectors":
                if info.get("sector","") != selected_sector:
                    continue
            cap_m = (info.get("marketCap") or 0) / 1_000_000
            if cap_bucket != "Any" and not (min_cap <= cap_m <= max_cap):
                continue
            # Dividend filter
            if div_filter and not (info.get("dividendYield") or 0) > 0:
                continue
            qs = quality_score(info)
            _roe  = qs["roe"]        or 0
            _gm   = qs["gross_margin"] or 0
            _nm   = (info.get("profitMargins") or 0)
            _de   = qs["debt_equity"]
            _cc   = qs["cash_conv"]  or 0
            _pe   = qs["pe"]         or 0
            _pb   = info.get("priceToBook") or 0
            _avg_vol_m  = (info.get("averageVolume") or 0) / 1_000_000
            _curr_price = info.get("currentPrice") or info.get("regularMarketPrice") or 0
            _ma50       = info.get("fiftyDayAverage") or 0
            _ma200      = info.get("twoHundredDayAverage") or 0
            _wk52_high  = info.get("fiftyTwoWeekHigh") or 0
            _today_vol  = info.get("volume") or info.get("regularMarketVolume") or 0
            _avg_vol_raw= info.get("averageVolume") or 0
            _pct_from_52h = ((_wk52_high - _curr_price) / _wk52_high * 100) if _wk52_high > 0 else 100
            _vol_spike_pct= (_today_vol / _avg_vol_raw * 100) if _avg_vol_raw > 0 else 0
            # ── Fundamental gate only ──
            passes = (
                _roe  >= min_roe and
                _gm   >= min_gm  and
                _nm   >= min_nm  and
                (_de is None or _de <= max_de) and
                _cc   >= min_cc  and
                (_pe  <= 0 or max_pe == 100 or _pe <= max_pe) and
                (min_pe == 0 or _pe >= min_pe) and
                (pb_max >= 20.0 or (0 < _pb <= pb_max)) and
                (min_vol_m == 0.0 or _avg_vol_m >= min_vol_m)
            )
            if passes:
                # ── Technical signal score (0–6, shown as indicator not filter) ──
                _tscore = 0
                if _ma200 > 0 and _curr_price >= _ma200: _tscore += 2
                if _ma50  > 0 and _curr_price >= _ma50:  _tscore += 2
                if _wk52_high > 0 and _pct_from_52h <= 10: _tscore += 1
                if _avg_vol_raw > 0 and _vol_spike_pct >= 150: _tscore += 1
                _tech_signal = ("✅ Strong" if _tscore >= 4
                                else "⚠️ Neutral" if _tscore >= 2
                                else "🔴 Weak")
                _ma50_str  = (f"▲ {((_curr_price/_ma50-1)*100):+.1f}%" if _curr_price >= _ma50
                              else f"▼ {((_curr_price/_ma50-1)*100):+.1f}%") if _ma50 > 0 else "–"
                _ma200_str = (f"▲ {((_curr_price/_ma200-1)*100):+.1f}%" if _curr_price >= _ma200
                              else f"▼ {((_curr_price/_ma200-1)*100):+.1f}%") if _ma200 > 0 else "–"
                curr     = _f(info.get("currentPrice") or info.get("regularMarketPrice"))
                sym_curr = get_currency_symbol(sym)
                _live_cur = info.get("financialCurrency") or info.get("currency","USD")
                _cur_map  = {"USD":"$","GBP":"£","GBp":"£","EUR":"€","JPY":"¥","CAD":"CA$",
                              "AUD":"A$","HKD":"HK$","SGD":"S$","CHF":"CHF ","INR":"₹"}
                mc_sym    = _cur_map.get(_live_cur, sym_curr)
                results.append({
                    "Ticker":          sym,
                    "Company":         info.get("longName", sym),
                    "Sector":          qs["sector"],
                    "Quality Score":   qs["score"],
                    "Tech Signal":     _tech_signal,
                    "ROE":             fmt_pct(qs["roe"])          if qs["roe"]         else "–",
                    "P/E":             fmt_number(qs["pe"],1)      if qs["pe"]          else "–",
                    "P/B":             fmt_number(info.get("priceToBook"),2) if info.get("priceToBook") else "–",
                    "Gross Margin":    fmt_pct(qs["gross_margin"]) if qs["gross_margin"] else "–",
                    "Net Margin":      fmt_pct(info.get("profitMargins")) if info.get("profitMargins") else "–",
                    "Debt/Equity":     fmt_number(qs["debt_equity"],2) if qs["debt_equity"] is not None else "–",
                    "Cash Conv":       fmt_number(qs["cash_conv"],2) if qs["cash_conv"]  else "–",
                    "Avg Vol (M)":     f"{_avg_vol_m:.1f}" if _avg_vol_m > 0 else "–",
                    "vs 50d MA":       _ma50_str,
                    "vs 200d MA":      _ma200_str,
                    "52w High %":      f"-{_pct_from_52h:.1f}%" if _wk52_high > 0 else "–",
                    "Market Cap":    fmt_currency(cap_m*1_000_000, mc_sym),
                    "Price":         get_price_display(curr, sym, info) if curr else "–",
                })

        progress.empty(); status.empty()

        if not results:
            st.warning("No stocks passed all filters. Try relaxing thresholds above.")
            # Clear stale results so old run doesn't show below
            for _k in ["screened_df", "screened_symbols"]:
                if _k in st.session_state: del st.session_state[_k]
        else:
            df_r = pd.DataFrame(results).sort_values("Quality Score", ascending=False)
            st.session_state["screened_df"]      = df_r
            st.session_state["screened_symbols"] = df_r["Ticker"].tolist()
            st.success(f"✅  **{len(df_r)} stocks** passed all quality filters")

    # ── Results table + deep-dive — always shown when results exist ──
    if "screened_df" in st.session_state and not st.session_state["screened_df"].empty:
        df_r = st.session_state["screened_df"]

        def colour_qs(v):
            if isinstance(v, (int,float)):
                if v >= 80: return "background-color:#14532D;color:#A7F3D0;font-weight:700;letter-spacing:0.5px"
                if v >= 60: return "background-color:#78350F;color:#FEF3C7;font-weight:700"
                return "background-color:#7F1D1D;color:#FECACA;font-weight:700"
            return ""

        def colour_tech(v):
            if "Strong" in str(v):  return "color:#22C55E;font-weight:700"
            if "Neutral" in str(v): return "color:#F59E0B;font-weight:600"
            if "Weak" in str(v):    return "color:#EF4444;font-weight:600"
            return ""

        st.caption(
            "📐 **Tech Signal** — ✅ Strong: above both MAs + near 52-wk high / volume spike  "
            "⚠️ Neutral: mixed signals  🔴 Weak: below key MAs. "
            "All stocks shown passed **fundamental** filters. Use Tech Signal to time your entry.")

        _style_cols = {"Quality Score": colour_qs}
        if "Tech Signal" in df_r.columns:
            _style_cols["Tech Signal"] = colour_tech

        # Table with single-row selection — use key to preserve across reruns
        _sel = st.dataframe(
            df_r.style.map(colour_qs, subset=["Quality Score"]).map(
                colour_tech, subset=["Tech Signal"]) if "Tech Signal" in df_r.columns
            else df_r.style.map(colour_qs, subset=["Quality Score"]),
            use_container_width=True, height=380,
            selection_mode="single-row",
            on_select="rerun",
            key="results_table",
        )
        # Auto-populate deep-dive from table row click
        try:
            _rows = _sel.selection.rows if _sel.selection else []
        except Exception:
            _rows = []
        if _rows:
            _clicked = df_r.iloc[_rows[0]]["Ticker"]
            if st.session_state.get("deepdive_pick") != _clicked:
                st.session_state["deepdive_pick"] = _clicked

        # ── Excel download ──
        _xl_bytes = build_fintiq_excel(df_r, "Fintiq Screen")
        _xl_ext   = "xlsx" if _OPENPYXL else "csv"
        st.download_button(
            label="📥 Download Results as Excel",
            data=_xl_bytes,
            file_name=f"fintiq_screen_{datetime.now().strftime('%Y%m%d_%H%M')}.{_xl_ext}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
            key="excel_dl",
        )
        st.info("💡 Screened stocks auto-loaded into Catalyst Alerts & Technical Setup tabs.")

        # ── Deep-Dive ─────────────────────────────────────────────
        st.markdown('<div class="section-header">🔬 Stock Deep-Dive + Intrinsic Valuation</div>',
                    unsafe_allow_html=True)

        _ticker_list = df_r["Ticker"].tolist()
        _default_idx = 0
        if "deepdive_pick" in st.session_state and st.session_state["deepdive_pick"] in _ticker_list:
            _default_idx = _ticker_list.index(st.session_state["deepdive_pick"])

        st.caption("💡 Click any row in the table above to auto-select stock below.")
        pick = st.selectbox(
            "Select stock for detailed analysis:",
            _ticker_list,
            index=_default_idx,
            format_func=lambda t: f"{t}  —  {df_r.loc[df_r['Ticker']==t,'Company'].values[0] if not df_r.loc[df_r['Ticker']==t,'Company'].empty else t}",
            key="deepdive_select",
        )
        # Keep session state in sync with manual selectbox choice
        if pick and st.session_state.get("deepdive_pick") != pick:
            st.session_state["deepdive_pick"] = pick

        if pick:
            info    = get_yf_info(pick)
            qs      = quality_score(info)
            sym_cur = get_currency_symbol(pick)
            curr_px = _f(info.get("currentPrice") or info.get("regularMarketPrice"))
            eps     = _f(info.get("trailingEps") or info.get("epsTrailingTwelveMonths"))
            bvps    = _f(info.get("bookValue"))
            sector_name = qs["sector"] or "Other"
            sector_pe   = SECTOR_PE_AVERAGES.get(sector_name, 17)

            def _mc(col, label, value, sub="", extra_class=""):
                col.markdown(f"""
                <div class="metric-card {extra_class}">
                  <div class="metric-label">{label}</div>
                  <div class="metric-value">{value}</div>
                  <div class="metric-sub">{sub}</div>
                </div>""", unsafe_allow_html=True)

            # ── Company description ───────────────────────────────
            company_name = info.get("longName", pick)
            long_biz     = info.get("longBusinessSummary","")
            website      = info.get("website","")
            country      = info.get("country","")
            employees    = info.get("fullTimeEmployees")
            exchange_disp= info.get("exchange","")

            st.markdown(f"""
            <div style="background:rgba(22,32,50,0.8);border:1px solid rgba(245,158,11,0.15);
                        border-radius:10px;padding:16px 22px;margin-bottom:14px">
              <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">
                <div>
                  <span style="color:#F59E0B;font-size:1.3rem;font-weight:800">{company_name}</span>
                  <span style="color:#64748B;font-size:0.88rem;margin-left:10px">{pick} · {qs['sector']} · {qs['industry']}</span>
                </div>
                <div style="text-align:right">
                  <span style="color:#64748B;font-size:0.82rem">{country} · {exchange_disp}</span>
                  {f'<br><a href="{website}" style="color:#3B82F6;font-size:0.82rem">{website}</a>' if website else ''}
                  {f'<br><span style="color:#64748B;font-size:0.82rem">{employees:,} employees</span>' if employees else ''}
                </div>
              </div>
              <div style="color:#94A3B8;font-size:0.88rem;line-height:1.6">
                {long_biz[:500] + '…' if len(long_biz) > 500 else long_biz or 'Business description not available.'}
              </div>
            </div>""", unsafe_allow_html=True)

            # ── Key metric cards ──────────────────────────────────
            price_display = get_price_display(curr_px, pick, info)
            raw_currency  = info.get("currency","")
            price_note    = "LSE price in pence (GBp) — shown as £/pence" if raw_currency=="GBp" else "Live market price"

            # Dividend info
            div_rate     = _f(info.get("dividendRate") or info.get("trailingAnnualDividendRate"))
            div_yield_v  = _f(info.get("dividendYield") or info.get("trailingAnnualDividendYield"))
            last_div_dt  = info.get("lastDividendDate") or info.get("exDividendDate")
            fy_end_month = info.get("lastFiscalYearEnd") or info.get("mostRecentQuarter")
            if last_div_dt:
                try:
                    import datetime as _dt
                    last_div_str = _dt.datetime.fromtimestamp(int(last_div_dt)).strftime("%d %b %Y")
                except Exception:
                    last_div_str = str(last_div_dt)
            else:
                last_div_str = "–"
            if fy_end_month:
                try:
                    import datetime as _dt
                    fy_str = _dt.datetime.fromtimestamp(int(fy_end_month)).strftime("%b %Y")
                except Exception:
                    fy_str = str(fy_end_month)
            else:
                fy_str = "–"
            div_display = get_price_display(div_rate, pick, info) if div_rate else "–"
            div_yield_display = f"{div_yield_v*100:.2f}%" if div_yield_v else "–"

            # Row 1 — core metrics
            m1,m2,m3,m4,m5,m6 = st.columns(6)
            _mc(m1,"Quality Score",f"{qs['score']}/100",
                "🟢 Strong" if qs['score']>=80 else "🟡 Good" if qs['score']>=60 else "🔴 Weak",
                "metric-card-gold")
            _mc(m2,"Current Price", price_display, price_note)
            _mc(m3,"ROE",
                fmt_pct(qs["roe"]) if qs["roe"] else "–","Return on Equity","metric-card-green")
            _mc(m4,"P/E Ratio",
                fmt_number(qs["pe"],1) if qs["pe"] else "–","Trailing P/E")
            _mc(m5,"Gross Margin",
                fmt_pct(qs["gross_margin"]) if qs["gross_margin"] else "–","Profitability")
            _mc(m6,"Debt/Equity",
                fmt_number(qs["debt_equity"],2) if qs["debt_equity"] is not None else "–",
                "Leverage","metric-card-red" if (qs["debt_equity"] or 0)>2 else "")

            # Row 2 — financial calendar + dividends
            d1,d2,d3,d4 = st.columns(4)
            _mc(d1,"FY End", fy_str, "Latest financial year end")
            _mc(d2,"Last Dividend Date", last_div_str, "Most recent ex-div date")
            _mc(d3,"Dividend / Share", div_display, f"Yield: {div_yield_display}")
            _fwd_pe  = _f(info.get("forwardPE"))
            _mc(d4,"Forward P/E",
                fmt_number(_fwd_pe,1) if _fwd_pe else "–",
                "Forward earnings estimate")

            # ─────────────────────────────────────────────────────
            # INTRINSIC VALUE CALCULATOR — Full Context + 3 Methods
            # ─────────────────────────────────────────────────────
            st.markdown('<div class="section-header">💎 Intrinsic Value Calculator</div>',
                        unsafe_allow_html=True)

            # ── Pull 5-year historical financials ─────────────────
            try:
                tk_val  = yf.Ticker(pick)
                fin_ann = tk_val.financials    # columns = dates (most recent first)
                cf_ann  = tk_val.cashflow      # for capex / reinvestment rate
                shares  = _f(info.get("sharesOutstanding") or info.get("impliedSharesOutstanding"))

                # Sector benchmarks
                SECTOR_OP_MARGINS = {
                    "Technology":20,"Healthcare":15,"Financials":25,
                    "Consumer Discretionary":8,"Consumer Staples":10,
                    "Energy":12,"Materials":14,"Industrials":12,
                    "Utilities":18,"Real Estate":35,
                    "Communication Services":20,"Other":12,
                }
                SECTOR_REV_GROWTH = {
                    "Technology":12,"Healthcare":8,"Financials":6,
                    "Consumer Discretionary":7,"Consumer Staples":5,
                    "Energy":5,"Materials":6,"Industrials":7,
                    "Utilities":3,"Real Estate":5,
                    "Communication Services":6,"Other":5,
                }
                SECTOR_INV_RATE = {
                    "Technology":15,"Healthcare":20,"Financials":10,
                    "Consumer Discretionary":20,"Consumer Staples":15,
                    "Energy":30,"Materials":25,"Industrials":25,
                    "Utilities":35,"Real Estate":20,
                    "Communication Services":18,"Other":20,
                }
                sect_op  = SECTOR_OP_MARGINS.get(sector_name, 12)
                sect_rg  = SECTOR_REV_GROWTH.get(sector_name, 6)
                sect_inv = SECTOR_INV_RATE.get(sector_name, 20)

                # Determine reference year for sector averages
                _cur_year = datetime.now().year
                _sect_yr  = f"{_cur_year-1}/{_cur_year}"  # e.g. "2024/2025 avg"

                hist_rows  = []
                raw_revs   = []   # for CAGR calc
                raw_eps_list = [] # for earnings growth calc
                if fin_ann is not None and not fin_ann.empty:
                    # Take up to 6 columns so we can compute growth for the first shown year
                    _all_cols = list(reversed(fin_ann.columns[:6]))  # oldest → newest

                    # Build capex lookup from cashflow statement
                    _capex_map = {}
                    if cf_ann is not None and not cf_ann.empty:
                        _capex_keys = ["Capital Expenditure", "Purchase Of PPE",
                                       "Capital Expenditures", "Purchases Of Property Plant And Equipment"]
                        for _ck in _capex_keys:
                            if _ck in cf_ann.index:
                                for _cd in cf_ann.columns:
                                    v = _f(cf_ann.loc[_ck, _cd])
                                    if v is not None:
                                        _capex_map[_cd.year] = abs(v)
                                break

                    def _fmt_pct(val):
                        """Format a percentage value without + sign. Negative shows negative."""
                        if val is None: return "–"
                        return f"{val:.1f}%"

                    prev_rev = None
                    prev_ni  = None
                    for i, col_date in enumerate(_all_cols):
                        yr   = str(col_date.year)
                        rev  = _f(fin_ann.loc["Total Revenue", col_date]) if "Total Revenue" in fin_ann.index else None
                        ebit = _f(fin_ann.loc["EBIT",          col_date]) if "EBIT"          in fin_ann.index else None
                        ni   = _f(fin_ann.loc["Net Income",    col_date]) if "Net Income"    in fin_ann.index else None

                        op_margin   = round(ebit/rev*100, 1) if ebit is not None and rev and rev != 0 else None
                        net_margin  = round(ni  /rev*100, 1) if ni   is not None and rev and rev != 0 else None
                        rev_growth_yoy = round((rev/prev_rev - 1)*100, 1) if prev_rev and prev_rev != 0 and rev else None
                        ni_growth_yoy  = round((ni /prev_ni  - 1)*100, 1) if prev_ni  and prev_ni  != 0 and ni  else None

                        # Reinvestment rate from cashflow capex / EBIT
                        capex_val = _capex_map.get(col_date.year)
                        if capex_val and ebit and ebit != 0:
                            inv_rate = round(capex_val / abs(ebit) * 100, 1)
                        else:
                            inv_rate = None   # will show sector avg in sector row only

                        raw_revs.append(rev)
                        if ni:
                            raw_eps_list.append(ni)

                        # First column is the base year for growth calculation — skip from display
                        if i == 0:
                            prev_rev = rev; prev_ni = ni
                            continue

                        hist_rows.append({
                            "Year":                  yr,
                            f"Revenue ({sym_cur})":  fmt_currency(rev, sym_cur) if rev else "–",
                            "Rev Growth %":          _fmt_pct(rev_growth_yoy),
                            "Earnings Growth %":     _fmt_pct(ni_growth_yoy),
                            "Operating Margin %":    _fmt_pct(op_margin),
                            "Net Margin %":          _fmt_pct(net_margin),
                            "Reinvestment Rate %":   f"{inv_rate:.1f}%" if inv_rate is not None else "–",
                            "EBIT":                  fmt_currency(ebit, sym_cur) if ebit else "–",
                        })
                        prev_rev = rev
                        prev_ni  = ni

                # Sector averages row — show reference year
                hist_rows.append({
                    "Year":                  f"⚖️ Sector Avg ({_sect_yr})",
                    f"Revenue ({sym_cur})":  "–",
                    "Rev Growth %":          f"~{sect_rg}%",
                    "Earnings Growth %":     "–",
                    "Operating Margin %":    f"~{sect_op}%",
                    "Net Margin %":          "–",
                    "Reinvestment Rate %":   f"~{sect_inv}%",
                    "EBIT":                  "–",
                })
            except Exception as _he:
                hist_rows    = []
                raw_revs     = []
                raw_eps_list = []
                shares       = _f(info.get("sharesOutstanding") or info.get("impliedSharesOutstanding"))
                sect_op = SECTOR_OP_MARGINS.get(sector_name, 12) if 'SECTOR_OP_MARGINS' in dir() else 12
                sect_rg = SECTOR_REV_GROWTH.get(sector_name, 6)  if 'SECTOR_REV_GROWTH'  in dir() else 6
                sect_inv = 20

            # ── Show historical context table ─────────────────────
            if hist_rows:
                st.markdown(f"""
                <div style="background:rgba(245,158,11,0.06);border:1px solid rgba(245,158,11,0.2);
                            border-radius:8px;padding:12px 18px;margin-bottom:12px;">
                  <span style="color:#F59E0B;font-weight:700;font-size:1rem;">
                    📋 {pick} — 5-Year Historical Financials + Sector Benchmarks
                  </span>
                  <span style="color:#64748B;font-size:0.85rem;margin-left:10px;">
                    Use these to anchor your DCF assumptions · ⚖️ row = sector average benchmark &nbsp;·&nbsp; * = sector avg used where capex data unavailable
                  </span>
                </div>""", unsafe_allow_html=True)
                df_hist = pd.DataFrame(hist_rows)

                # Styling: white for normal values, red for negatives, gold for sector row — NO green
                _growth_cols = ["Rev Growth %", "Earnings Growth %"]
                _margin_cols = ["Operating Margin %", "Net Margin %", "Reinvestment Rate %"]
                _all_style_cols = _growth_cols + _margin_cols

                def _colour_hist(v):
                    s = str(v)
                    if s.startswith("-"):
                        return "color:#DC2626;font-weight:600"   # red for negatives
                    if s in ("–", "nan%", ""):
                        return "color:#6B7280"                   # muted grey for missing
                    return "color:#111827;font-weight:500"       # dark/black for positive values

                def _colour_sector_row(row):
                    """Gold italic for the sector benchmark row, dark text for data cells."""
                    if "Sector Avg" in str(row.get("Year", "")):
                        return ["color:#92400E;font-style:italic;font-weight:600"] * len(row)
                    return ["color:#111827"] * len(row)

                def _colour_year_col(v):
                    s = str(v)
                    if "Sector Avg" in s:
                        return "color:#92400E;font-style:italic;font-weight:600"
                    return "color:#1E3A5F;font-weight:700"  # dark blue for year labels

                st.dataframe(
                    df_hist.style
                        .set_properties(**{"background-color": "#F8FAFC"})
                        .apply(_colour_sector_row, axis=1)
                        .map(_colour_hist, subset=[c for c in _all_style_cols if c in df_hist.columns])
                        .map(_colour_year_col, subset=["Year"]),
                    use_container_width=True, hide_index=True)

                raw_revs_clean = [r for r in raw_revs if r]
                if len(raw_revs_clean) >= 2:
                    avg_rev_growth = ((raw_revs_clean[-1]/raw_revs_clean[0])**(1/max(len(raw_revs_clean)-1,1))-1)*100
                else:
                    avg_rev_growth = sect_rg
                # Earnings CAGR
                raw_eps_clean = [e for e in raw_eps_list if e and e > 0]
                if len(raw_eps_clean) >= 2:
                    avg_earn_growth = ((raw_eps_clean[-1]/raw_eps_clean[0])**(1/max(len(raw_eps_clean)-1,1))-1)*100
                    earn_str = f" · Earnings CAGR: **{avg_earn_growth:.1f}%/yr**"
                else:
                    earn_str = ""
                st.caption(f"📈 Revenue CAGR: **{avg_rev_growth:.1f}%/yr** vs sector avg **~{sect_rg}%/yr**{earn_str} — "
                           f"use as baseline for your DCF sliders below")
            else:
                avg_rev_growth = sect_rg
                st.info("Historical financials not available for this ticker — DCF will use EPS-based estimates.")

            # ── Three valuation methods ────────────────────────────
            st.markdown("""
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin:16px 0 8px 0">
              <div style="background:rgba(30,58,95,0.3);border-radius:8px;padding:12px 16px;
                          border:1px solid rgba(245,158,11,0.15);font-size:0.88rem;color:#94A3B8">
                <div style="color:#F59E0B;font-weight:700;margin-bottom:6px">📊 Method 1: DCF (Revenue-Based)</div>
                Project future free cash flows based on revenue growth + margin assumptions.
                Discount them back to today. Best for growing businesses with predictable margins.
                <span style="color:#4ADE80;font-weight:600"> Set inputs below →</span>
              </div>
              <div style="background:rgba(30,58,95,0.3);border-radius:8px;padding:12px 16px;
                          border:1px solid rgba(245,158,11,0.15);font-size:0.88rem;color:#94A3B8">
                <div style="color:#F59E0B;font-weight:700;margin-bottom:6px">📐 Method 2: Graham Number</div>
                Benjamin Graham's formula: √(22.5 × EPS × Book Value).
                Gives a <strong style="color:#F1F5F9">conservative floor value</strong> — stock below Graham Number = classic value buy.
                Works best for profitable, asset-heavy businesses.
              </div>
              <div style="background:rgba(30,58,95,0.3);border-radius:8px;padding:12px 16px;
                          border:1px solid rgba(245,158,11,0.15);font-size:0.88rem;color:#94A3B8">
                <div style="color:#F59E0B;font-weight:700;margin-bottom:6px">🏭 Method 3: Industry P/E</div>
                What would this stock be worth if priced at its sector's average P/E?
                If the stock P/E &lt; sector P/E = <strong style="color:#4ADE80">potential re-rating opportunity</strong>.
                Sector avg used: {sector_name} = {sector_pe}x
              </div>
            </div>
            """.format(sector_name=sector_name, sector_pe=sector_pe), unsafe_allow_html=True)

            # ── DCF Inputs — three time horizons ──────────────────
            st.markdown('<div class="section-header" style="margin-top:8px">⚙️ DCF Forecast Assumptions — 3 Time Periods</div>',
                        unsafe_allow_html=True)

            # ── Smart DCF Pre-Seeding (programmatic + optional Claude AI) ──────────
            _ai_dcf_key      = f"ai_dcf_assumptions_{pick}"
            _ai_dcf_exp_key  = f"ai_dcf_explanation_{pick}"
            _smart_seeded_key = f"smart_seeded_{pick}"  # track auto-seed per stock

            # ── Step 1: Compute per-stock defaults from actual data ───────────────

            # Historical rev growth / op margin (needed here and below for sliders)
            try:
                _arv = float(avg_rev_growth) if 'avg_rev_growth' in dir() else float('nan')
                _hist_rg = max(-50, min(150, int(round(_arv)))) if _arv == _arv else 8
            except Exception:
                _hist_rg = 8
            try:
                _gm = qs.get("gross_margin") or 0
                _hist_om = max(-50, min(100, int(round(float(_gm) * 100 * 0.6)))) if _gm and float(_gm)==float(_gm) else 20
            except Exception:
                _hist_om = 20

            # CAPM discount rate: Re = Rf + β × ERP; then blend with cost of debt
            _beta_raw = _f(info.get("beta"))
            _rf_rate  = 4.5   # UK 10-yr gilt yield (2026)
            _erp      = 5.0   # UK equity risk premium
            _beta_val = _beta_raw if (_beta_raw and _beta_raw == _beta_raw and 0.1 <= _beta_raw <= 5.0) else 1.0
            _ke       = _rf_rate + _beta_val * _erp
            _de_ratio = _f(qs.get("debt_equity")) or 0.0
            if _de_ratio and _de_ratio > 0:
                _wd  = min(0.6, _de_ratio / (1 + _de_ratio))
                _kd  = _rf_rate + 1.5   # cost of debt = rf + 150bp credit spread
                _wacc_calc = _ke * (1 - _wd) + _kd * (1 - 0.25) * _wd
            else:
                _wd  = 0.0
                _wacc_calc = _ke
            _smart_dr = max(6, min(20, int(round(_wacc_calc))))
            if _beta_raw and _beta_raw == _beta_raw:
                _capm_line = (f"CAPM: {_rf_rate}% (Rf) + {_beta_val:.2f}β × {_erp}% (ERP) = {_ke:.1f}% cost of equity"
                              + (f"; blended WACC with {_de_ratio:.1f}x D/E = {_wacc_calc:.1f}%" if _wd > 0 else ""))
            else:
                _capm_line = f"Beta unavailable — used default WACC of {_smart_dr}%"

            # Effective tax rate from actual financials (net income / pretax income)
            _smart_tax = 25
            _tax_line  = "Used sector default of 25%"
            try:
                _eff_taxes = []
                if not fin_ann.empty:
                    for _fc in list(fin_ann.columns)[:4]:
                        _ni_v  = _f(fin_ann.loc["Net Income", _fc])    if "Net Income"    in fin_ann.index else None
                        _pt_v  = _f(fin_ann.loc["Pretax Income", _fc]) if "Pretax Income" in fin_ann.index else None
                        if _ni_v and _pt_v and _pt_v > 0 and _ni_v > 0 and _pt_v != 0:
                            _et = (1 - _ni_v / _pt_v) * 100
                            if 5 <= _et <= 45:
                                _eff_taxes.append(_et)
                if _eff_taxes:
                    _avg_tax  = sum(_eff_taxes) / len(_eff_taxes)
                    _smart_tax = max(10, min(40, int(round(_avg_tax))))
                    _tax_line  = f"Avg effective tax rate ({len(_eff_taxes)} yrs of financials): {_avg_tax:.1f}%"
            except Exception:
                pass

            # Reinvestment rate from capex vs EBIT (capex already in _capex_map)
            _smart_ir = 30
            _ir_line  = "Used sector default of 30%"
            try:
                _ir_vals = []
                if _capex_map and not fin_ann.empty:
                    for _yr, _cx in _capex_map.items():
                        for _fc in fin_ann.columns:
                            if hasattr(_fc, 'year') and _fc.year == _yr:
                                _ebit = (_f(fin_ann.loc["Operating Income", _fc])
                                         if "Operating Income" in fin_ann.index else None)
                                if _ebit and _ebit > 0:
                                    _ir_v = abs(_cx) / _ebit * 100
                                    if 0 <= _ir_v <= 200:
                                        _ir_vals.append(_ir_v)
                if _ir_vals:
                    _avg_ir   = sum(_ir_vals) / len(_ir_vals)
                    _smart_ir = max(-10, min(200, int(round(_avg_ir))))
                    _ir_line  = f"Avg CapEx/EBIT ({len(_ir_vals)} yrs): {_avg_ir:.1f}% → seeded at {_smart_ir}%"
            except Exception:
                pass

            # Phase-specific defaults
            _smart_rg_s = _hist_rg
            _smart_om_s = _hist_om
            _smart_rg_m = max(-30, min(100, int(round(_hist_rg * 0.6))))
            _smart_om_m = max(-30, min(100, _hist_om - 1))
            _smart_rg_l = max(-10, min(50, sect_rg))
            _smart_om_l = max(-10, min(100, _hist_om - 2))
            _smart_tg   = min(3, max(0, int(round(_smart_rg_l * 0.3))))

            _smart_all = {
                "tax_rate":    _smart_tax,
                "inv_short":   _smart_ir,
                "inv_med":     _smart_ir,
                "inv_long":    max(10, int(_smart_ir * 0.6)),
                "discount_r":  _smart_dr,
                "rg_short":    _smart_rg_s,
                "om_short":    _smart_om_s,
                "rg_med":      _smart_rg_m,
                "om_med":      _smart_om_m,
                "rg_long":     _smart_rg_l,
                "om_long":     _smart_om_l,
                "terminal_g":  _smart_tg,
            }
            _slider_key_map = [
                (f"tr_{pick}",   "tax_rate"),   (f"irs_{pick}", "inv_short"),
                (f"irm_{pick}",  "inv_med"),    (f"irl_{pick}", "inv_long"),
                (f"dr_{pick}",   "discount_r"), (f"rgs_{pick}", "rg_short"),
                (f"oms_{pick}",  "om_short"),   (f"rgm_{pick}", "rg_med"),
                (f"omm_{pick}",  "om_med"),     (f"rgl_{pick}", "rg_long"),
                (f"oml_{pick}",  "om_long"),    (f"tg_{pick}",  "terminal_g"),
            ]

            # Auto-seed slider session state keys on first load of this stock
            if not st.session_state.get(_smart_seeded_key, False):
                for _sk, _dk in _slider_key_map:
                    st.session_state[_sk] = _smart_all[_dk]
                st.session_state[_smart_seeded_key] = True

            # Build smart explanation text (always shown)
            _smart_exp = (
                f"<b>Discount Rate / WACC</b> — {_capm_line}. Final WACC = <b>{_smart_dr}%</b>.<br><br>"
                f"<b>Effective Tax Rate</b> — {_tax_line}. Using <b>{_smart_tax}%</b> in model.<br><br>"
                f"<b>Reinvestment Rate</b> — {_ir_line}. Using <b>{_smart_ir}%</b> of NOPAT.<br><br>"
                f"<b>Revenue Growth — Short Term</b> — 5-yr historical CAGR: {_hist_rg}% → seeded at <b>{_smart_rg_s}%</b> for Yrs 1–3.<br>"
                f"<b>Revenue Growth — Medium Term</b> — Moderates to <b>{_smart_rg_m}%</b> (60% of near-term trend) for Yrs 4–7.<br>"
                f"<b>Revenue Growth — Long Term</b> — Converges to sector average of <b>{sect_rg}%</b> for Yrs 8–10.<br><br>"
                f"<b>Operating Margin</b> — Estimated from gross margin (×0.6 adj for OpEx). "
                f"Sector avg: {sect_op}%. Starting at <b>{_smart_om_s}%</b>, fading to <b>{_smart_om_l}%</b> in steady-state.<br><br>"
                f"<b>Terminal Growth</b> — Set at <b>{_smart_tg}%</b> (~30% of long-term revenue growth, anchored to GDP)."
            )

            # ── Step 2: Button row ────────────────────────────────────────────────
            _btn_c1, _btn_c2, _ = st.columns([3, 2, 2])
            with _btn_c1:
                if st.button(f"🤖 Claude AI: Enhance DCF Assumptions for {pick}",
                             key=f"ai_dcf_btn_{pick}",
                             help="Uses Claude AI for deeper qualitative analysis on top of the smart defaults"):
                    with st.spinner("Claude is analysing financials, sector trends and economic context…"):
                        try:
                            import anthropic as _anth
                            _ai_c = _anth.Anthropic()
                            _hist_summary = ""
                            if hist_rows:
                                for _hr in hist_rows[:-1]:
                                    _hist_summary += (
                                        f"  {_hr.get('Year','?')}: Rev={_hr.get(f'Revenue ({sym_cur})','–')}, "
                                        f"RevGrowth={_hr.get('Rev Growth %','–')}, "
                                        f"EarnGrowth={_hr.get('Earnings Growth %','–')}, "
                                        f"OpMargin={_hr.get('Operating Margin %','–')}, "
                                        f"NetMargin={_hr.get('Net Margin %','–')}, "
                                        f"Reinvest={_hr.get('Reinvestment Rate %','–')}\n"
                                    )
                            _dcf_prompt = f"""You are a senior equity research analyst. Suggest intelligent DCF assumption inputs for a 3-phase DCF model.

COMPANY: {info.get('longName', pick)} ({pick})
SECTOR: {sector_name} | INDUSTRY: {qs.get('industry','?')}
DESCRIPTION: {info.get('longBusinessSummary','')[:600]}

KEY METRICS:
- Price: {get_price_display(curr_px, pick, info)} | Quality Score: {qs['score']}/100
- ROE: {fmt_pct(qs['roe']) if qs['roe'] else '?'} | Gross Margin: {fmt_pct(qs['gross_margin']) if qs['gross_margin'] else '?'}
- P/E: {fmt_number(qs['pe'],1) if qs['pe'] else '?'} | Beta: {f'{_beta_val:.2f}' if _beta_raw else 'N/A'} | D/E: {fmt_number(qs['debt_equity'],2) if qs['debt_equity'] is not None else '?'}

SMART DEFAULT CONTEXT (already computed):
- CAPM WACC: {_smart_dr}% ({_capm_line})
- Effective Tax: {_smart_tax}% ({_tax_line})
- Reinvestment Rate: {_smart_ir}% ({_ir_line})
- Historical Rev CAGR: {_hist_rg}% | Hist Op Margin: {_hist_om}%

5-YEAR HISTORICAL FINANCIALS:
{_hist_summary if _hist_summary else '  Not available'}

SECTOR BENCHMARKS ({sector_name}): Rev growth ~{sect_rg}%/yr | Op margin ~{sect_op}% | Reinvest ~{sect_inv}%

ECONOMIC CONTEXT (2026): UK rate 4.5%, US 4.3%, UK inflation 2.6%, GDP growth UK ~1%, US ~2.1%

YOUR TASK: Review the smart defaults above. Adjust if your qualitative analysis of the company warrants it.
Return ONLY valid JSON (no markdown):
{{
  "rg_short": <int -50 to 150>, "om_short": <int -50 to 100>,
  "rg_med": <int -30 to 100>, "om_med": <int -30 to 100>,
  "rg_long": <int -10 to 50>, "om_long": <int -10 to 100>,
  "tax_rate": <int 10 to 40>,
  "inv_short": <int -10 to 200>, "inv_med": <int -10 to 200>, "inv_long": <int -10 to 200>,
  "discount_r": <int 6 to 20>, "terminal_g": <int -2 to 8>,
  "reasoning": "<3-paragraph explanation citing specific data points, how you adjusted vs smart defaults, and why>"
}}"""
                            _resp = _ai_c.messages.create(
                                model="claude-sonnet-5", max_tokens=1500,
                                messages=[{"role": "user", "content": _dcf_prompt}]
                            )
                            import json as _json
                            _raw = _resp.content[0].text.strip()
                            if _raw.startswith("```"):
                                _raw = _raw.split("```")[1]
                                if _raw.startswith("json"): _raw = _raw[4:]
                            _parsed = _json.loads(_raw)
                            # Directly write into slider session-state keys so they update immediately
                            for _sk, _dk in _slider_key_map:
                                if _dk in _parsed:
                                    _v = _parsed[_dk]
                                    if isinstance(_v, (int, float)) and _v == _v:
                                        st.session_state[_sk] = int(round(_v))
                            st.session_state[_ai_dcf_key]     = _parsed
                            st.session_state[_ai_dcf_exp_key] = _parsed.get("reasoning", "")
                            st.rerun()
                        except ImportError:
                            st.toast("Claude AI package not installed — smart defaults are already applied.", icon="ℹ️")
                        except Exception as _e:
                            st.error(f"AI enhancement failed: {_e}")

            with _btn_c2:
                if st.button("↺ Reset Smart Defaults", key=f"rst_dcf_{pick}"):
                    for _sk, _dk in _slider_key_map:
                        st.session_state[_sk] = _smart_all[_dk]
                    for _k in [_ai_dcf_key, _ai_dcf_exp_key]:
                        if _k in st.session_state: del st.session_state[_k]
                    st.toast(f"Sliders reset to smart defaults for {pick}", icon="✅")
                    st.rerun()

            # ── Collapsible explanation panel ─────────────────────────────────────
            _using_ai  = bool(st.session_state.get(_ai_dcf_exp_key, ""))
            _exp_title = "🤖 Claude AI — DCF Assumption Rationale" if _using_ai else "📊 View Smart Defaults Rationale — CAPM + Historical Data"
            _exp_body  = st.session_state.get(_ai_dcf_exp_key, "").replace("\n", "<br>") if _using_ai else _smart_exp
            with st.expander(_exp_title, expanded=False):
                st.markdown(f"""
                <div style="color:#CBD5E1;font-size:0.87rem;line-height:1.9;padding:4px 0">
                  {_exp_body}
                </div>
                <div style="color:#475569;font-size:0.75rem;margin-top:10px;
                            border-top:1px solid rgba(255,255,255,0.06);padding-top:8px">
                  ⚠️ Starting-point estimates only. Adjust sliders to reflect your own view. Not financial advice.
                </div>""", unsafe_allow_html=True)

            st.caption("Adjust the sliders below. Changes apply immediately to the DCF valuation.")

            # ── Shared constants (per-stock via session state seeded above) ──────
            sh1, sh2 = st.columns(2)
            with sh1:
                tax_rate   = st.slider("Operating Tax Rate %", 10, 40,
                                       st.session_state.get(f"tr_{pick}", _smart_tax),
                                       key=f"tr_{pick}")
            with sh2:
                discount_r = st.slider("Discount Rate / WACC %", 6, 20,
                                       st.session_state.get(f"dr_{pick}", _smart_dr),
                                       key=f"dr_{pick}",
                                       help=f"CAPM-derived: Rf {_rf_rate}% + β{_beta_val:.2f} × ERP{_erp}% = {_ke:.1f}%. Adjust for leverage/risk.")

            # ── Phase-specific investment rates ──────────────────────────────────
            st.markdown(
                '<div style="color:#94A3B8;font-size:0.8rem;margin:10px 0 4px 0">'
                '<b>Reinvestment Rate % of NOPAT — by phase</b> &nbsp;·&nbsp; '
                'Set how much of NOPAT is reinvested back into the business in each phase. '
                '0% = all NOPAT is free cash flow. >100% = investing more than NOPAT (growth burn).'
                '</div>', unsafe_allow_html=True)
            _ir1, _ir2, _ir3 = st.columns(3)
            with _ir1:
                inv_short = st.slider("Inv Rate % — Yrs 1–3", -10, 200,
                                      st.session_state.get(f"irs_{pick}", _smart_ir),
                                      key=f"irs_{pick}",
                                      help="High-growth phase — often reinvests heavily. >100% if growth is funded externally.")
            with _ir2:
                inv_med   = st.slider("Inv Rate % — Yrs 4–7", -10, 200,
                                      st.session_state.get(f"irm_{pick}", _smart_ir),
                                      key=f"irm_{pick}",
                                      help="Maturing phase — reinvestment typically moderates as growth slows.")
            with _ir3:
                inv_long  = st.slider("Inv Rate % — Yrs 8–10", -10, 200,
                                      st.session_state.get(f"irl_{pick}", max(10, int(_smart_ir * 0.6))),
                                      key=f"irl_{pick}",
                                      help="Steady-state — lower reinvestment as company approaches terminal phase. 10–40% is typical.")

            st.markdown("""
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin:14px 0 6px 0">
              <div style="background:rgba(34,197,94,0.08);border:1px solid rgba(34,197,94,0.25);
                          border-radius:8px;padding:10px 14px">
                <div style="color:#4ADE80;font-weight:700;font-size:0.88rem;margin-bottom:4px">
                  🌱 SHORT TERM — Years 1–3
                </div>
                <div style="color:#64748B;font-size:0.8rem">
                  High-growth phase. Typically above-average revenue growth.
                  Margins may still be expanding. Use the most recent 1–2 years as baseline.
                </div>
              </div>
              <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.25);
                          border-radius:8px;padding:10px 14px">
                <div style="color:#F59E0B;font-weight:700;font-size:0.88rem;margin-bottom:4px">
                  📈 MEDIUM TERM — Years 4–7
                </div>
                <div style="color:#64748B;font-size:0.8rem">
                  Growth begins to moderate. Business matures.
                  Margins should be near sustainable levels. Use sector CAGR as reference.
                </div>
              </div>
              <div style="background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.25);
                          border-radius:8px;padding:10px 14px">
                <div style="color:#60A5FA;font-weight:700;font-size:0.88rem;margin-bottom:4px">
                  🏁 LONG TERM — Years 8–10 + Terminal
                </div>
                <div style="color:#64748B;font-size:0.8rem">
                  Steady-state. Growth converges toward GDP/sector average (2–4%).
                  Margins stabilise. Terminal value dominates total DCF.
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Short-term inputs
            st.markdown('<div style="color:#4ADE80;font-weight:700;font-size:0.88rem;margin:10px 0 4px 0">🌱 Short Term — Years 1–3</div>', unsafe_allow_html=True)
            s1c1, s1c2 = st.columns(2)
            with s1c1:
                rg_short = st.slider("Revenue Growth % (Yrs 1–3)", -50, 150,
                                     st.session_state.get(f"rgs_{pick}", _smart_rg_s),
                                     key=f"rgs_{pick}",
                                     help="Pre-seeded from 5-yr historical CAGR. Negative = contraction. >100% for hyper-growth.")
            with s1c2:
                om_short = st.slider("Operating Margin % (Yrs 1–3)", -50, 100,
                                     st.session_state.get(f"oms_{pick}", _smart_om_s),
                                     key=f"oms_{pick}",
                                     help="Pre-seeded from gross margin (×0.6). Negative = loss-making.")

            # Medium-term inputs
            st.markdown('<div style="color:#F59E0B;font-weight:700;font-size:0.88rem;margin:10px 0 4px 0">📈 Medium Term — Years 4–7</div>', unsafe_allow_html=True)
            s2c1, s2c2 = st.columns(2)
            with s2c1:
                rg_med = st.slider("Revenue Growth % (Yrs 4–7)", -30, 100,
                                   st.session_state.get(f"rgm_{pick}", _smart_rg_m),
                                   key=f"rgm_{pick}",
                                   help="Growth typically moderates in medium term.")
            with s2c2:
                om_med = st.slider("Operating Margin % (Yrs 4–7)", -30, 100,
                                   st.session_state.get(f"omm_{pick}", _smart_om_m),
                                   key=f"omm_{pick}",
                                   help="Margins approaching sustainable levels by this phase.")

            # Long-term inputs
            st.markdown('<div style="color:#60A5FA;font-weight:700;font-size:0.88rem;margin:10px 0 4px 0">🏁 Long Term — Years 8–10 + Terminal</div>', unsafe_allow_html=True)
            s3c1, s3c2, s3c3, s3c4 = st.columns(4)
            with s3c1:
                rg_long = st.slider("Revenue Growth % (Yrs 8–10)", -10, 50,
                                    st.session_state.get(f"rgl_{pick}", _smart_rg_l),
                                    key=f"rgl_{pick}",
                                    help="Converges toward GDP/sector average. Pre-seeded to sector avg.")
            with s3c2:
                om_long = st.slider("Operating Margin % (Yrs 8–10)", -10, 100,
                                    st.session_state.get(f"oml_{pick}", _smart_om_l),
                                    key=f"oml_{pick}",
                                    help="Steady-state margin — reflects long-run competitive position.")
            with s3c3:
                terminal_growth = st.slider("Terminal Growth % (beyond Yr 10)", -2, 8,
                                            st.session_state.get(f"tg_{pick}", _smart_tg),
                                            key=f"tg_{pick}",
                                            help="Perpetual growth rate (must be < WACC). Typically = inflation + GDP.")
            with s3c4:
                # Smart default RONIC: estimate from actual NOPAT / Invested Capital
                # Invested Capital = Total Equity + Total Debt (book value basis)
                _total_equity = _f(info.get("totalStockholderEquity")) or 0
                _total_debt   = _f(info.get("totalDebt")) or 0
                _ebit_raw     = _f(info.get("ebit")) or 0
                _tax_dec      = (_f(info.get("effectiveTaxRate")) or 25) / 100
                _nopat_raw    = _ebit_raw * (1 - _tax_dec)
                _ic_raw       = _total_equity + _total_debt
                if _ic_raw > 0 and _nopat_raw > 0:
                    _hist_roic_pct = round(_nopat_raw / _ic_raw * 100, 1)
                    # Reasonable bounds: floor at WACC, cap at 60%
                    _smart_ronic = max(discount_r, min(60, _hist_roic_pct))
                else:
                    _smart_ronic = max(discount_r, 12)   # default if data unavailable

                ronic = st.slider(
                    "RONIC — Return on New Invested Capital %",
                    int(discount_r), 60,
                    int(st.session_state.get(f"ronic_{pick}", _smart_ronic)),
                    key=f"ronic_{pick}",
                    help=(
                        "Return on New Invested Capital used in the Continuing Value formula: "
                        "CV = NOPAT(t+1) × (1 − g/RONIC) / (WACC − g). "
                        "If RONIC = WACC, growth adds no value. "
                        "If RONIC > WACC, growth creates value. "
                        f"Historical ROIC estimate: {_hist_roic_pct:.1f}%"
                        if _ic_raw > 0 and _nopat_raw > 0 else
                        "Historical ROIC not available — defaulted to WACC+2%."
                    )
                )

            # Show RONIC impact note
            _reinvest_rate_terminal = terminal_growth / ronic if ronic > 0 else 0
            st.markdown(
                f'<div style="background:rgba(96,165,250,0.08);border-left:3px solid #60A5FA;'
                f'border-radius:0 6px 6px 0;padding:8px 12px;font-size:0.8rem;color:#94A3B8;margin:6px 0">'
                f'<b style="color:#60A5FA">Continuing Value formula:</b> '
                f'CV = NOPAT(Yr11) × (1 − {terminal_growth}%/{ronic}%) / (WACC − {terminal_growth}%) &nbsp;·&nbsp; '
                f'Reinvestment rate in terminal = <b style="color:#F1F5F9">{_reinvest_rate_terminal*100:.1f}%</b> of NOPAT &nbsp;·&nbsp; '
                f'FCF to perpetuity = <b style="color:#F1F5F9">{(1-_reinvest_rate_terminal)*100:.1f}%</b> of NOPAT'
                f'</div>',
                unsafe_allow_html=True)

            # convenience alias for display
            rev_growth  = rg_short   # used in result card summary
            op_margin_f = om_short

            # ── Revenue-based 3-phase DCF ──────────────────────────
            def calc_revenue_dcf_3phase(curr_rev, rg_s, om_s, rg_m, om_m, rg_l, om_l,
                                         tax_r, inv_s, inv_m, inv_l, disc_r, term_g, ronic_pct=15):
                """3-phase DCF: Short (1-3), Medium (4-7), Long (8-10) + McKinsey Continuing Value.

                Continuing Value formula (McKinsey Value Driver Model):
                  CV = NOPAT(t+1) × (1 − g/RONIC) / (WACC − g)

                Where:
                  NOPAT(t+1)  = Year 11 NOPAT  = Rev_10 × (1+g) × long-run margin × (1−tax)
                  g           = terminal growth rate
                  RONIC       = Return on New Invested Capital (explicit user input)
                  WACC        = discount rate

                RONIC interpretation:
                  RONIC > WACC  → growth creates value (reinvestment earns more than cost)
                  RONIC = WACC  → growth neutral (CV simplifies to NOPAT/WACC)
                  RONIC < WACC  → growth destroys value (should not reinvest; return cash)

                Explicit forecast period reinvestment:
                  inv_r < 0  → company returns MORE than NOPAT (buybacks, balance sheet cash)
                  inv_r = 0  → 100% of NOPAT as FCF
                  0 < inv_r < 100 → normal reinvestment
                  inv_r > 100 → investing MORE than NOPAT (growth burn, funded by debt/equity)
                """
                try:
                    if not curr_rev or curr_rev == 0:
                        return None
                    om_s /= 100; om_m /= 100; om_l /= 100
                    rg_s /= 100; rg_m /= 100; rg_l /= 100
                    tax_r /= 100
                    _inv_s = inv_s / 100; _inv_m = inv_m / 100; _inv_l = inv_l / 100
                    disc_r /= 100; term_g /= 100
                    ronic = ronic_pct / 100   # RONIC as decimal

                    # WACC must exceed terminal growth for a finite valuation
                    if disc_r <= term_g or disc_r <= 0:
                        return None

                    # ── Explicit forecast: Years 1–10 ──────────────────────
                    pv_total = 0.0
                    revenue  = abs(curr_rev)

                    for t in range(1, 11):
                        if   t <= 3:  rg, om, inv_r = rg_s, om_s, _inv_s
                        elif t <= 7:  rg, om, inv_r = rg_m, om_m, _inv_m
                        else:         rg, om, inv_r = rg_l, om_l, _inv_l

                        revenue *= (1 + rg)
                        if revenue < 0:
                            revenue = 0
                        effective_tax = tax_r if om > 0 else 0
                        nopat = revenue * om * (1 - effective_tax)
                        fcf = nopat * (1 - inv_r)
                        pv_total += fcf / ((1 + disc_r) ** t)

                    # ── Continuing Value (McKinsey Value Driver Formula) ───
                    # CV = NOPAT(t+1) × (1 − g/RONIC) / (WACC − g)
                    term_effective_tax = tax_r if om_l > 0 else 0
                    term_nopat = revenue * (1 + term_g) * om_l * (1 - term_effective_tax)

                    # Reinvestment rate in terminal = g / RONIC
                    if ronic > 0 and term_g > 0:
                        _term_reinvest_rate = term_g / ronic
                    elif term_g <= 0:
                        # Zero/negative growth: minimal maintenance reinvestment only
                        _term_reinvest_rate = 0.05
                    else:
                        _term_reinvest_rate = 0.0

                    # Terminal FCF = NOPAT(t+1) × (1 − g/RONIC)
                    term_fcf = term_nopat * (1 - _term_reinvest_rate)

                    if (disc_r - term_g) == 0:
                        return None
                    tv    = term_fcf / (disc_r - term_g)
                    pv_tv = tv / ((1 + disc_r) ** 10)
                    result = pv_total + pv_tv
                    if abs(result) > 1e8:
                        return None
                    return result
                except Exception:
                    return None

            curr_rev_raw  = _f(info.get("totalRevenue") or info.get("revenue"))
            curr_rev_m    = (curr_rev_raw / 1e6) if curr_rev_raw else None

            # ── Phase Diagnostics expander ─────────────────────────
            with st.expander("🔬 Phase Diagnostics — Implied NOPAT Growth & Investment Rates", expanded=False):
                st.markdown(
                    '<div style="color:#94A3B8;font-size:0.82rem;margin-bottom:8px">'
                    'Shows what your revenue growth + margin assumptions <em>actually imply</em> '
                    'for NOPAT growth at each phase, the natural investment rate '
                    '(g<sub>NOPAT</sub> / RONIC), and whether your chosen investment rate is '
                    'above or below that natural rate.'
                    '</div>', unsafe_allow_html=True)

                _tx = tax_rate / 100
                _base_rev = abs(curr_rev_m) if curr_rev_m else 1.0

                def _phase_diag(rev_start, rg_pct, om_pct, inv_pct, n_yrs):
                    rg = rg_pct / 100
                    rev_end    = rev_start * ((1 + rg) ** n_yrs)
                    nopat_cagr = rg_pct   # within a phase margin is flat so NOPAT CAGR = rev CAGR
                    ronic_dec  = ronic / 100
                    nat_ir     = (rg / ronic_dec * 100) if ronic_dec > 0 else 0
                    return rev_end, nopat_cagr, nat_ir

                _rev_s_end, _ncagr_s, _nir_s = _phase_diag(_base_rev, rg_short, om_short, inv_short, 3)
                _rev_m_end, _ncagr_m, _nir_m = _phase_diag(_rev_s_end, rg_med,  om_med,  inv_med,  4)
                _rev_l_end, _ncagr_l, _nir_l = _phase_diag(_rev_m_end, rg_long, om_long, inv_long, 3)

                def _traffic(user_ir, nat_ir):
                    diff = user_ir - nat_ir
                    if abs(diff) <= 5:   return "🟡 Approx aligned"
                    elif diff > 5:       return f"🔴 Over-investing by {diff:.0f}pp"
                    else:                return f"🟢 Under-investing by {abs(diff):.0f}pp (returning cash)"

                _diag_rows = [
                    ("Yrs 1–3",  "#4ADE80", rg_short, om_short, _ncagr_s, inv_short, _nir_s),
                    ("Yrs 4–7",  "#F59E0B", rg_med,   om_med,   _ncagr_m, inv_med,   _nir_m),
                    ("Yrs 8–10", "#60A5FA", rg_long,  om_long,  _ncagr_l, inv_long,  _nir_l),
                ]
                _diag_html = (
                    '<table style="width:100%;border-collapse:collapse;font-size:0.82rem;color:#CBD5E1">'
                    '<thead><tr style="border-bottom:1px solid rgba(255,255,255,0.1)">'
                    '<th style="text-align:left;padding:6px 8px;color:#64748B">Phase</th>'
                    '<th style="text-align:right;padding:6px 8px;color:#64748B">Rev CAGR</th>'
                    '<th style="text-align:right;padding:6px 8px;color:#64748B">Op Margin</th>'
                    '<th style="text-align:right;padding:6px 8px;color:#64748B">Implied NOPAT CAGR</th>'
                    '<th style="text-align:right;padding:6px 8px;color:#64748B">Natural Inv Rate (g/RONIC)</th>'
                    '<th style="text-align:right;padding:6px 8px;color:#64748B">Your Inv Rate</th>'
                    '<th style="text-align:left;padding:6px 8px;color:#64748B">Signal</th>'
                    '</tr></thead><tbody>'
                )
                for _ph, _col, _rg, _om, _nc, _ir, _ni in _diag_rows:
                    _signal = _traffic(_ir, _ni)
                    _diag_html += (
                        f'<tr style="border-bottom:1px solid rgba(255,255,255,0.05)">'
                        f'<td style="padding:7px 8px;color:{_col};font-weight:700">{_ph}</td>'
                        f'<td style="text-align:right;padding:7px 8px">{_rg}%</td>'
                        f'<td style="text-align:right;padding:7px 8px">{_om}%</td>'
                        f'<td style="text-align:right;padding:7px 8px;font-weight:600;color:#F1F5F9">{_nc:.1f}%</td>'
                        f'<td style="text-align:right;padding:7px 8px;color:#A78BFA">{_ni:.1f}%</td>'
                        f'<td style="text-align:right;padding:7px 8px;color:#F1F5F9">{_ir}%</td>'
                        f'<td style="padding:7px 8px">{_signal}</td>'
                        f'</tr>'
                    )
                _term_nat_ir = terminal_growth / ronic * 100 if ronic > 0 else 0
                _diag_html += (
                    f'<tr style="border-top:2px solid rgba(255,255,255,0.15);background:rgba(167,139,250,0.06)">'
                    f'<td style="padding:7px 8px;color:#C084FC;font-weight:700">Terminal</td>'
                    f'<td style="text-align:right;padding:7px 8px">{terminal_growth}%</td>'
                    f'<td style="text-align:right;padding:7px 8px">{om_long}%</td>'
                    f'<td style="text-align:right;padding:7px 8px;font-weight:600;color:#F1F5F9">{terminal_growth:.1f}%</td>'
                    f'<td style="text-align:right;padding:7px 8px;color:#A78BFA">{_term_nat_ir:.1f}%</td>'
                    f'<td style="text-align:right;padding:7px 8px;color:#94A3B8">g/RONIC (auto)</td>'
                    f'<td style="padding:7px 8px;color:#94A3B8">McKinsey formula applied</td>'
                    f'</tr>'
                    '</tbody></table>'
                    '<div style="color:#64748B;font-size:0.78rem;margin-top:8px">'
                    '🟢 Under-investing = returning cash to shareholders — good if RONIC &lt; WACC.<br>'
                    '🔴 Over-investing = burning more than NOPAT — only value-creating if RONIC &gt; WACC.<br>'
                    '<b style="color:#A78BFA">Natural Inv Rate</b> = g/RONIC — the reinvestment rate implied by your growth and RONIC.'
                    '</div>'
                )
                st.markdown(_diag_html, unsafe_allow_html=True)

            dcf_total     = calc_revenue_dcf_3phase(
                curr_rev_m, rg_short, om_short, rg_med, om_med, rg_long, om_long,
                tax_rate, inv_short, inv_med, inv_long, discount_r, terminal_growth, ronic)

            # ── Net cash adjustment: EV → Equity Value ─────────────
            # Equity Value = Enterprise Value (DCF) + Cash – Total Debt
            _cash_raw       = _f(info.get("totalCash") or info.get("cash")) or 0
            _debt_raw       = _total_debt  # already fetched above for RONIC smart default
            _net_cash_m     = (_cash_raw - _debt_raw) / 1e6  # net cash in £M / $M
            _net_cash_ps    = _net_cash_m * 1e6 / shares if shares and shares > 0 else 0
            dcf_equity_m    = (dcf_total + _net_cash_m) if dcf_total is not None else None
            dcf_per_share_gbp = round(dcf_equity_m * 1e6 / shares, 2) if dcf_equity_m is not None and shares and shares > 0 else None
            graham_val_raw   = calc_graham_number(eps, bvps)
            pe_iv_val_raw    = calc_pe_intrinsic(eps, sector_pe)
            # For GBp stocks: yfinance returns financial data (revenue, EPS, BVPS) in GBP (£)
            # but currentPrice is in GBp (pence). Convert all IV estimates to pence for consistency.
            if raw_currency == "GBp":
                dcf_per_share = round(dcf_per_share_gbp * 100, 1) if dcf_per_share_gbp is not None else None
                graham_val    = round(graham_val_raw   * 100, 1) if graham_val_raw    is not None else None
                pe_iv_val     = round(pe_iv_val_raw    * 100, 1) if pe_iv_val_raw     is not None else None
            else:
                dcf_per_share = dcf_per_share_gbp
                graham_val    = graham_val_raw
                pe_iv_val     = pe_iv_val_raw

            def _verdict(iv, price, sym, raw_cur):
                """Clear verdict: Fair Value £X vs Current £X → Over/Undervalued by X%"""
                if not iv or not price or price <= 0:
                    return "<span style='color:#64748B'>Insufficient data</span>"
                pct = (iv - price) / price * 100
                if raw_cur == "GBp":
                    iv_str  = get_price_display(iv,    pick, info)
                    px_str  = get_price_display(price, pick, info)
                else:
                    iv_str  = fmt_currency(iv,    sym)
                    px_str  = fmt_currency(price, sym)
                if pct >= 0:
                    verdict = f'<span style="color:#4ADE80;font-weight:700">▲ UNDERVALUED by {pct:.1f}%</span>'
                    sub     = f"Fair value {iv_str} vs current {px_str} — potential upside"
                else:
                    verdict = f'<span style="color:#F87171;font-weight:700">▼ OVERVALUED by {abs(pct):.1f}%</span>'
                    sub     = f"Fair value {iv_str} vs current {px_str} — priced above estimate"
                return f'{verdict}<br><span style="color:#64748B;font-size:0.8rem">{sub}</span>'

            # ── Three result cards ─────────────────────────────────
            vc1, vc2, vc3 = st.columns(3)
            with vc1:
                _dcf_d = get_price_display(dcf_per_share, pick, info) if dcf_per_share else "–"
                st.markdown(f"""
                <div class="val-card">
                  <div class="val-method">📊 Method 1 — DCF Intrinsic Value / Share</div>
                  <div class="val-price">{_dcf_d}</div>
                  {_verdict(dcf_per_share, curr_px, sym_cur, raw_currency)}
                  <div style="font-size:0.8rem;color:#64748B;margin-top:10px;border-top:1px solid rgba(255,255,255,0.06);padding-top:8px">
                    <strong style="color:#94A3B8">3-Phase Assumptions:</strong><br>
                    Base Revenue: {fmt_currency(curr_rev_m,'',0) if curr_rev_m else '?'}M<br>
                    <span style="color:#4ADE80">🌱 Yrs 1–3:</span> {rg_short}% rev growth, {om_short}% margin<br>
                    <span style="color:#F59E0B">📈 Yrs 4–7:</span> {rg_med}% rev growth, {om_med}% margin<br>
                    <span style="color:#60A5FA">🏁 Yrs 8–10:</span> {rg_long}% rev growth, {om_long}% margin<br>
                    WACC {discount_r}% · Tax {tax_rate}% · Terminal {terminal_growth}%<br>
                    Reinvest: Yrs 1–3 <b style="color:#4ADE80">{inv_short}%</b> · Yrs 4–7 <b style="color:#F59E0B">{inv_med}%</b> · Yrs 8–10 <b style="color:#60A5FA">{inv_long}%</b><br>
                    Net cash adj: <b style="color:#A78BFA">{"+{:.2f}".format(_net_cash_ps) if _net_cash_ps >= 0 else "{:.2f}".format(_net_cash_ps)}</b>/share
                  </div>
                </div>""", unsafe_allow_html=True)
                if not dcf_per_share and curr_rev_raw:
                    st.caption("⚠️ Shares outstanding not available.")
                elif not curr_rev_raw:
                    st.caption("⚠️ Revenue data not available for this ticker.")

            with vc2:
                _gr_d = get_price_display(graham_val, pick, info) if graham_val else "–"
                curr_pe_str = fmt_number(qs["pe"],1) if qs["pe"] else "–"
                st.markdown(f"""
                <div class="val-card">
                  <div class="val-method">📐 Method 2 — Graham Number</div>
                  <div class="val-price">{_gr_d}</div>
                  {_verdict(graham_val, curr_px, sym_cur, raw_currency)}
                  <div style="font-size:0.8rem;color:#64748B;margin-top:10px;border-top:1px solid rgba(255,255,255,0.06);padding-top:8px">
                    Formula: √(22.5 × EPS × Book Value/Share)<br>
                    EPS: {get_price_display(eps, pick, info) if eps else '–'} · Book Value/sh: {get_price_display(bvps, pick, info) if bvps else '–'}<br><br>
                    <span style="color:#F59E0B">📌 How to read it:</span>
                    Graham Number is a <em>conservative floor value</em> for defensive stock pickers.
                    Stock price below the Graham Number = <strong style="color:#4ADE80">potential value buy</strong>.
                    Graham himself required a 33% margin of safety (buy at ≤66% of Graham Number).
                    Best suited to low-growth, asset-backed businesses (banks, industrials).
                  </div>
                </div>""", unsafe_allow_html=True)

            with vc3:
                _pe_d = get_price_display(pe_iv_val, pick, info) if pe_iv_val else "–"
                st.markdown(f"""
                <div class="val-card">
                  <div class="val-method">🏭 Method 3 — Industry P/E Fair Value</div>
                  <div class="val-price">{_pe_d}</div>
                  {_verdict(pe_iv_val, curr_px, sym_cur, raw_currency)}
                  <div style="font-size:0.8rem;color:#64748B;margin-top:10px;border-top:1px solid rgba(255,255,255,0.06);padding-top:8px">
                    Sector avg P/E ({sector_name}): <strong style="color:#F59E0B">{sector_pe}x</strong> ×
                    EPS {get_price_display(eps, pick, info) if eps else '–'}<br>
                    This company's P/E: <strong style="color:#F1F5F9">{curr_pe_str}x</strong><br><br>
                    <span style="color:#F59E0B">📌 How to read it:</span>
                    If company P/E &lt; sector P/E → <strong style="color:#4ADE80">undervalued vs peers</strong>
                    — potential re-rating catalyst as market corrects the discount.
                    If P/E &gt; sector avg → premium pricing needing strong growth justification.
                  </div>
                </div>""", unsafe_allow_html=True)

            # ── Valuation summary banner ───────────────────────────
            vals = [(v,"DCF") for v in [dcf_per_share] if v] + \
                   [(v,"Graham") for v in [graham_val] if v] + \
                   [(v,"Industry P/E") for v in [pe_iv_val] if v]
            if vals and curr_px:
                avg_iv = sum(v for v,_ in vals) / len(vals)
                # Save to session_state so Monte Carlo Decision Dashboard can read it
                st.session_state["fintiq_dcf"] = {
                    "ticker":   pick,
                    "dcf":      dcf_per_share,
                    "graham":   graham_val,
                    "pe_iv":    pe_iv_val,
                    "avg_iv":   avg_iv,
                    "curr_px":  curr_px,
                    "sym":      sym_cur,
                }
                tot_up = (avg_iv - curr_px) / curr_px * 100
                banner_col = "#4ADE80" if tot_up >= 0 else "#F87171"
                st.markdown(f"""
                <div style="background:rgba(30,58,95,0.4);border:1px solid {banner_col}40;
                            border-radius:10px;padding:16px 22px;margin-top:12px;
                            display:flex;align-items:center;gap:20px">
                  <div>
                    <div style="color:#64748B;font-size:0.78rem;font-weight:700;text-transform:uppercase">
                      Average Intrinsic Value (All Methods)
                    </div>
                    <div style="color:#F8FAFC;font-size:1.8rem;font-weight:800">
                      {get_price_display(avg_iv, pick, info)}
                    </div>
                  </div>
                  <div>
                    <div style="color:#64748B;font-size:0.78rem;font-weight:700;text-transform:uppercase">
                      vs Current Price {get_price_display(curr_px, pick, info)}
                    </div>
                    <div style="color:{banner_col};font-size:1.8rem;font-weight:800">
                      {'▲' if tot_up>=0 else '▼'} {abs(tot_up):.1f}% {'Upside' if tot_up>=0 else 'Downside'}
                    </div>
                  </div>
                  <div style="margin-left:auto;color:#64748B;font-size:0.85rem;max-width:280px">
                    {'🟢 Below average intrinsic value — potential value opportunity. Confirm with fundamentals and catalyst before acting.' if tot_up>=0 else '🔴 Above average intrinsic value — premium priced. Needs strong growth justification or wait for better entry.'}
                  </div>
                </div>""", unsafe_allow_html=True)

            # ── Rich Valuation Multiples ───────────────────────────
            st.markdown('<div class="section-header">📐 Valuation Multiples</div>',
                        unsafe_allow_html=True)

            # Collect multiples from yfinance info
            _ev          = _f(info.get("enterpriseValue"))
            _ebitda      = _f(info.get("ebitda"))
            _fcf         = _f(info.get("freeCashflow"))
            _tot_rev     = _f(info.get("totalRevenue"))
            _mktcap      = _f(info.get("marketCap"))
            _tot_debt    = _f(info.get("totalDebt")) or 0
            _tot_cash    = _f(info.get("totalCash")) or 0
            _p2b         = _f(info.get("priceToBook"))
            _peg         = _f(info.get("pegRatio"))
            _ps          = _f(info.get("priceToSalesTrailing12Months"))
            _ev_ebitda_r = _f(info.get("enterpriseToEbitda"))
            _ev_rev_r    = _f(info.get("enterpriseToRevenue"))
            # Compute P/FCF manually
            _pfcf = (_mktcap / _fcf) if (_mktcap and _fcf and _fcf > 0) else None
            # EV/EBITDA and EV/Revenue from yfinance or manual
            _ev_ebitda = _ev_ebitda_r if _ev_ebitda_r else ((_ev / _ebitda) if (_ev and _ebitda and _ebitda > 0) else None)
            _ev_rev    = _ev_rev_r    if _ev_rev_r    else ((_ev / _tot_rev) if (_ev and _tot_rev and _tot_rev > 0) else None)

            # Sector benchmark multiples (approximate by sector)
            _sect_benchmarks = {
                "Technology":             {"ev_ebitda": 22, "ev_rev": 6,  "pfcf": 30, "ps": 6,  "pb": 8},
                "Communication Services": {"ev_ebitda": 14, "ev_rev": 3,  "pfcf": 22, "ps": 3,  "pb": 4},
                "Consumer Discretionary": {"ev_ebitda": 16, "ev_rev": 1.5,"pfcf": 22, "ps": 1.5,"pb": 5},
                "Consumer Staples":       {"ev_ebitda": 14, "ev_rev": 1.2,"pfcf": 25, "ps": 1.2,"pb": 4},
                "Financials":             {"ev_ebitda": 12, "ev_rev": 3,  "pfcf": 14, "ps": 3,  "pb": 1.5},
                "Healthcare":             {"ev_ebitda": 18, "ev_rev": 4,  "pfcf": 25, "ps": 4,  "pb": 5},
                "Industrials":            {"ev_ebitda": 14, "ev_rev": 1.5,"pfcf": 20, "ps": 1.5,"pb": 4},
                "Energy":                 {"ev_ebitda": 8,  "ev_rev": 1.2,"pfcf": 12, "ps": 1.2,"pb": 2},
                "Materials":              {"ev_ebitda": 9,  "ev_rev": 1.5,"pfcf": 14, "ps": 1.5,"pb": 2},
                "Real Estate":            {"ev_ebitda": 20, "ev_rev": 8,  "pfcf": 28, "ps": 8,  "pb": 1.8},
                "Utilities":              {"ev_ebitda": 12, "ev_rev": 2.5,"pfcf": 18, "ps": 2.5,"pb": 1.5},
            }
            _sbm = _sect_benchmarks.get(sector_name, {"ev_ebitda": 14, "ev_rev": 2.5, "pfcf": 22, "ps": 2, "pb": 3})

            # Render one multiple card into a Streamlit column (no HTML concat needed)
            def _render_mult_col(col_obj, label, value, sector_val, fmt="x",
                                 lower_is_cheaper=True, note=""):
                # safe-escape angle brackets
                safe_note = note.replace("<", "&lt;").replace(">", "&gt;")
                with col_obj:
                    if value is None or value != value:
                        st.markdown(
                            f'<div style="background:rgba(15,25,50,0.7);border:1px solid rgba(255,255,255,0.08);'
                            f'border-radius:10px;padding:14px 16px;text-align:center;height:100%">'
                            f'<div style="color:#475569;font-size:0.78rem;font-weight:700;'
                            f'text-transform:uppercase;letter-spacing:0.5px">{label}</div>'
                            f'<div style="color:#475569;font-size:1.4rem;font-weight:800;margin:8px 0">—</div>'
                            f'<div style="color:#334155;font-size:0.72rem">No data</div></div>',
                            unsafe_allow_html=True)
                        return
                    v_str = f"{value:.1f}{fmt}"
                    if sector_val:
                        pct_diff = (value - sector_val) / sector_val * 100
                        cheap    = (value < sector_val) if lower_is_cheaper else (value > sector_val)
                        accent   = "#4ADE80" if cheap else ("#F87171" if abs(pct_diff) > 20 else "#F59E0B")
                        vs_txt   = f"{'▼' if value < sector_val else '▲'} {abs(pct_diff):.0f}% vs sector avg {sector_val}{fmt}"
                        verdict  = "Cheaper than peers" if cheap else ("Expensive vs peers" if abs(pct_diff) > 20 else "In line with peers")
                    else:
                        accent  = "#94A3B8"
                        vs_txt  = "Sector benchmark N/A"
                        verdict = ""
                    vd_html = (f'<div style="color:{accent};font-size:0.72rem;margin-top:1px">{verdict}</div>'
                               if verdict else "")
                    nt_html = (f'<div style="color:#475569;font-size:0.7rem;margin-top:6px;'
                               f'border-top:1px solid rgba(255,255,255,0.06);padding-top:5px">{safe_note}</div>'
                               if safe_note else "")
                    st.markdown(
                        f'<div style="background:rgba(15,25,50,0.7);border:1px solid rgba(255,255,255,0.08);'
                        f'border-radius:10px;padding:14px 16px;text-align:center">'
                        f'<div style="color:#94A3B8;font-size:0.78rem;font-weight:700;'
                        f'text-transform:uppercase;letter-spacing:0.5px">{label}</div>'
                        f'<div style="color:#F1F5F9;font-size:1.55rem;font-weight:800;margin:7px 0">{v_str}</div>'
                        f'<div style="color:{accent};font-size:0.75rem;font-weight:600">{vs_txt}</div>'
                        f'{vd_html}{nt_html}</div>',
                        unsafe_allow_html=True)

            # Row 1: EV multiples + cash-flow based
            _mc1, _mc2, _mc3, _mc4 = st.columns(4)
            _render_mult_col(_mc1, "EV / EBITDA",  _ev_ebitda, _sbm["ev_ebitda"], "x", True,
                             "Enterprise value / operating profit. Most-used cross-sector multiple.")
            _render_mult_col(_mc2, "EV / Revenue", _ev_rev,    _sbm["ev_rev"],    "x", True,
                             "Enterprise value / sales. Useful when EBITDA is negative.")
            _render_mult_col(_mc3, "P / FCF",      _pfcf,      _sbm["pfcf"],      "x", True,
                             "Price / free cash flow — the real earnings multiple.")
            _render_mult_col(_mc4, "P / Sales",    _ps,        _sbm["ps"],        "x", True,
                             "Market cap / revenue. Lower = potentially cheaper vs peers.")

            # Row 2: book-value + earnings multiples
            _mb1, _mb2, _mb3, _mb4 = st.columns(4)
            _render_mult_col(_mb1, "P / Book",   _p2b,          _sbm["pb"], "x", True,
                             "Price / book value. Below 1 = trading below net assets.")
            _render_mult_col(_mb2, "P / E",      qs.get("pe"),  None,       "x", True,
                             f"Sector avg P/E: {sector_pe}x")
            _render_mult_col(_mb3, "PEG Ratio",  _peg,          1.0,        "x", True,
                             "P/E / EPS growth rate. Below 1 = growth at a reasonable price.")
            _render_mult_col(_mb4, "Ent. Value",
                             (_ev / 1e9) if _ev else None, None, "B", False,
                             "Mkt Cap + Total Debt minus Cash & Equivalents.")

            st.caption(f"Sector benchmarks are approximate medians for {sector_name}. Use as directional guide only.")

            # ── DCF Sensitivity Matrix ─────────────────────────────
            st.markdown('<div class="section-header">🔢 DCF Sensitivity Analysis</div>',
                        unsafe_allow_html=True)
            st.markdown(
                "Each cell shows intrinsic value per share at that WACC × Terminal Growth combination. "
                "**Your current selection is highlighted in gold.** "
                "Green = undervalued vs current price. Red = overvalued.",
                unsafe_allow_html=False
            )

            _wacc_range  = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
            _tg_range    = [0, 1, 2, 3, 4, 5]
            _matrix      = {}  # populated below if revenue data available

            if curr_rev_m and shares and shares > 0:
                # Build matrix values
                _matrix = {}
                for _w in _wacc_range:
                    for _tg in _tg_range:
                        if _w <= _tg:
                            _matrix[(_w, _tg)] = None
                            continue
                        _raw = calc_revenue_dcf_3phase(
                            curr_rev_m, rg_short, om_short, rg_med, om_med, rg_long, om_long,
                            tax_rate, inv_short, inv_med, inv_long, _w, _tg, ronic
                        )
                        if _raw and shares > 0:
                            _iv_gbp = (_raw + _net_cash_m) * 1e6 / shares
                            _iv_disp = _iv_gbp * 100 if raw_currency == "GBp" else _iv_gbp
                        else:
                            _iv_disp = None
                        _matrix[(_w, _tg)] = _iv_disp

                # Find min/max for colour gradient
                _vals_flat = [v for v in _matrix.values() if v is not None]
                _v_min     = min(_vals_flat) if _vals_flat else 0
                _v_max     = max(_vals_flat) if _vals_flat else 1

                def _cell_colour(iv, price):
                    if iv is None: return "#1E293B", "#475569"
                    ratio = (iv - price) / price if price else 0
                    if   ratio >  0.30: bg = "rgba(34,197,94,0.30)";  fg = "#4ADE80"
                    elif ratio >  0.10: bg = "rgba(34,197,94,0.15)";  fg = "#86EFAC"
                    elif ratio >  0.00: bg = "rgba(34,197,94,0.07)";  fg = "#A7F3D0"
                    elif ratio > -0.10: bg = "rgba(248,113,113,0.07)"; fg = "#FCA5A5"
                    elif ratio > -0.30: bg = "rgba(248,113,113,0.18)"; fg = "#F87171"
                    else:               bg = "rgba(248,113,113,0.35)"; fg = "#EF4444"
                    return bg, fg

                # Build HTML table
                _unit = "p" if raw_currency == "GBp" else sym_cur
                _tbl  = f"""
                <div style="overflow-x:auto;margin:10px 0">
                <table style="width:100%;border-collapse:collapse;font-size:0.82rem">
                  <thead>
                    <tr>
                      <th style="background:#0D1F35;color:#F59E0B;padding:8px 10px;border:1px solid rgba(255,255,255,0.08);
                                 text-align:left;white-space:nowrap">
                        WACC ↓ / Term.g →
                      </th>"""
                for _tg in _tg_range:
                    _is_cur_tg = (_tg == terminal_growth)
                    _tg_col = "#F59E0B" if _is_cur_tg else "#94A3B8"
                    _tg_bg  = "rgba(245,158,11,0.12)" if _is_cur_tg else "#0D1F35"
                    _tbl += f'<th style="background:{_tg_bg};color:{_tg_col};padding:8px 10px;border:1px solid rgba(255,255,255,0.08);text-align:center;font-weight:{"800" if _is_cur_tg else "600"}">{_tg}%</th>'
                _tbl += "</tr></thead><tbody>"

                for _w in _wacc_range:
                    _is_cur_w = (_w == discount_r)
                    _row_bg   = "rgba(245,158,11,0.06)" if _is_cur_w else "transparent"
                    _w_col    = "#F59E0B" if _is_cur_w else "#94A3B8"
                    _tbl += f'<tr style="background:{_row_bg}">'
                    _tbl += f'<td style="background:#0D1F35;color:{_w_col};padding:7px 10px;border:1px solid rgba(255,255,255,0.08);font-weight:{"800" if _is_cur_w else "600"};white-space:nowrap">{_w}%</td>'
                    for _tg in _tg_range:
                        _iv   = _matrix.get((_w, _tg))
                        _bg, _fg = _cell_colour(_iv, curr_px)
                        _is_cur  = (_w == discount_r and _tg == terminal_growth)
                        _border  = "2px solid #F59E0B" if _is_cur else "1px solid rgba(255,255,255,0.06)"
                        if _iv is not None:
                            _disp = f"£{_iv/100:,.2f}" if raw_currency == "GBp" else f"{_unit}{_iv:,.2f}"
                        else:
                            _disp = "—"
                        _tbl += f'<td style="background:{_bg};color:{_fg};padding:7px 10px;border:{_border};text-align:center;font-weight:{"800" if _is_cur else "500"}">{_disp}</td>'
                    _tbl += "</tr>"
                _tbl += "</tbody></table></div>"

                # Legend
                _tbl += f"""
                <div style="display:flex;gap:16px;flex-wrap:wrap;margin:8px 0;font-size:0.75rem;color:#64748B">
                  <span>Current price: <strong style="color:#F1F5F9">{get_price_display(curr_px, pick, info)}</strong></span>
                  <span style="color:#4ADE80">■ >30% upside</span>
                  <span style="color:#86EFAC">■ 10–30% upside</span>
                  <span style="color:#A7F3D0">■ 0–10% upside</span>
                  <span style="color:#FCA5A5">■ 0–10% downside</span>
                  <span style="color:#F87171">■ 10–30% downside</span>
                  <span style="color:#EF4444">■ >30% downside</span>
                  <span style="color:#F59E0B">■ Your selection</span>
                </div>"""
                st.markdown(_tbl, unsafe_allow_html=True)
            else:
                st.info("Revenue data not available — cannot compute sensitivity matrix.")

            # ── AI Analyst Commentary ──────────────────────────────
            st.markdown('<div class="section-header">🤖 AI Analyst Commentary</div>',
                        unsafe_allow_html=True)

            # Track current DCF params so commentary knows when assumptions changed
            _cur_dcf_params = {
                "rg_short": rg_short, "om_short": om_short,
                "rg_med":   rg_med,   "om_med":   om_med,
                "rg_long":  rg_long,  "om_long":  om_long,
                "tax_rate": tax_rate,
                "inv_short": inv_short, "inv_med": inv_med, "inv_long": inv_long,
                "discount_r": discount_r, "terminal_growth": terminal_growth,
                "ronic": ronic,
                "dcf_result": round(dcf_per_share, 1) if dcf_per_share else None,
            }
            _saved_params = st.session_state.get(f"ai_commentary_params_{pick}", None)
            _params_changed = (_saved_params is not None and
                               f"ai_commentary_{pick}" in st.session_state and
                               _saved_params != _cur_dcf_params)

            # Show stale-data warning + refresh button
            if _params_changed:
                _changed_keys = [k for k in _cur_dcf_params if _saved_params.get(k) != _cur_dcf_params[k]]
                _change_desc  = ", ".join(
                    f"{k.replace('_',' ')} {_saved_params.get(k,'?')}→{_cur_dcf_params[k]}"
                    for k in _changed_keys
                )
                st.markdown(f"""
                <div style="background:rgba(239,68,68,0.12);border:1px solid rgba(239,68,68,0.4);
                            border-left:4px solid #EF4444;border-radius:8px;
                            padding:10px 16px;margin:6px 0 10px 0;
                            display:flex;align-items:center;gap:12px">
                  <span style="font-size:1.2rem">⚠️</span>
                  <div>
                    <span style="color:#FCA5A5;font-weight:700;font-size:0.87rem">
                      Assumptions changed — commentary is stale
                    </span>
                    <br>
                    <span style="color:#94A3B8;font-size:0.8rem">{_change_desc}</span>
                  </div>
                </div>""", unsafe_allow_html=True)

            _btn_row_a, _btn_row_b = st.columns([2, 1])
            with _btn_row_a:
                _gen_label = "🔄 Refresh Commentary with Current Assumptions" if _params_changed else f"Generate AI Commentary for {pick}"
                _do_gen = st.button(_gen_label, key=f"ai_btn_{pick}")
            with _btn_row_b:
                if f"ai_commentary_{pick}" in st.session_state:
                    if st.button("🗑 Clear Commentary", key=f"ai_clear_{pick}"):
                        del st.session_state[f"ai_commentary_{pick}"]
                        if f"ai_commentary_params_{pick}" in st.session_state:
                            del st.session_state[f"ai_commentary_params_{pick}"]
                        st.rerun()

            if _do_gen:
                with st.spinner("Analysing valuation…"):
                    try:
                        import anthropic as _anthropic_sdk
                        _ai_client = _anthropic_sdk.Anthropic()
                        _dcf_str   = get_price_display(dcf_per_share, pick, info) if dcf_per_share else "not computed"
                        _gr_str    = get_price_display(graham_val,    pick, info) if graham_val    else "not computed"
                        _pe_str    = get_price_display(pe_iv_val,     pick, info) if pe_iv_val     else "not computed"
                        _cur_str   = get_price_display(curr_px,       pick, info) if curr_px       else "unknown"
                        _avg_rev_growth_str = f"{avg_rev_growth:.1f}%/yr" if 'avg_rev_growth' in dir() and avg_rev_growth else "unknown"
                        _prompt    = f"""You are a senior equity analyst at a top-tier investment bank. Write a substantive investment commentary on {info.get('longName', pick)} ({pick}) — the kind of analysis you'd find in a sell-side initiation note.

COMPANY DATA:
- Sector: {sector_name} | Industry: {qs.get('industry','?')}
- Business: {info.get('longBusinessSummary','')[:500]}
- Current price: {_cur_str} | Mkt Cap: {fmt_currency(_f(info.get('marketCap')), sym_cur) if info.get('marketCap') else '?'}
- Quality Score: {qs['score']}/100 | ROE: {fmt_pct(qs['roe']) if qs['roe'] else '?'} | Gross Margin: {fmt_pct(qs['gross_margin']) if qs['gross_margin'] else '?'}
- P/E: {fmt_number(qs['pe'],1) if qs['pe'] else '?'} | Debt/Equity: {fmt_number(qs['debt_equity'],2) if qs['debt_equity'] is not None else '?'}
- Dividend/share: {div_display} (Yield: {div_yield_display}) | FY end: {fy_str}
- Revenue CAGR: {_avg_rev_growth_str} vs sector avg ~{sect_rg}%/yr
- Sector avg operating margin: ~{sect_op}% | This company gross margin: {fmt_pct(qs['gross_margin']) if qs['gross_margin'] else '?'}

VALUATION ANALYSIS:
- DCF 3-phase intrinsic value: {_dcf_str} (user assumptions: ST {rg_short}% growth/{om_short}% margin, MT {rg_med}%/{om_med}%, LT {rg_long}%/{om_long}%, WACC {discount_r}%)
- Graham Number (conservative floor): {_gr_str}
- Industry P/E fair value ({sector_name} avg {sector_pe}x): {_pe_str}
- Current price: {_cur_str}

Write a rigorous 5-paragraph investment note:

**Paragraph 1 — Business Quality & Competitive Position**
Assess the company's competitive moat, pricing power, and why (or whether) it deserves a quality premium. Reference the margins and ROE vs sector benchmarks.

**Paragraph 2 — Valuation Disconnect Analysis**
Explain the gap between all three intrinsic value estimates and the current price. Which valuation method is most appropriate for this business type and why? Are the models likely understating or overstating value for this specific company?

**Paragraph 3 — Why Is the Market Pricing It This Way?**
Give 2-3 concrete reasons: e.g. sector re-rating, growth expectations baked in, quality premium, macro tailwinds/headwinds, competitive threats, or irrational exuberance/pessimism. Be specific to this company.

**Paragraph 4 — DCF Assumption Critique**
Are the user's DCF assumptions (growth rates, margins, WACC) conservative, aggressive, or fair relative to this company's historical track record and industry norms? What adjustment would you make?

**Paragraph 5 — Analyst Verdict**
Is this worth deeper research now, or is there a better entry point? What catalyst (earnings, re-rating, sector rotation) could change the picture? Give a clear directional view.

**Paragraph 6 — How to Use This Analysis (for retail investors)**
Write a plain-English guide explaining how a retail investor should use the information on this screen. Cover: (a) which of the three valuation methods is most relevant for THIS specific company and why; (b) how to interpret the sensitivity table — what it tells you about risk; (c) which valuation multiples matter most for this sector and how to compare them; (d) what next steps to take — e.g. check the news, look at earnings dates, compare to a competitor. Make this practical and jargon-free.

Be direct, analytical, and specific. Avoid generic statements. Write like you're accountable to a portfolio manager. End the entire commentary with exactly this line on a new paragraph: "⚠️ AI-generated analysis — may contain errors. Not financial advice. Always conduct independent due diligence before any investment decision."
"""
                        _resp = _ai_client.messages.create(
                            model="claude-sonnet-5",
                            max_tokens=2000,
                            messages=[{"role":"user","content":_prompt}]
                        )
                        _commentary = _resp.content[0].text
                        st.session_state[f"ai_commentary_{pick}"]        = _commentary
                        st.session_state[f"ai_commentary_params_{pick}"] = _cur_dcf_params
                        st.rerun()
                    except ImportError:
                        # Fallback: rich rules-based commentary (anthropic package not installed)
                        _co = info.get('longName', pick)
                        _sect = sector_name
                        _roe_v  = qs.get('roe') or 0
                        _gm_v   = qs.get('gross_margin') or 0
                        _pe_v   = qs.get('pe') or 0
                        _de_v   = qs.get('debt_equity')
                        _sc     = qs['score']

                        # Para 1 — Business Quality & Moat
                        if _gm_v > 0.5:
                            _moat = f"exceptional gross margins of {_gm_v*100:.1f}% — well above typical {_sect} norms — point to genuine pricing power and a durable competitive moat"
                        elif _gm_v > 0.3:
                            _moat = f"solid gross margins of {_gm_v*100:.1f}% suggesting moderate pricing power within the {_sect} sector"
                        else:
                            _moat = f"compressed gross margins of {_gm_v*100:.1f}%, indicating a commoditised or highly competitive operating environment"
                        _roe_comment = f"ROE of {_roe_v*100:.1f}% is {'exceptional, implying the business generates substantial returns on shareholder capital without excessive leverage' if _roe_v > 0.2 else 'modest, suggesting either reinvestment drag or limited pricing power' if _roe_v > 0.08 else 'weak, raising questions about capital allocation effectiveness'}."
                        _p1 = f"**Business Quality & Competitive Position:** {_co} exhibits {_moat}. {_roe_comment} A Quality Score of {_sc}/100 {'places this firmly in the top tier of screened companies — a hallmark of businesses with defendable market positions' if _sc>=80 else 'reflects decent but not exceptional fundamentals, typical of improving-quality mid-caps' if _sc>=60 else 'flags concerns on one or more fundamental dimensions that warrant caution'}."

                        # Para 2 — Valuation Disconnect
                        _v_comments = []
                        if dcf_per_share and curr_px:
                            _d_pct = (dcf_per_share - curr_px)/curr_px*100
                            if _d_pct < -50:
                                _v_comments.append(f"The 3-Phase DCF signals {abs(_d_pct):.0f}% overvaluation vs current price — but DCF models are highly sensitive to terminal growth and WACC assumptions. For capital-light or asset-light businesses with strong intangibles, DCF chronically understates intrinsic value by ignoring brand, network effects, and optionality.")
                            elif _d_pct > 30:
                                _v_comments.append(f"The DCF model implies {_d_pct:.0f}% upside, which is a meaningful margin of safety if revenue growth and margin assumptions prove conservative. However, DCF estimates must be stress-tested — a 1% rise in WACC can eliminate apparent undervaluation.")
                            else:
                                _v_comments.append(f"The DCF model prices the stock close to fair value ({abs(_d_pct):.0f}% {'above' if _d_pct<0 else 'below'} estimate), suggesting the market is broadly in agreement with consensus growth and margin forecasts.")
                        if graham_val and curr_px:
                            _g_pct = (graham_val - curr_px)/curr_px*100
                            if _g_pct < -40:
                                _v_comments.append(f"The Graham Number at {get_price_display(graham_val, pick, info)} sits far below market price — expected for growth or asset-light companies since Graham's formula weights book value heavily and was designed for capital-intensive defensive businesses. Its understatement here is a feature, not a bug.")
                            elif _g_pct > 0:
                                _v_comments.append(f"Notably, the stock trades below its Graham Number ({get_price_display(graham_val, pick, info)}), historically associated with a margin of safety for value investors — rare for companies with this quality profile.")
                        if pe_iv_val and curr_px and _pe_v:
                            _p_pct = (pe_iv_val - curr_px)/curr_px*100
                            if _p_pct < -20:
                                _v_comments.append(f"The Industry P/E method implies {abs(_p_pct):.0f}% overvaluation vs sector peers (sector avg {sector_pe}x vs this company's {curr_pe_str}x), suggesting the market is assigning a deliberate growth premium. Whether that premium is justified depends on whether forward earnings growth can re-rate the multiple.")
                            elif _p_pct > 20:
                                _v_comments.append(f"At {curr_pe_str}x vs sector average {sector_pe}x, the stock trades at a discount to peers — Industry P/E fair value implies {_p_pct:.0f}% upside, which is a classic mean-reversion opportunity if operational momentum continues.")
                        _p2 = "**Valuation Disconnect Analysis:** " + " ".join(_v_comments) if _v_comments else "**Valuation Disconnect Analysis:** Insufficient valuation data to form a view."

                        # Para 3 — Why market prices it this way
                        _reasons = []
                        if _pe_v and _pe_v > sector_pe * 1.3:
                            _reasons.append(f"the market is pricing in durable above-average growth (P/E {curr_pe_str}x vs sector {sector_pe}x), likely driven by strong historical earnings momentum")
                        if _pe_v and _pe_v < sector_pe * 0.8:
                            _reasons.append(f"the stock trades at a sector discount (P/E {curr_pe_str}x vs {sector_pe}x average), potentially reflecting near-term earnings concern, sector rotation out of {_sect}, or simply market neglect of a less-covered name")
                        if _sc >= 85:
                            _reasons.append("quality scarcity premium — high-quality compounders attract institutional buyers willing to pay above intrinsic value estimates to secure exposure to a rare, durable business")
                        if _de_v is not None and _de_v > 2:
                            _reasons.append(f"elevated leverage (D/E {_de_v:.1f}x) creates perception risk — rising rate environments discount leveraged businesses more aggressively")
                        if not _reasons:
                            _reasons.append("macro sector rotation and liquidity dynamics may be the dominant pricing driver rather than fundamental mispricing")
                        _p3 = f"**Why Is the Market Pricing It This Way?** The most likely explanations are: {'; '.join(_reasons)}. Investors should assess whether these factors are structural or cyclical before positioning."

                        # Para 4 — DCF assumption critique (uses CURRENT slider values)
                        _wacc_comment = f"The WACC of {discount_r}% {'appears conservative for a quality business and may understate intrinsic value' if discount_r >= 12 else 'is broadly reasonable for a mid-cap in the current rate environment' if discount_r >= 9 else 'may be optimistic in a higher-for-longer rate environment — consider sensitivity testing at 10-11%'}."
                        _tg_comment   = f"The terminal growth rate of {terminal_growth}% {'is at the high end — verify against long-run nominal GDP for the relevant market' if terminal_growth >= 4 else 'is conservative, which is prudent for a mature business' if terminal_growth <= 1 else 'is reasonable, roughly in line with nominal GDP growth expectations'}."
                        _p4 = f"**DCF Assumption Critique:** {_wacc_comment} Short-term revenue growth at {rg_short}% {'is aggressive — confirm whether the company has catalysts (contract wins, market expansion) to sustain this trajectory' if rg_short > 15 else 'appears sensible given historical trends' if rg_short > 5 else 'is conservative — if the business is re-accelerating, intrinsic value may be materially higher'}. {_tg_comment} A 0.5% change in terminal growth typically moves DCF output by 10–15%. Operating margin assumptions ({om_short}% ST / {om_med}% MT / {om_long}% LT) should be benchmarked against the last 3 years of actuals."

                        # Para 5 — Verdict
                        if dcf_per_share and curr_px and (dcf_per_share - curr_px)/curr_px*100 > 20 and _sc >= 70:
                            _verdict = f"**Analyst Verdict:** This appears worthy of deeper research now. The combination of quality fundamentals and apparent DCF undervaluation is a constructive setup. Key catalysts to watch: next earnings release for margin confirmation, any analyst consensus upgrades, and sector multiple expansion. Risk: if revenue growth disappoints relative to DCF assumptions, the investment case unravels quickly. Suggested entry discipline: build a position only if the stock holds key support; avoid chasing strength."
                        elif dcf_per_share and curr_px and (dcf_per_share - curr_px)/curr_px*100 < -40:
                            _verdict = f"**Analyst Verdict:** The current price embeds significant growth optimism that the DCF does not corroborate. A better entry point may emerge after the next earnings catalyst — either growth re-accelerates and justifies the premium, or it disappoints and the multiple compresses, creating a more attractive risk-reward. Patience is the discipline here."
                        else:
                            _verdict = f"**Analyst Verdict:** The stock is in a mixed-signal zone. Quality score is {'strong' if _sc>=75 else 'moderate'}, but valuation signals are not unanimous. This is a 'monitor' rather than 'act' situation — wait for a clearer entry signal from either a pullback to DCF fair value or a positive earnings catalyst that forces a re-rating."

                        # Para 6 — How to use this screen (plain English for retail investors)
                        _best_method = ("DCF" if _gm_v > 0.3 and _roe_v > 0.1
                                        else "Industry P/E" if _pe_v and _pe_v > 0
                                        else "Graham Number")
                        _key_multiples = ("EV/EBITDA and P/FCF" if sector_name in ["Technology","Communication Services","Healthcare"]
                                          else "EV/EBITDA and P/Book" if sector_name in ["Financials","Real Estate"]
                                          else "EV/Revenue and P/Sales")
                        _p6 = (
                            f"**How to Use This Analysis:**\n"
                            f"*For retail investors — a plain-English guide to acting on what you see here.*\n\n"
                            f"**Which valuation method to trust most:** For {_co}, the **{_best_method}** is likely the most relevant. "
                            f"{'DCF works best for profitable, cash-generative businesses — focus on the sensitivity table to understand the range of outcomes rather than a single number.' if _best_method=='DCF' else 'The Industry P/E compares this stock directly to peers — if it trades at a discount with similar or better fundamentals, that gap may close.' if _best_method=='Industry P/E' else 'The Graham Number provides a conservative floor — stocks trading below it are rare and often signal deep value opportunities.'}\n\n"
                            f"**Reading the sensitivity table:** Each cell shows what the DCF value would be at a different WACC and terminal growth combination. The gold-highlighted cell is your current assumption. "
                            f"If most of the table is green, the stock looks undervalued across a wide range of scenarios — that is a robust thesis. "
                            f"If only 1-2 cells are green, the investment case is fragile and depends heavily on precise assumptions being correct.\n\n"
                            f"**Key multiples for {sector_name}:** Focus on **{_key_multiples}**. "
                            f"A green card means this company is cheaper than its sector peers on that measure — but always check *why* it is cheaper before concluding it is an opportunity.\n\n"
                            f"**Suggested next steps:** (1) Check the next earnings release date and any recent news. "
                            f"(2) Compare these multiples to one direct competitor. "
                            f"(3) Look at the 5-year financial trend — is the operating margin expanding or contracting? "
                            f"(4) Only if the DCF sensitivity table shows green in most scenarios AND fundamentals are improving, consider this for further due diligence. "
                            f"(5) Never act on a single tool — use this as a starting filter, then read the annual report."
                        )

                        _commentary = f"{_p1}\n\n{_p2}\n\n{_p3}\n\n{_p4}\n\n{_verdict}\n\n{_p6}\n\n⚠️ AI-generated analysis — may contain errors. Not financial advice. Always conduct independent due diligence before any investment decision.\n\n*(Note: For full Claude AI commentary, run: `pip install anthropic` and set `ANTHROPIC_API_KEY` environment variable)*"
                        st.session_state[f"ai_commentary_{pick}"]        = _commentary
                        st.session_state[f"ai_commentary_params_{pick}"] = _cur_dcf_params
                        st.rerun()
                    except Exception as e:
                        st.session_state[f"ai_commentary_{pick}"] = f"Commentary unavailable: {e}\n\n⚠️ Always do your own research before making any investment decision."

            if f"ai_commentary_{pick}" in st.session_state:
                _p = st.session_state.get(f"ai_commentary_params_{pick}", {})
                _params_footer = (
                    f"Generated with: WACC {_p.get('discount_r','?')}% | "
                    f"Terminal growth {_p.get('terminal_growth','?')}% | "
                    f"ST {_p.get('rg_short','?')}%/{_p.get('om_short','?')}% | "
                    f"MT {_p.get('rg_med','?')}%/{_p.get('om_med','?')}% | "
                    f"LT {_p.get('rg_long','?')}%/{_p.get('om_long','?')}% | "
                    f"Tax {_p.get('tax_rate','?')}% | Inv S/M/L {_p.get('inv_short','?')}/{_p.get('inv_med','?')}/{_p.get('inv_long','?')}%"
                ) if _p else ""
                _border_col = "rgba(239,68,68,0.5)" if _params_changed else "rgba(245,158,11,0.25)"
                _left_col   = "#EF4444" if _params_changed else "#F59E0B"
                _stale_tag  = ' ⚠️ STALE — assumptions changed' if _params_changed else ''
                # Render header as HTML, body as markdown (avoids pre-wrap/HTML-tag bug)
                st.markdown(f"""
                <div style="background:rgba(13,31,53,0.92);border:1px solid {_border_col};
                            border-radius:10px;padding:16px 22px 4px 22px;margin:12px 0 0 0;
                            border-left:4px solid {_left_col}">
                  <div style="color:#F59E0B;font-weight:700;font-size:0.85rem;
                              text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">
                    🤖 AI Equity Analyst Commentary — {info.get('longName', pick)}
                    <span style="color:#EF4444;font-size:0.75rem;font-weight:400;
                                 text-transform:none;letter-spacing:0;margin-left:12px">{_stale_tag}</span>
                  </div>
                </div>""", unsafe_allow_html=True)
                # Commentary text rendered as markdown (handles **bold**, paragraphs correctly)
                st.markdown(
                    f'<div style="background:rgba(13,31,53,0.92);border:1px solid {_border_col};'
                    f'border-top:none;border-left:4px solid {_left_col};'
                    f'border-radius:0 0 10px 10px;padding:12px 22px 16px 22px;margin:0 0 12px 0;">'
                    f'</div>',
                    unsafe_allow_html=True
                )
                # Use Streamlit container for the actual text so markdown renders properly
                with st.container():
                    st.markdown(st.session_state[f"ai_commentary_{pick}"])
                    if _params_footer:
                        st.caption(f"📌 {_params_footer}")

            # ── Financial Trend Charts ─────────────────────────────
            st.markdown('<div class="section-header">📊 Financial Trends — Multi-Year Fundamentals</div>',
                        unsafe_allow_html=True)
            try:
                tk_fin  = yf.Ticker(pick)
                fin     = tk_fin.financials
                bs      = tk_fin.balance_sheet
                cf      = tk_fin.cashflow
                _LAYOUT = dict(height=280, template="plotly_dark",
                               paper_bgcolor="#0D1F35", plot_bgcolor="#0F1923",
                               legend=dict(orientation="h", y=-0.22,
                                           font=dict(color="#CBD5E1", size=10)),
                               font=dict(color="#CBD5E1", size=11),
                               title_font=dict(color="#E2E8F0", size=12),
                               margin=dict(t=36, b=40, l=60, r=40))

                def _scale_m(vals_m):
                    """Auto-scale: if max > 1000M use B, else M."""
                    valid = [v for v in vals_m if v is not None]
                    if valid and max(abs(v) for v in valid) >= 1000:
                        return [v/1000 if v is not None else None for v in vals_m], "B"
                    return vals_m, "M"

                if fin is not None and not fin.empty:
                    _yrs = sorted([d.year for d in fin.columns])
                    _yr_labels = [str(y) for y in _yrs]

                    def _row(df, *keys):
                        for k in keys:
                            if k in df.index:
                                vals = df.loc[k].reindex(sorted(df.columns)).dropna()
                                return [str(d.year) for d in vals.index], vals.values
                        return _yr_labels, [None]*len(_yr_labels)

                    tc1, tc2 = st.columns(2)

                    # Chart 1: Revenue + Revenue Growth %
                    with tc1:
                        r_yrs, r_vals = _row(fin, "Total Revenue")
                        r_vals_m_raw = [v/1e6 if v else None for v in r_vals]
                        r_vals_scaled, r_unit = _scale_m(r_vals_m_raw)
                        r_growth = [None] + [
                            round((r_vals[i]-r_vals[i-1])/r_vals[i-1]*100,1)
                            if r_vals[i] and r_vals[i-1] else None
                            for i in range(1, len(r_vals))
                        ]
                        fig1 = go.Figure()
                        fig1.add_bar(x=r_yrs, y=r_vals_scaled,
                                     name=f"Revenue ({sym_cur}{r_unit})",
                                     marker_color="#1E3A8A", opacity=0.85)
                        fig1.add_scatter(x=r_yrs, y=r_growth, name="YoY Growth %",
                                         yaxis="y2", mode="lines+markers+text",
                                         line=dict(color="#F59E0B", width=2.5),
                                         marker=dict(size=7),
                                         text=[f"{v:.0f}%" if v else "" for v in r_growth],
                                         textposition="top center",
                                         textfont=dict(size=9, color="#F59E0B"))
                        fig1.update_layout(**_LAYOUT,
                            title=dict(text=f"Revenue & Growth ({sym_cur}{r_unit})", font=dict(color="#F1F5F9", size=12)),
                            yaxis=dict(title=f"{sym_cur}{r_unit}", gridcolor="#1A2840",
                                       tickformat=",.1f",
                                       title_font=dict(color="#94A3B8", size=10),
                                       tickfont=dict(color="#94A3B8", size=10)),
                            yaxis2=dict(title="Growth %", overlaying="y", side="right",
                                        showgrid=False, zeroline=False,
                                        ticksuffix="%",
                                        title_font=dict(color="#94A3B8", size=10),
                                        tickfont=dict(color="#94A3B8", size=10)))
                        st.plotly_chart(fig1, use_container_width=True)

                    # Chart 2: EPS trend
                    with tc2:
                        ni_yrs, ni_vals = _row(fin, "Net Income")
                        sh_count = _f(info.get("sharesOutstanding"))
                        if sh_count and sh_count > 0:
                            eps_vals = [round(v/sh_count,4) if v else None for v in ni_vals]
                            if raw_currency == "GBp":
                                eps_vals_d = [round(v*100,2) if v else None for v in eps_vals]
                                eps_unit = "pence"
                            else:
                                eps_vals_d = eps_vals
                                eps_unit = sym_cur
                        else:
                            eps_vals_d = [None]*len(ni_yrs)
                            eps_unit = sym_cur
                        fig2 = go.Figure()
                        fig2.add_scatter(x=ni_yrs, y=eps_vals_d, name=f"EPS ({eps_unit})",
                                         mode="lines+markers+text",
                                         line=dict(color="#4ADE80", width=2.5),
                                         marker=dict(size=8, color="#4ADE80"),
                                         fill="tozeroy",
                                         fillcolor="rgba(74,222,128,0.08)",
                                         text=[f"{eps_unit}{v:.2f}" if v else "" for v in eps_vals_d],
                                         textposition="top center",
                                         textfont=dict(size=9, color="#4ADE80"))
                        fig2.update_layout(**_LAYOUT,
                            title=dict(text=f"Earnings Per Share ({eps_unit})", font=dict(color="#F1F5F9", size=12)),
                            yaxis=dict(gridcolor="#1A2840",
                                       tickfont=dict(color="#94A3B8", size=10),
                                       title_font=dict(color="#94A3B8", size=10)))
                        st.plotly_chart(fig2, use_container_width=True)

                    tc3, tc4 = st.columns(2)

                    # Chart 3: Operating Margin + Gross Margin
                    with tc3:
                        rev_yrs, rev_v = _row(fin, "Total Revenue")
                        gp_yrs,  gp_v  = _row(fin, "Gross Profit")
                        op_yrs,  op_v  = _row(fin, "EBIT", "Operating Income")
                        gm_vals = [round(gp/rv*100,1) if gp and rv else None for gp,rv in zip(gp_v,rev_v)]
                        om_vals = [round(op/rv*100,1) if op and rv else None for op,rv in zip(op_v,rev_v)]
                        fig3 = go.Figure()
                        fig3.add_scatter(x=rev_yrs, y=gm_vals, name="Gross Margin %",
                                         mode="lines+markers", line=dict(color="#38BDF8", width=2.5))
                        fig3.add_scatter(x=rev_yrs, y=om_vals, name="Operating Margin %",
                                         mode="lines+markers", line=dict(color="#F59E0B", width=2.5))
                        fig3.add_hline(y=sect_op, line_dash="dot", line_color="#64748B",
                                       annotation_text=f"Sector avg ~{sect_op}%",
                                       annotation_font=dict(color="#94A3B8", size=9))
                        fig3.update_layout(**_LAYOUT,
                            title=dict(text="Gross & Operating Margin Trend (%)", font=dict(color="#F1F5F9", size=12)),
                            yaxis=dict(title="%", gridcolor="#1A2840",
                                       ticksuffix="%",
                                       title_font=dict(color="#94A3B8", size=10),
                                       tickfont=dict(color="#94A3B8", size=10)))
                        st.plotly_chart(fig3, use_container_width=True)

                    # Chart 4: Net Income + Free Cash Flow
                    with tc4:
                        ni_yrs2, ni_v2 = _row(fin, "Net Income")
                        ni_m_raw = [v/1e6 if v else None for v in ni_v2]
                        ni_m_scaled, ni_unit = _scale_m(ni_m_raw)
                        fcf_scaled = None
                        if cf is not None and not cf.empty:
                            oc_yrs, oc_v = _row(cf, "Operating Cash Flow", "Total Cash From Operating Activities")
                            cp_yrs, cp_v = _row(cf, "Capital Expenditure", "Capital Expenditures")
                            fcf_raw = [round((oc-abs(cp))/1e6,1) if oc and cp else None
                                       for oc,cp in zip(oc_v,cp_v)]
                            fcf_scaled = [v/1000 if v is not None and ni_unit=="B" else v for v in fcf_raw]
                        fig4 = go.Figure()
                        fig4.add_bar(x=ni_yrs2, y=ni_m_scaled,
                                     name=f"Net Income ({sym_cur}{ni_unit})",
                                     marker_color="rgba(59,130,246,0.6)")
                        if fcf_scaled:
                            fig4.add_scatter(x=oc_yrs, y=fcf_scaled,
                                             name=f"Free Cash Flow ({sym_cur}{ni_unit})",
                                             mode="lines+markers",
                                             line=dict(color="#34D399", width=2.5))
                        fig4.update_layout(**_LAYOUT,
                            title=dict(text=f"Net Income & Free Cash Flow ({sym_cur}{ni_unit})", font=dict(color="#F1F5F9", size=12)),
                            yaxis=dict(gridcolor="#1A2840",
                                       tickformat=",.1f",
                                       title_font=dict(color="#94A3B8", size=10),
                                       tickfont=dict(color="#94A3B8", size=10)))
                        st.plotly_chart(fig4, use_container_width=True)

            except Exception as _fe:
                st.caption(f"Financial charts unavailable. ({_fe})")

            # ── Key stats ─────────────────────────────────────────
            st.markdown('<div class="section-header">📋 Key Statistics</div>',
                        unsafe_allow_html=True)
            s1, s2 = st.columns(2)
            with s1:
                st.dataframe(pd.DataFrame({
                    "Metric": ["52-Week High","52-Week Low","Dividend Yield","Beta",
                               "Current Ratio","Quick Ratio"],
                    "Value":  [
                        fmt_currency(info.get("fiftyTwoWeekHigh"), sym_cur),
                        fmt_currency(info.get("fiftyTwoWeekLow"),  sym_cur),
                        fmt_pct(info.get("dividendYield")),
                        fmt_number(info.get("beta"),2),
                        fmt_number(info.get("currentRatio"),2),
                        fmt_number(info.get("quickRatio"),2),
                    ]
                }), use_container_width=True, hide_index=True)
            with s2:
                st.dataframe(pd.DataFrame({
                    "Metric": ["Revenue Growth","Earnings Growth","Net Margin",
                               "Operating Cashflow","Free Cashflow","EPS (TTM)"],
                    "Value":  [
                        fmt_pct(info.get("revenueGrowth")),
                        fmt_pct(info.get("earningsGrowth")),
                        fmt_pct(qs["net_margin"]),
                        fmt_currency(info.get("operatingCashflow"), sym_cur),
                        fmt_currency(info.get("freeCashflow"), sym_cur),
                        fmt_currency(eps, sym_cur) if eps else "–",
                    ]
                }), use_container_width=True, hide_index=True)

            # ── WATCHLIST + EXPORT ROW ─────────────────────────────
            st.divider()
            _wl_col, _exp_col = st.columns(2)

            # Add / Remove watchlist
            _wl_now  = st.session_state.get("fintiq_watchlist", {})
            _in_wl   = pick in _wl_now
            with _wl_col:
                if _in_wl:
                    st.markdown(
                        f'<div style="background:rgba(74,222,128,0.1);border:1px solid rgba(74,222,128,0.35);'
                        f'border-radius:8px;padding:8px 14px;color:#4ADE80;font-weight:600;font-size:0.9rem;'
                        f'margin-bottom:6px">⭐ {pick} is in your watchlist — scroll to top to view</div>',
                        unsafe_allow_html=True
                    )
                    if st.button(f"✕ Remove {pick} from Watchlist", key=f"wl_rem_{pick}", use_container_width=True):
                        del st.session_state["fintiq_watchlist"][pick]
                        _wl_save(st.session_state["fintiq_watchlist"])
                        st.rerun()
                else:
                    if st.button(f"⭐ Add {pick} to Watchlist", key=f"wl_add_{pick}", use_container_width=True):
                        _wl_entry = {
                            "name":          info.get("longName", pick),
                            "price":         curr_px,
                            "currency":      info.get("currency", ""),
                            "dcf_iv":        dcf_per_share if dcf_per_share else None,
                            "quality_score": qs.get("score"),
                            "added":         str(pd.Timestamp.now().date()),
                            "sector":        sector_name,
                        }
                        st.session_state["fintiq_watchlist"][pick] = _wl_entry
                        _wl_save(st.session_state["fintiq_watchlist"])
                        st.rerun()

            # Export Deep-Dive to Excel
            with _exp_col:
                if _OPENPYXL:
                    # Build sensitivity dict from already-computed _matrix / _wacc_range / _tg_range
                    _sens_export = {}
                    try:
                        _sens_export = {
                            "wacc_range": _wacc_range,
                            "tg_range":   _tg_range,
                            "matrix":     {str(k): v for k, v in _matrix.items()},
                            "price":      curr_px or 0,
                            "unit":       "p" if raw_currency == "GBp" else sym_cur,
                        }
                        # Convert string keys back to tuples in export function
                    except Exception:
                        _sens_export = {}

                    # Build multiples dict
                    _mults_export = {"items": [
                        {"label": "EV/EBITDA",  "value": _ev_ebitda, "benchmark": _sbm.get("ev_ebitda"), "fmt": "x", "lower_is_cheaper": True},
                        {"label": "EV/Revenue", "value": _ev_rev,    "benchmark": _sbm.get("ev_rev"),    "fmt": "x", "lower_is_cheaper": True},
                        {"label": "P/FCF",      "value": _pfcf,      "benchmark": _sbm.get("pfcf"),      "fmt": "x", "lower_is_cheaper": True},
                        {"label": "P/Sales",    "value": _ps,        "benchmark": _sbm.get("ps"),        "fmt": "x", "lower_is_cheaper": True},
                        {"label": "P/Book",     "value": _p2b,       "benchmark": _sbm.get("pb"),        "fmt": "x", "lower_is_cheaper": True},
                        {"label": "P/E",        "value": qs.get("pe"), "benchmark": sector_pe,           "fmt": "x", "lower_is_cheaper": True},
                        {"label": "PEG",        "value": _peg,       "benchmark": 1.0,                   "fmt": "x", "lower_is_cheaper": True},
                    ]}

                    _assumptions_exp = {
                        "discount_r": discount_r, "terminal_growth": terminal_growth,
                        "ronic": ronic,
                        "rg_short": rg_short, "om_short": om_short,
                        "rg_med": rg_med, "om_med": om_med,
                        "rg_long": rg_long, "om_long": om_long,
                        "tax_rate": tax_rate,
                        "inv_short": inv_short, "inv_med": inv_med, "inv_long": inv_long,
                    }

                    # Fix sensitivity matrix keys (tuples serialised to strings above)
                    _sens_fixed = dict(_sens_export)
                    if _sens_export.get("matrix"):
                        import ast as _ast
                        _sens_fixed["matrix"] = {}
                        for _k, _v in _sens_export["matrix"].items():
                            try:
                                _sens_fixed["matrix"][_ast.literal_eval(_k)] = _v
                            except Exception:
                                pass

                    _dd_bytes = build_deepdive_excel(
                        ticker=pick,
                        company=info.get("longName", pick),
                        curr_px=curr_px,
                        dcf_val=dcf_per_share if 'dcf_per_share' in locals() else None,
                        graham_val=graham_val if 'graham_val' in locals() else None,
                        pe_val=pe_iv_val if 'pe_iv_val' in locals() else None,
                        avg_iv=avg_iv if 'avg_iv' in locals() else None,
                        multiples=_mults_export,
                        sensitivity=_sens_fixed,
                        assumptions=_assumptions_exp,
                        hist_rows=[],
                        commentary=st.session_state.get(f"ai_commentary_{pick}", ""),
                    )
                    st.download_button(
                        label="📊 Export Deep-Dive to Excel",
                        data=_dd_bytes,
                        file_name=f"Fintiq_{pick}_deepdive.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"dl_deepdive_{pick}",
                    )
                else:
                    st.caption("Install openpyxl to enable Excel export.")


# ═══════════════════════════════════════════════════════════════
# TAB 2 — CATALYST ALERTS  (rebuilt)
# ═══════════════════════════════════════════════════════════════

with tab2:
    st.markdown('<div class="section-header">⚡ Catalyst Alerts — Strategy 2 Layered In</div>',
                unsafe_allow_html=True)
    st.caption("Earnings dates · Director dealings · Analyst ratings · Short interest · News — all in one view.")

    # ── Build universe: screened symbols + Screen 1 pick ──
    _wl_tickers   = list(st.session_state.get("fintiq_watchlist", {}).keys())
    _scr_tickers  = st.session_state.get("screened_symbols",
                    ["LLOY.L","BARC.L","AZN.L","HSBA.L","BP.L","RIO.L","GSK.L","VOD.L"])
    _cat_universe = list(dict.fromkeys(_wl_tickers + _scr_tickers))

    # Always sync to whatever stock is open in Screen 1 deep-dive
    _s1_pick = st.session_state.get("deepdive_pick")
    if _s1_pick:
        if _s1_pick not in _cat_universe:
            _cat_universe.insert(0, _s1_pick)
        # Force Screen 2 to follow Screen 1
        st.session_state["cat_pick_select"] = _s1_pick

    _src_label = []
    if _wl_tickers:  _src_label.append(f"⭐ {len(_wl_tickers)} from watchlist")
    if _scr_tickers: _src_label.append(f"🔍 {len(_scr_tickers)} from last screen")
    st.caption("  ·  ".join(_src_label) if _src_label else "Run the Fundamental Screen to populate.")

    if not _cat_universe:
        st.info("No stocks in universe — run the Fundamental Screen first.")
        st.stop()

    # ── Stock selector — follows Screen 1 ──────────────────────
    st.markdown('<div class="section-header">🔬 Select Stock for Catalyst Analysis</div>',
                unsafe_allow_html=True)

    _cat_pick = st.selectbox(
        "Stock:", _cat_universe,
        key="cat_pick_select",
        format_func=lambda t: f"⭐ {t}" if t in _wl_tickers else t
    )

    if _cat_pick:
        with st.spinner(f"Loading catalyst data for {_cat_pick}…"):
            try:
                _ctk   = yf.Ticker(_cat_pick)
                _cinfo = _ctk.info or {}
            except Exception:
                _cinfo = {}

            # ── Pull all data upfront ──
            _c_name      = _cinfo.get("longName", _cat_pick)
            _c_price     = _cinfo.get("currentPrice") or _cinfo.get("regularMarketPrice")
            _c_currency  = _cinfo.get("currency", "")
            _c_is_gbp    = _c_currency == "GBp"   # pence — prices are in pence, display as Xp
            _c_sym       = "" if _c_is_gbp else ("£" if _c_currency=="GBP" else ("$" if _c_currency=="USD" else _c_currency+" "))
            _c_unit      = "p" if _c_is_gbp else ""  # suffix for pence

            def _fmt_cat_px(v):
                """Format a price value for Screen 2, consistent dual £/p for GBp stocks."""
                if v is None: return "—"
                if _c_is_gbp:
                    return f"£{v/100:,.2f}  ({v:,.0f}p)" if abs(v) >= 100 else f"{v:.2f}p"
                return f"{_c_sym}{v:,.2f}"

            # Earnings calendar
            _earn_date   = None
            _earn_eps_est= None
            _earn_days   = None
            try:
                _cal = _ctk.calendar
                if isinstance(_cal, dict):
                    _earn_date = _cal.get("Earnings Date") or _cal.get("earningsDate")
                    if isinstance(_earn_date, list): _earn_date = _earn_date[0]
                    _earn_eps_est = _cal.get("Earnings Average") or _cal.get("EPS Estimate")
                elif _cal is not None and not _cal.empty:
                    _cal_d = dict(zip(_cal.iloc[:,0], _cal.iloc[:,1])) if _cal.shape[1]>=2 else {}
                    _earn_date = _cal_d.get("Earnings Date") or _cal_d.get("earningsDate")
                if _earn_date:
                    try:
                        _ed = pd.Timestamp(_earn_date)
                        _earn_days = (_ed - pd.Timestamp.now()).days
                    except Exception:
                        pass
            except Exception:
                pass

            # Earnings history
            _earn_hist = None
            try:
                _eh = _ctk.earnings_dates
                if _eh is not None and not _eh.empty:
                    _earn_hist = _eh.head(8).reset_index()
            except Exception:
                pass

            # Insider transactions
            _ins_df      = None
            _ins_buys    = 0
            _ins_sells   = 0
            _ins_signal  = "neutral"
            try:
                _ins_raw = _ctk.insider_transactions
                if _ins_raw is not None and not _ins_raw.empty:
                    _ins_df = _ins_raw.reset_index()
                    _ins_df.columns = [str(c) for c in _ins_df.columns]
                    for _col in _ins_df.columns:
                        if _ins_df[_col].dtype == object:
                            _ins_buys  = _ins_df[_col].str.contains("Buy|Purchase|Acqui",case=False,na=False).sum()
                            _ins_sells = _ins_df[_col].str.contains("Sale|Sell|Dispose",case=False,na=False).sum()
                            if _ins_buys > 0 or _ins_sells > 0:
                                break
                    if _ins_buys > _ins_sells * 2:
                        _ins_signal = "buying"
                    elif _ins_sells > _ins_buys * 2:
                        _ins_signal = "selling"
            except Exception:
                pass

            # Analyst ratings
            _analyst_rec  = _cinfo.get("recommendationKey","").replace("_"," ").title()
            _analyst_mean = _cinfo.get("recommendationMean")   # 1=Strong Buy … 5=Strong Sell
            _target_price = _cinfo.get("targetMeanPrice")
            _target_high  = _cinfo.get("targetHighPrice")
            _target_low   = _cinfo.get("targetLowPrice")
            _num_analysts = _cinfo.get("numberOfAnalystOpinions", 0)
            _target_upside= None
            if _target_price and _c_price:
                _target_upside = (_target_price - _c_price) / _c_price * 100

            # Short interest
            _short_pct    = _cinfo.get("shortPercentOfFloat")
            _short_ratio  = _cinfo.get("shortRatio")          # days to cover
            _shares_short = _cinfo.get("sharesShort")

            # Institutional ownership
            _inst_pct     = _cinfo.get("heldPercentInstitutions")
            _insider_pct  = _cinfo.get("heldPercentInsiders")

            # News
            _news_items = []
            try:
                _raw_news = _ctk.news or []
                for _art in _raw_news[:10]:
                    _t  = (_art.get("title") or (_art.get("content") or {}).get("title") or "")
                    _u  = (_art.get("link") or (_art.get("content") or {}).get("canonicalUrl",{}).get("url") or "#")
                    _p  = (_art.get("publisher") or (_art.get("content") or {}).get("provider",{}).get("displayName") or "")
                    _ts = _art.get("providerPublishTime") or _art.get("pubDate")
                    try:
                        _d = datetime.fromtimestamp(_ts).strftime("%d %b %Y") if isinstance(_ts,(int,float)) and _ts>0 else (str(_ts)[:10] if _ts else "")
                    except Exception:
                        _d = ""
                    if _t:
                        _news_items.append({"title":_t,"url":_u,"publisher":_p,"date":_d})
            except Exception:
                pass

        # ═══════════════════════════════════════════════════════
        # COMPANY HEADER
        # ═══════════════════════════════════════════════════════
        _px_str = _fmt_cat_px(_c_price)
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0D1F35,#162032);
                    border:1px solid rgba(245,158,11,0.25);border-radius:12px;
                    padding:16px 22px;margin:10px 0 18px 0;
                    display:flex;align-items:center;justify-content:space-between">
          <div>
            <div style="color:#F59E0B;font-weight:800;font-size:1.25rem">{_cat_pick}</div>
            <div style="color:#94A3B8;font-size:0.88rem">{_c_name}</div>
          </div>
          <div style="text-align:right">
            <div style="color:#F1F5F9;font-weight:700;font-size:1.35rem">{_px_str}</div>
            <div style="color:#64748B;font-size:0.8rem">{_c_currency} · Yahoo Finance</div>
          </div>
        </div>""", unsafe_allow_html=True)

        # ═══════════════════════════════════════════════════════
        # SIGNAL SUMMARY CARDS (4 across)
        # ═══════════════════════════════════════════════════════
        st.markdown('<div class="section-header">📡 Signal Summary</div>', unsafe_allow_html=True)

        def _sig_card(col, icon, title, value, sub, accent):
            with col:
                st.markdown(f"""
                <div style="background:#0D1F35;border:1px solid {accent}40;border-left:4px solid {accent};
                            border-radius:10px;padding:14px 16px;min-height:110px">
                  <div style="font-size:1.4rem;margin-bottom:4px">{icon}</div>
                  <div style="color:#94A3B8;font-size:0.75rem;text-transform:uppercase;
                              letter-spacing:1px;margin-bottom:4px">{title}</div>
                  <div style="color:{accent};font-weight:700;font-size:1.0rem;
                              line-height:1.2">{value}</div>
                  <div style="color:#64748B;font-size:0.76rem;margin-top:4px">{sub}</div>
                </div>""", unsafe_allow_html=True)

        _sc1, _sc2, _sc3, _sc4 = st.columns(4)

        # Card 1 — Earnings countdown
        if _earn_days is not None:
            if   _earn_days < 0:   _earn_v = "Earnings passed"; _earn_sub = f"{abs(_earn_days)}d ago"; _earn_col = "#64748B"
            elif _earn_days <= 14: _earn_v = f"⚠️ {_earn_days} days away"; _earn_sub = "Imminent — review position size"; _earn_col = "#F87171"
            elif _earn_days <= 45: _earn_v = f"📅 {_earn_days} days away"; _earn_sub = "Watch closely"; _earn_col = "#F59E0B"
            else:                  _earn_v = f"{_earn_days} days away"; _earn_sub = str(pd.Timestamp(_earn_date).strftime("%d %b %Y")) if _earn_date else ""; _earn_col = "#4ADE80"
        else:
            _earn_v = "No date found"; _earn_sub = "Check IR website"; _earn_col = "#64748B"
        _sig_card(_sc1, "📅", "Next Earnings", _earn_v, _earn_sub, _earn_col)

        # Card 2 — Insider signal
        if _ins_signal == "buying":
            _ins_v = f"🟢 Net Buying ({_ins_buys} buys)"; _ins_sub = "Directors buying own shares"; _ins_col = "#4ADE80"
        elif _ins_signal == "selling":
            _ins_v = f"🔴 Net Selling ({_ins_sells} sells)"; _ins_sub = "Caution — insider exits"; _ins_col = "#F87171"
        else:
            _ins_v = "🟡 Neutral / No data"; _ins_sub = "No clear directional signal"; _ins_col = "#F59E0B"
        _sig_card(_sc2, "👔", "Director Dealings", _ins_v, _ins_sub, _ins_col)

        # Card 3 — Analyst consensus
        if _analyst_rec:
            _rec_map = {"strong buy":"#4ADE80","buy":"#86EFAC","hold":"#F59E0B",
                        "underperform":"#F87171","sell":"#EF4444","strong sell":"#EF4444"}
            _rec_col = _rec_map.get(_analyst_rec.lower(), "#94A3B8")
            _an_sub  = f"Target: {_fmt_cat_px(_target_price)}" if _target_price else f"{_num_analysts} analysts"
            if _target_upside is not None:
                _an_sub += f"  ({_target_upside:+.0f}%)"
            _sig_card(_sc3, "🎯", "Analyst Consensus", _analyst_rec, _an_sub, _rec_col)
        else:
            _sig_card(_sc3, "🎯", "Analyst Consensus", "No coverage", "Not covered by analysts", "#64748B")

        # Card 4 — Short interest
        if _short_pct is not None:
            _sp = _short_pct * 100
            if   _sp > 15: _sh_v = f"🔴 {_sp:.1f}% of float"; _sh_sub = "High — squeeze potential"; _sh_col = "#F87171"
            elif _sp > 8:  _sh_v = f"🟡 {_sp:.1f}% of float"; _sh_sub = f"Days to cover: {_short_ratio:.1f}" if _short_ratio else "Moderate short interest"; _sh_col = "#F59E0B"
            else:          _sh_v = f"🟢 {_sp:.1f}% of float"; _sh_sub = "Low short interest"; _sh_col = "#4ADE80"
        else:
            _sh_v = "No data"; _sh_sub = "Short data unavailable"; _sh_col = "#64748B"
        _sig_card(_sc4, "📉", "Short Interest", _sh_v, _sh_sub, _sh_col)

        st.markdown("<br>", unsafe_allow_html=True)

        # ═══════════════════════════════════════════════════════
        # ROW 2: EARNINGS + INSIDER TRADES
        # ═══════════════════════════════════════════════════════
        _r2a, _r2b = st.columns(2)

        with _r2a:
            st.markdown('<div class="section-header">📅 Earnings Detail</div>', unsafe_allow_html=True)

            # EPS estimate card
            if _earn_eps_est or _earn_date:
                _ed_str  = pd.Timestamp(_earn_date).strftime("%d %b %Y") if _earn_date else "Unknown"
                _eps_str = _fmt_cat_px(_earn_eps_est) if _earn_eps_est else "Not available"
                st.markdown(f"""
                <div style="background:#0D1F35;border:1px solid rgba(245,158,11,0.2);
                            border-radius:8px;padding:12px 16px;margin-bottom:10px">
                  <div style="display:flex;justify-content:space-between;align-items:center">
                    <div>
                      <div style="color:#94A3B8;font-size:0.75rem;text-transform:uppercase;letter-spacing:1px">Next Earnings Date</div>
                      <div style="color:#F1F5F9;font-weight:700;font-size:1.05rem">{_ed_str}</div>
                    </div>
                    <div style="text-align:right">
                      <div style="color:#94A3B8;font-size:0.75rem;text-transform:uppercase;letter-spacing:1px">EPS Estimate</div>
                      <div style="color:#F59E0B;font-weight:700;font-size:1.05rem">{_eps_str}</div>
                    </div>
                  </div>
                </div>""", unsafe_allow_html=True)

            # Earnings history table
            if _earn_hist is not None:
                st.caption("Recent earnings history — EPS reported vs estimated:")
                _eh_disp = _earn_hist.copy()
                _eh_disp.columns = [str(c) for c in _eh_disp.columns]
                # Highlight beats in green, misses in red
                def _beat_style(val):
                    try:
                        return "color: #4ADE80" if float(val) > 0 else "color: #F87171"
                    except Exception:
                        return ""
                _surprise_col = next((c for c in _eh_disp.columns if "surprise" in c.lower()), None)
                if _surprise_col:
                    st.dataframe(
                        _eh_disp.head(8).style.applymap(_beat_style, subset=[_surprise_col]),
                        use_container_width=True, hide_index=True
                    )
                else:
                    st.dataframe(_eh_disp.head(8), use_container_width=True, hide_index=True)

                # Count beats vs misses
                if _surprise_col:
                    try:
                        _beats  = (_eh_disp[_surprise_col].astype(float) > 0).sum()
                        _misses = (_eh_disp[_surprise_col].astype(float) < 0).sum()
                        _b_col  = "#4ADE80" if _beats >= _misses else "#F87171"
                        st.markdown(f'<div style="color:{_b_col};font-size:0.82rem;margin-top:4px">'
                                    f'Last 8 quarters: <strong>{_beats} beats</strong> · '
                                    f'<strong>{_misses} misses</strong></div>', unsafe_allow_html=True)
                    except Exception:
                        pass
            else:
                st.caption("No earnings history available from Yahoo Finance.")

            # Why earnings matter
            with st.expander("ℹ️ How to use earnings data", expanded=False):
                st.markdown("""
**Before earnings (2–4 weeks out):**
- Decide whether to hold through or trim position
- Check options market implied move (if available) to gauge expected volatility
- High-quality businesses with consistent beat history are safer holds

**The surprise is more important than the result:**
- A stock can beat on EPS but *fall* if guidance disappoints
- The market reprices future expectations, not past results
- First profit warning typically causes 20–30% drop — treat as a serious red flag
""")

        with _r2b:
            st.markdown('<div class="section-header">👔 Director & Insider Dealings</div>',
                        unsafe_allow_html=True)

            if _ins_df is not None:
                # Colour-code buy vs sell rows
                _ins_show = _ins_df.head(12).copy()
                _ins_show.columns = [str(c) for c in _ins_show.columns]

                # Find the transaction type column
                _tx_col = next((c for c in _ins_show.columns
                                if _ins_show[c].dtype==object and
                                _ins_show[c].str.contains("Buy|Sell|Sale|Purchase",case=False,na=False).any()), None)

                if _tx_col:
                    def _tx_style(val):
                        if isinstance(val, str):
                            if any(w in val for w in ["Buy","Purchase","Acqui"]):
                                return "background-color: rgba(74,222,128,0.12); color: #4ADE80"
                            elif any(w in val for w in ["Sale","Sell","Dispos"]):
                                return "background-color: rgba(248,113,113,0.12); color: #F87171"
                        return ""
                    st.dataframe(
                        _ins_show.style.applymap(_tx_style, subset=[_tx_col]),
                        use_container_width=True, hide_index=True
                    )
                else:
                    st.dataframe(_ins_show, use_container_width=True, hide_index=True)

                # Ownership summary
                _own_cols = st.columns(2)
                if _insider_pct is not None:
                    _own_cols[0].metric("Insider Ownership", f"{_insider_pct*100:.1f}%",
                                        help="% of shares held by insiders. >5% is meaningful.")
                if _inst_pct is not None:
                    _own_cols[1].metric("Institutional Ownership", f"{_inst_pct*100:.1f}%",
                                        help="% held by funds. >50% = well-followed stock.")
            else:
                st.info("No insider transaction data available from Yahoo Finance.")
                st.caption("For UK stocks, director dealings are filed at Companies House "
                           "and the FCA within 3 business days of the transaction.")

            with st.expander("ℹ️ How to interpret director dealings", expanded=False):
                st.markdown("""
**Why insider buying is the strongest signal:**
- Directors buy their own shares for one reason only: they believe the price will go higher
- A CEO buying £500k+ of stock on the open market (not options) is the clearest conviction signal available
- **Cluster buying** — multiple directors buying within weeks of each other — is especially powerful

**Insider selling is noisier:**
- Directors sell for many reasons: tax bills, divorces, diversification
- *Don't* treat isolated selling as a red flag
- **Exception:** heavy selling by multiple insiders at once near all-time highs is a warning

**For UK stocks:**
- All director dealings above £5,000 must be disclosed within 3 business days
- Look for dealings on Companies House or RNS (Regulatory News Service)
""")

        # ═══════════════════════════════════════════════════════
        # ROW 3: ANALYST RATINGS + SHORT INTEREST
        # ═══════════════════════════════════════════════════════
        _r3a, _r3b = st.columns(2)

        with _r3a:
            st.markdown('<div class="section-header">🎯 Analyst Ratings & Price Targets</div>',
                        unsafe_allow_html=True)

            if _analyst_rec or _target_price:
                # Visual consensus bar
                _mean = _analyst_mean or 3.0
                _bar_pct = int((5 - _mean) / 4 * 100)   # 1=Strong Buy→100%, 5=Strong Sell→0%
                _bar_col = "#4ADE80" if _bar_pct > 60 else ("#F59E0B" if _bar_pct > 40 else "#F87171")
                st.markdown(f"""
                <div style="background:#0D1F35;border:1px solid rgba(245,158,11,0.2);
                            border-radius:8px;padding:14px 16px;margin-bottom:12px">
                  <div style="display:flex;justify-content:space-between;margin-bottom:8px">
                    <span style="color:#94A3B8;font-size:0.8rem">STRONG SELL</span>
                    <span style="color:{_bar_col};font-weight:700">{_analyst_rec or "—"}</span>
                    <span style="color:#94A3B8;font-size:0.8rem">STRONG BUY</span>
                  </div>
                  <div style="background:#1E293B;border-radius:4px;height:10px;overflow:hidden">
                    <div style="background:{_bar_col};width:{_bar_pct}%;height:100%;
                                border-radius:4px;transition:width 0.5s"></div>
                  </div>
                  <div style="color:#64748B;font-size:0.75rem;margin-top:6px;text-align:center">
                    {_num_analysts} analyst{'' if _num_analysts==1 else 's'} covering this stock
                  </div>
                </div>""", unsafe_allow_html=True)

                # Price target range
                if _target_price:
                    _tgt_col = "#4ADE80" if (_target_upside or 0) > 10 else ("#F87171" if (_target_upside or 0) < -10 else "#F59E0B")
                    st.markdown(f"""
                    <div style="background:#0D1F35;border:1px solid rgba(245,158,11,0.15);
                                border-radius:8px;padding:12px 16px">
                      <div style="color:#94A3B8;font-size:0.75rem;text-transform:uppercase;
                                  letter-spacing:1px;margin-bottom:10px">Analyst Price Targets</div>
                      <div style="display:flex;justify-content:space-between;align-items:center">
                        <div style="text-align:center">
                          <div style="color:#F87171;font-size:0.75rem">Low</div>
                          <div style="color:#F87171;font-weight:600">{_fmt_cat_px(_target_low)}</div>
                        </div>
                        <div style="text-align:center">
                          <div style="color:#94A3B8;font-size:0.75rem">Current</div>
                          <div style="color:#F1F5F9;font-weight:700;font-size:1.05rem">{_fmt_cat_px(_c_price)}</div>
                        </div>
                        <div style="text-align:center">
                          <div style="color:#94A3B8;font-size:0.75rem">Consensus</div>
                          <div style="color:{_tgt_col};font-weight:700;font-size:1.05rem">{_fmt_cat_px(_target_price)}</div>
                          <div style="color:{_tgt_col};font-size:0.8rem">{f"{_target_upside:+.1f}%" if _target_upside else ""}</div>
                        </div>
                        <div style="text-align:center">
                          <div style="color:#4ADE80;font-size:0.75rem">High</div>
                          <div style="color:#4ADE80;font-weight:600">{_fmt_cat_px(_target_high)}</div>
                        </div>
                      </div>
                    </div>""", unsafe_allow_html=True)

            else:
                st.info("No analyst coverage data available for this stock.")

            # Ratings history
            try:
                _rec_hist = _ctk.recommendations
                if _rec_hist is not None and not _rec_hist.empty:
                    st.caption("Recent rating changes:")
                    _rh = _rec_hist.reset_index().tail(8)
                    _rh.columns = [str(c) for c in _rh.columns]
                    st.dataframe(_rh, use_container_width=True, hide_index=True)
            except Exception:
                pass

            with st.expander("ℹ️ How to use analyst ratings", expanded=False):
                st.markdown("""
**Use as a sentiment cross-check, not a primary signal:**
- Analyst consensus alone is a weak signal — they are often late and conflict-of-interest prone
- A stock rated "Strong Buy" by 15 analysts with 40% upside to target is interesting, but verify with your own DCF
- **Upgrades matter more than the current rating** — a stock moving from Sell → Buy often moves 5–10% on the day
- **Lone dissenting Sell** among many Buys can be the most useful opinion — find out why

**Price target reliability:**
- Consensus target is typically set 12 months out
- Targets cluster around current price (anchoring bias) — treat with healthy scepticism
- A target far above/below consensus from a reputable house is worth reading the full note
""")

        with _r3b:
            st.markdown('<div class="section-header">📉 Short Interest & Squeeze Risk</div>',
                        unsafe_allow_html=True)

            if _short_pct is not None:
                _sp_val = _short_pct * 100
                # Short interest gauge
                _gauge_segments = [
                    (5, "#4ADE80", "Low"),
                    (10, "#F59E0B", "Moderate"),
                    (15, "#F87171", "High"),
                    (100, "#EF4444", "Very High"),
                ]
                _seg_label = "Low"
                _seg_col   = "#4ADE80"
                for _thresh, _col, _lbl in _gauge_segments:
                    if _sp_val <= _thresh:
                        _seg_col = _col; _seg_label = _lbl; break

                _bar_fill = min(100, int(_sp_val / 20 * 100))   # scale: 20%+ = full bar
                st.markdown(f"""
                <div style="background:#0D1F35;border:1px solid rgba(245,158,11,0.2);
                            border-radius:8px;padding:14px 16px;margin-bottom:10px">
                  <div style="display:flex;justify-content:space-between;margin-bottom:6px">
                    <span style="color:#94A3B8;font-size:0.8rem">Short % of Float</span>
                    <span style="color:{_seg_col};font-weight:700;font-size:1.1rem">{_sp_val:.1f}%
                      <span style="font-size:0.8rem;font-weight:400"> — {_seg_label}</span>
                    </span>
                  </div>
                  <div style="background:#1E293B;border-radius:4px;height:10px;overflow:hidden">
                    <div style="background:{_seg_col};width:{_bar_fill}%;height:100%;border-radius:4px"></div>
                  </div>
                  <div style="display:flex;justify-content:space-between;margin-top:4px">
                    <span style="color:#64748B;font-size:0.72rem">0% (no shorts)</span>
                    <span style="color:#64748B;font-size:0.72rem">20%+ (extreme)</span>
                  </div>
                </div>""", unsafe_allow_html=True)

                _si_cols = st.columns(2)
                if _short_ratio:
                    _si_cols[0].metric("Days to Cover", f"{_short_ratio:.1f} days",
                                       help="How many days of average volume needed for shorts to cover. >5 = elevated squeeze risk.")
                if _shares_short:
                    _si_cols[1].metric("Shares Short", f"{_shares_short/1e6:.1f}M")

                # Squeeze risk assessment
                _squeeze_risk = "Low"
                _sq_col       = "#4ADE80"
                if _sp_val > 15 and (_short_ratio or 0) > 5:
                    _squeeze_risk = "High — potential short squeeze setup"; _sq_col = "#F87171"
                elif _sp_val > 10:
                    _squeeze_risk = "Moderate — worth monitoring"; _sq_col = "#F59E0B"
                st.markdown(f'<div style="color:{_sq_col};font-size:0.85rem;margin-top:6px">'
                            f'⚡ Squeeze Risk: <strong>{_squeeze_risk}</strong></div>',
                            unsafe_allow_html=True)
            else:
                st.info("Short interest data not available for this stock.")
                st.caption("Short interest is typically less available for UK stocks on Yahoo Finance. "
                           "Check the London Stock Exchange or FCA disclosures for UK short positions.")

            # Institutional holders
            st.markdown('<div class="section-header" style="margin-top:16px">🏛️ Institutional Holders</div>',
                        unsafe_allow_html=True)
            try:
                _inst_df = _ctk.institutional_holders
                if _inst_df is not None and not _inst_df.empty:
                    st.dataframe(_inst_df.head(8), use_container_width=True, hide_index=True)
                else:
                    st.caption("No institutional holder data available.")
            except Exception:
                st.caption("Not available.")

            with st.expander("ℹ️ How to use short interest", expanded=False):
                st.markdown("""
**Short interest = % of shares borrowed and sold short by traders betting the price falls.**

**High short interest (>10%) is a double-edged sword:**
- On one hand, smart money thinks the stock is overvalued or has problems
- On the other hand, if *good* news hits a heavily shorted stock, shorts must buy to cover, amplifying the move
- This is a **short squeeze** — how GameStop went from $20 to $500 in January 2021

**What to look for:**
- Short interest *rising* = growing bearish sentiment — be cautious
- Short interest *falling* on a rising stock = shorts covering = strong confirmation
- >15% short float + improving fundamentals = potential squeeze candidate

**Days to cover (short ratio):**
- How many days of normal trading volume it would take all shorts to exit
- >5 days = elevated squeeze risk; >10 days = extreme
""")

        # ═══════════════════════════════════════════════════════
        # ROW 4: NEWS FEED
        # ═══════════════════════════════════════════════════════
        st.markdown('<div class="section-header">📰 Recent News & Press Releases</div>',
                    unsafe_allow_html=True)

        if _news_items:
            # Keyword signal scan
            _BULLISH_KW  = ["beat","record","upgrade","buy","acquisition","contract","dividend","growth","profit"]
            _BEARISH_KW  = ["miss","warning","downgrade","sell","loss","debt","cut","fine","probe","investigation"]
            for _ni in _news_items:
                _tl = _ni["title"].lower()
                _bull = sum(1 for w in _BULLISH_KW if w in _tl)
                _bear = sum(1 for w in _BEARISH_KW if w in _tl)
                if _bull > _bear:     _sent_dot = "🟢"; _sent_bg = "rgba(74,222,128,0.06)"
                elif _bear > _bull:   _sent_dot = "🔴"; _sent_bg = "rgba(248,113,113,0.06)"
                else:                 _sent_dot = "⚪"; _sent_bg = "rgba(30,41,59,0.4)"

                st.markdown(f"""
                <div style="background:{_sent_bg};border:1px solid rgba(255,255,255,0.06);
                            border-radius:8px;padding:10px 14px;margin-bottom:8px">
                  <div style="display:flex;align-items:flex-start;gap:10px">
                    <span style="font-size:1.0rem;margin-top:2px">{_sent_dot}</span>
                    <div>
                      <a href="{_ni['url']}" target="_blank"
                         style="color:#CBD5E1;font-weight:600;font-size:0.9rem;text-decoration:none">
                        {_ni['title']}
                      </a>
                      <div style="color:#64748B;font-size:0.75rem;margin-top:3px">
                        {_ni['publisher']}{'  ·  ' if _ni['publisher'] and _ni['date'] else ''}{_ni['date']}
                      </div>
                    </div>
                  </div>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No recent news found for this stock.")

        st.caption("🟢 Headline contains bullish keywords  ·  🔴 Bearish keywords  ·  ⚪ Neutral — always read the full article before acting.")

        # ═══════════════════════════════════════════════════════
        # CATALYST VERDICT — Combined Signal Tier
        # ═══════════════════════════════════════════════════════
        st.markdown('<div class="section-header">🏆 Catalyst Verdict</div>', unsafe_allow_html=True)

        # Score the signals
        _cv_score = 0
        _cv_notes = []
        if _ins_signal == "buying":          _cv_score += 2; _cv_notes.append("✅ Director buying")
        if _earn_days and 10 <= _earn_days <= 30: _cv_score += 1; _cv_notes.append("📅 Earnings approaching — catalyst window")
        if _analyst_mean and _analyst_mean <= 2.0: _cv_score += 1; _cv_notes.append("✅ Strong analyst consensus")
        if _target_upside and _target_upside > 20: _cv_score += 1; _cv_notes.append(f"✅ {_target_upside:.0f}% analyst upside")
        if _short_pct and _short_pct * 100 > 10:  _cv_score += 1; _cv_notes.append("⚡ High short interest — squeeze potential")
        if _ins_signal == "selling":         _cv_score -= 1; _cv_notes.append("⚠️ Insider selling detected")
        if _earn_days and _earn_days <= 7:   _cv_score -= 1; _cv_notes.append("⚠️ Earnings in <7 days — elevated risk")

        if _cv_score >= 3:
            _tier = "🟢 TIER 3 — High Conviction Catalyst"
            _tier_col = "#4ADE80"
            _tier_bg  = "rgba(74,222,128,0.08)"
            _tier_bdr = "rgba(74,222,128,0.35)"
            _tier_desc = "Multiple positive signals aligned. Combine with DCF upside from Screen 1 and technical confirmation from Screen 3 before acting."
        elif _cv_score >= 1:
            _tier = "🟡 TIER 2 — Moderate Catalyst Signal"
            _tier_col = "#F59E0B"
            _tier_bg  = "rgba(245,158,11,0.08)"
            _tier_bdr = "rgba(245,158,11,0.35)"
            _tier_desc = "Some positive signals present. Monitor closely. Wait for a clearer setup or additional confirmation."
        elif _cv_score <= -1:
            _tier = "🔴 CAUTION — Negative Signals Present"
            _tier_col = "#F87171"
            _tier_bg  = "rgba(248,113,113,0.08)"
            _tier_bdr = "rgba(248,113,113,0.35)"
            _tier_desc = "Negative signals detected. Review your thesis carefully before adding or holding this position."
        else:
            _tier = "⚪ TIER 1 — No Strong Catalyst"
            _tier_col = "#94A3B8"
            _tier_bg  = "rgba(148,163,184,0.06)"
            _tier_bdr = "rgba(148,163,184,0.2)"
            _tier_desc = "No significant catalyst signals either way. Monitor for changes — a catalyst could emerge at any time."

        _notes_html = "".join(f'<div style="color:#CBD5E1;font-size:0.83rem;margin:3px 0">{n}</div>' for n in _cv_notes) if _cv_notes else '<div style="color:#64748B;font-size:0.83rem">No signals detected</div>'

        st.markdown(f"""
        <div style="background:{_tier_bg};border:1px solid {_tier_bdr};border-left:4px solid {_tier_col};
                    border-radius:10px;padding:16px 22px;margin:12px 0">
          <div style="color:{_tier_col};font-weight:800;font-size:1.05rem;margin-bottom:8px">{_tier}</div>
          <div style="margin-bottom:10px">{_notes_html}</div>
          <div style="color:#94A3B8;font-size:0.82rem;border-top:1px solid rgba(255,255,255,0.06);
                      padding-top:8px;margin-top:8px">{_tier_desc}</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("""
        <div style="background:rgba(239,68,68,0.06);border:1px solid rgba(239,68,68,0.2);
                    border-radius:8px;padding:8px 14px;margin-top:16px;font-size:0.78rem;color:#94A3B8">
          <strong style="color:#F87171">⚠️ Disclaimer:</strong>
          For educational purposes only. Not financial advice. Not FCA authorised or regulated.
          Always verify independently and read the full source before acting on any signal.
        </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# TAB 3 — TECHNICAL SETUP
# ═══════════════════════════════════════════════════════════════

with tab3:
    st.markdown('<div class="section-header">📈 Technical Setup — Entry & Exit Timing</div>',
                unsafe_allow_html=True)

    # Always sync to Screen 1 deep-dive selection
    _t3_default = st.session_state.get("deepdive_pick") or "LLOY.L"
    st.session_state["tech_sym_input"] = _t3_default

    t3c1, t3c2, t3c3 = st.columns([3, 1, 1])
    with t3c1:
        tech_sym = st.text_input(
            "Ticker  (auto-filled from Fundamental Screen):",
            value=_t3_default, placeholder="e.g. LLOY.L", key="tech_sym_input")
    with t3c2:
        price_period = st.selectbox("Period:", ["1mo","3mo","6mo","1y","2y"], index=3)
    with t3c3:
        chart_style = st.selectbox("Chart:", ["Candlestick","Line"], index=0)

    if tech_sym:
        _tsym = tech_sym.strip().upper()
        with st.spinner(f"Loading {_tsym}…"):
            pdf = get_price_history(_tsym, price_period)

        if pdf.empty:
            st.error(f"No price data for '{_tsym}'. Examples: LLOY.L · AAPL · SAP.DE · BHP.AX")
        else:
            df_i       = calc_indicators(pdf)
            sigs       = detect_signals(df_i)
            score      = sigs.get("setup_score", 0)

            # Detect GBp (LSE pence) — prices from yFinance for .L tickers are in pence
            try:
                _t3_currency = yf.Ticker(_tsym).fast_info.currency or ""
            except Exception:
                _t3_currency = "GBp" if _tsym.endswith(".L") else ""
            _t3_is_gbp = (_t3_currency == "GBp")
            sym_c      = "£"  # always £ for GBp stocks; get_currency_symbol for others
            if not _t3_is_gbp:
                sym_c = get_currency_symbol(_tsym)

            # Raw prices (pence for GBp stocks); convert to £ for display
            _raw_close  = float(df_i["Close"].iloc[-1])
            _raw_prev   = float(df_i["Close"].iloc[-2]) if len(df_i) > 1 else _raw_close
            _px_div     = 100.0 if _t3_is_gbp else 1.0
            curr_price  = _raw_close / _px_div   # display price (£ or native)
            prev_price  = _raw_prev  / _px_div
            day_chg     = curr_price - prev_price
            day_pct     = day_chg / prev_price * 100 if prev_price else 0
            rsi_val     = sigs.get("rsi") or 0
            _52w_hi     = float(df_i["High"].max())  / _px_div
            _52w_lo     = float(df_i["Low"].min())   / _px_div

            # ── Score verdict ───────────────────────────────────
            if score >= 4:
                _vdict_col  = "#4ADE80"; _vdict_bg = "rgba(74,222,128,0.12)"
                _vdict_icon = "🟢"; _vdict_txt = f"Strong Setup ({score}/5)"
                _vdict_sub  = "Technical signals are aligning. Confirm fundamentals + catalyst before entering."
            elif score >= 2:
                _vdict_col  = "#F59E0B"; _vdict_bg = "rgba(245,158,11,0.10)"
                _vdict_icon = "🟡"; _vdict_txt = f"Moderate Setup ({score}/5)"
                _vdict_sub  = "Some signals present. Monitor for further confirmation before acting."
            else:
                _vdict_col  = "#60A5FA"; _vdict_bg = "rgba(96,165,250,0.10)"
                _vdict_icon = "🔵"; _vdict_txt = f"Weak Setup ({score}/5)"
                _vdict_sub  = "Conditions not yet favourable. Wait for better entry signals."

            _chg_col = "#4ADE80" if day_chg >= 0 else "#F87171"
            _chg_arr = "▲" if day_chg >= 0 else "▼"
            _rsi_col = "#F87171" if rsi_val > 70 else ("#4ADE80" if rsi_val < 30 else "#F59E0B")
            _ma_ok   = sigs.get("ma50_above_200", False)
            _vol_ok  = sigs.get("volume_spike", False)
            _pf52    = sigs.get("pct_from_52w_high")

            # KPI cards using Streamlit columns (no CSS grid)
            _kc = st.columns(6)
            def _kpi(col, label, value, sub, border_col="#334155"):
                col.markdown(f"""
<div style="background:rgba(13,31,53,0.8);border:1px solid {border_col};
            border-radius:10px;padding:14px 10px;text-align:center;height:90px">
  <div style="color:#64748B;font-size:0.68rem;font-weight:700;letter-spacing:.07em;margin-bottom:4px">{label}</div>
  <div style="font-size:1.35rem;font-weight:900;color:#F1F5F9;line-height:1.1">{value}</div>
  <div style="font-size:0.69rem;color:#64748B;margin-top:4px">{sub}</div>
</div>""", unsafe_allow_html=True)

            _kpi(_kc[0], "SETUP SCORE", f"{score}/5",
                 f"{_vdict_icon} {_vdict_txt.split('(')[0].strip()}", _vdict_col)
            _kpi(_kc[1], "PRICE", f"{sym_c}{curr_price:,.2f}",
                 f'<span style="color:{_chg_col}">{_chg_arr} {abs(day_pct):.2f}% today</span>')
            _kpi(_kc[2], "RSI (14)", f"{rsi_val:.1f}",
                 f'<span style="color:{_rsi_col}">{"Overbought" if rsi_val>70 else ("Oversold" if rsi_val<30 else "Neutral zone")}</span>')
            _kpi(_kc[3], "TREND",
                 "Bullish" if _ma_ok else "Bearish",
                 f'MA50 {"above" if _ma_ok else "below"} MA200',
                 "#4ADE80" if _ma_ok else "#F87171")
            _kpi(_kc[4], "VOLUME",
                 "Spike ⚡" if _vol_ok else "Normal",
                 "vs 20-day average",
                 "#4ADE80" if _vol_ok else "#334155")
            _kpi(_kc[5], "FROM 52W HIGH",
                 f"{_pf52:.1f}%" if _pf52 is not None else "–",
                 f"Hi:{sym_c}{_52w_hi:,.0f} Lo:{sym_c}{_52w_lo:,.0f}",
                 "#4ADE80" if (_pf52 or 100) < 10 else "#F59E0B")

            st.markdown(
                f'<div style="background:{_vdict_bg};border-left:4px solid {_vdict_col};'
                f'border-radius:0 8px 8px 0;padding:10px 18px;margin:10px 0 18px">'
                f'<span style="color:{_vdict_col};font-weight:700">{_vdict_icon} {_vdict_txt}</span>'
                f'<span style="color:#94A3B8;font-size:0.88rem;margin-left:12px">{_vdict_sub}</span>'
                f'</div>', unsafe_allow_html=True)

            # ── Shared chart theme ──────────────────────────────
            _PAPER = "#0D1F35"
            _PLOT  = "#0A1929"
            _GRID  = "rgba(255,255,255,0.04)"
            _FONT  = dict(family="Inter, sans-serif", color="#94A3B8", size=11)

            def _base_layout(title, height):
                return dict(
                    title=dict(text=title, font=dict(color="#CBD5E1", size=12,
                               family="Inter, sans-serif"), x=0.01, xanchor="left", y=0.97),
                    height=height,
                    paper_bgcolor=_PAPER, plot_bgcolor=_PLOT, font=_FONT,
                    xaxis=dict(showgrid=True, gridcolor=_GRID, zeroline=False,
                               tickfont=dict(color="#475569", size=10),
                               showspikes=True, spikecolor="#334155", spikethickness=1,
                               spikedash="dot"),
                    yaxis=dict(showgrid=True, gridcolor=_GRID, zeroline=False,
                               tickfont=dict(color="#475569", size=10),
                               showspikes=True, spikecolor="#334155", spikethickness=1),
                    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                                font=dict(color="#94A3B8", size=11), bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10, r=10, t=52, b=10),
                    hovermode="x unified",
                    hoverlabel=dict(bgcolor="#1E3A52", bordercolor="#334155",
                                    font=dict(color="#F1F5F9", size=11)),
                )

            # ── Scale prices for GBp ─────────────────────────────
            _df_plot = df_i.copy()
            if _t3_is_gbp:
                for _col in ["Open","High","Low","Close","MA50","MA200","BB_upper","BB_lower"]:
                    if _col in _df_plot.columns:
                        _df_plot[_col] = _df_plot[_col] / 100.0

            # ── Auto-detect Support / Resistance levels ──────────
            def _sr_levels(df, n_levels=3, window=10):
                """Pivot-point support/resistance: find local highs and lows."""
                highs = df["High"].values if not _t3_is_gbp else df["High"].values / 100.0
                lows  = df["Low"].values  if not _t3_is_gbp else df["Low"].values  / 100.0
                res, sup = [], []
                for i in range(window, len(df) - window):
                    if highs[i] == max(highs[i-window:i+window+1]):
                        res.append(highs[i])
                    if lows[i] == min(lows[i-window:i+window+1]):
                        sup.append(lows[i])
                # cluster nearby levels (within 0.5%)
                def _cluster(vals):
                    if not vals: return []
                    vals = sorted(set(vals))
                    clusters, grp = [], [vals[0]]
                    for v in vals[1:]:
                        if abs(v - grp[-1]) / grp[-1] < 0.005:
                            grp.append(v)
                        else:
                            clusters.append(sum(grp)/len(grp)); grp = [v]
                    clusters.append(sum(grp)/len(grp))
                    return clusters
                res_cl = _cluster(res); sup_cl = _cluster(sup)
                # return levels closest to current price
                cp = float(_df_plot["Close"].iloc[-1])
                res_cl = sorted(res_cl, key=lambda x: abs(x-cp))[:n_levels]
                sup_cl = sorted(sup_cl, key=lambda x: abs(x-cp))[:n_levels]
                return res_cl, sup_cl

            # ── Auto-detect Trendlines ───────────────────────────
            def _trendline(df, use_highs=True, lookback=60):
                """Linear regression trendline through recent highs or lows."""
                series = (_df_plot["High"] if use_highs else _df_plot["Low"]).tail(lookback)
                if len(series) < 10: return None, None
                x = np.arange(len(series))
                y = series.values
                m, b = np.polyfit(x, y, 1)
                x_dates = series.index
                y_fit = m * x + b
                return x_dates, y_fit

            _res_levels, _sup_levels = _sr_levels(_df_plot if not _t3_is_gbp else df_i)
            _tl_dates_hi, _tl_vals_hi = _trendline(_df_plot, use_highs=True)
            _tl_dates_lo, _tl_vals_lo = _trendline(_df_plot, use_highs=False)

            # ── MACD calculation ─────────────────────────────────
            _close_s = _df_plot["Close"]
            _ema12   = _close_s.ewm(span=12, adjust=False).mean()
            _ema26   = _close_s.ewm(span=26, adjust=False).mean()
            _macd    = _ema12 - _ema26
            _signal  = _macd.ewm(span=9, adjust=False).mean()
            _hist    = _macd - _signal

            # ── 1. Main price chart (subplot with volume) ─────────
            from plotly.subplots import make_subplots
            fig = make_subplots(
                rows=2, cols=1, shared_xaxes=True,
                row_heights=[0.75, 0.25],
                vertical_spacing=0.02,
            )

            # Candlestick or line
            if chart_style == "Candlestick":
                fig.add_trace(go.Candlestick(
                    x=_df_plot.index, open=_df_plot["Open"], high=_df_plot["High"],
                    low=_df_plot["Low"], close=_df_plot["Close"], name="Price",
                    increasing=dict(line=dict(color="#22C55E", width=1), fillcolor="#16A34A"),
                    decreasing=dict(line=dict(color="#EF4444", width=1), fillcolor="#991B1B"),
                    whiskerwidth=0.4), row=1, col=1)
            else:
                fig.add_trace(go.Scatter(
                    x=_df_plot.index, y=_df_plot["Close"], mode="lines", name="Close",
                    line=dict(color="#60A5FA", width=2),
                    fill="tozeroy", fillcolor="rgba(96,165,250,0.05)"), row=1, col=1)

            # MAs and Bollinger Bands
            for nm, color, dash, wid, lbl in [
                ("MA50",     "#F59E0B", "solid", 1.8, "MA 50"),
                ("MA200",    "#60A5FA", "solid", 1.8, "MA 200"),
                ("BB_upper", "#475569", "dot",   1.0, "BB Upper"),
                ("BB_lower", "#475569", "dot",   1.0, "BB Lower"),
            ]:
                if nm in _df_plot.columns:
                    fig.add_trace(go.Scatter(
                        x=_df_plot.index, y=_df_plot[nm], mode="lines", name=lbl,
                        line=dict(color=color, width=wid, dash=dash),
                        fill="tonexty" if nm == "BB_lower" else None,
                        fillcolor="rgba(71,85,105,0.06)" if nm == "BB_lower" else None,
                        hovertemplate=f"{lbl}: {sym_c}%{{y:,.2f}}<extra></extra>"),
                        row=1, col=1)

            # Trendlines
            if _tl_dates_hi is not None:
                fig.add_trace(go.Scatter(
                    x=_tl_dates_hi, y=_tl_vals_hi, mode="lines", name="Resistance Trendline",
                    line=dict(color="#F87171", width=1.5, dash="dash"),
                    hovertemplate="Res trendline: %{y:,.2f}<extra></extra>"),
                    row=1, col=1)
            if _tl_dates_lo is not None:
                fig.add_trace(go.Scatter(
                    x=_tl_dates_lo, y=_tl_vals_lo, mode="lines", name="Support Trendline",
                    line=dict(color="#4ADE80", width=1.5, dash="dash"),
                    hovertemplate="Sup trendline: %{y:,.2f}<extra></extra>"),
                    row=1, col=1)

            # Support / Resistance horizontal lines
            # De-duplicate levels within 2% of each other to avoid label overlap
            def _dedup_levels(levels, pct=0.02):
                if not levels:
                    return []
                srt = sorted(levels)
                out = [srt[0]]
                for lv in srt[1:]:
                    if abs(lv - out[-1]) / max(abs(out[-1]), 1e-9) > pct:
                        out.append(lv)
                return out

            _res_clean = _dedup_levels(_res_levels)
            _sup_clean = _dedup_levels(_sup_levels)

            # Alternate annotation sides to prevent overlap when levels are close
            for _i, _lvl in enumerate(_res_clean):
                _pos = "right" if _i % 2 == 0 else "top right"
                fig.add_hline(y=_lvl, line_color="rgba(248,113,113,0.5)", line_dash="dot",
                              line_width=1,
                              annotation_text=f"R {sym_c}{_lvl:,.2f}",
                              annotation_position=_pos,
                              annotation_font=dict(color="#F87171", size=9), row=1, col=1)
            for _i, _lvl in enumerate(_sup_clean):
                _pos = "right" if _i % 2 == 0 else "bottom right"
                fig.add_hline(y=_lvl, line_color="rgba(74,222,128,0.5)", line_dash="dot",
                              line_width=1,
                              annotation_text=f"S {sym_c}{_lvl:,.2f}",
                              annotation_position=_pos,
                              annotation_font=dict(color="#4ADE80", size=9), row=1, col=1)

            # Volume bars (row 2)
            _vol_raw  = np.array(df_i["Volume"]).flatten().astype(float)
            _cls_raw  = np.array(df_i["Close"]).flatten().astype(float)
            _opn_raw  = np.array(df_i["Open"]).flatten().astype(float)
            _vol_cols = ["rgba(34,197,94,0.6)" if c >= o else "rgba(239,68,68,0.6)"
                         for c, o in zip(_cls_raw, _opn_raw)]
            fig.add_trace(go.Bar(
                x=df_i.index, y=_vol_raw, name="Volume",
                marker_color=_vol_cols, marker_line_width=0,
                hovertemplate="Vol: %{y:,.0f}<extra></extra>"), row=2, col=1)
            if "Vol20" in df_i.columns:
                fig.add_trace(go.Scatter(
                    x=df_i.index, y=np.array(df_i["Vol20"]).flatten().astype(float),
                    mode="lines", name="Vol MA20",
                    line=dict(color="#F59E0B", width=1.5, dash="dash"),
                    hovertemplate="Vol MA20: %{y:,.0f}<extra></extra>"), row=2, col=1)

            fig.update_layout(
                height=640,
                paper_bgcolor=_PAPER, plot_bgcolor=_PLOT, font=_FONT,
                title=dict(
                    text=f"<b>{_tsym}</b>  ·  {price_period.upper()}  ·  Last: <b>{sym_c}{curr_price:,.2f}</b>"
                         f"  <span style='color:{'#4ADE80' if day_chg>=0 else '#F87171'}'>"
                         f"{'▲' if day_chg>=0 else '▼'} {abs(day_pct):.2f}%</span>",
                    font=dict(color="#F1F5F9", size=13, family="Inter, sans-serif"),
                    x=0.01, xanchor="left"),
                xaxis=dict(showgrid=True, gridcolor=_GRID, zeroline=False,
                           tickfont=dict(color="#64748B", size=10),
                           showspikes=True, spikecolor="#475569", spikethickness=1, spikedash="dot",
                           rangeslider_visible=False),
                xaxis2=dict(showgrid=True, gridcolor=_GRID, zeroline=False,
                            tickfont=dict(color="#64748B", size=10)),
                yaxis=dict(showgrid=True, gridcolor=_GRID, zeroline=False,
                           tickprefix=sym_c, tickfont=dict(color="#94A3B8", size=10),
                           showspikes=True, spikecolor="#475569", spikethickness=1, side="right"),
                yaxis2=dict(showgrid=False, zeroline=False,
                            tickfont=dict(color="#64748B", size=9), side="right",
                            title=dict(text="Volume", font=dict(color="#64748B", size=9))),
                legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                            font=dict(color="#94A3B8", size=10), bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=10, r=80, t=52, b=10),
                hovermode="x unified",
                hoverlabel=dict(bgcolor="#1E293B", bordercolor="#334155",
                                font=dict(color="#F1F5F9", size=11)),
                bargap=0.1,
            )
            st.plotly_chart(fig, use_container_width=True)

            # ── 2. RSI chart ─────────────────────────────────────
            rsi_vals = np.array(df_i["RSI"]).flatten()
            fig_r = go.Figure()
            fig_r.add_hrect(y0=70, y1=100, fillcolor="rgba(239,68,68,0.06)",  line_width=0)
            fig_r.add_hrect(y0=0,  y1=30,  fillcolor="rgba(34,197,94,0.06)",  line_width=0)
            fig_r.add_trace(go.Scatter(
                x=df_i.index, y=rsi_vals, mode="lines", name="RSI (14)",
                line=dict(color="#A78BFA", width=2),
                fill="tozeroy", fillcolor="rgba(167,139,250,0.05)",
                hovertemplate="RSI: %{y:.1f}<extra></extra>"))
            for y_lvl, col, lbl in [(70,"#EF4444","Overbought 70"),(30,"#22C55E","Oversold 30"),(50,"#475569","50")]:
                fig_r.add_hline(y=y_lvl, line_color=col, line_dash="dash", line_width=1,
                                annotation_text=lbl, annotation_position="right",
                                annotation_font=dict(color=col, size=9))
            lay_r = _base_layout("RSI (14)  ·  <70 overbought  ·  <30 oversold  ·  Ideal entry zone 40–65", 200)
            lay_r["yaxis"]["range"] = [0, 100]
            lay_r["yaxis"]["tickfont"] = dict(color="#94A3B8", size=10)
            lay_r["margin"] = dict(l=10, r=80, t=40, b=10)
            fig_r.update_layout(**lay_r)
            st.plotly_chart(fig_r, use_container_width=True)

            # ── 3. MACD chart ─────────────────────────────────────
            _hist_colors = ["rgba(74,222,128,0.7)" if v >= 0 else "rgba(248,113,113,0.7)"
                            for v in _hist.values]
            fig_m = go.Figure()
            fig_m.add_trace(go.Bar(
                x=_df_plot.index, y=_hist.values, name="MACD Histogram",
                marker_color=_hist_colors, marker_line_width=0,
                hovertemplate="Hist: %{y:,.4f}<extra></extra>"))
            fig_m.add_trace(go.Scatter(
                x=_df_plot.index, y=_macd.values, mode="lines", name="MACD",
                line=dict(color="#60A5FA", width=1.8),
                hovertemplate="MACD: %{y:,.4f}<extra></extra>"))
            fig_m.add_trace(go.Scatter(
                x=_df_plot.index, y=_signal.values, mode="lines", name="Signal (9)",
                line=dict(color="#F59E0B", width=1.8),
                hovertemplate="Signal: %{y:,.4f}<extra></extra>"))
            fig_m.add_hline(y=0, line_color="#334155", line_width=1)
            lay_m = _base_layout("MACD (12, 26, 9)  ·  Blue crosses amber = bullish  ·  Histogram above 0 = momentum building", 200)
            lay_m["yaxis"]["tickfont"] = dict(color="#94A3B8", size=10)
            lay_m["margin"] = dict(l=10, r=80, t=40, b=10)
            lay_m["bargap"] = 0.15
            fig_m.update_layout(**lay_m)
            st.plotly_chart(fig_m, use_container_width=True)

            # ── INTERPRETATION GUIDE (collapsible) ───────────────
            st.divider()
            with st.expander("📖 How to Read These Charts", expanded=False):
                g1, g2, g3 = st.columns(3)
                with g1:
                    st.markdown("##### 🕯️ Candlestick & Moving Averages")
                    st.markdown("""
- **Green candle** = closed higher than open (bullish day)
- **Red candle** = closed lower than open (bearish day)
- **MA 50 (amber line)** = 50-day moving average — short-term trend direction. Price above MA50 = bullish momentum
- **MA 200 (blue line)** = 200-day moving average — long-term trend. When MA50 crosses above MA200 it's a **Golden Cross** — a major bullish signal
- **Bollinger Bands (grey dots)** = volatility envelope. Price near the lower band may signal an oversold bounce; near upper band = potential resistance
""")
                with g2:
                    st.markdown("##### 📊 RSI — Relative Strength Index")
                    st.markdown("""
Momentum oscillator on a **0–100 scale**.

- 🔴 **Above 70** = Overbought — avoid new entries, pullback likely
- 🟢 **Below 30** = Oversold — potential reversal / buying opportunity
- 🟡 **40–65 (ideal zone)** = Positive momentum without being stretched — best entry window
- *RSI Divergence*: price makes a new high but RSI doesn't — warns of weakening momentum and possible reversal ahead
""")
                with g3:
                    st.markdown("##### 📦 Volume")
                    st.markdown("""
Volume **confirms** price moves — never ignore it.

- 🟢 **Green bar above amber line** = High-volume up-day — institutional buying, strong bullish signal
- 🔴 **Red bar above amber line** = High-volume selling — distribution / smart money exiting
- 🟡 **Amber dashed line** = 20-day average volume baseline
- A price breakout on **low volume** is unreliable and often fails. Always look for volume confirmation before entering
""")

            with st.expander("🔗 How to use Technical Setup with Fundamental Screen & Catalyst Alerts", expanded=False):
                s1, s2, s3 = st.columns(3)
                with s1:
                    st.markdown("##### 🔵 Step 1 — Qualify on Fundamentals")
                    st.markdown("""
Start with the **Fundamental Screen**. Only proceed with stocks that score **≥ 70/100** and show positive DCF upside.

You must be investing in a genuinely good business at a fair or cheap price. Technical analysis on a poor-quality stock is speculation, not investing. Fundamentals define **what** to buy.
""")
                with s2:
                    st.markdown("##### 🟡 Step 2 — Find a Catalyst")
                    st.markdown("""
Go to **Catalyst Alerts**. Look for Tier 1 signals: earnings in 10–30 days, director buying, positive analyst revisions.

A catalyst gives the market a specific reason to re-rate the stock upward. Without one, even the most undervalued quality stocks can stay cheap for years. Catalysts define **when** to act.
""")
                with s3:
                    st.markdown("##### 🟢 Step 3 — Time Your Entry")
                    st.markdown("""
Only now use this screen to time your entry precisely.

**Best setup — all of these together:**
- RSI between 40–65
- MA50 above MA200 (bullish trend)
- Price pulling back to MA50 support
- Rising volume on up-days

A **Setup Score ≥ 4/5** with a confirmed catalyst and strong fundamentals = high-conviction entry.

> ⚠️ Technicals improve *timing* but never replace fundamental analysis.
""")



# ═══════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════
# TAB MC — GBM MONTE CARLO RISK SIMULATOR
# ═══════════════════════════════════════════════════════════════

with tab_mc:
    st.markdown('<div class="section-header">🎲 Monte Carlo Risk Simulator — GBM Price Path Analysis</div>',
                unsafe_allow_html=True)
    st.caption(
        "Geometric Brownian Motion · Jump Diffusion · EWMA Volatility · "
        "Used by investment banks & hedge funds to map the full probability distribution of future prices"
    )

    # ── Ticker input ─────────────────────────────────────────────
    _mc_universe = list(st.session_state.get("screened_symbols", []))
    _mc_col1, _mc_col2 = st.columns([2, 3])
    with _mc_col1:
        if _mc_universe:
            _mc_ticker_sel = st.selectbox(
                "Stock (from your screened universe):",
                ["— type below —"] + _mc_universe, key="mc_ticker_sel"
            )
            _mc_ticker = st.text_input("Or enter any ticker:", key="mc_ticker_manual",
                                        placeholder="e.g. AAPL, HSBA.L")
            _mc_ticker = (_mc_ticker.strip().upper()
                          if _mc_ticker.strip()
                          else (None if _mc_ticker_sel == "— type below —" else _mc_ticker_sel))
        else:
            _mc_ticker = st.text_input("Ticker symbol:", key="mc_ticker_manual",
                                        placeholder="e.g. AAPL, HSBA.L, VOD.L").strip().upper() or None
            st.caption("💡 Run the Fundamental Screen first to auto-populate your universe here.")

    with _mc_col2:
        _mc_target = st.number_input(
            "Target price (e.g. your DCF intrinsic value):",
            min_value=0.0, value=0.0, step=0.5, format="%.2f", key="mc_target",
            help="Set this to your DCF estimate. The simulator will calculate the probability of reaching it."
        )
        _mc_stop = st.number_input(
            "Stop-loss / downside floor:",
            min_value=0.0, value=0.0, step=0.5, format="%.2f", key="mc_stop",
            help="Price below which you'd exit. Used to calculate probability of loss."
        )

    # ── Simulation settings ───────────────────────────────────────
    st.markdown("#### ⚙️ Simulation Settings")
    _ms1, _ms2, _ms3, _ms4 = st.columns(4)
    with _ms1:
        _mc_horizon = st.selectbox("Time horizon:", ["30 days","60 days","90 days",
                                                      "6 months","1 year","2 years","5 years"],
                                    index=4, key="mc_horizon")
        _mc_T = {"30 days":30/252,"60 days":60/252,"90 days":90/252,
                  "6 months":0.5,"1 year":1.0,"2 years":2.0,"5 years":5.0}[_mc_horizon]
    with _ms2:
        _mc_n_sims = st.selectbox("Simulations:", [5_000, 10_000, 50_000, 100_000],
                                   index=1, key="mc_nsims",
                                   format_func=lambda x: f"{x:,}")
    with _ms3:
        _mc_model = st.selectbox("Model:", ["Standard GBM","Jump Diffusion","EWMA Volatility"],
                                  key="mc_model",
                                  help="Standard GBM: classic model  |  "
                                       "Jump Diffusion: adds crash/spike events (Merton)  |  "
                                       "EWMA Volatility: volatility clustering (RiskMetrics)")
    with _ms4:
        _mc_ret_src = st.selectbox("Expected return source:",
                                    ["Historical average","CAPM","Custom"],
                                    key="mc_ret_src")
        _mc_custom_mu = None
        if _mc_ret_src == "Custom":
            _mc_custom_mu = st.number_input("Annual return (%):", value=10.0,
                                             step=0.5, key="mc_custom_mu") / 100

    _run_mc = st.button("⚡ Run Simulation", type="primary", key="mc_run_btn",
                         disabled=not _mc_ticker)

    if _run_mc and _mc_ticker:

        @st.cache_data(ttl=300, show_spinner=False)
        def _mc_fetch(ticker: str):
            import yfinance as _yf_mc
            _t = _yf_mc.Ticker(ticker)
            _hist = _t.history(period="3y", auto_adjust=True)
            _info = {}
            try: _info = _t.info
            except Exception: pass
            return _hist, _info

        with st.spinner(f"Fetching {_mc_ticker} price history…"):
            _mc_hist, _mc_info = _mc_fetch(_mc_ticker)

        if _mc_hist.empty or len(_mc_hist) < 60:
            st.error(f"Not enough price data for {_mc_ticker}. Check the ticker symbol.")
        else:
            import numpy as _np_mc

            _mc_close = _mc_hist["Close"].dropna()
            # yfinance returns LSE (.L) prices in pence — convert to pounds
            _mc_is_gbx = _mc_ticker.upper().endswith(".L")
            if _mc_is_gbx:
                _mc_close = _mc_close / 100
            _mc_S0    = float(_mc_close.iloc[-1])
            _mc_log_r = _np_mc.log(_mc_close / _mc_close.shift(1)).dropna().values
            _mc_hist_vol = float(_mc_log_r.std() * _np_mc.sqrt(252))

            # ── Risk-free rate from ^TNX ──────────────────────────
            try:
                import yfinance as _yf_rf
                _rf_data = _yf_rf.Ticker("^TNX").history(period="5d")
                _mc_rf = float(_rf_data["Close"].iloc[-1]) / 100
            except Exception:
                _mc_rf = 0.0425

            # ── Expected return μ ─────────────────────────────────
            if _mc_ret_src == "Historical average":
                _mc_mu = float(_mc_log_r.mean() * 252)
            elif _mc_ret_src == "CAPM":
                # β from 3y regression vs S&P500
                try:
                    import yfinance as _yf_spy
                    _spy = _yf_spy.Ticker("^GSPC").history(period="3y", auto_adjust=True)["Close"]
                    _spy_r = _np_mc.log(_spy / _spy.shift(1)).dropna().values
                    _min_len = min(len(_spy_r), len(_mc_log_r))
                    _cov = _np_mc.cov(_mc_log_r[-_min_len:], _spy_r[-_min_len:])
                    _beta = _cov[0,1] / _cov[1,1]
                    _mkt_ret = float(_spy_r.mean() * 252)
                    _mc_mu = _mc_rf + _beta * (_mkt_ret - _mc_rf)
                except Exception:
                    _mc_mu = float(_mc_log_r.mean() * 252)
            else:
                _mc_mu = _mc_custom_mu if _mc_custom_mu is not None else float(_mc_log_r.mean() * 252)

            # ── Model-specific volatility ─────────────────────────
            if _mc_model == "EWMA Volatility":
                # RiskMetrics EWMA: λ=0.94
                _lam = 0.94
                _ewma_var = float(_mc_log_r[-1]**2)
                for _r in _mc_log_r[-252:]:
                    _ewma_var = _lam * _ewma_var + (1 - _lam) * _r**2
                _mc_sigma = float(_np_mc.sqrt(_ewma_var * 252))
            else:
                _mc_sigma = _mc_hist_vol

            # ── Jump parameters (Merton) ──────────────────────────
            _mc_lam_j, _mc_mu_j, _mc_sig_j = 0.0, 0.0, 0.0
            if _mc_model == "Jump Diffusion":
                # Estimate jump frequency: days where |return| > 3σ
                _thresh = 3 * _mc_hist_vol / _np_mc.sqrt(252)
                _jumps  = _mc_log_r[_np_mc.abs(_mc_log_r) > _thresh]
                _mc_lam_j  = len(_jumps) / (len(_mc_log_r) / 252)  # jumps/year
                _mc_mu_j   = float(_jumps.mean()) if len(_jumps) > 0 else 0.0
                _mc_sig_j  = float(_jumps.std())  if len(_jumps) > 1 else 0.01
                # Drift adjustment so total expected return = μ
                _mc_mu_adj = _mc_mu - _mc_lam_j * (_np_mc.exp(_mc_mu_j + 0.5*_mc_sig_j**2) - 1)
            else:
                _mc_mu_adj = _mc_mu

            # ── Run simulation ────────────────────────────────────
            _mc_steps = max(int(_mc_T * 252), 1)
            _mc_dt    = _mc_T / _mc_steps

            _np_mc.random.seed(42)
            with st.spinner(f"Running {_mc_n_sims:,} simulations…"):
                _mc_Z     = _np_mc.random.normal(0, 1, (_mc_n_sims, _mc_steps))
                _mc_paths = _np_mc.zeros((_mc_n_sims, _mc_steps + 1))
                _mc_paths[:, 0] = _mc_S0

                if _mc_model == "Jump Diffusion":
                    _mc_N_j = _np_mc.random.poisson(_mc_lam_j * _mc_dt, (_mc_n_sims, _mc_steps))
                    _mc_J   = _np_mc.random.normal(_mc_mu_j, _mc_sig_j, (_mc_n_sims, _mc_steps))
                    for _t in range(_mc_steps):
                        _mc_paths[:, _t+1] = _mc_paths[:, _t] * _np_mc.exp(
                            (_mc_mu_adj - 0.5 * _mc_sigma**2) * _mc_dt
                            + _mc_sigma * _np_mc.sqrt(_mc_dt) * _mc_Z[:, _t]
                            + _mc_N_j[:, _t] * _mc_J[:, _t]
                        )
                else:
                    _mc_paths[:, 1:] = _np_mc.exp(
                        (_mc_mu_adj - 0.5 * _mc_sigma**2) * _mc_dt
                        + _mc_sigma * _np_mc.sqrt(_mc_dt) * _mc_Z
                    )
                    _mc_paths = _mc_paths[:, 0:1] * _np_mc.cumprod(
                        _np_mc.concatenate([_np_mc.ones((_mc_n_sims, 1)), _mc_paths[:, 1:]], axis=1),
                        axis=1
                    )

            _mc_final = _mc_paths[:, -1]
            _mc_mean  = float(_mc_final.mean())
            _mc_med   = float(_np_mc.median(_mc_final))
            _mc_std   = float(_mc_final.std())
            _mc_p5    = float(_np_mc.percentile(_mc_final, 5))
            _mc_p10   = float(_np_mc.percentile(_mc_final, 10))
            _mc_p25   = float(_np_mc.percentile(_mc_final, 25))
            _mc_p75   = float(_np_mc.percentile(_mc_final, 75))
            _mc_p90   = float(_np_mc.percentile(_mc_final, 90))
            _mc_p95   = float(_np_mc.percentile(_mc_final, 95))
            _mc_var95 = float(_mc_S0 - _mc_p5)
            _mc_var99 = float(_mc_S0 - _np_mc.percentile(_mc_final, 1))
            # Expected Shortfall (CVaR) at 95%
            _mc_es95  = float(_mc_S0 - _mc_final[_mc_final <= _mc_p5].mean()) if (_mc_final <= _mc_p5).any() else 0.0
            _mc_cagr  = float((_mc_mean / _mc_S0) ** (1 / max(_mc_T, 0.08)) - 1)
            # Max drawdown across paths (sample 5000 for speed)
            _mc_mdd_sample = _mc_paths[:5000]
            _mc_running_max = _np_mc.maximum.accumulate(_mc_mdd_sample, axis=1)
            _mc_dd = (_mc_mdd_sample - _mc_running_max) / _mc_running_max
            _mc_max_dd = float(_mc_dd.min())

            # Probabilities
            _mc_p_profit  = float((_mc_final > _mc_S0).mean())
            _mc_p_loss    = float((_mc_final < _mc_S0).mean())
            _mc_p_target  = float((_mc_final >= _mc_target).mean()) if _mc_target > 0 else None
            _mc_p_stop    = float((_mc_final <= _mc_stop).mean())   if _mc_stop > 0 else None

            # ── Current price symbol ──────────────────────────────
            _mc_sym = "£" if _mc_ticker.endswith(".L") else "$"
            _mc_name = _mc_info.get("shortName", _mc_ticker)

            st.markdown("---")

            # ── KPI summary row ───────────────────────────────────
            st.markdown(f"#### 📊 Results — {_mc_name} · {_mc_n_sims:,} simulations · {_mc_horizon}")
            _mk1, _mk2, _mk3, _mk4, _mk5, _mk6 = st.columns(6)
            def _mc_kpi(col, label, val, color="#F1F5F9"):
                col.markdown(
                    f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.25);'
                    f'border-radius:10px;padding:10px;text-align:center">'
                    f'<div style="font-size:0.58rem;color:#64748B;text-transform:uppercase;'
                    f'letter-spacing:0.5px;margin-bottom:3px">{label}</div>'
                    f'<div style="font-size:1.05rem;font-weight:800;color:{color}">{val}</div>'
                    f'</div>', unsafe_allow_html=True)

            _mc_kpi(_mk1, "Current Price", f"{_mc_sym}{_mc_S0:,.2f}", "#F1F5F9")
            _mc_kpi(_mk2, "Mean Outcome", f"{_mc_sym}{_mc_mean:,.2f}",
                    "#22C55E" if _mc_mean > _mc_S0 else "#EF4444")
            _mc_kpi(_mk3, "Median Outcome", f"{_mc_sym}{_mc_med:,.2f}",
                    "#22C55E" if _mc_med > _mc_S0 else "#EF4444")
            _mc_kpi(_mk4, "Expected CAGR", f"{_mc_cagr:+.1%}",
                    "#22C55E" if _mc_cagr > 0 else "#EF4444")
            _mc_kpi(_mk5, "VaR 95%", f"-{_mc_sym}{_mc_var95:,.2f}", "#EF4444")
            _mc_kpi(_mk6, "Max Sim Drawdown", f"{_mc_max_dd:.1%}", "#991B1B")

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            # ── Probability row ───────────────────────────────────
            _mp1, _mp2, _mp3, _mp4 = st.columns(4)
            _mc_kpi(_mp1, "Prob. of Profit", f"{_mc_p_profit:.1%}",
                    "#22C55E" if _mc_p_profit > 0.5 else "#EF4444")
            _mc_kpi(_mp2, "Prob. of Loss", f"{_mc_p_loss:.1%}",
                    "#EF4444" if _mc_p_loss > 0.3 else "#F59E0B")
            if _mc_p_target is not None:
                _mc_kpi(_mp3, f"P(≥ {_mc_sym}{_mc_target:,.0f})", f"{_mc_p_target:.1%}",
                        "#22C55E" if _mc_p_target > 0.4 else "#F59E0B" if _mc_p_target > 0.2 else "#EF4444")
            else:
                _mc_kpi(_mp3, "P(≥ Target)", "Set target above", "#64748B")
            if _mc_p_stop is not None:
                _mc_kpi(_mp4, f"P(≤ {_mc_sym}{_mc_stop:,.0f})", f"{_mc_p_stop:.1%}",
                        "#EF4444" if _mc_p_stop > 0.15 else "#F59E0B" if _mc_p_stop > 0.05 else "#22C55E")
            else:
                _mc_kpi(_mp4, "P(≤ Stop Loss)", "Set stop above", "#64748B")

            st.markdown("---")

            # ── Fan Chart ─────────────────────────────────────────
            st.markdown("#### 📈 Price Path Fan Chart")
            _mc_x = _np_mc.linspace(0, _mc_T * 252, _mc_steps + 1)

            _fan_fig = go.Figure()

            # Confidence bands (shaded fans)
            _bands = [
                (_mc_paths, 1, 99, "rgba(239,68,68,0.08)"),
                (_mc_paths, 5, 95, "rgba(245,158,11,0.12)"),
                (_mc_paths, 10, 90, "rgba(245,158,11,0.16)"),
                (_mc_paths, 25, 75, "rgba(34,197,94,0.18)"),
            ]
            _band_labels = ["1–99%","5–95%","10–90%","25–75%"]
            for (_paths_b, _plo, _phi, _fc), _lbl in zip(_bands, _band_labels):
                _lo = _np_mc.percentile(_paths_b, _plo, axis=0)
                _hi = _np_mc.percentile(_paths_b, _phi, axis=0)
                _fan_fig.add_trace(go.Scatter(
                    x=_np_mc.concatenate([_mc_x, _mc_x[::-1]]),
                    y=_np_mc.concatenate([_hi, _lo[::-1]]),
                    fill="toself", fillcolor=_fc,
                    line=dict(width=0), name=_lbl, showlegend=True,
                    hoverinfo="skip"
                ))

            # Median path
            _mc_p50 = _np_mc.percentile(_mc_paths, 50, axis=0)
            _fan_fig.add_trace(go.Scatter(
                x=_mc_x, y=_mc_p50, mode="lines",
                line=dict(color="#F59E0B", width=2.5, dash="solid"),
                name="Median path",
                hovertemplate="Day %{x:.0f}: " + _mc_sym + "%{y:,.2f}<extra></extra>"
            ))

            # Target price line
            if _mc_target > 0:
                _fan_fig.add_hline(y=_mc_target, line_color="#22C55E", line_dash="dash",
                                   line_width=1.5,
                                   annotation_text=f"Target {_mc_sym}{_mc_target:,.2f}",
                                   annotation_font=dict(color="#22C55E", size=10),
                                   annotation_position="right")
            # Stop loss line
            if _mc_stop > 0:
                _fan_fig.add_hline(y=_mc_stop, line_color="#EF4444", line_dash="dash",
                                   line_width=1.5,
                                   annotation_text=f"Stop {_mc_sym}{_mc_stop:,.2f}",
                                   annotation_font=dict(color="#EF4444", size=10),
                                   annotation_position="right")
            # Current price
            _fan_fig.add_hline(y=_mc_S0, line_color="#94A3B8", line_dash="dot",
                                line_width=1,
                                annotation_text=f"Now {_mc_sym}{_mc_S0:,.2f}",
                                annotation_font=dict(color="#94A3B8", size=10),
                                annotation_position="right")

            _fan_fig.update_layout(
                height=420,
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=40, r=120, t=20, b=50),
                legend=dict(font=dict(color="#94A3B8", size=10),
                            bgcolor="rgba(13,31,51,0.8)", bordercolor="rgba(100,116,139,0.3)",
                            borderwidth=1, orientation="h", x=0, y=1.08),
                xaxis=dict(title="Trading Days", gridcolor="rgba(100,116,139,0.15)",
                           tickfont=dict(color="#64748B")),
                yaxis=dict(title=f"Price ({_mc_sym})", tickprefix=_mc_sym,
                           gridcolor="rgba(100,116,139,0.15)", tickfont=dict(color="#64748B")),
                hovermode="x unified"
            )
            st.plotly_chart(_fan_fig, use_container_width=True, config={"displayModeBar": False})
            st.caption(
                "Shaded bands show the probability range of outcomes. "
                "Green band = 25–75% (most likely). "
                "Outer red = 1–99% (nearly all scenarios). "
                "Gold line = median simulated path."
            )

            # ── Distribution Histogram ────────────────────────────
            st.markdown("#### 📊 Final Price Distribution")
            _hist_fig = go.Figure()
            _hist_fig.add_trace(go.Histogram(
                x=_mc_final, nbinsx=120,
                marker=dict(
                    color=[
                        "#EF4444" if v < _mc_S0 else
                        ("#22C55E" if (_mc_target > 0 and v >= _mc_target) else "#F59E0B")
                        for v in _mc_final
                    ],
                    opacity=0.75,
                    line=dict(width=0)
                ),
                name="Simulated outcomes",
                hovertemplate="Price: " + _mc_sym + "%{x:,.2f}<br>Count: %{y}<extra></extra>"
            ))
            # Mark key percentiles
            for _pct_val, _pct_lbl, _pct_clr in [
                (_mc_p5, "VaR 95%", "#EF4444"),
                (_mc_med, "Median", "#F59E0B"),
                (_mc_mean, "Mean", "#60A5FA"),
                (_mc_p95, "P95", "#22C55E"),
            ]:
                _hist_fig.add_vline(x=_pct_val, line_color=_pct_clr, line_dash="dash",
                                     line_width=1.5,
                                     annotation_text=f"{_pct_lbl} {_mc_sym}{_pct_val:,.0f}",
                                     annotation_font=dict(color=_pct_clr, size=9),
                                     annotation_position="top")
            if _mc_target > 0:
                _hist_fig.add_vline(x=_mc_target, line_color="#22C55E", line_dash="solid",
                                     line_width=2,
                                     annotation_text=f"Target {_mc_sym}{_mc_target:,.0f}",
                                     annotation_font=dict(color="#22C55E", size=9),
                                     annotation_position="top right")

            _hist_fig.update_layout(
                height=320,
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=40, r=40, t=40, b=50),
                showlegend=False,
                bargap=0.02,
                xaxis=dict(title=f"Price at end of {_mc_horizon} ({_mc_sym})",
                           tickprefix=_mc_sym, gridcolor="rgba(100,116,139,0.15)",
                           tickfont=dict(color="#64748B")),
                yaxis=dict(title="Simulation count", gridcolor="rgba(100,116,139,0.15)",
                           tickfont=dict(color="#64748B")),
            )
            st.plotly_chart(_hist_fig, use_container_width=True, config={"displayModeBar": False})

            # ── Full statistics table ─────────────────────────────
            st.markdown("---")
            st.markdown("#### 📋 Full Statistics")
            _sc1, _sc2 = st.columns(2)
            with _sc1:
                st.markdown("**Price Distribution**")
                _stats_df = pd.DataFrame([
                    ("Current Price",        f"{_mc_sym}{_mc_S0:,.2f}",     ""),
                    ("Mean (expected)",       f"{_mc_sym}{_mc_mean:,.2f}",    ""),
                    ("Median (50th pct)",     f"{_mc_sym}{_mc_med:,.2f}",     ""),
                    ("Std Deviation",         f"{_mc_sym}{_mc_std:,.2f}",     ""),
                    ("5th percentile",        f"{_mc_sym}{_mc_p5:,.2f}",      "worst 5%"),
                    ("10th percentile",       f"{_mc_sym}{_mc_p10:,.2f}",     ""),
                    ("25th percentile",       f"{_mc_sym}{_mc_p25:,.2f}",     ""),
                    ("75th percentile",       f"{_mc_sym}{_mc_p75:,.2f}",     ""),
                    ("90th percentile",       f"{_mc_sym}{_mc_p90:,.2f}",     ""),
                    ("95th percentile",       f"{_mc_sym}{_mc_p95:,.2f}",     "best 5%"),
                    ("Expected CAGR",         f"{_mc_cagr:+.1%}",             "annualised"),
                ], columns=["Metric", "Value", "Note"])
                st.dataframe(_stats_df, hide_index=True, use_container_width=True)

            with _sc2:
                st.markdown("**Risk Metrics**")
                _risk_rows = [
                    ("Model",               _mc_model,                                  ""),
                    ("Annual Volatility (σ)", f"{_mc_sigma:.1%}",                        "historical" if _mc_model != "EWMA Volatility" else "EWMA"),
                    ("Annual Drift (μ)",     f"{_mc_mu:+.1%}",                           _mc_ret_src),
                    ("Risk-Free Rate",        f"{_mc_rf:.2%}",                            "10Y Treasury"),
                    ("VaR 95%",              f"-{_mc_sym}{_mc_var95:,.2f}",              "worst 5% scenario"),
                    ("VaR 99%",              f"-{_mc_sym}{_mc_var99:,.2f}",              "worst 1% scenario"),
                    ("Expected Shortfall",   f"-{_mc_sym}{_mc_es95:,.2f}",              "avg loss beyond VaR"),
                    ("Max Sim Drawdown",     f"{_mc_max_dd:.1%}",                        "peak-to-trough"),
                    ("Prob. Profit",         f"{_mc_p_profit:.1%}",                      f"> {_mc_sym}{_mc_S0:,.2f}"),
                    ("Prob. Loss",           f"{_mc_p_loss:.1%}",                        f"< {_mc_sym}{_mc_S0:,.2f}"),
                ]
                if _mc_p_target is not None:
                    _risk_rows.append(("Prob. ≥ Target", f"{_mc_p_target:.1%}", f"≥ {_mc_sym}{_mc_target:,.0f}"))
                if _mc_p_stop is not None:
                    _risk_rows.append(("Prob. ≤ Stop", f"{_mc_p_stop:.1%}", f"≤ {_mc_sym}{_mc_stop:,.0f}"))
                if _mc_model == "Jump Diffusion":
                    _risk_rows.extend([
                        ("Jump Frequency (λ)", f"{_mc_lam_j:.1f}/year", "estimated from history"),
                        ("Avg Jump Size",      f"{_mc_mu_j:+.2%}",     "mean log-return of outliers"),
                    ])
                _risk_df = pd.DataFrame(_risk_rows, columns=["Metric", "Value", "Note"])
                st.dataframe(_risk_df, hide_index=True, use_container_width=True)

            # ══════════════════════════════════════════════════════
            # DECISION DASHBOARD
            # ══════════════════════════════════════════════════════
            st.markdown("---")
            st.markdown("#### 🎯 Investment Decision Dashboard")
            st.caption("Synthesises Fundamental, Monte Carlo and Technical signals into one decision framework")

            # Gather signals — primary source: Fundamental tab DCF session state
            _saved_dcf   = st.session_state.get("fintiq_dcf", {})
            _screened_df = st.session_state.get("screened_df")
            _tech_signal = None
            _dcf_val     = None
            _graham_val  = None
            _pe_iv_val   = None
            _avg_iv      = None
            _fund_score  = None
            _mc_signal_colour = "#F59E0B"
            _mc_signal_label  = "Neutral"

            # Read from Fundamental DCF if ticker matches
            if _saved_dcf.get("ticker", "").upper() == _mc_ticker.upper():
                _dcf_val   = _saved_dcf.get("dcf")
                _graham_val = _saved_dcf.get("graham")
                _pe_iv_val  = _saved_dcf.get("pe_iv")
                _avg_iv     = _saved_dcf.get("avg_iv")
                # Use avg_iv as the primary comparison value
                if _avg_iv:
                    _dcf_val = _avg_iv

            # Fallback: try screened_df
            if not _dcf_val and _screened_df is not None and not _screened_df.empty:
                try:
                    _match = _screened_df[_screened_df["Ticker"].str.upper() == _mc_ticker.upper()]
                    if not _match.empty:
                        _row = _match.iloc[0]
                        _tech_signal = _row.get("Tech Signal", None)
                        for _dcf_col in ["DCF Value", "DCF", "Intrinsic Value"]:
                            if _dcf_col in _row and pd.notna(_row[_dcf_col]):
                                try: _dcf_val = float(str(_row[_dcf_col]).replace("£","").replace("$","").replace(",",""))
                                except: pass
                                break
                except Exception:
                    pass

            # Tech signal from screened_df regardless
            if _screened_df is not None and not _screened_df.empty:
                try:
                    _match2 = _screened_df[_screened_df["Ticker"].str.upper() == _mc_ticker.upper()]
                    if not _match2.empty:
                        _tech_signal = _match2.iloc[0].get("Tech Signal", None)
                except Exception:
                    pass

            # Monte Carlo signal
            if _mc_p_target is not None:
                if _mc_p_target >= 0.45 and _mc_p_profit >= 0.55:
                    _mc_signal_colour, _mc_signal_label = "#22C55E", "Favourable"
                elif _mc_p_target >= 0.25:
                    _mc_signal_colour, _mc_signal_label = "#F59E0B", "Moderate"
                else:
                    _mc_signal_colour, _mc_signal_label = "#EF4444", "Unfavourable"
            elif _mc_p_profit >= 0.60:
                _mc_signal_colour, _mc_signal_label = "#22C55E", "Favourable"
            elif _mc_p_profit >= 0.45:
                _mc_signal_colour, _mc_signal_label = "#F59E0B", "Moderate"
            else:
                _mc_signal_colour, _mc_signal_label = "#EF4444", "Unfavourable"

            # Build signal cards
            _dd_html = '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px">'

            # 1. Fundamental
            if _dcf_val and _dcf_val > 0:
                _margin = (_dcf_val - _mc_S0) / _mc_S0
                _fund_col = "#22C55E" if _margin > 0.1 else "#F59E0B" if _margin > -0.1 else "#EF4444"
                _fund_label = "Undervalued" if _margin > 0.1 else "Fair Value" if _margin > -0.1 else "Overvalued"
                _fund_detail = ""
                if _saved_dcf.get("ticker", "").upper() == _mc_ticker.upper():
                    _parts = []
                    if _saved_dcf.get("dcf"):    _parts.append(f"DCF {_mc_sym}{_saved_dcf['dcf']:,.0f}")
                    if _saved_dcf.get("graham"): _parts.append(f"Graham {_mc_sym}{_saved_dcf['graham']:,.0f}")
                    if _saved_dcf.get("pe_iv"):  _parts.append(f"P/E {_mc_sym}{_saved_dcf['pe_iv']:,.0f}")
                    _fund_detail = " · ".join(_parts) if _parts else ""
                _dd_html += (
                    f'<div style="background:#0D1F33;border:1px solid {_fund_col}40;border-top:3px solid {_fund_col};'
                    f'border-radius:10px;padding:14px;text-align:center">'
                    f'<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">① Fundamental (DCF)</div>'
                    f'<div style="font-size:1.1rem;font-weight:800;color:{_fund_col}">{_fund_label}</div>'
                    f'<div style="font-size:0.72rem;color:#94A3B8;margin-top:4px">'
                    f'Avg IV {_mc_sym}{_dcf_val:,.0f} vs {_mc_sym}{_mc_S0:,.0f}<br>'
                    f'Margin: <b style="color:{_fund_col}">{_margin:+.1%}</b><br>'
                    f'<span style="font-size:0.65rem;color:#64748B">{_fund_detail}</span>'
                    f'</div></div>'
                )
            else:
                _dd_html += (
                    '<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.3);border-top:3px solid #475569;'
                    'border-radius:10px;padding:14px;text-align:center">'
                    '<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">① Fundamental (DCF)</div>'
                    '<div style="font-size:0.85rem;color:#64748B">Open this ticker in<br>Fundamental tab first</div>'
                    '<div style="font-size:0.7rem;color:#475569;margin-top:4px">Runs DCF · Graham · P/E</div>'
                    '</div>'
                )

            # 2. Monte Carlo
            _dd_html += (
                f'<div style="background:#0D1F33;border:1px solid {_mc_signal_colour}40;border-top:3px solid {_mc_signal_colour};'
                f'border-radius:10px;padding:14px;text-align:center">'
                f'<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">② Monte Carlo</div>'
                f'<div style="font-size:1.1rem;font-weight:800;color:{_mc_signal_colour}">{_mc_signal_label}</div>'
                f'<div style="font-size:0.75rem;color:#94A3B8;margin-top:4px">'
                f'P(profit): <b style="color:{_mc_signal_colour}">{_mc_p_profit:.0%}</b><br>'
            )
            if _mc_p_target is not None:
                _dd_html += f'P(≥ target): <b style="color:{_mc_signal_colour}">{_mc_p_target:.0%}</b>'
            _dd_html += '</div></div>'

            # 3. Technical
            if _tech_signal:
                _tc = "#22C55E" if "Strong" in str(_tech_signal) else "#F59E0B" if "Neutral" in str(_tech_signal) else "#EF4444"
                _dd_html += (
                    f'<div style="background:#0D1F33;border:1px solid {_tc}40;border-top:3px solid {_tc};'
                    f'border-radius:10px;padding:14px;text-align:center">'
                    f'<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">③ Technical</div>'
                    f'<div style="font-size:1.1rem;font-weight:800;color:{_tc}">{_tech_signal}</div>'
                    f'<div style="font-size:0.75rem;color:#94A3B8;margin-top:4px">From screener</div></div>'
                )
            else:
                _dd_html += (
                    '<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.3);border-top:3px solid #475569;'
                    'border-radius:10px;padding:14px;text-align:center">'
                    '<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">③ Technical</div>'
                    '<div style="font-size:0.85rem;color:#64748B">Check Technical<br>screen for signals</div></div>'
                )

            # 4. Overall Decision
            _signals_present = sum([
                1 if _dcf_val and _dcf_val > _mc_S0 * 1.1 else 0,
                1 if _mc_signal_label == "Favourable" else 0,
                1 if _tech_signal and "Strong" in str(_tech_signal) else 0,
            ])
            if _signals_present == 3:
                _dec_col, _dec_label, _dec_icon = "#22C55E", "HIGH CONVICTION BUY", "🟢"
            elif _signals_present == 2:
                _dec_col, _dec_label, _dec_icon = "#F59E0B", "CAUTIOUS BUY", "🟡"
            elif _signals_present == 1:
                _dec_col, _dec_label, _dec_icon = "#F59E0B", "WATCH / WAIT", "⚠️"
            else:
                _dec_col, _dec_label, _dec_icon = "#EF4444", "AVOID / REVIEW", "🔴"

            _dd_html += (
                f'<div style="background:linear-gradient(135deg,#0D2137,#0A1628);'
                f'border:2px solid {_dec_col};border-radius:10px;padding:14px;text-align:center">'
                f'<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;margin-bottom:6px">④ Decision</div>'
                f'<div style="font-size:1rem;font-weight:900;color:{_dec_col}">{_dec_icon} {_dec_label}</div>'
                f'<div style="font-size:0.72rem;color:#94A3B8;margin-top:4px">'
                f'{_signals_present}/3 signals aligned</div></div>'
            )
            _dd_html += '</div>'
            st.markdown(_dd_html, unsafe_allow_html=True)

            # Signal breakdown
            st.markdown(
                '<div style="background:rgba(13,31,53,0.7);border-left:3px solid #F59E0B;'
                'border-radius:8px;padding:14px 18px;font-size:0.85rem;color:#94A3B8;line-height:1.8">'
                '<b style="color:#F1F5F9">How to read this:</b> '
                'Signal ① checks whether the business is priced below intrinsic value (margin of safety). '
                'Signal ② checks whether the statistical distribution of future prices favours upside. '
                'Signal ③ checks whether the technical trend confirms momentum. '
                'All three aligning = highest conviction setup. '
                '<b style="color:#F59E0B">This dashboard does not constitute financial advice.</b>'
                '</div>',
                unsafe_allow_html=True
            )

            with st.expander("📚 Model Guide — How to read this simulator (start here if you're new)", expanded=False):
                st.markdown(f"""
## 🎲 What is Monte Carlo Simulation?

Imagine flipping a coin 10,000 times. You can't predict each individual flip, but you can say with confidence that roughly half will be heads. Monte Carlo simulation applies this same logic to stock prices.

Instead of predicting **one** future price, it simulates **{_mc_n_sims:,} different scenarios** — each representing a possible path the stock price could take over the next {_mc_horizon}. Some paths go up, some go down, most land somewhere in the middle. The result is a **probability map** of the future.

> **In plain English:** The fan chart shows the range of outcomes the maths considers plausible. The green middle band is where the price is most likely to end up. The red outer bands show the extreme scenarios.

---

## 📊 How to Read the Fan Chart

The coloured bands on the fan chart represent probability ranges:

| Band | Colour | Meaning |
|------|--------|---------|
| 25–75% | Green | The **most likely** 50% of outcomes — if you had to bet, this is where |
| 10–90% | Amber | 80% of simulations landed here |
| 5–95% | Yellow | 90% of simulations — only 1 in 10 paths went outside this zone |
| 1–99% | Red edge | Almost all simulations — the rarest outcomes live in this thin outer band |

The **gold median line** is the middle path — exactly half the simulations ended higher, half lower. Think of it as the "best guess" path (though not a prediction).

---

## 📈 How to Read the Numbers

| Term | What it means in plain English |
|------|-------------------------------|
| **Mean outcome** | The average price across all {_mc_n_sims:,} simulations |
| **Median outcome** | The middle outcome — half above, half below. Often more useful than mean |
| **Expected CAGR** | The annualised return the maths implies, based on past behaviour |
| **P(profit)** | Probability the stock is higher than today at the end of the period |
| **P(≥ target)** | If you set a target price, this is the chance of hitting it |
| **VaR 95%** | "Value at Risk" — in a bad month (worst 5% of scenarios), how much could you lose? |
| **Expected Shortfall** | If you land in that worst 5%, what's the *average* loss? Worse than VaR |
| **Max Sim Drawdown** | The deepest peak-to-trough fall seen across simulations — the gut-punch test |

---

## 🔬 Which Simulation Model Should I Use?

Fintiq offers three models. Each has different assumptions about how markets behave:

### 1. Standard GBM (Geometric Brownian Motion)
**Best for:** Blue-chip stable companies (AAPL, Unilever, HSBA)

This is the classic textbook model — the same one used in the famous Black-Scholes options pricing formula. It assumes the stock wanders randomly with a steady average drift (like walking randomly on a slope).

- ✅ Simple and well-understood
- ✅ Works well for stable, mature businesses
- ❌ Assumes volatility is constant (it isn't — markets get more volatile in crashes)
- ❌ Doesn't model sudden crashes or spikes

**Current parameters (this run):** Annual drift μ = {_mc_mu:+.1%} · Volatility σ = {_mc_sigma:.1%}

---

### 2. Jump Diffusion (Merton Model)
**Best for:** Volatile stocks, small-caps, biotech, commodity-sensitive companies

Developed by Nobel Prize winner Robert Merton. Takes Standard GBM and adds **random jump events** — sudden large moves (crashes and spikes) that happen occasionally, modelled using a Poisson process.

- ✅ Captures the "fat tails" of real markets (crashes happen more than normal distribution predicts)
- ✅ Better for stocks prone to sudden news-driven moves
- ❌ Jump parameters are estimated from history — past jumps may not predict future ones
{'- 📌 **Jump parameters detected from history:** ' + f'{_mc_lam_j:.1f} jumps/year · avg jump size {_mc_mu_j:+.2%} · jump volatility {_mc_sig_j:.2%}' if _mc_model == 'Jump Diffusion' else ''}

---

### 3. EWMA Volatility (RiskMetrics)
**Best for:** Periods of market stress, earnings season, highly volatile stocks

"EWMA" stands for Exponentially Weighted Moving Average. Rather than assuming volatility is constant, it updates the volatility estimate in real-time, giving **more weight to recent price swings** (λ=0.94, the RiskMetrics standard used by major banks).

- ✅ Captures **volatility clustering** — calm periods stay calm, stormy periods stay stormy
- ✅ More realistic than Standard GBM in stressed markets
- ❌ Still doesn't model structural breaks or regime changes

**Current EWMA vol:** {_mc_sigma:.1%} annualised (vs historical avg {_mc_hist_vol:.1%})

---

## 🧠 How to Use This Tool in Practice

**Step 1 — Value the business first (Fundamental tab)**
Run the DCF valuation in the Fundamental screen. Your DCF intrinsic value becomes your **target price** here.

**Step 2 — Enter the target and stop-loss**
Put your DCF target in the "Target price" box. Set a stop-loss at the price where you'd accept the trade hasn't worked. The simulator will tell you the probability of hitting each.

**Step 3 — Run all 3 models**
Start with Standard GBM. Then run Jump Diffusion. If the probabilities are broadly similar, confidence is higher. Wide divergence means the result is model-sensitive — treat with caution.

**Step 4 — Read the Decision Dashboard**
The dashboard aggregates your DCF signal, Monte Carlo signal and Technical signal. Look for all 3 aligning before acting.

---

## ⚠️ Limitations — What This Tool Cannot Do

- **Cannot predict the future.** No model can. This gives you probabilities, not certainties.
- **Garbage in, garbage out.** If the historical period was a bull market, drift μ will be optimistic. Always sanity-check the μ parameter.
- **Does not model:** earnings surprises, management changes, regulatory shocks, geopolitical events, liquidity crises.
- **Past volatility ≠ future volatility.** A stock that was calm for 3 years may become highly volatile.
- **This is not financial advice.** Use alongside your own research, not as a substitute for it.

---
*Model used: **{_mc_model}** · Simulations: **{_mc_n_sims:,}** · Horizon: **{_mc_horizon}** · Return source: **{_mc_ret_src}***
""")

    elif not _mc_ticker and not _run_mc:
        st.info("Enter a ticker above and click **Run Simulation** to begin.")


# ═══════════════════════════════════════════════════════════════
# TAB 4 — PAIRS DASHBOARD
# ═══════════════════════════════════════════════════════════════

with tab4:
    st.markdown('<div class="section-header">⚖️ Pairs Trading Dashboard — Strategy 3</div>',
                unsafe_allow_html=True)
    st.caption("Market-neutral strategy. Long the underperformer, short the outperformer. "
               "Profit from mean-reversion regardless of market direction.")

    # ── Pairs Watchlist (collapsible) ───────────────────────────
    _pwl = st.session_state.get("fintiq_pairs_watchlist", [])
    with st.expander(f"⭐ My Pairs Watchlist  ({len(_pwl)} saved)", expanded=False):
        if not _pwl:
            st.info("No saved pairs yet. Analyse a pair below, then click **⭐ Save to My Watchlist**.")
        else:
            for _pi, _pe in enumerate(_pwl):
                _pw1, _pw2, _pw3, _pw4, _pw5 = st.columns([2, 1, 1, 2, 1])
                _pw1.markdown(
                    f'<div style="background:rgba(13,31,53,0.8);border:1px solid #1E3A5F;'
                    f'border-radius:8px;padding:8px 12px;font-weight:700;color:#F1F5F9">'
                    f'{_pe.get("name","")}</div>', unsafe_allow_html=True)
                _pw2.markdown(
                    f'<div style="background:rgba(13,31,53,0.6);border:1px solid #1E3A5F;'
                    f'border-radius:8px;padding:8px 12px;text-align:center;color:#7DD3FC;font-weight:700">'
                    f'{_pe.get("ticker_a","")}</div>', unsafe_allow_html=True)
                _pw3.markdown(
                    f'<div style="background:rgba(13,31,53,0.6);border:1px solid #1E3A5F;'
                    f'border-radius:8px;padding:8px 12px;text-align:center;color:#F59E0B;font-weight:700">'
                    f'{_pe.get("ticker_b","")}</div>', unsafe_allow_html=True)
                _pw4.markdown(
                    f'<div style="background:rgba(13,31,53,0.4);border-radius:8px;padding:8px 12px;'
                    f'font-size:0.75rem;color:#64748B">'
                    f'Added {_pe.get("added","")}</div>', unsafe_allow_html=True)
                with _pw5:
                    _load_col, _del_col = st.columns(2)
                    if _load_col.button("📂", key=f"pwl_load_{_pi}",
                                        help=f"Load {_pe.get('ticker_a')} / {_pe.get('ticker_b')}"):
                        st.session_state["pwl_load_a"] = _pe.get("ticker_a","")
                        st.session_state["pwl_load_b"] = _pe.get("ticker_b","")
                        st.session_state["pwl_load_name"] = _pe.get("name","")
                        st.rerun()
                    if _del_col.button("🗑️", key=f"pwl_del_{_pi}",
                                       help="Remove from watchlist"):
                        _pwl.pop(_pi)
                        st.session_state["fintiq_pairs_watchlist"] = _pwl
                        _pwl_save(_pwl)
                        st.rerun()

    # ── Controls row ────────────────────────────────────────────
    t4c1, t4c2, t4c3, t4c4 = st.columns([2, 1, 1, 1])
    with t4c1:
        pair_mode = st.radio("Pair source:", ["Preset pair", "Custom pair", "My Watchlist pair"], horizontal=True)
    with t4c2:
        lookback = st.slider("Lookback (days)", 20, 120, 60)
    with t4c3:
        entry_z = st.slider("Entry Z-Score", 1.0, 3.0, 2.0, 0.25)
    with t4c4:
        hist_period = st.selectbox("History:", ["6mo", "1y", "2y"], index=1)

    if pair_mode == "Preset pair":
        pair_name = st.selectbox("Select pair:", list(PRESET_PAIRS.keys()))
        ticker_a, ticker_b = PRESET_PAIRS[pair_name]
        pair_label = pair_name

    elif pair_mode == "My Watchlist pair":
        _pwl_now = st.session_state.get("fintiq_pairs_watchlist", [])
        if not _pwl_now:
            st.info("Your pairs watchlist is empty. Save a pair first using the Preset or Custom builder below.")
            st.stop()
        _pwl_options = {f"{p['name']}  ({p['ticker_a']} / {p['ticker_b']})": p for p in _pwl_now}
        _pwl_sel_key = st.selectbox("Select from My Watchlist:", list(_pwl_options.keys()))
        _pwl_entry   = _pwl_options[_pwl_sel_key]
        ticker_a     = _pwl_entry["ticker_a"]
        ticker_b     = _pwl_entry["ticker_b"]
        pair_label   = _pwl_entry["name"]

    else:
        st.markdown('<div class="section-header">✏️ Custom Pair Builder</div>',
                    unsafe_allow_html=True)
        # Pre-fill from watchlist load button if pressed
        _pre_a = st.session_state.pop("pwl_load_a", "")
        _pre_b = st.session_state.pop("pwl_load_b", "")
        _pre_n = st.session_state.pop("pwl_load_name", "Custom Pair")
        cc1, cc2, cc3 = st.columns(3)
        with cc1: ticker_a  = st.text_input("Ticker A:", value=_pre_a).strip().upper()
        with cc2: ticker_b  = st.text_input("Ticker B:", value=_pre_b).strip().upper()
        with cc3: pair_desc = st.text_input("Description:", value=_pre_n)
        pair_label = f"{ticker_a} / {ticker_b} ({pair_desc})" if ticker_a and ticker_b else ""
        if not ticker_a or not ticker_b:
            st.info("Enter both tickers above to analyse the pair.")
            st.stop()

    # ── Save to Pairs Watchlist ──────────────────────────────────
    _sv1, _sv2 = st.columns([3, 1])
    with _sv1:
        _save_name = st.text_input("Watchlist name:", value=pair_label,
                                    placeholder="e.g. My Bank Pair", key="pwl_save_name",
                                    label_visibility="collapsed")
    with _sv2:
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        if st.button("⭐ Save to My Watchlist", use_container_width=True):
            _pwl_cur = st.session_state.get("fintiq_pairs_watchlist", [])
            # Avoid exact duplicates (same ticker_a + ticker_b)
            _exists = any(p["ticker_a"] == ticker_a and p["ticker_b"] == ticker_b for p in _pwl_cur)
            if _exists:
                st.info("This pair is already in your watchlist.")
            else:
                _pwl_cur.append({
                    "ticker_a": ticker_a, "ticker_b": ticker_b, "name": _save_name or pair_label
                })
                st.session_state["fintiq_pairs_watchlist"] = _pwl_cur
                st.success(f"✅ Saved: {_save_name or pair_label}")

# ═══════════════════════════════════════════════════════════════
# TAB 5 — TRADING JOURNAL  (preserved — full code on Desktop copy)
# ═══════════════════════════════════════════════════════════════

with tab5:
    st.markdown('<div class="section-header">📒 Trading Journal & P&L Account</div>',
                unsafe_allow_html=True)

    _juser = st.session_state.get("fintiq_user", {})
    if not _juser:
        st.info("Please log in to use the Trading Journal.")
    else:
        # ── Add new trade ────────────────────────────────────────
        with st.expander("➕ Log New Trade", expanded=False):
            jc1, jc2, jc3 = st.columns(3)
            with jc1:
                j_date   = st.date_input("Date", key="j_date")
                j_ticker = st.text_input("Ticker", key="j_ticker").strip().upper()
                j_co     = st.text_input("Company name", key="j_co")
            with jc2:
                j_dir    = st.selectbox("Direction", ["Long", "Short"], key="j_dir")
                j_strat  = st.selectbox("Strategy", ["Value", "Growth", "Momentum", "Pairs", "Other"], key="j_strat")
                j_status = st.selectbox("Status", ["Open", "Closed"], key="j_status")
            with jc3:
                j_entry  = st.number_input("Entry price", min_value=0.0, step=0.01, key="j_entry")
                j_exit   = st.number_input("Exit price (0 if open)", min_value=0.0, step=0.01, key="j_exit")
                j_shares = st.number_input("Shares / units", min_value=0.0, step=1.0, key="j_shares")
                j_ccy    = st.selectbox("Currency", ["GBP","USD","EUR","JPY","CHF"], key="j_ccy")
            j_notes = st.text_area("Notes / thesis", key="j_notes")
            if st.button("💾 Save Trade", use_container_width=True):
                db_add_trade(str(j_date), j_ticker, j_co, j_dir, j_strat,
                             j_entry, j_exit or None, j_shares, j_ccy, j_status, j_notes)
                st.success(f"Trade logged: {j_ticker}")
                st.rerun()

        # ── Trade log ───────────────────────────────────────────
        _jtrades = db_get_trades()
        if _jtrades.empty:
            st.info("No trades logged yet. Add your first position above.")
        else:
            st.markdown(f"**{len(_jtrades)} trades logged**")

            # P&L calc
            _jtrades["P&L"] = _jtrades.apply(
                lambda r: (r["exit_price"] - r["entry_price"]) * r["shares"]
                if r["exit_price"] and r["exit_price"] > 0 else None, axis=1)

            # Summary stats
            _closed = _jtrades[_jtrades["status"] == "Closed"].copy()
            if not _closed.empty and _closed["P&L"].notna().any():
                _total_pnl  = _closed["P&L"].sum()
                _win_rate   = (_closed["P&L"] > 0).mean() * 100
                _best       = _closed["P&L"].max()
                _worst      = _closed["P&L"].min()
                sc1, sc2, sc3, sc4 = st.columns(4)
                _pnl_col = "#22C55E" if _total_pnl >= 0 else "#EF4444"
                sc1.markdown(f'<div style="text-align:center"><div style="font-size:0.7rem;color:#64748B">TOTAL P&L</div>'
                             f'<div style="font-size:1.4rem;font-weight:800;color:{_pnl_col}">'
                             f'{"+" if _total_pnl>=0 else ""}{_total_pnl:,.2f}</div></div>', unsafe_allow_html=True)
                sc2.markdown(f'<div style="text-align:center"><div style="font-size:0.7rem;color:#64748B">WIN RATE</div>'
                             f'<div style="font-size:1.4rem;font-weight:800;color:#F59E0B">{_win_rate:.0f}%</div></div>', unsafe_allow_html=True)
                sc3.markdown(f'<div style="text-align:center"><div style="font-size:0.7rem;color:#64748B">BEST TRADE</div>'
                             f'<div style="font-size:1.4rem;font-weight:800;color:#22C55E">+{_best:,.2f}</div></div>', unsafe_allow_html=True)
                sc4.markdown(f'<div style="text-align:center"><div style="font-size:0.7rem;color:#64748B">WORST TRADE</div>'
                             f'<div style="font-size:1.4rem;font-weight:800;color:#EF4444">{_worst:,.2f}</div></div>', unsafe_allow_html=True)
                st.markdown("---")

            # Trade table
            _display_cols = ["date","ticker","company","direction","strategy",
                             "entry_price","exit_price","shares","currency","status","P&L","notes"]
            _show_cols = [c for c in _display_cols if c in _jtrades.columns or c == "P&L"]
            st.dataframe(_jtrades[_show_cols].rename(columns={
                "date":"Date","ticker":"Ticker","company":"Company",
                "direction":"Dir","strategy":"Strategy","entry_price":"Entry",
                "exit_price":"Exit","shares":"Shares","currency":"CCY",
                "status":"Status","P&L":"P&L","notes":"Notes"
            }), use_container_width=True, hide_index=True)

            # Delete
            _del_id = st.number_input("Delete trade by ID:", min_value=0, step=1, key="j_del_id")
            if st.button("🗑 Delete trade", key="j_del_btn") and _del_id:
                db_delete_trade(int(_del_id))
                st.success(f"Trade #{_del_id} deleted.")
                st.rerun()

# ═══════════════════════════════════════════════════════════════
# TAB OPT — PORTFOLIO OPTIMIZER (MPT)
# ═══════════════════════════════════════════════════════════════

with tab_opt:
    st.markdown('<div class="section-header">📐 Portfolio Optimizer — Modern Portfolio Theory</div>',
                unsafe_allow_html=True)
    st.caption("Mean-Variance Optimization · Efficient Frontier · Value at Risk · Sharpe Maximization")

    _opt_user = st.session_state.get("fintiq_user", {})
    _opt_pro  = _opt_user.get("is_pro", False)

    # If is_pro not yet loaded (user went straight to optimizer without a search),
    # fetch from Supabase now so Pro users aren't blocked
    if not _opt_pro and _opt_user.get("id"):
        _opt_prof = st.session_state.get("fintiq_profile") or _get_profile(_opt_user["id"])
        if _opt_prof.get("is_pro"):
            st.session_state["fintiq_user"]["is_pro"] = True
            st.session_state["fintiq_profile"] = _opt_prof
            _opt_pro = True

    # ── Pro gate with LIVE demo portfolio ────────────────────────
    if not _opt_pro:
        import numpy as _np_demo
        from scipy.optimize import minimize as _min_demo

        # ── Hero pitch ───────────────────────────────────────────
        st.markdown("""
<div style="background:linear-gradient(135deg,#0D2137,#0A1628);border:1px solid rgba(245,158,11,0.35);
border-radius:14px;padding:24px 28px;margin-bottom:20px">
  <div style="font-size:1.6rem;font-weight:900;color:#F59E0B;margin-bottom:6px">
    📐 Portfolio Optimiser — powered by Nobel Prize-winning finance</div>
  <div style="color:#CBD5E1;font-size:0.95rem;line-height:1.7">
    Stop guessing how much to put in each stock. The Fintiq Optimiser uses
    <b style="color:#F1F5F9">Modern Portfolio Theory</b> to find the exact allocation that
    maximises your return per unit of risk — the same mathematics used by hedge funds and
    institutional asset managers every day.
  </div>
  <div style="display:flex;gap:24px;margin-top:16px;flex-wrap:wrap">
    <div style="color:#94A3B8;font-size:0.85rem">✅ Efficient Frontier chart</div>
    <div style="color:#94A3B8;font-size:0.85rem">✅ 6 optimisation objectives</div>
    <div style="color:#94A3B8;font-size:0.85rem">✅ Sharpe · Sortino · VaR · Max Drawdown</div>
    <div style="color:#94A3B8;font-size:0.85rem">✅ Correlation matrix</div>
    <div style="color:#94A3B8;font-size:0.85rem">✅ Live weight drift monitor</div>
    <div style="color:#94A3B8;font-size:0.85rem">✅ AI portfolio commentary</div>
  </div>
</div>""", unsafe_allow_html=True)

        st.markdown(
            '<div style="font-size:0.85rem;color:#F59E0B;font-weight:700;'
            'text-transform:uppercase;letter-spacing:1px;margin-bottom:12px">'
            '🔴 Live Demo — Real Data · 10 Blue-Chip Stocks · Results blurred for Pro subscribers</div>',
            unsafe_allow_html=True)

        # ── Run REAL optimisation on demo portfolio ───────────────
        _DEMO_TICKERS = ("AAPL", "MSFT", "NVDA", "GOOGL", "HSBA.L",
                         "SHEL.L", "BP.L", "ULVR.L", "GSK.L", "RIO.L")
        _DEMO_NAMES   = {"AAPL":"Apple","MSFT":"Microsoft","NVDA":"Nvidia",
                         "GOOGL":"Alphabet","HSBA.L":"HSBC","SHEL.L":"Shell",
                         "BP.L":"BP","ULVR.L":"Unilever","GSK.L":"GSK","RIO.L":"Rio Tinto"}

        @st.cache_data(ttl=86400, show_spinner=False)
        def _demo_opt_data():
            try:
                import yfinance as _yfd
                _px = _yfd.download(list(_DEMO_TICKERS), period="2y",
                                    auto_adjust=True, progress=False)["Close"]
                if isinstance(_px, pd.Series):
                    _px = _px.to_frame()
                _px = _px.ffill().dropna()
                _ret = _px.pct_change().dropna()
                _mr  = _ret.mean()
                _cv  = _ret.cov()
                _n   = len(_mr)
                def _ps(w):
                    r = float(_np_demo.dot(w, _mr)*252)
                    v = float(_np_demo.sqrt(w @ _cv.values @ w * 252))
                    s = (r - 0.0425) / v if v > 0 else 0
                    return r, v, s
                cons = ({"type":"eq","fun":lambda w: _np_demo.sum(w)-1},)
                bnds = tuple((0.02,0.40) for _ in range(_n))
                w0   = _np_demo.array([1/_n]*_n)
                res  = _min_demo(lambda w: -_ps(w)[2], w0, method="SLSQP",
                                 bounds=bnds, constraints=cons)
                _ow  = res.x if res.success else w0
                _or, _ov, _os = _ps(_ow)
                # Simulate portfolios for frontier
                _np_demo.random.seed(42)
                _svols,_srets,_sshrp = [],[],[]
                for _ in range(1500):
                    _w = _np_demo.random.dirichlet(_np_demo.ones(_n))
                    _r,_v,_s = _ps(_w); _svols.append(_v); _srets.append(_r); _sshrp.append(_s)
                # Efficient frontier line
                _min_r = float(_mr.min()*252); _max_r = float(_mr.max()*252)
                _ef_v,_ef_r = [],[]
                for _t in _np_demo.linspace(_min_r,_max_r,200):
                    _c = ({"type":"eq","fun":lambda w: _np_demo.sum(w)-1},
                          {"type":"eq","fun":lambda w,t=_t: _ps(w)[0]-t})
                    _b = tuple((0.0,1.0) for _ in range(_n))
                    _r2 = _min_demo(lambda w: _ps(w)[1], w0, method="SLSQP",
                                    bounds=_b, constraints=_c)
                    if _r2.success:
                        _ef_v.append(_ps(_r2.x)[1]); _ef_r.append(_t)
                return dict(weights=_ow, ret=_or, vol=_ov, sharpe=_os,
                            tickers=list(_mr.index), svols=_svols, srets=_srets,
                            sshrp=_sshrp, ef_v=_ef_v, ef_r=_ef_r,
                            corr=_ret[list(_mr.index)].corr().values.tolist(),
                            var95=float(-_np_demo.percentile(_ret.dot(_ow), 5)))
            except Exception as _e:
                return None

        with st.spinner("Loading live demo data (real market prices)…"):
            _dd = _demo_opt_data()

        if _dd is None:
            st.info("Demo data temporarily unavailable. Subscribe to run optimisations on your own portfolio.")
        else:
            _d_tickers = _dd["tickers"]
            _d_weights = _dd["weights"]
            _d_ret, _d_vol, _d_sharpe = _dd["ret"], _dd["vol"], _dd["sharpe"]

            # ── Efficient Frontier — FULL, unblurred ─────────────
            st.markdown("**Efficient Frontier** — 1,500 random portfolios + optimal allocation (real data, 2 years)")
            _df2 = go.Figure()
            _df2.add_trace(go.Scatter(
                x=_dd["svols"], y=_dd["srets"], mode="markers",
                marker=dict(color=_dd["sshrp"], colorscale="Viridis", size=4, opacity=0.45,
                            colorbar=dict(title="Sharpe", thickness=10,
                                          tickfont=dict(color="#64748B"), len=0.7)),
                name="Random portfolios",
                hovertemplate="Vol: %{x:.1%}<br>Return: %{y:.1%}<extra></extra>"
            ))
            if _dd["ef_v"]:
                _df2.add_trace(go.Scatter(
                    x=_dd["ef_v"], y=_dd["ef_r"], mode="lines",
                    line=dict(color="#F59E0B", width=2.5), name="Efficient Frontier"))
            _df2.add_trace(go.Scatter(
                x=[_d_vol], y=[_d_ret], mode="markers",
                marker=dict(color="#F59E0B", size=18, symbol="star",
                            line=dict(color="#fff", width=1.5)),
                name="Optimal Portfolio",
                hovertemplate=f"<b>Optimal</b><br>Vol:{_d_vol:.1%} Return:{_d_ret:.1%} Sharpe:{_d_sharpe:.2f}<extra></extra>"
            ))
            _df2.add_annotation(x=_d_vol, y=_d_ret, text="⭐ Optimal",
                                xanchor="left", yanchor="bottom", xshift=8, yshift=4,
                                font=dict(color="#F59E0B", size=11), showarrow=False,
                                bgcolor="rgba(13,31,51,0.8)", borderpad=3)
            _df2.update_layout(
                height=380, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=40, r=30, t=10, b=50),
                legend=dict(font=dict(color="#94A3B8", size=10), bgcolor="rgba(13,31,51,0.85)",
                            bordercolor="rgba(100,116,139,0.3)", borderwidth=1,
                            orientation="h", x=0, y=1.08),
                xaxis=dict(title="Annual Volatility (Risk)", tickformat=".0%",
                           gridcolor="rgba(100,116,139,0.15)", tickfont=dict(color="#64748B")),
                yaxis=dict(title="Expected Annual Return", tickformat=".0%",
                           gridcolor="rgba(100,116,139,0.15)", tickfont=dict(color="#64748B")),
            )
            st.plotly_chart(_df2, use_container_width=True, config={"displayModeBar": False})

            # ── KPIs — blurred ───────────────────────────────────
            st.markdown("**Optimisation Results** — upgrade to see your portfolio's numbers")
            _kc1, _kc2, _kc3, _kc4 = st.columns(4)
            for _col, _lbl, _val, _clr in [
                (_kc1, "Expected Return", f"{_d_ret:.1%}", "#22C55E"),
                (_kc2, "Portfolio Volatility", f"{_d_vol:.1%}", "#F59E0B"),
                (_kc3, "Sharpe Ratio", f"{_d_sharpe:.2f}", "#22C55E"),
                (_kc4, "Daily VaR 95%", f"{_dd['var95']:.2%}", "#EF4444"),
            ]:
                _col.markdown(
                    f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.25);'
                    f'border-radius:10px;padding:12px;text-align:center;filter:blur(4px);'
                    f'user-select:none;pointer-events:none">'
                    f'<div style="font-size:0.6rem;color:#64748B;text-transform:uppercase;'
                    f'margin-bottom:4px">{_lbl}</div>'
                    f'<div style="font-size:1.2rem;font-weight:800;color:{_clr}">{_val}</div>'
                    f'</div>', unsafe_allow_html=True)

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            # ── Optimal Allocation — blurred ──────────────────────
            st.markdown("**Optimal Allocation** — exact weights hidden until you subscribe")
            _al1, _al2 = st.columns([1, 1])
            with _al1:
                _alloc_html = '<div style="filter:blur(5px);user-select:none;pointer-events:none">'
                for _tk, _wt in sorted(zip(_d_tickers, _d_weights), key=lambda x: -x[1]):
                    _nm = _DEMO_NAMES.get(_tk, _tk)
                    _alloc_html += (
                        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">'
                        f'<div style="width:72px;font-size:0.78rem;color:#F1F5F9;font-weight:700">{_tk}</div>'
                        f'<div style="flex:1;background:#1E3A5F;border-radius:4px;height:8px">'
                        f'<div style="background:#F59E0B;width:{_wt*100:.0f}%;height:8px;border-radius:4px"></div></div>'
                        f'<div style="width:38px;text-align:right;font-size:0.78rem;color:#F59E0B;font-weight:700">{_wt:.1%}</div>'
                        f'</div>'
                    )
                _alloc_html += '</div>'
                st.markdown(_alloc_html, unsafe_allow_html=True)
            with _al2:
                # Correlation heatmap — partially blurred (go already imported at top of file)
                _demo_corr = pd.DataFrame(_dd["corr"], index=_d_tickers, columns=_d_tickers)
                _ch = go.Figure(go.Heatmap(
                    z=_demo_corr.values, x=_d_tickers, y=_d_tickers,
                    colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
                    text=[[f"{v:.2f}" for v in row] for row in _demo_corr.values],
                    texttemplate="%{text}", textfont=dict(size=8),
                    showscale=False,
                ))
                _ch.update_layout(
                    height=260, margin=dict(l=0,r=0,t=10,b=0),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(tickfont=dict(color="#94A3B8", size=8)),
                    yaxis=dict(tickfont=dict(color="#94A3B8", size=8)),
                )
                st.markdown('<div style="filter:blur(3px);pointer-events:none">', unsafe_allow_html=True)
                st.plotly_chart(_ch, use_container_width=True, config={"displayModeBar": False})
                st.markdown('</div>', unsafe_allow_html=True)

        # ── Unlock CTA ────────────────────────────────────────────
        st.markdown("---")
        st.markdown(
            '<div style="text-align:center;padding:8px 0 4px 0;font-size:0.9rem;color:#94A3B8">'
            'The chart above is <b style="color:#F1F5F9">real data</b> from your demo portfolio. '
            'Subscribe to run this on <b style="color:#F1F5F9">your own stocks</b> — any ticker, any market.'
            '</div>', unsafe_allow_html=True)

        _u_email = _opt_user.get("email", "")
        _, _cta_col, _ = st.columns([1, 2, 1])
        with _cta_col:
            if _opt_user:
                if st.button("🚀 Upgrade to Pro — Unlock Portfolio Optimiser",
                             use_container_width=True, type="primary", key="opt_upgrade_btn"):
                    _co_url = _create_checkout("monthly", _u_email, _opt_user.get("id", ""))
                    if _co_url:
                        st.session_state["_checkout_ready"] = _co_url
                if "_checkout_ready" in st.session_state:
                    st.link_button("🔒 Proceed to Payment →",
                                   st.session_state["_checkout_ready"],
                                   use_container_width=True, type="primary")
                st.caption("£10/month · Cancel anytime · Unlock all 8 screens")
            else:
                st.link_button("🔑 Create free account / Log in →", "?action=login",
                               use_container_width=True, type="primary")

        st.stop()

    # ═══════════════════════════════════════════════════════════════
    # PRO USERS — FULL OPTIMIZER
    # ═══════════════════════════════════════════════════════════════

    import numpy as _np
    from scipy.optimize import minimize as _minimize

    # ── Helper functions ────────────────────────────────────────

    @st.cache_data(ttl=3600)
    def _opt_fetch_prices(tickers: tuple, period: str = "2y") -> pd.DataFrame:
        """Download adjusted close prices for a list of tickers."""
        try:
            raw = yf.download(list(tickers), period=period, interval="1d",
                              progress=False, auto_adjust=True)
            if isinstance(raw.columns, pd.MultiIndex):
                prices = raw["Close"].dropna(how="all")
            else:
                prices = raw[["Close"]].rename(columns={"Close": tickers[0]})
            prices = prices.dropna(how="all")
            return prices
        except Exception as e:
            return pd.DataFrame()

    def _portfolio_stats(weights, mean_returns, cov_matrix, rf=0.04):
        """Return (annual_return, annual_vol, sharpe)."""
        weights = _np.array(weights)
        ret  = float(_np.dot(weights, mean_returns) * 252)
        vol  = float(_np.sqrt(_np.dot(weights.T, _np.dot(cov_matrix * 252, weights))))
        shrp = (ret - rf) / vol if vol > 0 else 0.0
        return ret, vol, shrp

    def _max_sharpe(mean_returns, cov_matrix, rf=0.04, n=None):
        """Find weights that maximise Sharpe Ratio."""
        if n is None:
            n = len(mean_returns)
        constraints = ({"type": "eq", "fun": lambda w: _np.sum(w) - 1},)
        bounds = tuple((0.02, 0.40) for _ in range(n))
        init   = _np.array([1/n]*n)
        result = _minimize(
            lambda w: -_portfolio_stats(w, mean_returns, cov_matrix, rf)[2],
            init, method="SLSQP", bounds=bounds, constraints=constraints,
            options={"maxiter": 1000, "ftol": 1e-9}
        )
        return result.x if result.success else init

    def _min_volatility(mean_returns, cov_matrix, n=None):
        """Find weights that minimise volatility."""
        if n is None:
            n = len(mean_returns)
        constraints = ({"type": "eq", "fun": lambda w: _np.sum(w) - 1},)
        bounds = tuple((0.02, 0.40) for _ in range(n))
        init   = _np.array([1/n]*n)
        result = _minimize(
            lambda w: _portfolio_stats(w, mean_returns, cov_matrix)[1],
            init, method="SLSQP", bounds=bounds, constraints=constraints,
            options={"maxiter": 1000, "ftol": 1e-9}
        )
        return result.x if result.success else init

    def _max_return(mean_returns, cov_matrix, n=None):
        """Maximise expected return (concentrated, high-conviction)."""
        if n is None:
            n = len(mean_returns)
        constraints = ({"type": "eq", "fun": lambda w: _np.sum(w) - 1},)
        bounds = tuple((0.02, 0.60) for _ in range(n))  # allow higher concentration
        init   = _np.array([1/n]*n)
        result = _minimize(
            lambda w: -_portfolio_stats(w, mean_returns, cov_matrix)[0],
            init, method="SLSQP", bounds=bounds, constraints=constraints,
            options={"maxiter": 1000, "ftol": 1e-9}
        )
        return result.x if result.success else init

    def _risk_parity(mean_returns, cov_matrix, n=None):
        """Risk Parity — equal risk contribution from each asset (Ray Dalio All Weather style)."""
        if n is None:
            n = len(mean_returns)
        cov = cov_matrix.values
        def _risk_contributions(w):
            port_var = float(w @ cov @ w)
            marginal = cov @ w
            return w * marginal / port_var
        def _obj(w):
            rc = _risk_contributions(w)
            target = 1.0 / n
            return float(_np.sum((rc - target)**2))
        constraints = ({"type": "eq", "fun": lambda w: _np.sum(w) - 1},)
        bounds = tuple((0.01, 0.60) for _ in range(n))
        init   = _np.array([1/n]*n)
        result = _minimize(_obj, init, method="SLSQP", bounds=bounds,
                           constraints=constraints, options={"maxiter": 2000, "ftol": 1e-12})
        return result.x if result.success else init

    def _max_diversification(mean_returns, cov_matrix, n=None):
        """Maximum Diversification — maximises ratio of weighted avg vol to portfolio vol."""
        if n is None:
            n = len(mean_returns)
        cov = cov_matrix.values
        asset_vols = _np.sqrt(_np.diag(cov))
        def _obj(w):
            weighted_avg_vol = float(_np.dot(w, asset_vols))
            port_vol = float(_np.sqrt(w @ cov @ w))
            return -weighted_avg_vol / port_vol if port_vol > 0 else 0.0
        constraints = ({"type": "eq", "fun": lambda w: _np.sum(w) - 1},)
        bounds = tuple((0.02, 0.40) for _ in range(n))
        init   = _np.array([1/n]*n)
        result = _minimize(_obj, init, method="SLSQP", bounds=bounds,
                           constraints=constraints, options={"maxiter": 1000, "ftol": 1e-9})
        return result.x if result.success else init

    def _target_return_min_vol(mean_returns, cov_matrix, target_ret, n=None):
        """Minimum volatility portfolio that achieves a target annual return."""
        if n is None:
            n = len(mean_returns)
        constraints = (
            {"type": "eq", "fun": lambda w: _np.sum(w) - 1},
            {"type": "eq", "fun": lambda w: _portfolio_stats(w, mean_returns, cov_matrix)[0] - target_ret},
        )
        bounds = tuple((0.0, 0.60) for _ in range(n))
        init   = _np.array([1/n]*n)
        result = _minimize(
            lambda w: _portfolio_stats(w, mean_returns, cov_matrix)[1],
            init, method="SLSQP", bounds=bounds, constraints=constraints,
            options={"maxiter": 2000, "ftol": 1e-9}
        )
        return result.x if result.success else None

    def _efficient_frontier_points(mean_returns, cov_matrix, n_points=400):
        """Generate efficient frontier by sweeping target returns."""
        n = len(mean_returns)
        min_ret = float(mean_returns.min() * 252)
        max_ret = float(mean_returns.max() * 252)
        targets = _np.linspace(min_ret, max_ret, n_points)
        ef_vols, ef_rets = [], []
        for t in targets:
            constraints = (
                {"type": "eq", "fun": lambda w: _np.sum(w) - 1},
                {"type": "eq", "fun": lambda w, t=t: _portfolio_stats(w, mean_returns, cov_matrix)[0] - t},
            )
            bounds = tuple((0.0, 1.0) for _ in range(n))
            res = _minimize(
                lambda w: _portfolio_stats(w, mean_returns, cov_matrix)[1],
                _np.array([1/n]*n), method="SLSQP",
                bounds=bounds, constraints=constraints,
                options={"maxiter": 500, "ftol": 1e-8}
            )
            if res.success:
                ef_vols.append(_portfolio_stats(res.x, mean_returns, cov_matrix)[1])
                ef_rets.append(t)
        return ef_vols, ef_rets

    def _calc_var(daily_returns_series, weights, confidence=0.95):
        """Historical VaR at given confidence level (daily %)."""
        port_daily = daily_returns_series.dot(weights)
        return float(-_np.percentile(port_daily, (1 - confidence) * 100))

    # ── Ticker Input ─────────────────────────────────────────────
    st.markdown("#### 📋 Portfolio Input")

    _oi1, _oi2 = st.columns([3, 1])
    with _oi1:
        _opt_tickers_raw = st.text_input(
            "Enter tickers (comma-separated):",
            placeholder="e.g. AAPL, MSFT, HSBA.L, VOD.L, SHEL.L",
            key="opt_tickers_input",
            help="Use Yahoo Finance format — UK stocks need .L suffix (e.g. BARC.L)"
        )
    with _oi2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("📒 Import from Journal", use_container_width=True, key="opt_import_journal"):
            _jdf = db_get_trades()
            if not _jdf.empty:
                _open_tickers = _jdf[_jdf["status"] == "Open"]["ticker"].dropna().unique().tolist()
                if _open_tickers:
                    st.session_state["opt_journal_tickers"] = ", ".join(_open_tickers)
                    st.success(f"Imported {len(_open_tickers)} open positions from journal.")
                    st.rerun()
                else:
                    st.warning("No open positions found in journal.")
            else:
                st.warning("Journal is empty.")

    # Use journal-imported tickers if available
    if "opt_journal_tickers" in st.session_state and not _opt_tickers_raw:
        _opt_tickers_raw = st.session_state["opt_journal_tickers"]

    _oc_left, _oc_right = st.columns([2, 2])
    with _oc_left:
        _opt_period = st.selectbox("Historical data period:", ["1y", "2y", "3y", "5y"],
                                   index=1, key="opt_period",
                                   help="Longer periods give more robust estimates but include older market regimes")
    with _oc_right:
        _OBJ_OPTIONS = [
            "Maximise Sharpe Ratio",
            "Minimise Volatility",
            "Maximise Return",
            "Risk Parity",
            "Maximum Diversification",
            "Target Return (Min Risk)",
        ]
        _opt_objective = st.selectbox(
            "Optimisation objective:",
            _OBJ_OPTIONS,
            key="opt_objective",
            help=(
                "Sharpe: best return per unit of risk  |  "
                "Min Vol: lowest possible volatility  |  "
                "Max Return: highest expected return  |  "
                "Risk Parity: equal risk from each asset (All Weather style)  |  "
                "Max Diversification: maximise spread across uncorrelated assets  |  "
                "Target Return: you set the annual return, optimizer minimises risk to achieve it"
            )
        )

    _opt_target_ret = None
    if _opt_objective == "Target Return (Min Risk)":
        _tr_col1, _tr_col2 = st.columns([2, 1])
        with _tr_col1:
            _opt_target_ret = st.slider(
                "Target annual return (%):", min_value=2.0, max_value=50.0, value=12.0, step=0.5,
                key="opt_target_ret",
                help="The optimizer will find the lowest-risk portfolio that achieves this return"
            ) / 100

    _opt_rf = st.slider("Risk-free rate (%):", min_value=0.0, max_value=8.0, value=4.25, step=0.25,
                        key="opt_rf", help="Current UK base rate ~4.25%") / 100

    _run_opt = st.button("⚡ Run Optimisation", type="primary", use_container_width=False, key="run_opt_btn")

    if not _run_opt:
        st.info("Enter your tickers and click **Run Optimisation** to generate the efficient frontier and optimal weights.")
        st.stop()

    # ── Parse & validate tickers ──────────────────────────────────
    if not _opt_tickers_raw.strip():
        st.error("Please enter at least 2 ticker symbols.")
        st.stop()

    _opt_tickers = [t.strip().upper() for t in _opt_tickers_raw.split(",") if t.strip()]
    if len(_opt_tickers) < 2:
        st.error("Portfolio optimizer requires at least 2 assets.")
        st.stop()
    if len(_opt_tickers) > 20:
        st.error("Maximum 20 tickers supported. Please reduce your list.")
        st.stop()

    with st.spinner(f"Fetching {len(_opt_tickers)} assets over {_opt_period}…"):
        _prices = _opt_fetch_prices(tuple(_opt_tickers), _opt_period)

    if _prices.empty or len(_prices.columns) < 2:
        st.error("Could not download price data. Check your ticker symbols and try again.")
        st.stop()

    # Drop assets with >10% missing data
    _threshold = int(len(_prices) * 0.9)
    _prices = _prices.dropna(thresh=_threshold, axis=1)
    _valid_tickers = list(_prices.columns)
    _removed = [t for t in _opt_tickers if t not in _valid_tickers]
    if _removed:
        st.warning(f"Removed due to insufficient data: {', '.join(_removed)}")
    if len(_valid_tickers) < 2:
        st.error("Need at least 2 assets with sufficient data.")
        st.stop()

    _prices = _prices.ffill().dropna()
    _returns = _prices.pct_change().dropna()
    _mean_ret   = _returns.mean()
    _cov_matrix = _returns.cov()
    _n = len(_valid_tickers)

    _target_ret_feasible = True
    with st.spinner("Optimising portfolio…"):
        if _opt_objective == "Maximise Sharpe Ratio":
            _opt_weights = _max_sharpe(_mean_ret, _cov_matrix, rf=_opt_rf, n=_n)
        elif _opt_objective == "Minimise Volatility":
            _opt_weights = _min_volatility(_mean_ret, _cov_matrix, n=_n)
        elif _opt_objective == "Maximise Return":
            _opt_weights = _max_return(_mean_ret, _cov_matrix, n=_n)
        elif _opt_objective == "Risk Parity":
            _opt_weights = _risk_parity(_mean_ret, _cov_matrix, n=_n)
        elif _opt_objective == "Maximum Diversification":
            _opt_weights = _max_diversification(_mean_ret, _cov_matrix, n=_n)
        elif _opt_objective == "Target Return (Min Risk)":
            _tr = _opt_target_ret if _opt_target_ret is not None else 0.12
            _tr_weights = _target_return_min_vol(_mean_ret, _cov_matrix, _tr, n=_n)
            if _tr_weights is None:
                st.warning(f"⚠️ Target return of {_tr:.0%} is not achievable with this portfolio. "
                           f"Falling back to Maximise Sharpe. Try a lower target.")
                _target_ret_feasible = False
                _opt_weights = _max_sharpe(_mean_ret, _cov_matrix, rf=_opt_rf, n=_n)
            else:
                _opt_weights = _tr_weights
        else:
            _opt_weights = _max_sharpe(_mean_ret, _cov_matrix, rf=_opt_rf, n=_n)

        _opt_ret, _opt_vol, _opt_sharpe = _portfolio_stats(_opt_weights, _mean_ret, _cov_matrix, _opt_rf)
        _ef_vols, _ef_rets = _efficient_frontier_points(_mean_ret, _cov_matrix)

        # Random portfolios for frontier scatter
        _np.random.seed(42)
        _n_sim = 2000
        _sim_vols, _sim_rets, _sim_shrp = [], [], []
        for _ in range(_n_sim):
            _w = _np.random.dirichlet(_np.ones(_n))
            _r, _v, _s = _portfolio_stats(_w, _mean_ret, _cov_matrix, _opt_rf)
            _sim_rets.append(_r); _sim_vols.append(_v); _sim_shrp.append(_s)

        # VaR
        _var_95  = _calc_var(_returns, _opt_weights, 0.95)
        _var_99  = _calc_var(_returns, _opt_weights, 0.99)
        _port_daily = _returns.dot(_opt_weights)
        _sortino_denom = float(_port_daily[_port_daily < 0].std() * _np.sqrt(252))
        _sortino = (_opt_ret - _opt_rf) / _sortino_denom if _sortino_denom > 0 else 0.0
        _cum = (1 + _port_daily).cumprod()
        _rolling_max = _cum.cummax()
        _max_dd = float(((_cum - _rolling_max) / _rolling_max).min())

    # ── Results ──────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 Optimisation Results")

    # KPI row
    _k1, _k2, _k3, _k4, _k5 = st.columns(5)
    def _opt_kpi(col, label, value, color="#F1F5F9", suffix=""):
        col.markdown(
            f'<div style="background:#0D1F33;border:1px solid rgba(100,116,139,0.25);'
            f'border-radius:10px;padding:12px;text-align:center">'
            f'<div style="font-size:0.65rem;color:#64748B;text-transform:uppercase;'
            f'letter-spacing:0.5px;margin-bottom:4px">{label}</div>'
            f'<div style="font-size:1.25rem;font-weight:800;color:{color}">{value}{suffix}</div>'
            f'</div>', unsafe_allow_html=True)

    _ret_col = "#22C55E" if _opt_ret >= 0 else "#EF4444"
    _opt_kpi(_k1, "Expected Return", f"{_opt_ret:.1%}", _ret_col)
    _opt_kpi(_k2, "Portfolio Volatility", f"{_opt_vol:.1%}", "#F59E0B")
    _opt_kpi(_k3, "Sharpe Ratio", f"{_opt_sharpe:.2f}",
             "#22C55E" if _opt_sharpe > 1 else "#F59E0B" if _opt_sharpe > 0.5 else "#EF4444")
    _opt_kpi(_k4, "Sortino Ratio", f"{_sortino:.2f}",
             "#22C55E" if _sortino > 1 else "#F59E0B" if _sortino > 0.5 else "#EF4444")
    _opt_kpi(_k5, "Max Drawdown", f"{_max_dd:.1%}", "#EF4444")

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # VaR row
    _v1, _v2, _v3 = st.columns(3)
    _port_value = 10000  # illustrative £10k portfolio
    _opt_kpi(_v1, "Daily VaR 95%", f"{_var_95:.2%}", "#EF4444")
    _opt_kpi(_v2, "Daily VaR 99%", f"{_var_99:.2%}", "#991B1B")
    _opt_kpi(_v3, f"VaR 95% on £10,000",
             f"£{_port_value * _var_95:,.0f}", "#EF4444")

    st.caption("VaR = Value at Risk. At 95% confidence, daily losses should not exceed the VaR figure on 19 out of 20 trading days.")

    st.markdown("---")

    # ── Efficient Frontier chart ──────────────────────────────────
    st.markdown("#### 📈 Efficient Frontier")
    _fig_ef = go.Figure()

    # Simulated portfolios scatter
    _fig_ef.add_trace(go.Scatter(
        x=_sim_vols, y=_sim_rets, mode="markers",
        marker=dict(color=_sim_shrp, colorscale="Viridis", size=4, opacity=0.5,
                    colorbar=dict(title="Sharpe", thickness=12, len=0.7,
                                  tickfont=dict(color="#64748B"))),
        name="Random portfolios",
        hovertemplate="Vol: %{x:.1%}<br>Return: %{y:.1%}<extra></extra>"
    ))

    # Efficient frontier line
    if _ef_vols:
        _fig_ef.add_trace(go.Scatter(
            x=_ef_vols, y=_ef_rets, mode="lines",
            line=dict(color="#F59E0B", width=2.5, dash="solid"),
            name="Efficient Frontier",
            hovertemplate="Vol: %{x:.1%}<br>Return: %{y:.1%}<extra></extra>"
        ))

    # Optimal portfolio star — use annotation (not text trace) to avoid legend overlap
    _eq_w = _np.array([1/_n]*_n)
    _eq_ret, _eq_vol, _eq_sharpe = _portfolio_stats(_eq_w, _mean_ret, _cov_matrix, _opt_rf)

    _fig_ef.add_trace(go.Scatter(
        x=[_opt_vol], y=[_opt_ret], mode="markers",
        marker=dict(color="#F59E0B", size=16, symbol="star",
                    line=dict(color="#FFFFFF", width=1.5)),
        name="Optimal Portfolio",
        hovertemplate=f"<b>Optimal</b><br>Vol: {_opt_vol:.1%}<br>Return: {_opt_ret:.1%}<br>Sharpe: {_opt_sharpe:.2f}<extra></extra>"
    ))

    # Equal-weight reference
    _fig_ef.add_trace(go.Scatter(
        x=[_eq_vol], y=[_eq_ret], mode="markers",
        marker=dict(color="#64748B", size=12, symbol="diamond"),
        name="Equal Weight",
        hovertemplate=f"<b>Equal Weight</b><br>Vol: {_eq_vol:.1%}<br>Return: {_eq_ret:.1%}<br>Sharpe: {_eq_sharpe:.2f}<extra></extra>"
    ))

    # Use annotations instead of inline text — no overlap
    _ef_annotations = [
        dict(x=_opt_vol, y=_opt_ret, text="⭐ Optimal",
             xanchor="left", yanchor="bottom", xshift=8, yshift=4,
             font=dict(color="#F59E0B", size=11), showarrow=False,
             bgcolor="rgba(13,31,51,0.7)", borderpad=3),
        dict(x=_eq_vol, y=_eq_ret, text="Equal Wt",
             xanchor="left", yanchor="top", xshift=8, yshift=-4,
             font=dict(color="#94A3B8", size=10), showarrow=False,
             bgcolor="rgba(13,31,51,0.7)", borderpad=3),
    ]

    _fig_ef.update_layout(
        height=440,
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=40, r=30, t=20, b=50),
        legend=dict(
            font=dict(color="#94A3B8", size=10),
            bgcolor="rgba(13,31,51,0.85)",
            bordercolor="rgba(100,116,139,0.3)",
            borderwidth=1,
            orientation="h",
            x=0, y=1.08, xanchor="left", yanchor="bottom"
        ),
        annotations=_ef_annotations,
        xaxis=dict(title=dict(text="Annual Volatility (Risk)", font=dict(color="#64748B")),
                   tickformat=".0%", gridcolor="rgba(100,116,139,0.15)",
                   tickfont=dict(color="#64748B")),
        yaxis=dict(title=dict(text="Expected Annual Return", font=dict(color="#64748B")),
                   tickformat=".0%", gridcolor="rgba(100,116,139,0.15)",
                   tickfont=dict(color="#64748B")),
    )
    st.plotly_chart(_fig_ef, use_container_width=True, config={"displayModeBar": False})
    st.caption("⭐ Star = optimal portfolio. ◆ Diamond = equal-weight baseline. Colour = Sharpe Ratio. Hover any point for details.")

    st.markdown("---")

    # Optimal Weights
    st.markdown("#### Optimal Allocation")
    _wt_left, _wt_right = st.columns([1, 1])
    with _wt_left:
        _weights_df = pd.DataFrame({
            "Ticker": _valid_tickers,
            "Weight": _opt_weights,
            "Allocation %": [f"{w:.1%}" for w in _opt_weights],
        }).sort_values("Weight", ascending=False).reset_index(drop=True)
        _weights_df.index += 1
        st.dataframe(_weights_df[["Ticker", "Allocation %"]],
                     use_container_width=True, hide_index=False)
    with _wt_right:
        _pie_fig = go.Figure(go.Pie(
            labels=_valid_tickers, values=_opt_weights, hole=0.45,
            marker=dict(colors=["#F59E0B","#22C55E","#3B82F6","#EC4899","#8B5CF6",
                                 "#06B6D4","#EF4444","#84CC16","#F97316","#A78BFA",
                                 "#10B981","#FBBF24","#6366F1","#14B8A6","#FB923C",
                                 "#E11D48","#7C3AED","#0EA5E9","#4ADE80","#FCD34D"][:_n]),
            textfont=dict(size=11, color="#F1F5F9"),
            hovertemplate="%{label}: %{percent}<extra></extra>"
        ))
        _pie_fig.update_layout(
            height=280, margin=dict(l=0,r=0,t=10,b=0),
            paper_bgcolor="rgba(0,0,0,0)",
            legend=dict(font=dict(color="#94A3B8", size=10), bgcolor="rgba(0,0,0,0)"),
            annotations=[dict(text=_opt_objective.split()[0], x=0.5, y=0.5,
                              font=dict(size=11, color="#94A3B8"), showarrow=False)]
        )
        st.plotly_chart(_pie_fig, use_container_width=True, config={"displayModeBar": False})

    st.markdown("---")

    # Correlation Matrix
    st.markdown("#### Correlation Matrix")
    _corr = _returns.corr()
    _corr_fig = go.Figure(go.Heatmap(
        z=_corr.values, x=_valid_tickers, y=_valid_tickers,
        colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
        text=[[f"{v:.2f}" for v in row] for row in _corr.values],
        texttemplate="%{text}", textfont=dict(size=10),
        hovertemplate="%{y} / %{x}: %{z:.2f}<extra></extra>",
        colorbar=dict(tickfont=dict(color="#64748B"))
    ))
    _corr_fig.update_layout(
        height=max(250, _n * 35),
        margin=dict(l=0, r=0, t=10, b=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(tickfont=dict(color="#94A3B8")),
        yaxis=dict(tickfont=dict(color="#94A3B8")),
    )
    st.plotly_chart(_corr_fig, use_container_width=True, config={"displayModeBar": False})
    st.caption("Low or negative correlations (blue) reduce portfolio volatility. "
               "High correlations (red) mean assets move together - less diversification benefit.")

    # ── Weight Drift Monitor ──────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📡 Live Weight Drift Monitor")
    st.caption("Compares your optimal target weights to where the portfolio sits today based on live prices.")

    with st.spinner("Fetching live prices for drift analysis…"):
        try:
            import yfinance as _yf_drift
            _live_data = _yf_drift.download(_valid_tickers, period="5d", auto_adjust=True, progress=False)["Close"]
            if isinstance(_live_data, pd.Series):
                _live_data = _live_data.to_frame()
            _live_prices = _live_data.ffill().iloc[-1]

            # Reconstruct current market-value weights from live prices
            # Assume we started with equal capital allocation, then let prices drift
            _start_prices = _live_data.ffill().iloc[0]
            _shares = _opt_weights / _start_prices.values  # notional shares per unit capital
            _current_values = _shares * _live_prices.values
            _current_weights = _current_values / _current_values.sum()

            _drift_df = pd.DataFrame({
                "Ticker": _valid_tickers,
                "Target %": [f"{w:.1%}" for w in _opt_weights],
                "Current %": [f"{w:.1%}" for w in _current_weights],
                "Drift": _current_weights - _opt_weights,
            })
            _drift_df["Action"] = _drift_df["Drift"].apply(
                lambda d: "🔴 Trim" if d > 0.03 else ("🟢 Top Up" if d < -0.03 else "✅ On Track")
            )
            _drift_df["Drift %"] = _drift_df["Drift"].apply(lambda d: f"{d:+.1%}")
            _drift_df = _drift_df.drop(columns=["Drift"]).reset_index(drop=True)
            _drift_df.index += 1

            _needs_rebal = (_drift_df["Action"] != "✅ On Track").any()
            if _needs_rebal:
                st.warning("⚠️ Portfolio has drifted from target. Review positions flagged below.")
            else:
                st.success("✅ Portfolio is within tolerance of target weights (±3%).")

            st.dataframe(_drift_df, use_container_width=True)
            st.caption("Drift > +3%: position has grown too large → consider trimming. Drift < -3%: position has shrunk → consider topping up.")
        except Exception as _drift_err:
            st.info(f"Live drift data unavailable: {_drift_err}")

    # ── AI Commentary & Glossary ──────────────────────────────────
    st.markdown("---")
    with st.expander("🤖 AI Commentary — What This Portfolio Is Telling You", expanded=False):
        # Generate plain-English interpretation of the numbers
        _sharpe_comment = (
            "excellent — this portfolio generates strong return per unit of risk, comparable to top institutional funds"
            if _opt_sharpe > 1.5 else
            "good — above 1.0 is considered solid by professional standards"
            if _opt_sharpe > 1.0 else
            "moderate — acceptable but there may be room to improve through diversification"
            if _opt_sharpe > 0.5 else
            "low — the portfolio is taking significant risk for the return it generates"
        )
        _vol_comment = (
            "very low volatility — suitable for conservative investors"
            if _opt_vol < 0.10 else
            "moderate volatility — typical of a diversified equity portfolio"
            if _opt_vol < 0.18 else
            "above-average volatility — expect meaningful swings in portfolio value"
            if _opt_vol < 0.25 else
            "high volatility — this portfolio can move sharply; ensure this matches your risk appetite"
        )
        _dd_comment = (
            "minimal drawdown — strong capital preservation"
            if _max_dd > -0.10 else
            "moderate drawdown — manageable for most investors"
            if _max_dd > -0.20 else
            "significant drawdown — the portfolio has experienced meaningful peak-to-trough losses historically"
        )
        _corr_avg = float(_corr.values[_np.triu_indices_from(_corr.values, k=1)].mean())
        _div_comment = (
            "excellent diversification — assets move largely independently"
            if _corr_avg < 0.3 else
            "reasonable diversification — some correlation but meaningful spread"
            if _corr_avg < 0.5 else
            "moderate correlation — assets tend to move together; consider adding uncorrelated assets like bonds or commodities"
        )

        st.markdown(f"""
**Portfolio Summary — {_opt_objective}**

Your optimised portfolio of **{_n} assets** targets an expected annual return of **{_opt_ret:.1%}**
with a volatility of **{_opt_vol:.1%}**.

**Risk-adjusted performance:** The Sharpe Ratio is **{_opt_sharpe:.2f}** — {_sharpe_comment}.
The Sortino Ratio is **{_sortino:.2f}**, which measures return per unit of *downside* risk only —
a Sortino above 1.0 is considered strong.

**Volatility:** {_vol_comment.capitalize()}. In practical terms, a portfolio with {_opt_vol:.0%} annual
volatility could swing by roughly **±{_opt_vol/16:.1%}** on a typical trading day.

**Drawdown:** The worst historical peak-to-trough decline was **{_max_dd:.1%}** — {_dd_comment}.

**Diversification:** Average pairwise correlation is **{_corr_avg:.2f}** — {_div_comment}.

**Risk (VaR):** On 95% of trading days, daily losses on a £10,000 portfolio should not exceed
**£{10000*_var_95:,.0f}**. On the worst 1% of days (VaR 99%), losses could reach **£{10000*_var_99:,.0f}**.

**Compared to equal weight:** The optimised allocation delivers
{'**better**' if _opt_sharpe > _eq_sharpe else '**similar**'} risk-adjusted returns
(Sharpe {_opt_sharpe:.2f} vs {_eq_sharpe:.2f} for equal weight).
""")

    with st.expander("📚 Glossary — What Every Term Means & How to Read This Screen", expanded=False):
        st.markdown("""
**Efficient Frontier**
The curved line on the chart represents every *mathematically optimal* portfolio —
meaning no other combination of your assets gives higher return for the same level of risk.
Portfolios below the curve are inefficient (you're taking more risk than necessary).
The gold star marks your optimised portfolio on this frontier.

---

**Sharpe Ratio**
Measures how much return you earn per unit of total risk.
Formula: `(Portfolio Return − Risk-Free Rate) ÷ Volatility`
- Above 1.0 = good  |  Above 1.5 = excellent  |  Below 0.5 = poor
Think of it as: "am I being compensated fairly for the risk I'm taking?"

---

**Sortino Ratio**
Like the Sharpe Ratio but only counts *downside* volatility (bad days).
Upward volatility (good days) is not penalised. Preferred by many professional investors.
- Above 1.0 = solid  |  Above 2.0 = strong

---

**Volatility (Annual)**
The standard deviation of portfolio returns, annualised.
A 15% volatility means returns fluctuate roughly ±15% around the average in a typical year.
Lower is calmer; higher means bigger swings — up *and* down.

---

**Max Drawdown**
The largest peak-to-trough fall in portfolio value over the historical period.
A drawdown of -20% means at some point the portfolio fell 20% from its recent high before recovering.
This is the number that tests whether you'd panic-sell.

---

**VaR (Value at Risk)**
A statistical estimate of potential daily loss.
*Daily VaR 95%* = on 95% of days, losses should not exceed this amount.
*Daily VaR 99%* = the worst 1% of days — the "tail risk."
Important: VaR does *not* tell you how bad losses can get on those worst days, only that they exceed the threshold.

---

**Correlation Matrix**
Shows how closely pairs of assets move together (scale: −1 to +1).
- **+1.0** = perfectly correlated — they move identically (no diversification benefit)
- **0.0** = uncorrelated — they move independently (good diversification)
- **−1.0** = perfectly inverse — one rises when the other falls (best diversification)
Blue cells = low/negative correlation = good for reducing portfolio risk.

---

**Risk Parity**
An objective where each asset contributes *equally* to total portfolio risk —
not equal capital allocation. Made famous by Ray Dalio's All Weather fund.
Assets with lower volatility get higher weights to equalise their risk contribution.

---

**Maximum Diversification**
Maximises the ratio of the weighted-average asset volatility to total portfolio volatility.
A higher ratio means you're getting more diversification benefit from your asset mix.

---

**Weight Drift**
Over time, assets that perform well grow as a share of your portfolio,
and underperformers shrink. This "drift" moves you away from your target allocation.
The drift monitor shows where each position stands today vs. your optimal targets.
A ±3% tolerance is standard; beyond that, rebalancing is worth considering.

---

**How to use this screen**
1. Enter your tickers (or import from Journal)
2. Choose an objective matching your goal — Sharpe for balanced growth, Min Vol for capital preservation, Risk Parity for institutional-style balance
3. Review the Efficient Frontier — your portfolio should be *on* the curve, not below it
4. Check the Correlation Matrix — if everything is dark blue, you need more diversification
5. Monitor Weight Drift regularly — rebalance when any position drifts more than ±5%
""")

    st.markdown(
        '<div style="font-size:0.75rem;color:#475569;padding:8px 0">'
        'Disclaimer: Portfolio optimization uses historical price data. Past performance'
        ' does not guarantee future results. For educational purposes only,'
        ' not financial advice. Always do your own research.'
        '</div>', unsafe_allow_html=True)

# ── Global footer — bottom of page, visible on scroll ──
st.markdown("""
<div class="fintiq-global-footer">
  <div><span class="fg-brand">Fintiq</span> &nbsp;·&nbsp; © 2025 Fintiq Ltd &nbsp;·&nbsp; Registered in England &amp; Wales</div>
  <div style="display:flex;gap:18px;align-items:center;flex-wrap:wrap">
    <a href="mailto:contactfintiq@gmail.com">✉ contactfintiq@gmail.com</a>
    <span style="color:#2D3F55">|</span>
    <a href="mailto:contactfintiq@gmail.com?subject=Feature%20Suggestion">💡 Suggest a feature</a>
    <span style="color:#2D3F55">|</span>
    <a href="https://fintiq.uk" target="_blank">fintiq.uk</a>
    <span style="color:#2D3F55">|</span>
    <span style="color:#334155">Not financial advice. For educational use only.</span>
  </div>
</div>
""", unsafe_allow_html=True)
